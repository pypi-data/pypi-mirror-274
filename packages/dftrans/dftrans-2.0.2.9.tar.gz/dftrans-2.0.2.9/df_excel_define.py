# -*- coding: utf-8 -*-
import copy
import pandas as pd
import numpy as np
import sqlite3 as sl
import datacompy
import os
import sys
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.sql import default_comparator           # 必须加上否则nuikta编译后，找不到模块。并且需要好安装1.4.40版本
import shutil
from pathlib import Path
import time
from datetime import datetime
from hash_password import dftrans_simple_decrypt, dftrans_password_decrypt_str, dftrans_password_encrypt_str
from psgspecialelements import *
import inspect
import PySimpleGUI as Sg
import ast
from pypinyin import lazy_pinyin
import jsonlines
from df_excel_plot import DfPlot, PlotType, draw_figure, canvas_clear
from df_stock_define import *
DEF_FT_L = ('宋体', 10)
DEF_PD_L = (5, 5)
PARAM_FT = ('黑体', 10)
DEF_PLOT_WIDTH = 12
DEF_PLOT_HEIGHT = 8

# LOCAL_SYSTEM_ICON = "config/img/dftrans06.ico"
LOCAL_SYSTEM_ICON = "config/img/table_arrow9_transparent_64.ico"
LOCAL_SYSTEM_IMAGE = "config/img/table_arrow9_transparent_256.png"

HIGHLIST_SELECTED_COLOR = 'dark blue'
ARG_PARAM_VALUE_COLOR = 'blue'

TEXT_COLOR = 'black'
TEXT_BACKGROUND_COLOR = 'white'
INPUT_TEXT_COLOR = 'blue'
INPUT_TEXT_BACKGROUND_COLOR = 'white'
COMMAND_EXECUTED_COLOR = 'white'
COMMAND_EXECUTED_BACKGROUND_COLOR = 'green'

RAM_FILE_SUFFIX = 'ram'
RAM_FILE_PREFIX = 'mem:'

DATABASE_TYPE_SQLITE = 'sqlite'
DATABASE_TYPE_MYSQL = 'mysql'


def dfreaddatabysql(ip: str = '', port: int = 8000, database: str = '',
                    user: str = '', password: str = '', sqlstr: str = '', databasetype: str = 'mysql'):
    """
    用sql语句读取数据
    :param ip:
    :param port:
    :param database:
    :param user:
    :param password:
    :param sqlstr:
    :param databasetype:   数据库类型 sqlite, mysql
    :return:
    """

    if databasetype == DATABASE_TYPE_SQLITE:
        mycon = sl.connect(database)
    else:
        # mycon = 'mysql+mysqlconnector://用户名:密码@主机名或IP地址:端口/库名称?charset=utf8'
        mycon = "mysql+mysqlconnector://{0}:{1}@{2}:{3}/{4}?charset=utf8".format(
            user, dftrans_simple_decrypt(password), ip, port, database)
    try:
        # sql读取数据
        specialreaddf = SpecialDf(dfdata=pd.read_sql_query(sqlstr, mycon))
    except Exception as err:
        return errprintwithname("sql:{}读取数据失败:{}！".format(sqlstr, err))

    return [True, specialreaddf]


def dfloaddatafromdbbysql(specialdflist: list, tableindex: list, positionlist: list,
                          ip: str, port: int, database: str, user: str, password: str, sqlstr: str, databasetype: str):
    """
    读取db数据到df
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param ip:
    :param port:
    :param database:
    :param user:
    :param password:
    :param sqlstr:
    :param databasetype:                  str: 数据库类型 sqlite, mysql
    :return:
    """
    result = dfreaddatabysql(ip=ip, port=port, database=database,
                             user=user, password=password, sqlstr=sqlstr, databasetype=databasetype)
    if result[0] is False:
        return result
    # 判断是新增还是覆盖原有df
    targetdfid = positionlist[0]
    try:
        if targetdfid >= len(specialdflist):
            # 新增df
            specialdflist.append(result[1])
            # 记录数据来源
            specialdflist[-1].dataimportsource = "{}@{}:{}/{}".format(user, ip, port, database)
        else:
            # 覆盖旧表
            specialdflist[targetdfid] = result[1]
            # 记录数据来源
            specialdflist[targetdfid].dataimportsource = "{}@{}:{}/{}".format(user, ip, port, database)
    except Exception as err:
        return errprintwithname("sql：{}导入数据到表格{}错误:{}！".format(sqlstr, positionlist, err))

    return [True, ""]


def dfsavedatatodatabasebysql(specialdflist: list, tableindex: list, ip='',
                              port: int = 8000, database='', user='', password='', name="", schema=None,
                              if_exists='replace', index=False, index_label=None,
                              chunksize=None, dtype=None, method=None,  databasetype: str = 'mysql'):
    """
    保存表格数据到database table
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param ip:
    :param port:
    :param database:
    :param user:
    :param password:
    :param name:
    :param schema:
    :param if_exists:
    :param index:
    :param index_label:
    :param chunksize:
    :param dtype:
    :param method:
    :param databasetype:                  str: 数据库类型 sqlite, mysql
    :return:
    """
    # 当表格为空或者-1，使用最后一张表格
    if tableindex in [[], [-1], None]:
        tableindex = [len(specialdflist) - 1]
    specialdf = specialdflist[tableindex[0]]
    if databasetype == DATABASE_TYPE_SQLITE:
        mycon = sl.connect(database)
    else:
        # mycon = 'mysql+mysqlconnector://用户名:密码@主机名或IP地址:端口/库名称?charset=utf8'
        mycon = "mysql+mysqlconnector://{0}:{1}@{2}:{3}/{4}".format(
            user, dftrans_simple_decrypt(password), ip, port, database)
    print("save engine：{}".format(mycon))
    try:
        if databasetype == DATABASE_TYPE_SQLITE:
            if if_exists.lower() == 'replace':
                # 'replace'方式下，表格如果存在会报错，实际没有效果。因此先删除表格
                deleltestr = "DROP TABLE IF EXISTS {};".format(name)
                mycon.execute(deleltestr)
            specialdf.dfdata.to_sql(name=name, con=mycon)
        else:
            engine = create_engine(mycon, echo=False,
                                   connect_args={'auth_plugin': 'mysql_native_password', 'charset': 'utf8'})
            if if_exists.lower() == 'replace':
                # 'replace'方式下，表格如果存在会报错，实际没有效果。因此先删除表格
                deleltestr = "DROP TABLE IF EXISTS {};".format(name)
                engine.execute(deleltestr)
            specialdf.dfdata.to_sql(name=name, con=engine, schema=schema,
                                    if_exists=if_exists, index=index, index_label=index_label,
                                    chunksize=chunksize, dtype=dtype, method=method)
        # 记录数据来源
        specialdf.dataexporttarget = "{}@{}:{}/{}/{}".format(user, ip, port, database, name)
    except Exception as err:
        return errprintwithname("写入数据失败:{}！".format(err))

    return [True, ""]


def dfreaddatafile(filepath: str, headline: int = 0, tailline: int = 0, sheetnumber: int = 0,
                   headnameline: int = 0, na_filter: bool = False, colforcestring=None) -> list:
    """
    从文件导入数据到df
    :param filepath:                                str： 文件全路径名
    :param headline:                                int: 导出表首丢弃行数
    :param tailline:                                int： 导出表尾丢弃行数
    :param sheetnumber:                                int： 导出的子表号
    :param headnameline:                                int: 丢弃表首后，列名所在行，>=0有效
    :param na_filter:                               bool: 是否检测各种空值并且替换为NaN
    :param colforcestring:                          list 强制作为字符串读取的列名的列表
    :return:                    [bool True:成功，False:失败, errstr]
    """
    # 读取指定的文件
    if colforcestring is None:
        colforcestring = []
    fileresult = os.path.splitext(filepath)
    # print(fileresult[-1])
    # 如果是xlsx文件
    # 2、pandas读取csv文件
    # import pandas as pd
    # data = pd.read_csv('path', sep=',', header=0, names=["第一列"，"第二列"，"第三列"]，encoding = 'utf-8')
    # path: 要读取的文件的绝对路径
    # sep: 指定列和列的间隔符，默认sep =‘, ’
    # 若sep =‘’\t",即列与列之间用制表符\t分割，相当于tab——四个空格
    # header: 列名行，默认为0
    # names: 列名命名或重命名
    # encoding: 指定用于unicode文本编码格式
    # 3、pandas读取txt文件
    # read_csv
    # data = pd.read_table('path', sep='\t', header=None, names=['第一列', '第二列', '第三列'])

    myhead = headnameline if headnameline >= 0 else None
    if colforcestring not in [None, []] and isinstance(colforcestring, list):
        dtype = dict(zip(colforcestring, [str for _ in range(len(colforcestring))]))
    else:
        dtype = None
    if is_mem_virtual_file_exist(filepath):   # 首先判断是否内存虚拟对象文件
        readdf = load_object_from_virtual_file(filepath)
        if readdf is None:
            return errprintwithname("读取{}失败: 文件不存在！".format(filepath))
        # head需要转换为str以便后续处理
        readdf.columns = readdf.columns.map(str)
        specialreaddf = SpecialDf(dfdata=pd.DataFrame(readdf))
        specialreaddf.dataimportsource = filepath

    elif fileresult[-1] == ".xlsx":
        try:
            # 读取时跳过表头行和表尾行
            readdf = pd.read_excel(filepath, header=myhead, dtype=dtype,
                                   na_filter=na_filter, skiprows=headline,
                                   skipfooter=tailline, sheet_name=sheetnumber)
            # head需要转换为str以便后续处理
            readdf.columns = readdf.columns.map(str)
            specialreaddf = SpecialDf(dfdata=readdf)
            specialreaddf.dataimportsource = filepath
        except Exception as err:
            return errprintwithname("读取{}失败:{}！".format(filepath, err))
    
    # TODO 增加read_csv时分隔符设置。
    elif fileresult[-1] == ".csv":
        try:
            # 读取时跳过表头行和表尾行
            readdf = pd.read_csv(filepath, header=myhead, dtype=dtype,
                                 na_filter=na_filter, skiprows=headline,
                                 skipfooter=tailline, encoding='utf-8')
            # head需要转换为str以便后续处理
            readdf.columns = readdf.columns.map(str)
            specialreaddf = SpecialDf(dfdata=readdf)
            specialreaddf.dataimportsource = filepath
        except Exception as err:
            return errprintwithname("读取{}失败:{}！".format(filepath, err))

    elif fileresult[-1] == ".txt":
        try:
            # # 读取时跳过表头行和表尾行
            readdf = pd.read_csv(filepath, header=myhead,  dtype=dtype,
                                 na_filter=na_filter, skiprows=headline, sep=',',
                                 skipfooter=tailline, encoding='utf-8')
            # head需要转换为str以便后续处理
            readdf.columns = readdf.columns.map(str)
            specialreaddf = SpecialDf(dfdata=readdf)
            specialreaddf.dataimportsource = filepath
        except Exception as err:
            return errprintwithname("读取{}失败:{}！".format(filepath, err))

    elif fileresult[-1] == ".jsonl":
        try:
            datalist = []
            # 逐行读取jsonl
            with jsonlines.open(filepath) as fp:
                for items in fp:
                    datalist.append(items)
            readdf = pd.DataFrame(datalist)
            # head需要转换为str以便后续处理
            readdf.columns = readdf.columns.map(str)
            specialreaddf = SpecialDf(dfdata=readdf)
            specialreaddf.dataimportsource = filepath
        except Exception as err:
            return errprintwithname("读取{}失败:{}！".format(filepath, err))
    else:
        return errprintwithname("无法识别后缀的文件:{}！".format(filepath))

    return [True, specialreaddf]


def dfloaddatafromfile(specialdflist: list, tableindex: list, positionlist: list,
                       filepath: str, headline: int = 0, tailline: int = 0, sheetnumber: int = 0,
                       headnameline: int = 0, na_filter: bool = False, colforcestring=None) -> list:
    """
    从文件导入数据到df
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param filepath:                                str： 文件全路径名
    :param headline:                                int: 导出表首丢弃行数
    :param tailline:                                int： 导出表尾丢弃行数
    :param sheetnumber:                                int： 导出的子表号
    :param headnameline:                                int: 丢弃表首后，列名所在行，>=0有效
    :param na_filter:                               bool: 是否检测各种空值并且替换为NaN
    :param colforcestring:                          list 强制作为字符串读取的列名的列表
    :return:                    [bool True:成功，False:失败, errstr]
    """

    if colforcestring is None:
        colforcestring = []
    result = dfreaddatafile(filepath=filepath, headline=headline, tailline=tailline, colforcestring=colforcestring,
                            sheetnumber=sheetnumber, headnameline=headnameline, na_filter=na_filter)
    if result[0] is False:
        return result
    # 判断是新增还是覆盖原有df
    targetdfid = positionlist[0]
    try:
        if targetdfid >= len(specialdflist):
            # 新增df
            specialdflist.append(result[1])
        else:
            # 覆盖旧表
            specialdflist[targetdfid] = result[1]
    except Exception as err:
        return errprintwithname("导入文件{}到表格{}错误:{}！".format(filepath, positionlist, err))

    return [True, ""]


def dfloaddatafromfiletonewtablespecialdflist(specialdflist: list, tableindex: list,
                                              filepath: str, headline: int = 0,
                                              tailline: int = 0, sheetnumber: int = 0, headnameline: int = 0,
                                              na_filter: bool = False, colforcestring=None) -> list:
    """
    从文件导入数据到df
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param filepath:                                str： 文件全路径名
    :param headline:                                int: 导出表首丢弃行数
    :param tailline:                                int： 导出表尾丢弃行数
    :param sheetnumber:                                int： 导出的子表号
    :param headnameline:                                int: 丢弃表首后，列名所在行，>=0有效
    :param na_filter:                               bool: 是否检测各种空值并且替换为NaN
    :param colforcestring:                          list 强制作为字符串读取的列名的列表
    :return:                    [bool True:成功，False:失败, errstr]
    """

    if colforcestring is None:
        colforcestring = []
    result = dfreaddatafile(filepath=filepath, headline=headline, tailline=tailline, colforcestring=colforcestring,
                            sheetnumber=sheetnumber, headnameline=headnameline, na_filter=na_filter)
    if result[0] is False:
        return result
    # 新表号
    targetdfid = len(specialdflist)
    try:
        specialdflist.append(result[1])
    except Exception as err:
        return errprintwithname("导入文件{}到新表{}错误:{}！".format(filepath, targetdfid, err))

    return [True, ""]


def dfsavedatatofile(specialdflist: list, tableindex: list, outputfilename, sheetnumber: int = 0,
                     headline: int = 0, tailline: int = 0, modelfilename='', index: bool = True,
                     styleout: bool = False, sumrownumber: int = None, processingbarevent: tuple = None) -> list:
    """
    将df内容保存到文件
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param outputfilename:                  str：输出文件全路径名
    :param sheetnumber:                     int: 输入模板的子表号
    :param headline:                        int：输出模板表首跳过行数
    :param tailline:                        int: 输出模板表尾跳过行数
    :param modelfilename:                   str：输出模板文件全路径名
    :param index:                           bool: 是否输出index
    :param styleout:                        bool: 无模板输出时，选择是否带df。style的格式输出
    :param sumrownumber:                    int: 合计行号，None表示不存在合计行
    :param processingbarevent:              tuple (window, eventkey) key 用于发送当前命令进度的事件，该命令运行较慢需提示时使用
    :return:                    [bool True:成功，False:失败, errstr]
    """
    # 当表格为空或者-1，使用最后一张表格
    if tableindex in [[], [-1], None]:
        tableindex = [len(specialdflist) - 1]
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 首先判断是否使用虚拟内存对象文件
    if is_mem_virtual_file_name(outputfilename):
        result = save_object_to_virtual_file(outputfilename, df)
        if result[0]:
            # 记录输出target
            specialdf.dataexporttarget = outputfilename
            return [True, ""]
        else:
            return errprintwithname(result[1])
    # 判断是否有模板
    if modelfilename in [None, '']:  # 没有模板文件，直接导出到excel
        try:
            starttime = datetime.now()
            if styleout:
                # 无模板带df.style输出
                # 根据specialdf的row\column styleinfo，形成当前dfdata的style df然后输出 目前支持rowcolor
                # getrowoutputstyle 需要传入行号！ 首先获取行号的list，每次调用弹出最前面的行号供使用
                rownumberlist = list(range(df.shape[0]))

                def getrowcolor(row, rownumber):
                    currentrownumber = rownumber.pop(0)
                    return specialdf.getrowoutputstyle(row, currentrownumber)
                # 设置条件格式
                df_style = df.style.apply(getrowcolor, rownumber=rownumberlist, axis=1)
                # 引用style做带格式输出
                df_style.to_excel(outputfilename, engine="openpyxl", index=index)
                # 记录输出target
                specialdf.dataexporttarget = outputfilename
                print("无模板带格式保存文件用时：{}".format(datetime.now() - starttime))
            else:
                # 无模板带无格式输出
                df.to_excel(outputfilename, engine="openpyxl", index=index)
                specialdf.dataexporttarget = outputfilename
                print("无模板保存文件用时：{}".format(datetime.now() - starttime))
        except Exception as err:
            return errprintwithname("无模板导出数据df到文件{}失败：{}！".format(outputfilename, err))
    else:
        # model 方式导出
        return dfsavedftomodelexcelfile(specialdflist, tableindex, outputfilename, modelfilename=modelfilename,
                                        sheetnumber=sheetnumber, headline=headline, tailline=tailline,
                                        sumrownumber=sumrownumber, processingbarevent=processingbarevent)

    print("导出数据df到文件{}子表{}完成！".format(outputfilename, sheetnumber))
    return [True, ""]


def dfsavedftomodelexcelfile(
        specialdflist: list,
        tableindex: list,
        outputfilename: str = '',
        modelfilename: str = "",
        sheetnumber: int = 0,
        headline: int = 0,
        tailline: int = 0,
        sumrownumber: int = None,
        processingbarevent: tuple = None
) -> list:
    """
    将df内容保存到文件, 该方式下，涉及到使用模板行的格式复制到其他所有数据行，由于数据行格式每列都相同，而且只涉及到数据结果输出到文件，
    不存在单独编辑格式，因此没有使用传统的copy.copy格式对象，而是直接将模板行的格式对象指针交给其他数据行！
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param outputfilename:                  str：输出文件全路径名
    :param modelfilename:                   str：输出模板文件全路径名
    :param sheetnumber:                     int: 输出到文件的子表号
    :param headline:                        int：输出模板表首跳过行数
    :param tailline:                        int: 输出模板表尾跳过行数
    :param sumrownumber:                    int: 合计行号，None表示不存在合计行
    :param processingbarevent:              tuple (window, eventkey) key 用于发送当前名另进度的事件，该命令运行较慢，需提示
    :return:                    [bool True:成功，False:失败, errstr]

    """
    # 打开模板文件，处理结束后保存到目标文件
    try:
        wb = load_workbook(modelfilename)
        ws = wb.worksheets[sheetnumber]
    except Exception as err:
        return errprintwithname("打开模板文件{}子表{}失败：{}！".format(modelfilename, sheetnumber, err))

    # 一次读取一行所有列的格式信息
    def getwsrowformat(i) -> list:
        """
        获取整行的单元格格式列表
        :param i:
        :return:
        """
        # 获取一行的格式
        # return list(map(getwscellformat,
        #                 [ws.cell(i, j) for j in range(1, ws.max_column + 1)]))
        return [
            [ws.cell(i, j).number_format] + list(map(copy.copy, [ws.cell(i, j).font, ws.cell(i, j).alignment,
                                                                 ws.cell(i, j).fill, ws.cell(i, j).border]))
            for j in range(1, ws.max_column + 1)
        ]

    # 设置一个单元格的格式信息
    def setwscellformat(cell, cellformat) -> list:
        """
        设置一个单元格的格式
        :param cell:                        单元格object
        :param cellformat:                  单元格格式list表
        :return:                            list : [True, ""] or [Fasle, errstr]
        """
        try:
            # 设置单元格的格式, number_format 直接复制，其他需要深度拷贝，共5项
            # cell.number_format, cell.font, cell.alignment, cell.fill, cell.border = \
            #     [cellformat[0]] + list(map(copy.copy, cellformat[1:5]))
            # 对应本函数，就是复制模板行的所有格式，所以尝试不使用copy.copy,直接使用对象指针
            cell.number_format, cell.font, cell.alignment, cell.fill, cell.border = cellformat
            return [True, ""]
        except Exception as err:
            return errprintwithname("设置单元格格式错误:{}".format(err))

    # 设置整行的格式信息
    def setwsrowformat(i):
        """
        设置workbook ws 的整行格式
        :param i:                       int:行数
        :return:                        list : [True, ""] or [Fasle, errstr]
        """
        # 该行每列单独调用setwscellformat？能否map？pool不能嵌套调用
        try:
            resultlist = list(map(setwscellformat, [ws.cell(i, j) for j in range(1, ws.max_column + 1)], rowformatlist))
            if False in [x[0] for x in resultlist]:
                print("{}".format(resultlist))
                return errprintwithname("设置行格式错误：{}".format([x[1] for x in resultlist if x[0] is False]))
            # 涉及非直接读取变量赋值给变量，使用list（map)效率更高。
            # for j in range(1, ws.max_column + 1):
            #     cell = ws.cell(i, j)
            #     cell.number_format, cell.font, cell.alignment, cell.fill, cell.border = rowformatlist[j - 1]
            if i % 200 == 0:
                print("行格式设置累计{}行，累积用时：{}".format(i, datetime.now() - starttime))
                if processingbarevent is not None:
                    # 刷新季度条和说明
                    bartext.update(Sg.SYMBOL_SQUARE * (6 + int((51 - 6 - 8) * i/len(rowscale))))
                    barcommenttext.update("行格式设置累计{}行，累积用时：{}".format(i, datetime.now() - starttime))
                    barwindow.refresh()
        except Exception as err:
            return errprintwithname("设置行格式错误：{}".format(err))
        return [True, ""]

    # 检查行所有单元格是否有条件格式， 尝试
    def checkrowcellhasconditionalformat(i):
        for j in range(1, ws.max_column + 1):
            for conditional_formatting in ws.conditional_formatting._cf_rules:
                for cell_range in conditional_formatting.cells.ranges:
                    if ws.cell(i, j).coordinate in cell_range:
                        print('cell({}, {}) contains a conditional formatting'.format(i, j))

    # 当表格为空或者-1，使用最后一张表格
    if tableindex in [[], [-1], None]:
        tableindex = [len(specialdflist) - 1]
    specialdf = specialdflist[tableindex[0]]
    resultdf = specialdf.dfdata

    # 将结果数据填入模板
    # 逐行\逐列处理，笨办法
    # ws 起始行
    rownumber = headline + 1  # 跳过表头
    tailoldstart = ws.max_row - tailline + 1
    tailoldend = ws.max_row
    tailnewstart = headline + resultdf.shape[0] + 1

    # 能够拷贝的列数，小于模板列数时，使用源数据列数
    maxcolnumber = resultdf.shape[1] if resultdf.shape[1] < ws.max_column else ws.max_column  # 获取模板列数
    # 逐行获取源数据信息进行处理
    try:
        # global barwindow
        if processingbarevent is not None:
            barwindow = processingbarevent[0]
            bartext = processingbarevent[0][processingbarevent[1]]
            barcommenttext = processingbarevent[0][processingbarevent[2]]
        # 首先判断sumrow是否需要拷贝，将sumrow行值填写到（源数据默认sumrow是最后一行，-1行，模板行数从0开始),
        # 如果有sumrow，实际后续拷贝的数据行需要减少最后一行
        sumrowflag = False
        if sumrownumber is not None and isinstance(sumrownumber, int):
            sumrowflag = True
            realsumrownumber = sumrownumber + 1 if sumrownumber >= 0 \
                else ws.max_row + sumrownumber + 1 if ws.max_row + sumrownumber >= 0 else ws.max_row
            sumrowvalue = resultdf.iloc[-1, :]
            for colnumber in range(1, maxcolnumber + 1):
                # ws的列数从1开始，但是dataframe从0 开始，所以此处需要-1
                ws.cell(realsumrownumber, colnumber, sumrowvalue[colnumber - 1])
            # tail 移动的目标起始行-1，因为合计行使原数据减少一行
            tailnewstart -= 1

        # 首先把headling、tailline行分开，中间空出数据行数
        if tailline > 0:
            tailmovednumber = 0
            starttime = datetime.now()
            for i in range(tailoldstart, tailoldend + 1):
                # tail 每行向下移动数据的总行数
                currentnewrow = tailnewstart + tailmovednumber
                # 复制到目标行格式
                rowformatlist = getwsrowformat(i)
                setwsrowformat(currentnewrow)
                # 赋值
                for j in range(1, maxcolnumber + 1):
                    ws.cell(currentnewrow, j, ws.cell(i, j).value)
                tailmovednumber += 1
                print("移动{}行到{}行累积用时：{}".format(tailmovednumber, currentnewrow, datetime.now() - starttime))
                if processingbarevent is not None:
                    # 刷新季度条和说明
                    bartext.update(Sg.SYMBOL_SQUARE * 6)
                    barcommenttext.update("移动{}行到{}行累积用时：{}".format(
                        tailmovednumber, currentnewrow, datetime.now() - starttime))
                    barwindow.refresh()

        # 复制数据行的格式信息
        # 方案1、获取headling后的第一行作为模板行，后续数据行使用该行的格式
        dataformatrow = headline + 1
        # 被复制格式的行范围从模板格式行的下一行到df的总数最后（不算模板格式行，少一行）
        rowscale = [i for i in range(headline + 2, headline + resultdf.shape[0] + (0 if sumrowflag else 1))]

        rowformatlist = getwsrowformat(dataformatrow)

        starttime = datetime.now()
        # 方案1、map方式
        resultlist = list(map(setwsrowformat, rowscale))
        print("行方式设置方案1(map)所有行格式时间：{}".format(datetime.now() - starttime))

        # 方案2 直接用for循环给每个单元赋值格式指针, 最简单的方式，逐个单元格拷贝指针而不是copycopy，因为只是为了输出到文件，不做编辑
        # 从headline + 2开始
        # starttime = datetime.now()
        # for i in range(headline + 2, headline + resultdf.shape[0] + 2):
        #     for j in range(1, ws.max_column + 1):
        #         ws.cell(i, j).number_format,  ws.cell(i, j).font, \
        #         ws.cell(i, j).alignment, ws.cell(i, j).fill, ws.cell(i, j).border = rowformatlist[j - 1]
        #     if i % 200 == 0:
        #         print("行格式设置累计{}行，用时：{}".format(i, datetime.now() - starttime))
        #         if processingbarevent is not None:
        #             # 刷新季度条和说明
        #             bartext.update(Sg.SYMBOL_SQUARE * (6 + int((51 - 6 - 8) * i/len(rowscale))))
        #             barcommenttext.update("行格式设置累计{}行，累积用时：{}".format(i, datetime.now() - starttime))
        #             barwindow.refresh()
        # print("行方式设置方案2(for)所有行格式时间：{}".format(datetime.now() - starttime))

        # 直接for循环赋值比用函数调用快很多倍！！！！！当表格有无数据，对前面的格式设置速度影响好像极大，必须先设置格式，后赋值？？？！！！
        starttime = datetime.now()
        # 记录最后合计行的index
        sumindex = resultdf.index[-1]
        for index, rowseries in resultdf.iterrows():
            # print("处理第{}/{}行\r".format(index,  resultdf.shape[0]), end='', flush=False if index % 100 else True)
            # 当有合计时，最后一行需要跳过
            if sumrowflag and index == sumindex:
                break
            row = rowseries.values
            # ws的列数从1开始，但是dataframe从0 开始，所以此处range如此设定
            for colnumber in range(1, maxcolnumber + 1):
                # ws的列数从1开始，但是dataframe从0 开始，所以此处需要-1
                ws.cell(rownumber, colnumber, row[colnumber - 1])
            # 本行处理完成，行数+1
            rownumber += 1
        print("拷贝{}行数据累积用时：{}".format(resultdf.shape[0], datetime.now() - starttime))

    except Exception as err:
        return errprintwithname("按模板{}填写出错：{}".format(modelfilename, err))

    # 保存处理好的数据表到输出目标excel文件
    try:
        starttime = datetime.now()
        wb.save(outputfilename)
        print("wb保存文件用时：{}".format(datetime.now() - starttime))
        specialdf.dataexporttarget = outputfilename
    except Exception as err:
        return errprintwithname("输出文件{}错误：{}".format(outputfilename, err))

    print("{}处理完成！".format(outputfilename))
    return [True, ""]


def dfcopydfblocktomodelexcelfile(specialdflist: list, tableindex: list, outputfilename: str,
                                  modelfilename: str, sheetnumber: int = 0,
                                  sourcestartendrows: list = (0, None),
                                  sourcestartendcols=None,
                                  targetstartrow=None,
                                  targetstartcol=None) -> list:
    """
    将df矩形区域内容保存到文件的指定区域, sourcestartendcols, targetstartcol采用新column格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param outputfilename:                  str：输出文件全路径名
    :param modelfilename:                   str：输出模板文件全路径名
    :param sheetnumber:                     int: 输出到文件的子表号
    :param sourcestartendrows:              list: 拷贝开始、结束行号，从0开始 [0, n]
    :param sourcestartendcols:              list: 拷贝开始、结束列号，从0开始 [[0, n],[开始列名，结束列名]]
    :param targetstartrow:                  list:  目标起始行list
    :param targetstartcol:                  list: 目标文件起始列list [n]
    :return:                                [bool True:成功，False:失败, errstr]
    """
    if targetstartcol is None:
        targetstartcol = [0]
    if targetstartrow is None:
        targetstartrow = [0]
    if sourcestartendcols is None:
        sourcestartendcols = [[0, None], ["", ""]]

    # 当表格为空或者-1，使用最后一张表格
    if tableindex in [[], [-1], None]:
        tableindex = [len(specialdflist) - 1]
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 判读源数据是否空, 如果为空，直接复制模板到结果文件并返回
    if df.shape[0] == 0:
        # 打开模板文件，处理结束后保存到目标文件
        try:
            # 保存处理好的数据表到输出目标excel文件
            # wb = load_workbook(modelfilename)
            # wb.save(outputfilename)
            shutil.copy(modelfilename, outputfilename)
        except Exception as err:
            return errprintwithname("输出文件{}错误：{}".format(outputfilename, err))

        print("源数据为空，直接生成{}处理完成！".format(outputfilename))
        return [True, ""]

    # 数据合法性检查
    if len(sourcestartendrows) < 1 or len(sourcestartendcols[0]) < 1:
        return errprintwithname("源数据行数：{0}、列数{1}， 选择的行范围{2} 列范围{3}不合法！".format(
            df.shape[0] + 1, df.shape[1] + 1, sourcestartendrows, sourcestartendcols[0]))

    # 如果最后写填写的是-1， 表示最大行数、列数
    if sourcestartendrows[-1] == -1:
        sourcestartendrows[-1] = df.shape[0] - 1
    if sourcestartendcols[0][-1] == -1:
        sourcestartendcols[0][-1] = df.shape[1] - 1
    # 首先排序以便选取最小、最大值
    sourcestartendrows.sort()
    sourcestartendcols[0].sort()
    # 取开始，结束行列，如果是1个行列，通过-1解决
    sourcestartrow, sourceendrow = sourcestartendrows[0], sourcestartendrows[-1] + 1
    sourcestartcol, sourceendcol = sourcestartendcols[0][0], sourcestartendcols[0][-1] + 1
    if sourceendrow > df.shape[0]:  # 填写的是-1， 表示最大行数
        sourceendrow = df.shape[0]
    if sourceendcol > df.shape[1]:  # 填写的是-1， 表示最大列数
        sourceendcol = df.shape[1]

    # 数据合法性检查
    if False in [x in range(df.shape[0] + 1) for x in [sourcestartrow, sourceendrow]] + \
            [x in range(df.shape[1] + 1) for x in [sourcestartcol, sourceendcol]]:
        return errprintwithname("源数据行数：{0}、列数{1}， 选择的行范围{2} 列范围{3}不合法！".format(
            df.shape[0] + 1, df.shape[1] + 1, sourcestartendrows, sourcestartendcols))

    # 打开模板文件，处理结束后保存到目标文件
    try:
        wb = load_workbook(modelfilename)
        ws = wb.worksheets[sheetnumber]
    except Exception as err:
        return errprintwithname("打开模板文件{}子表{}失败：{}！".format(modelfilename, sheetnumber, err))

    # 将结果数据填入模板
    # 逐行\逐列处理，笨办法
    # ws 起始行
    rowstart = targetstartrow[0] + 1  # 目标的起始填写行
    colstart = targetstartcol[0] + 1  # 目标的起始填写列
    totalcol = sourceendcol - sourcestartcol
    # 是否在处理第一行，从数据第二行开始，复制第一行的格式
    # 逐行获取源数据信息进行处理
    try:
        # 逐行填写到ws
        starttime = datetime.now()
        for processedrow, currentrow in enumerate(range(sourcestartrow, sourceendrow)):
            # 读取源数据行内容
            # print("处理第{}/{}行\r".format(processedrow, totalrows), end='', flush=False if processedrow % 100 else True)
            currentrowvaluelist = df.iloc[currentrow, sourcestartcol: sourceendcol].to_list()
            # 逐个赋值给目标文件
            for currentvalueindex, currentfillcol in enumerate(range(colstart, colstart + totalcol)):
                ws.cell(rowstart + processedrow, currentfillcol, currentrowvaluelist[currentvalueindex])
        print("模块数据拷贝用时：{}".format(datetime.now() - starttime))
    except Exception as err:
        return errprintwithname("按模板{}填写出错：{}".format(modelfilename, err))

    # 保存处理好的数据表到输出目标excel文件
    try:
        starttime = datetime.now()
        wb.save(outputfilename)
        specialdf.dataexporttarget = outputfilename
        endtime = datetime.now()
        print("wb 保存用时：{}".format(endtime - starttime))
    except Exception as err:
        return errprintwithname("输出文件{}错误：{}".format(outputfilename, err))

    print("{}处理完成！".format(outputfilename))
    return [True, ""]


def dffillna(specialdflist: list, tableindex: list, fillvalue: str = "") -> list:
    """
    调用fillna
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param fillvalue:           str:  用于填充的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # # 使用系统赋值对全表所有列根据列信息转换数据
        # df.fillna(value=fillvalue, inplace=True)
        # 自定义类型赋值
        columns = [[x for x in range(df.shape[1])], list(df)]
        fillvaluewithtype = value2columntype(specialdf, columns, [fillvalue for _ in range(len(columns[0]))])
        valuedict = dict(zip(df.columns[columns[0]], fillvaluewithtype))
        df.fillna(value=valuedict, inplace=True)

        specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("dffillna 错误：{}".format(err))
    return [True, ""]


def dfcolfillna(specialdflist: list, tableindex: list,
                columns: list = None, fillvalue: str = '') -> list:
    """
    调用fillna， columns空值替换
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:             list: 用于填充的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param fillvalue:           str:  用于填充的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if columns is None:
        # 填充整张表
        valuedict = fillvalue
    else:
        # # 根据列信息转换数据使用系统自有赋值，
        # valuedict = {x: fillvalue for x in df.columns[columns[0]]}
        # 使用自定义type方式
        fillvaluewithtype = value2columntype(specialdf, columns, [fillvalue for _ in range(len(columns[0]))])
        valuedict = dict(zip(df.columns[columns[0]], fillvaluewithtype))
    try:
        df.fillna(value=valuedict, inplace=True)
        specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
    except Exception as err:
        return errprintwithname("列空值替换错误：{}".format(err))
    return [True, ""]


NA_VALUES_DEF = {
    '1.#QNAN', '', '<NA>', '1.#IND', '#NA', 'n/a', '#N/A N/A',
    '-1.#QNAN', '-1.#IND', 'NULL', 'NaN', 'nan',
    'N/A', '-nan', '#N/A', 'NA', 'null', '-NaN'}


def dfcolcheckna(specialdflist: list, tableindex: list, columns: list = None) -> list:
    """
    检测列是否为空值，并且替换为指定空值，默认使用np.nan
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:             list: 检测的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if columns is None:
        return errprintwithname("列空值检测错误：没有指定要检测空值的列")
    try:
        # for currentcolumns in columns[0]:
        #     df[df.columns[[currentcolumns]]] = list(map(
        #         lambda x: [np.nan] if x[0] in NA_VALUES_DEF
        #         else x, df[df.columns[[currentcolumns]]].values.tolist()))
        df[df.columns[columns[0]]] = df[df.columns[columns[0]]].applymap(lambda x: np.nan if x in NA_VALUES_DEF else x)
        specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
    except Exception as err:
        return errprintwithname("列{}检测空值替换为{}错误：{}".format(columns, np.nan, err))
    return [True, ""]


def dfallcheckna(specialdflist: list, tableindex: list) -> list:
    """
    检测全表所有空值，并且替换为指定空值，默认使用np.nan
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 形成全表所有列的列信息列表
    columns = [[x for x in range(df.shape[1])], df.columns.to_list()]

    return dfcolcheckna(specialdflist, tableindex, columns=columns)


def dfcolvaluereplace(specialdflist: list, tableindex: list, columns: list = None,
                      oldvalue: object = None, fillvalue: object = None) -> list:
    """
    对df 的column列用fillvalue替换所有等于oldvalue的值，columns调整为新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:             list: 用于替换值的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param oldvalue:            object:  要替换的值
    :param fillvalue:           object:  用于填充的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if columns is None:
        return errprintwithname("没有指定要替换值的列")
    try:
        # # 使用系统方式赋值
        # df[df.columns[columns[0]]] = df[df.columns[columns[0]]].applymap(lambda x: fillvalue if x == oldvalue else x)
        # 使用自定义类型赋值
        filltype = value2columntype(specialdf, columns, [fillvalue])[0]
        oldtype = value2columntype(specialdf, columns, [oldvalue])[0]
        df[df.columns[columns[0]]] = df[df.columns[columns[0]]].applymap(lambda x: filltype if x == oldtype else x)

        specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("列{}以{}替换{}错误：{}".format(columns, fillvalue, oldvalue, err))
    return [True, ""]


def dffillrowstylebymark(specialdflist: list, tableindex: list, markcolumns: list, markvalues: list,
                         textcolor: str = "blue", backgroundcolor: str = "white") -> list:
    """
    对行按索引列指定值填充style，目前支持文本、背景颜色
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param markcolumns:              提供标记的列list, 可以多列
    :param markvalues:               标记对应的值list，个数必须与标记列数一致
    :param textcolor:               str：文本颜色
    :param backgroundcolor:         str：背景颜色
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 逐行根据索引列值和mark值的对应关系，确定是否配置rowstyleinfo rowcolor
    try:
        for row in range(df.shape[0]):
            if df.iloc[row, markcolumns[0]].to_list() == markvalues:
                specialdf.rowstyleinfo[row][ROWSTYLE_ROWCOLOR] = (textcolor, backgroundcolor)
    except Exception as err:
        return errprintwithname("按标记列{}标记值{}填充颜色失败：{}".format(markcolumns, markvalues, err))
    return [True, ""]


def dfrowreindex(specialdflist: list, tableindex: list) -> list:
    """
    重建整张表的行索引
    :param specialdflist:
    :param tableindex:
    :return:
    """
    specialdf = specialdflist[tableindex[0]]
    try:
        # 重置行索引
        specialdf.dfreindex()
        # 初始化所有行行相关的style以及选中项
        # specialdf.dfbuildrowstyleinfo()               # 由于行索引只是修改行号，实际物理位置无变化，可以考虑不删除行格式信息
    except Exception as err:
        return errprintwithname("重建行索引失败：{}".format(err))
    return [True, ""]


def rownumber2rawrownumber(df: pd.DataFrame, targetrow) -> int:
    """
    将输入行号转化为df实际的行号：
    :param df:                      处理的dataframe
    :param targetrow:               原始targetrow
    :return:                        int：实际行号
                                    真实目标行号， targetrow 负数表示从表尾倒数，计算结果大于当前表行数表数在表尾新增，小于0则计算为0
    """
    oldrowcount = len(df.index)
    realtargetrow = targetrow if 0 <= targetrow < oldrowcount \
        else oldrowcount if targetrow >= oldrowcount \
        else oldrowcount + targetrow if 0 <= -targetrow <= oldrowcount else 0
    return realtargetrow


def value2columntype(specialdf: SpecialDf, columns: list, valuelist: list) -> list:
    """
    将行值list按对应的列值类型进行转换
    :param specialdf:           SpecialDf
    :param columns:             list： 要处理的列序号列表
    :param valuelist:           list:  要处理的值
    :return:                    list： 处理后的value list
    """
    # 判断数据有效性。
    # 如果值和列的信息不一致，或者个数大于df的列宽，直接返回原值
    df = specialdf.dfdata
    if len(valuelist) != len(columns[0]) or len(valuelist) > df.shape[1]:
        return valuelist
    # 首先获取主类型
    columnsname = df.columns[columns[0]]
    # 根据第一行数据将其中的"object"类型更换为更细的子类型
    typesubnamelist = specialdf.getcolumnstype(columnsname)

    # object --》 <class 'str'> , <class 'int'> , <class 'list'>........,
    # 如果列类型不含str，并且当前value是str，则运行ast.literal_eval转换类型，如果列类型是str，当前值不是str，则用str转换
    newvaluelist = []
    # 逐项转换，不成功就用原值
    for x in zip(typesubnamelist, valuelist):
        try:
            # 转换后的类型,
            if 'str' not in x[0] and isinstance(x[1], str):
                resultvalue = ast.literal_eval(x[1])
                print("columns type: {} value: {} trans by literal_eval -> type: {}".format(
                    x[0], x[1], type(resultvalue)))
            elif 'str' in x[0]:
                resultvalue = str(x[1])
                print("column type: {} value: {} to str: {}".format(x[0], x[1], resultvalue))
            else:
                resultvalue = x[1]
                print("column type: {} value: {} not changed".format(x[0], x[1]))
        except Exception as err:
            print("column type: {} value: {} not changed, exception: {}".format(x[0], x[1], err))
            resultvalue = x[1]
        newvaluelist.append(resultvalue)

    return newvaluelist


def dfaddrow(specialdflist: list, tableindex: list, rowvaluelist: list = None) -> list:
    """
    添加行
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rowvaluelist:        添加行的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        if rowvaluelist == [] or rowvaluelist is None:
            df.loc[len(df.index)] = [np.nan for _ in range(len(df.columns))]
        else:
            if len(rowvaluelist) != len(df.columns):
                return errprintwithname("添加行错误：输入行值个数{}与表格列数{}不一致".format(len(rowvaluelist), len(df.columns)))
            else:
                # 使用自定义类型
                rowvaluetype = value2columntype(specialdf, [[x for x in range(df.shape[1])], list(df)], rowvaluelist)
                df.loc[len(df.index)] = rowvaluetype
                # # 使用系统类型
                # df.loc[len(df.index)] = rowvaluelist
        # 建立该行的rowstyleinfo
        specialdf.rowstyleinfo.append({})
        # 删除原行索引并重建
        specialdf.dfreindex()
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("添加行错误：{}".format(err))
    return [True, ""]


def dfupdaterow(specialdflist: list, tableindex: list,
                rownumber: list = None, rowvaluelist: list = None) -> list:
    """
    修改一行的值
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rownumber:            修改行的行序号list,只用rownumber[0]
    :param rowvaluelist:        修改行的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 将输入原行号转换为实际行号
        realrowlist = [rownumber2rawrownumber(df, x) for x in rownumber]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行修改错误，指定行序号不存在:{}".format(rownumber))
        if realrowlist in [None, []] or rowvaluelist is None or not isinstance(rownumber, list):
            return errprintwithname("行修改错误，输入行序号{}或修改值：{}为空".format(rownumber, rowvaluelist))
        elif len(rowvaluelist) != len(df.columns):
            return errprintwithname("行修改错误，输入行值个数{}与表格列数{}不一致".format(len(rowvaluelist), len(df.columns)))
        else:
            # # 使用系统赋值，将输入值按列类型转换
            # df.iloc[realrowlist[0]] = rowvaluelist
            # 使用自定义类型赋值
            rowvaluetype = value2columntype(specialdf, [[x for x in range(df.shape[1])], list(df)], rowvaluelist)
            df.iloc[realrowlist[0]] = rowvaluetype
            # 刷新列类型
            specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("行修改错误， 行号:{}, 行值：{} 错误：{}".format(rownumber, rowvaluelist, err))

    return [True, ""]


def dfrowcellfill(specialdflist: list, tableindex: list,
                  rowslist: list, columns: list, fillvalue: object = None) -> list:
    """
    多行根据指定列（字符串列）进行赋值, columns使用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rowslist:            需赋值的行号list
    :param columns:             赋值涉及的列,[[],[]]
    :param fillvalue:           填充的值
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 将输入原行号转换为实际行号
        realrowlist = [rownumber2rawrownumber(df, x) for x in rowslist]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行单元赋值错误，指定行序号不存在:{}".format(rowslist))
        elif len(rowslist) < 1:  # 没有要合并的行，直接返回
            return [True, ""]
        else:
            # # 使用系统方式赋值
            # df.loc[realrowlist, df.columns[columns[0]]] = [fillvalue for _ in range(len(realrowlist))]
            # 自定义type方式赋值
            # 形成赋值的一行数据并且根据对应列的列类型分别转换
            fillvaluewithtype = value2columntype(specialdf, columns, [fillvalue for _ in range(len(columns[0]))])
            # 根据要填值得行列形成相应的二维list赋值
            # 当系统适应df单元格可以填写对象是，无法直接用list矩阵赋值（因为list本事也可以作为对象），
            # 1、没有list的可以用常规方式
            # 2、对于单行，可以逐列赋值
            # 3、剩余是大于1*1的矩阵可以使用子df给df赋值
            rowcount, columncount = len(realrowlist), len(fillvaluewithtype)
            if True not in [isinstance(x, list) for x in fillvaluewithtype]:
                df.loc[realrowlist, df.columns[columns[0]]] = [fillvaluewithtype for _ in range(rowcount)]
            elif df.shape[0] == 1:
                # 表格本身只有一行，可以逐列单独赋值
                for i, columname in enumerate(df.columns[columns[0]]):
                    df[columname] = [fillvaluewithtype[i]]
            else:
                # 包含list，当前赋值范围大于1*1范围 ，创建临时df与赋值范围一致
                dftemp = pd.DataFrame(
                    [[1 for _ in range(columncount)] for _ in range(rowcount)],
                    columns=[str(x) for x in range(columncount)])
                for i in range(columncount):
                    dftemp[str(i)] = [fillvaluewithtype[i] for _ in range(rowcount)]
                df.iloc[realrowlist, columns[0]] = dftemp

            # 刷新列类型
            specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])

    except Exception as err:
        return errprintwithname("行单元格赋值错误：{}".format(err))

    return [True, ""]


def dfrowcombine(specialdflist: list, tableindex: list,
                 rowslist: list, columns: list, targetrow: int,
                 seperate='\n', nanreplace='-') -> list:
    """
    多行根据指定列（字符串列）进行合并生成新行, columns使用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rowslist:            需合并的行号list
    :param columns:             合并涉及的str类型列,[[],[]]
    :param targetrow:           目标行行号
    :param seperate:            str:  合并行之间的分隔符
    :param nanreplace:          str:  nan的替换符号
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 将输入原行号转换为实际行号
        oldrowcount = len(df.index)
        realrowlist = [rownumber2rawrownumber(df, x) for x in rowslist]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行合并错误，指定行序号不存在:{}".format(rowslist))
        elif len(rowslist) < 1:  # 没有要合并的行，直接返回
            return [True, ""]
        else:
            realtargetrow = rownumber2rawrownumber(df, targetrow)

            # 以下行转列然后用列str.cat多列的方式
            df.loc[realtargetrow, df.columns[columns[0]]] = \
                df.iloc[rowslist[0], columns[0]].str.cat(df.iloc[rowslist[1:], columns[0]].T,
                                                         sep=seperate, na_rep=nanreplace)
            # 增加该行的rowstyleinfo
            if len(df.index) > oldrowcount:
                specialdf.rowstyleinfo.append({})
            # 删除原行索引并重建
            specialdf.dfreindex()
            # 刷新列类型
            specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
    except Exception as err:
        return errprintwithname("行合并错误：{}".format(err))

    return [True, ""]


def dfrowsdel(specialdflist: list, tableindex: list, rows: list) -> list:
    """
    删除df指定的行
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rows:                     list: 需要删除的列序号列表 [num1, num2, ....]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    try:
        # 将输入原行号转换为实际行号
        realrowlist = [rownumber2rawrownumber(df, x) for x in rows]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行删除错误，指定行序号不存在:{}".format(rows))
        elif len(realrowlist) < 1:  # 没有要删除的行，直接返回
            return [True, ""]
        df.drop(realrowlist, axis=0, inplace=True)
        # 清空被删除行的styleinfo（list需要逐项pop，由大到小）
        for item in sorted(realrowlist, reverse=True):
            specialdf.rowstyleinfo.pop(item)
        # 删除原行索引并重建
        specialdf.dfreindex()
        # 清除选中的行
        specialdf.selectedrawrow = set()
    except Exception as err:
        return errprintwithname("行删除错误，行:{}错误:{}".format(rows, err))
    return [True, ""]


def dfrowsduplicatedel(specialdflist: list, tableindex: list, columns: list = None,
                       keep: str = 'first', inplace: bool = True, ignore_index: bool = True):
    """
    根据指定列索引，根据删除方式保留重复值的某行，
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list： 指定的对比列序号（列名）， None或者[[], []]代表所有列
    :param keep:                    str：保留重复行中某一行采用的方式：last ：重复值最后一行, first: 重复值第一行
    :param inplace:                 bool: 结果是否保存到原df
    :param ignore_index:            bool: 是否重建索引
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    try:
        # 记录原index
        olddfindex = df.shape[0]
        # 将输入原行号转换为实际行号
        currentcolumns = [x for x in range(len(df.columns))] if columns in [None, [[], []]] else columns[0]
        df.drop_duplicates(subset=df.columns[currentcolumns], keep=keep, inplace=inplace, ignore_index=False)
        # 清空被删除行的styleinfo（list需要逐项pop，由大到小）
        # 获取所有被删除行的index
        deletedrowslist = [x for x in range(olddfindex) if x not in df.index]
        # 将所有被删除index(由大到小反向）对应的rowstyleinfo删除
        for item in sorted(deletedrowslist, reverse=True):
            specialdf.rowstyleinfo.pop(item)
        # 删除原行索引并重建
        if ignore_index:
            specialdf.dfreindex()
        # 清除选中的行
        specialdf.selectedrawrow = set()
    except Exception as err:
        return errprintwithname("行重复值删除保留方式（{}）错误:{}".format(keep, err))
    return [True, ""]


def dfrowsclear(specialdflist: list, tableindex: list, rows: list) -> list:
    """
    df指定的行全部赋值为空：np.nan
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rows:                     list: 需要清空的列序号列表 [num1, num2, ....]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    try:
        # 将输入原行号转换为实际行号
        realrowlist = [rownumber2rawrownumber(df, x) for x in rows]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行删除错误，指定行序号不存在:{}".format(rows))
        elif len(realrowlist) < 1:  # 没有要清空的行，直接返回
            return [True, ""]
        # 逐行赋值
        df.iloc[realrowlist, :] = [[np.nan for x in range(df.shape[1])] for _ in range(len(realrowlist))]
    except Exception as err:
        return errprintwithname("行清空错误，行:{}错误:{}".format(rows, err))
    return [True, ""]


def dfrowmove(specialdflist: list, tableindex: list, sourcerowlist: list, targetrowlist: list) -> list:
    """
    移动df指定的行到某行
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param sourcerowlist:           list: 源数据序号列表[num1,num2,.....]，仅第一个有效， 负数表示从最后行倒数
    :param targetrowlist:           list: 目标行序号列表 [num1, num2, ....]，仅第一个有效， 负数表示从最后行倒数
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    try:
        # 判断输入合法性
        if len(targetrowlist) < 1 or len(sourcerowlist) < 1:
            return errprintwithname("行移动错误，行{}移动到行{}:数据不合法:行序号范围不正确".format(targetrowlist, sourcerowlist))

        # 行号转换为正数
        sourcerow = [rownumber2rawrownumber(df, x) for x in sourcerowlist]
        targetrow = [rownumber2rawrownumber(df, x) for x in targetrowlist]

        # 判断数据量是否合法，
        if False in [x in range(0, df.shape[0]) for x in sourcerow] or \
                False in [x in range(0, df.shape[0]) for x in targetrow]:
            return errprintwithname("行移动错误，实际行{}移动到实际行{}:数据不合法:行序号范围不在有效范围{}内".format(
                sourcerow, targetrow, df.shape[0]))

        # 保留headings以备恢复
        tempheadings = df.columns
        # 获取当前行的值
        movevalueseries = df.iloc[sourcerow[0], :]
        # 获取当前行的styleinfo, 并删除
        temprowstyleinfo = specialdf.rowstyleinfo.pop(sourcerow[0])
        # movevalue = movevalueseries.tolist() 使用series 和 list都可以？
        # 删除当前行
        df.drop(sourcerow[0], axis=0, inplace=True)
        # 当前行值插入到指定行，该操作影响了df的columns，需要复原
        df = pd.DataFrame(np.insert(df.values, targetrow[0], values=movevalueseries, axis=0))
        # 将当前行rowstyleinfo插入到目标行位置
        specialdf.rowstyleinfo.insert(targetrow[0], temprowstyleinfo)
        df.columns = tempheadings
        specialdf.dfdata = df
        # 处理过滤，清理过滤
        specialdf.dfreindex()
        # 清除选中的行
        specialdf.selectedrawrow = set()
    except Exception as err:
        return errprintwithname("行移动错误，行{}移动到行{}失败：{}".format(targetrowlist, sourcerowlist, err))

    return [True, ""]


def dfrowmarkna(specialdflist: list, tableindex: list, allnaflag: bool = True, markcolumnname: str = "空值行标记"):
    """

    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]],
                                    元素从0到多个， 本命令使用第一个
    :param allnaflag:               bool； True，全行为空才标记为True，False，行有空值就标记为空
    :param markcolumnname:          存贮标记结果的行
    :return:
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    def partnulljudge(rowseries):
        return True if True in rowseries.values else False

    def allnuljudge(rowseries):
        return False if False in rowseries.values else True

    try:
        df[markcolumnname] = df.isnull().apply(allnuljudge if allnaflag else partnulljudge, axis=1)
        # 更新列信息
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列的类型
        specialdf.refreshcolumntype(columnnamelist=[markcolumnname])
    except Exception as err:
        return errprintwithname("行空值标记到列{}错误:{}".format(markcolumnname, err))
    return [True, ""]


def dfrowsmarkdel(specialdflist: list, tableindex: list, markcolumns=None,
                  delmarkvalue: bool = True) -> list:
    """
    删除df指定的列, markcolumn使用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param markcolumns:             list: 删除行使用的标记列列表，只有第一项有效[[],[]]
    :param delmarkvalue:            bool: True/False
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if markcolumns is None:
        markcolumns = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # demo: df1.drop(df1[df1[df1.columns[0]] == 5].index, inplace=True)
        olddfindex = df.shape[0]
        df.drop(df[df[df.columns[markcolumns[0][0]]] == delmarkvalue].index, inplace=True)
        # 获取所有被删除行的index
        deletedrowslist = [x for x in range(olddfindex) if x not in df.index]
        # 将所有被删除index(由大到小反向）对应的rowstyleinfo删除
        for item in sorted(deletedrowslist, reverse=True):
            specialdf.rowstyleinfo.pop(item)
        # 删除原行索引并重建
        df.reset_index(inplace=True, drop=True)
        # 清除选中的行
        specialdf.selectedrawrow = set()
    except Exception as err:
        return errprintwithname("按标记列{}标记值{}删除行错误:{}".format(markcolumns, delmarkvalue, err))
    return [True, ""]


def dfrowsmarkfill(specialdflist: list, tableindex: list,
                   columns: list, fillvalue: object = None,
                   markcolumns=None, fillmarkvalue: bool = True) -> list:
    """
    按指定的标记列markcolumn，对相应行的columns赋值
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 赋值涉及的列,[[],[]]
    :param fillvalue:               填充的值
    :param markcolumns:             list: 删除行使用的标记列列表，只有第一项有效[[],[]]
    :param fillmarkvalue:            bool: True/False
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if markcolumns is None:
        markcolumns = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 获取满足标记行的行号list《后续调用cellfill，不用处理填充列值的类型）
        rowslist = df[df[df.columns[markcolumns[0][0]]] == fillmarkvalue].index.tolist()
    except Exception as err:
        return errprintwithname("按标记列{}标记值{}行值填充错误:{}".format(markcolumns, fillmarkvalue, err))
    # 调用dfrowcellfill
    return dfrowcellfill(specialdflist, tableindex, rowslist=rowslist, columns=columns, fillvalue=fillvalue)


def dfrowsum(specialdflist: list, tableindex: list, columns: list, targetrow: int = 999999) -> list:
    """
    对df的指定列全部行求和, columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 用于求和的列序号、列名列表 [[num1, num2, ....],[name1,name2,......]]
    :param targetrow:               int:  生成的求和行号
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata

    try:
        oldrowcount = len(df)
        realtargetrow = rownumber2rawrownumber(df, targetrow)
        df.loc[realtargetrow] = df[df.columns[columns[0]]].sum(axis=0)
        # 增加该行的rowstyleinfo
        if len(df) > oldrowcount:
            specialdf.rowstyleinfo.append({})
        # 删除原行索引并重建
        specialdf.dfreindex()
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
    except Exception as err:
        return errprintwithname("行求和错误，求和:{} 生成列:{} 错误:{}".format(columns, targetrow, err))

    return [True, ""]


def dfrowmultilevelcopy(specialdflist: list, tableindex: list, columns: list, 
                        fillna: str = '', rownanflag: bool = False) -> list:
    """
    对df的指定列按多级方式填写空值，columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 用于复制的列序号、列名列表 [[num1, num2, ....],[name,name2....]]
    :param fillna:                  str: 处理时用于临时替换空值的值， reversed 保留项，目前不采用
    :param rownanflag:              bool: True， 只有整行为空值是，才复制本组的第一行（不全为空）的数据，
                                          False： 不考虑列之间的关系，只要是本行本列为空，就复制前行本列（不为空）数据
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 首先对目标列做一次空值检测，以保证所有空值均为np.nan
        result = dfcolcheckna(specialdflist, tableindex, columns)
        if result[0] is False:
            return errprintwithname("指定列多级拷贝错误，进行列{}空值检测出错:{}".format(columns, result[1]))
        # 需要进行覆盖的空值列表，个数和指定的列一致
        currentfillcontent = [np.nan for _ in range(len(df.columns))]
        for row in range(df.shape[0]):
            # 获取当前行内容
            currentrowdf = df.iloc[row, columns[0]]
            currentnullflag = currentrowdf.isnull().tolist()
            # rownanflag True：只有本行所有列为空，才复制本组最近一次（行值不全为空）的数据
            if rownanflag is True:
                # 如果当前行内容不全部为空，说明是分组的最近非全空数据，保留
                if False in currentnullflag:
                    currentfillcontent = currentrowdf.tolist()
                else:
                    # 如果当前行全部为空，赋值本组的最近非全空数据
                    df.iloc[row, columns[0]] = currentfillcontent
            else:
                # 如果当前行内容不为空，用当前行内容替换np.nan的内容，（不考虑列之间的关系，只要是本行本列为空，就复制前行本列（不为空）数据）
                currentfillcontent =\
                    [z if x else y for x, y, z in zip(currentnullflag, currentrowdf.tolist(), currentfillcontent)]
                # 然后回填全行内容
                df.iloc[row, columns[0]] = currentfillcontent
    except Exception as err:
        return errprintwithname("指定列多级拷贝错误， 列：{}多级行复制错误:{}".format(columns, err))

    return [True, ""]


def dfrowsubtract(specialdflist: list, tableindex: list, rowslist: list = None,
                  columns: list = None, targetrow: int = 999999) -> list:
    """
    对df的指定两行求差（数值）并生成新行,columns 采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rowslist:            求差的行号list
    :param columns:             求差涉及的数值类型列序号、列名列表
    :param targetrow:           int:生成的求差行号
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 将输入原行号转换为实际行号
    oldrowcount = len(df.index)
    realrowlist = [rownumber2rawrownumber(df, x) for x in rowslist]
    if False in [x in range(0, df.shape[0]) for x in realrowlist]:
        return errprintwithname("行求差错误，指定行序号不存在:{}".format(rowslist))
    elif len(realrowlist) < 1:  # 没有要求差的行，直接返回
        return [True, ""]
    # 行相减至少应该有两行（本处理只适用前两行）
    if len(realrowlist) >= 2:
        realtargetrow = rownumber2rawrownumber(df, targetrow)
        try:
            df.loc[realtargetrow, df.columns[columns[0]]] = \
                list(map(lambda x, y: (x - y), df.iloc[rowslist[0], columns[0]], df.iloc[rowslist[1], columns[0]]))
            # 增加该行的rowstyleinfo
            if len(df) > oldrowcount:
                specialdf.rowstyleinfo.append({})
            # 删除原行索引并重建
            specialdf.dfreindex()
            # 刷新列类型
            specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
        except Exception as err:
            return errprintwithname("行求差错误， 按列{} 行求差:{} 生成行:{} 错误:{}".format(
                columns, rowslist, targetrow, err))
    else:
        return errprintwithname("行求差错误 错误：{} 行数不到2个，无法求差".format(rowslist))

    return [True, ""]


def dfrowspecialprocess(specialdflist: list, tableindex: list,
                        rowslist: list, targetrow: int, processstr: str) -> list:
    """
    对df的指定行进行特殊处理生成新行，特殊处理为lambda函数字符串或其他处理函数字符串
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param rowslist:                 指定进行处理的行序号list
    :param targetrow:                目标行号， 负数表示从表尾倒数，大于当前表行数表数在表尾新增
    :param processstr:               用于处理函数或lambda函数字符串
    :return:                         [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 将输入原行号转换为实际行号
        realrowlist = [rownumber2rawrownumber(df, x) for x in rowslist]
        if False in [x in range(0, df.shape[0]) for x in realrowlist]:
            return errprintwithname("行特殊处理错误，指定行序号不存在:{}".format(rowslist))
        # 行抽取后要进行行列转置，lambda的输入数和行数才能一致。
        commandstr = 'list(map({}, df.iloc[{}].T.values))'.format(processstr, realrowlist)
        # 真实目标行号， 负数表示从表尾倒数，大于当前表行数表数在表尾新增
        oldrowcount = len(df.index)
        realtargetrow = rownumber2rawrownumber(df, targetrow)
        print("对行[{}]执行特殊处理到[{}]-实际[{}]：{} ...".format(
            df.iloc[realrowlist].T.values, targetrow, realtargetrow, commandstr))
        # print(eval(commandstr))
        df.loc[realtargetrow] = eval(commandstr)
        # 如果是新增行，增加该行的rowstyleinfo
        if len(df.index) > oldrowcount:
            specialdf.rowstyleinfo.append({})
        # 删除原行索引并重建
        specialdf.dfreindex()
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("行特殊处理错误，行处理生成行:{} 错误:{}".format(targetrow, err))

    return [True, ""]


def dfcolumnswap(specialdflist: list, tableindex: list, columns: list) -> list:
    """
    交换df两个指定列
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:             list: 用于交换的列序号、列名列表 ，仅适用前两个[[num1, num2, ....],[name1,name2.....]]
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 判断是否有两列
    if len(columns[0]) >= 2:
        # 交换列名
        try:
            cols = list(df)
            cols[columns[0][0]], cols[columns[0][1]] = cols[columns[0][1]], cols[columns[0][0]]
            # 两行代码作用相同
            df.iloc[:, [columns[0][1], columns[0][0]]] = df[df.columns[columns[0]]].values
            df.columns = cols
            specialdf.headings = list(df)
            specialdf.colnamelist = specialdf.headings
            specialdf.colnameusedlist = [""] + specialdf.headings
            specialdf.colnamefilterlist = [""] + specialdf.headings
        except Exception as err:
            return errprintwithname("列交换 {} 错误：{}".format(columns, err))
    else:
        return errprintwithname("列交换 {} 错误：列数不到2个，无法交换列".format(columns))

    return [True, ""]


def dfcolumnmove(specialdflist: list, tableindex: list,
                 sourcecolumnlist: list, targetcolumnlist: list) -> list:
    """
    移动df指定的列到某列，sourcecolumnlist ,targetcolumnlist采用新column格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param sourcecolumnlist:           list: 源数据列序号、列名列表[num1,num2,.....]，仅第一个有效， 负数表示从最后列倒数
    :param targetcolumnlist:           list: 目标列序号、列名列表 [num1, num2, ....]，仅第一个有效， 负数表示从最后列倒数
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 判断输入合法性
    if len(targetcolumnlist[0]) < 1 or len(sourcecolumnlist[0]) < 1:
        return errprintwithname("列{}移动到列{}:数据不合法:列序号范围不正确".format(targetcolumnlist, sourcecolumnlist))

    # 列号转换为正数
    sourcecolumn = sourcecolumnlist[0][0] if sourcecolumnlist[0][0] >= 0 else sourcecolumnlist[0][0] + df.shape[1]
    targetcolumn = targetcolumnlist[0][0] if targetcolumnlist[0][0] >= 0 else targetcolumnlist[0][0] + df.shape[1]
    # 判断数据量是否合法，
    if sourcecolumn not in range(0, df.shape[1]) or targetcolumn not in range(0, df.shape[1]):
        return errprintwithname("实际列{}移动到实际列{}:数据不合法:列序号范围不在有效范围{}内".format(
            sourcecolumn, targetcolumn, df.shape[1]))

    try:
        # 获取当前列名
        currentcolname = df.columns[sourcecolumn]
        # 弹出当前列的值
        currentcolumnvalue = df.pop(currentcolname)
        # 插入当前移动列到指定的列位置（附带列名)
        df.insert(loc=targetcolumn, column=currentcolname, value=currentcolumnvalue)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 清除选中的列
        specialdf.selectedrawcolumn = set()
    except Exception as err:
        return errprintwithname("列{}移动到列{}失败：{}".format(targetcolumnlist, sourcecolumnlist, err))

    return [True, ""]


def dfcolumnstoheadortail(specialdflist: list, tableindex: list,
                          sourcecolumnlist: list, tailflag: True) -> list:
    """
    移动df指定的一批列顺序放到列尾，sourcecolumnlist 采用新column格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param sourcecolumnlist:           list: 源数据列序号、列名列表[num1,num2,.....]，仅第一个有效， 负数表示从最后列倒数
    :param tailflag:            bool: True,移动到列尾，False：移动到列头
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 判断输入合法性
    if len(sourcecolumnlist[0]) < 1:
        return errprintwithname("列{}移动到列{}数据不合法:列序号范围不正确".format(
            sourcecolumnlist, "尾" if tailflag else "头"))

    # 列号转换为正数
    sourcecolumns = [x if x >= 0 else x + df.shape[1] for x in sourcecolumnlist[0]]
    # 判断数据量是否合法，
    if False in [x in range(0, df.shape[1]) for x in sourcecolumns]:
        return errprintwithname("将列{}移动到实际列{}数据不合法:列序号范围不在有效范围{}内".format(
            sourcecolumns, "尾" if tailflag else "头", df.shape[1]))

    sourceclumnnames = [df.columns[x] for x in (sourcecolumns if tailflag else sourcecolumns[::-1])]

    if tailflag:
        targetcolumn = df.shape[1] - 1
    else:
        targetcolumn = 0
    try:
        for currentcolname in sourceclumnnames:
            # 弹出当前列的值
            currentcolumnvalue = df.pop(currentcolname)
            # 插入当前移动列到指定的列位置（附带列名)
            df.insert(loc=targetcolumn, column=currentcolname, value=currentcolumnvalue)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 清除选中的列
        specialdf.selectedrawcolumn = set()
    except Exception as err:
        return errprintwithname("将列{}移动到列{}失败：{}".format(sourcecolumnlist, "尾" if tailflag else "头", err))

    return [True, ""]


def dfcolumncopy(specialdflist: list, tableindex: list,
                 targetcolumns: list, sourcecolumns: list) -> list:
    """
    两表之间列拷贝，快速拷贝， targetcolumn, sourcecolumns采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param targetcolumns:           list: 目标列序号、列名列表,个数与源数据列相同
    :param sourcecolumns:           list: 源数据列号、列名列表,个数与目标列相同
    :return:                    [bool True:成功，False:失败, errstr]
    """

    # 判断是否有两列
    if len(tableindex) >= 1:
        targetdf = specialdflist[tableindex[-1]].dfdata
        sourcedf = specialdflist[tableindex[0]].dfdata
    else:
        return errprintwithname("{} 错误：表格数目不够，无法拷贝列".format(len(tableindex)))

    if targetdf.shape[0] != sourcedf.shape[0]:
        return errprintwithname("对表格{} 目标列{} 源列 {} 拷贝错误：两表格行数不一致".format(
            tableindex, targetcolumns, sourcecolumns))

    try:
        targetdf.iloc[:, targetcolumns[0]] = sourcedf.iloc[:, sourcecolumns[0]]
        # 刷新目标列类型
        specialdflist[tableindex[-1]].refreshcolumntype(columnnamelist=targetdf.columns[targetcolumns[0]])
    except Exception as err:
        return errprintwithname("对表格{} 目标列{} 源列 {} 拷贝错误：{}".format(
            tableindex, targetcolumns, sourcecolumns, err))

    return [True, ""]


def dfexplode(specialdflist: list, tableindex: list, positionlist: list,
              columns=None, ignore_index: bool = True) -> list:
    """
    交换df两个指定列, columns采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:        list： [新生成的表号列表]
    :param columns:             list: 用于处理的列序号列表 ，，只处理第一个[[num1, num2, ....],[name1,name2....]]
    :param ignore_index:        False: 保留原行索引，True：删除原一级索引
    :return:                    [bool True:成功，False:失败, errstr]
    """
    if columns is None:
        columns = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("表格列展开，生成表{}与实际位置{}不符".format(positionlist, len(specialdflist)))

    # 判断是否有两列
    if len(columns[0]) >= 1:
        # 扩展列名
        try:
            specialdflist.append(SpecialDf(dfdata=df.explode(df.columns[columns[0][0]], ignore_index=ignore_index)))
        except Exception as err:
            return errprintwithname("处理列{}错误：{}".format(columns[0], err))
    else:
        return errprintwithname("列数{}不到1个，无法处理".format(columns))

    return [True, ""]


def dfcolumnsum(specialdflist: list, tableindex: list, columns: list, sumcolname: str) -> list:
    """
    对df的指定列求和,columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 用于求和的列序号、列名列表 [[num1, num2, ....],[name1,name2,....]]
    :param sumcolname:              str:  生成的求和列列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        if sumcolname in [None, ""]:
            sumcolname = "列求和结果"
        df[sumcolname] = df[df.columns[columns[0]]].sum(axis=1)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=[sumcolname])
    except Exception as err:
        return errprintwithname("求和:{} 生成列:{} 错误:{}".format(columns, sumcolname, err))

    return [True, ""]


def dfcolumnshift(specialdflist: list, tableindex: list, columns: list, shiftrows=1,
                  circleflag=False, fillnan=None, newcolname="移位列名" ) -> list:
    """
    对df的指定列进行上下行移位，并对
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
                                          本命令仅仅第一个有效
    :param columns:                 list: 用于合并的列序号、列名列表 [[num1, num2, ....],[name1,name2,....]]，
                                         本命令仅仅第一个有效
    :param shiftrows:               int： 移位的行数，可以为正、负数
    :param circleflag:              bool：是否将猎头列尾衔接进行列值循环移位，
    :param fillnan:                 circleflag为False有效，移位空出的行用于填充的值
    :param newcolname:              str: 移位后输出结果列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if shiftrows == 0:
        # 移动行数为0， 如果新列名与源列名同名，直接返回，不相同再拷贝再返回（相当于复制列）
        if newcolname == df.columns[columns[0][0]]:
            return [True, ""]
        else:
            try:
                df[newcolname] = df[df.columns[columns[0][0]]]
                return [True, ""]
            except Exception as err:
                return errprintwithname("列移位:{} 生成列:{} 错误:{}".format(columns, newcolname, err))

    try:
        # 根据circleflag确定整个要移位的列值（将列值list做扩展）然后截取列长复制（此处未使用dataframe shift
        # 取列值list
        templist = list(df[df.columns[columns[0][0]]])
        # 获取要填充的值list
        # 当shiftrows大于行数时，circle方式下，只需要循环的实际行数
        shiftrows = shiftrows if not circleflag else shiftrows % df.shape[0]
        # 不循环时，直接填充fillna，正循环，取尾部行，负循环，取头部行
        addlist = [fillnan for _ in range(abs(shiftrows))] if not circleflag \
            else templist[-shiftrows:] \
            if shiftrows > 0 else templist[0: abs(shiftrows)]
        # 正循环补充值放置在猎头，否则追加到列尾
        resultlist = addlist + templist if shiftrows > 0 else templist + addlist
        # 截取列长给新列赋值
        df[newcolname] = resultlist[0: df.shape[0]] if shiftrows > 0 else resultlist[-df.shape[0]:]
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=[newcolname])
    except Exception as err:
        return errprintwithname("列移位:{} 生成列:{} 错误:{}".format(columns, newcolname, err))

    return [True, ""]


def dfcolumncombine(specialdflist: list, tableindex: list, columns: list,
                    combinecolname: str, seperate='\n', nanreplace='-') -> list:
    """
    对df的指定列进行字符串合并,columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 用于合并的列序号、列名列表 [[num1, num2, ....],[name1,name2,....]]
    :param combinecolname:          str:  生成的合并列列名
    :param seperate:                str:  合并列之间的分隔符
    :param nanreplace:              str:  nan的替换符号
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if len(columns[0]) == 0:
        return errprintwithname("合并:{} 生成列:{} 错误:未指定合并列".format(columns, combinecolname))
    # df1['combine1'] = df1[df1.columns[0]].str.cat(df1[df1.columns[[1]]], sep='\n', na_rep='-')
    try:
        if combinecolname in [None, ""]:
            combinecolname = "列合并结果"
        df[combinecolname] = df[df.columns[columns[0][0]]].str.cat(df[df.columns[columns[0][1:]]],
                                                                   sep=seperate, na_rep=nanreplace)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=[combinecolname])

    except Exception as err:
        return errprintwithname("合并:{} 生成列:{} 错误:{}".format(columns, combinecolname, err))

    return [True, ""]


def dfcolumnsubtract(specialdflist: list, tableindex: list, columns: list, subcolname: str) -> list:
    """
    对df的指定两列求差,columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 用于求差的列序号、列名列表 [[num1, num2],[name1,name2]]
    :param subcolname:              str:  生成的求差列列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 列相减至少应该有两列（本处理只适用前两列）
    if len(columns[0]) >= 2:
        try:
            if subcolname in [None, ""]:
                subcolname = "列求差结果"
            df[subcolname] = list(map(lambda x, y: (x - y),
                                      df.iloc[:, columns[0][0]], df.iloc[:, columns[0][1]]))
            specialdf.headings = list(df)
            specialdf.colnamelist = specialdf.headings
            specialdf.colnameusedlist = [""] + specialdf.headings
            specialdf.colnamefilterlist = [""] + specialdf.headings
            # 刷新列类型
            specialdf.refreshcolumntype(columnnamelist=[subcolname])
        except Exception as err:
            return errprintwithname("{}求差生成列:{} 错误:{}".format(columns, subcolname, err))
    else:
        return errprintwithname("{}列求差错误：列数不到2个，无法求差".format(columns))

    return [True, ""]


def dfcolumnspecialprocess(specialdflist: list, tableindex: list, columns: list,
                           newcolumnname: str, processstr: str) -> list:
    """
    对df的指定列进行特殊处理生成新列，特殊处理为lambda函数字符串或其他处理函数字符串,columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                         指定进行处理的列序号、列名list
    :param newcolumnname:                   生成的新列名
    :param processstr:                      用于处理函数或lambda函数字符串
    :return:                                [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # commandstr = 'list(map({}, df[df.columns[{}]].values))'.format(processstr, columns)
    commandstr = 'list(map({}, df.iloc[:, {}].values))'.format(processstr, columns[0])
    try:
        print("dfcolumnspecialprocess 处理：{} ...".format(commandstr))
        if newcolumnname in [None, ""]:
            newcolumnname = "列处理结果"
        df[newcolumnname] = eval(commandstr)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=[newcolumnname])
    except Exception as err:
        return errprintwithname("处理{} 生成列:{} 错误:{}".format(commandstr, newcolumnname, err))

    return [True, ""]


def dfcolumnsdel(specialdflist: list, tableindex: list, columns: list) -> list:
    """
    删除df指定的列, columns采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 需要删除的列序号、列名列表 [[num1, num2, ....],[name1,name2....]]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        df.drop(df.columns[columns[0]], axis=1, inplace=True)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 清除选中的列
        specialdf.selectedrawcolumn = set()
        # 清除删除列的列stryle
        for columnname in columns[1]:
            if columnname in specialdf.columnstyleinfo.keys():
                specialdf.columnstyleinfo.pop(columnname)
    except Exception as err:
        return errprintwithname("删除列:{}错误:{}".format(columns, err))
    return [True, ""]


def dfcolumnsclear(specialdflist: list, tableindex: list, columns: list) -> list:
    """
    清空df指定的列, 赋值为np.nan
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 需要清空的列序号、列名列表 [[num1, num2, ....],[name1,name2....]]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        df[df.columns[columns[0]]] = df[df.columns[columns[0]]].applymap(lambda x: np.nan)
        # 不做列的类型刷新，保持原有类型
    except Exception as err:
        return errprintwithname("清空列:{}错误:{}".format(columns, err))
    return [True, ""]


def dfcolumnsmarkdel(specialdflist: list, tableindex: list,
                     markrows=None, delmarkvalue: bool = True) -> list:
    """
    删除df指定的列
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param markrows:                list: 删除列使用的标记行列表，只有第一项有效， -1表示最后一行
    :param delmarkvalue:            bool: True/False
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if markrows is None:
        markrows = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        rawmarkrows = df.shape[0] + markrows[0] if markrows[0] < 0 else markrows[0]
        if rawmarkrows not in range(0, df.shape[0]):
            return errprintwithname("标记行{}超出表格范围！".format(markrows))
        # 获取行值
        markrowcontentlist = df.iloc[rawmarkrows].values.tolist()
        # 首先通过列表转set，去掉重复值，新列表中只包含要删除的列和None
        delcolumnlist = list(set([x if y == 1 else None for x, y in enumerate(markrowcontentlist)]))
        delcolumnnames = df.columns[delcolumnlist]
        # 列表中删除None，剩下要删除的列清单
        delcolumnlist.remove(None)
        # 获取标记对应的列序号list
        df.drop(df.columns[delcolumnlist], axis=1, inplace=True)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 清除选中的列
        specialdf.selectedrawcolumn = set()
        # 清除删除列的列stryle
        for columnname in delcolumnnames:
            if columnname in specialdf.columnstyleinfo.keys():
                specialdf.columnstyleinfo.pop(columnname)
    except Exception as err:
        return errprintwithname("按标记行{}标记值{}删除列错误:{}".format(markrows, delmarkvalue, err))
    return [True, ""]


def dfextractcolumns(specialdflist: list, tableindex: list, columns: list) -> list:
    """
    抽取df指定的列, columns 采用新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 需要抽取保留的列序号，列名列表 [[num1, num2, ....],[name1,name2.....]]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        delcolumnslist = [x for x in list(range(len(df.columns))) if x not in columns[0]]
        delcolumnsname = df.columns[delcolumnslist]
        df.drop(df.columns[delcolumnslist], axis=1, inplace=True)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 清除选中的列
        specialdf.selectedrawcolumn = set()
        # 清除删除列的列stryle
        for columnname in delcolumnsname:
            if columnname in specialdf.columnstyleinfo.keys():
                specialdf.columnstyleinfo.pop(columnname)
    except Exception as err:
        return errprintwithname("抽取列:{}错误:{}".format(columns, err))
    return [True, ""]


def dfcolumnrename(specialdflist: list, tableindex: list, headings: list) -> list:
    """
    重命名df的列名
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param headings:                list: 新的列名列表 [num1, num2, ....] ， 和df的列数要相同
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if len(df.columns) != len(headings) or len(df.columns) != len(set(headings)):
        return errprintwithname("设置列名错误:列名{}个数{}与表列数{}不一致或重名！".format(
            headings, len(headings), len(df.columns)))
    try:
        df.columns = headings
        specialdf.headings = headings
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 列名变更，需要改变key值 TODO， 暂时刷新所有type
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=None)
    except Exception as err:
        return errprintwithname("设置列名:{}错误:{}".format(headings, err))
    return [True, ""]


def dfcolumntypeset(specialdflist: list, tableindex: list, columns: list, columntype: str) -> list:
    """
    设置df指定列的类型, columns修改为新的格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 指定类型的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param columntype:              str： 指定的类型
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    if columntype not in list(VALIE_DATA_TYPE.values()):
        return errprintwithname("列:{}, 列类型：{}，类型不在{}合法范围内！".format(
            columns, columntype,  list(VALIE_DATA_TYPE.values())))
    try:
        if columntype in ["datetime", "date"]:
            df[df.columns[columns[0]]] = df[df.columns[columns[0]]].apply(pd.to_datetime)
        elif columntype in ["str2object"]:
            if df.shape[0] > 0:
                # 必须是str类型才能做ast.literal_eval转换
                # 同时处理多列，首先判断列类型是否都是string
                temptype = [isinstance(x, str) for x in df.iloc[0, columns[0]]]
                if False not in temptype:
                    # 类型符合要求，都是string，分别调用literal_eval转换
                    df[df.columns[columns[0]]] = df[df.columns[columns[0]]].applymap(ast.literal_eval)
                else:
                    return errprintwithname("{}列类型string判断：{}不符合全True".format(columns, temptype))
            else:
                return errprintwithname("{}列无有效数据，无法转换".format(columns))
        else:
            df[df.columns[columns[0]]] = df[df.columns[columns[0]]].astype(columntype)
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=df.columns[columns[0]])
    except Exception as err:
        return errprintwithname("{}列类型设置为：{} 错误:{}".format(columns, columntype, err))
    return [True, ""]


def dfsortbycolumns(specialdflist: list, tableindex: list,
                    columns: list, hancolumns: list = None,
                    ascending: bool or list = None,
                    hanascending: bool or list = None) -> list:
    """
    把df按指定索引列排序, columns修改为新的格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columns:                 list: 指定的排序索引列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param hancolumns:              list: 指定按汉字排序索引列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param ascending:               bool | list:  True 正序，False：倒序, |[True, False.....] 个数必须和columns一致
    :param hanascending:            bool | list:  True 正序，False：倒序, |[True, False.....] 个数必须和hancolumns一致
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if hanascending is None:
        hanascending = []
    if ascending is None:
        ascending = []
    if ascending in [True, False]:
        ascending = [ascending for _ in range(len(columns[0]))]
    if hanascending in [True, False]:
        hanascending = [hanascending for _ in range(len(hancolumns[0]))]
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        if isinstance(ascending, list) and isinstance(hanascending, list):
            if len(ascending) + len(hanascending) != len(columns[0]) + len(hancolumns[0]):
                return errprintwithname("列:{}+{}排序, 顺序：{}+{}，错误:列数与排序规则数不一致".format(
                    columns, hancolumns, ascending, hanascending))
        else:
            return errprintwithname("列排序顺序：{}+{}，错误:顺序必须全部为列表方式".format(ascending, hanascending))
        # 如果有汉字列，需要单独转换为对应的拼音列再统一排序
        hancurrentcolumns = []
        for hanitem in hancolumns[1]:
            newcolumnprefix = hanitem + "_pinyin"
            df[newcolumnprefix] = df[hanitem].apply(lambda x: "".join(lazy_pinyin(x)))
            hancurrentcolumns.append(newcolumnprefix)
        # 使用新列排序(非汉字列+新生成的汉字列)
        df.sort_values(by=list(df.columns[columns[0]]) + hancurrentcolumns,
                       inplace=True,
                       ascending=ascending + hanascending)
        # 排序后删除拼音列
        df.drop(columns=hancurrentcolumns, inplace=True)
        # 需要处理排序后rowstyleinfo的对应变化
        # 将rowstyleinfo 重新排序
        newrowstyleinfo = [specialdf.rowstyleinfo[x] for x in df.index.tolist()]
        specialdf.rowstyleinfo = newrowstyleinfo
        # 删除原行索引并重建
        df.reset_index(inplace=True, drop=True)
    except Exception as err:
        return errprintwithname("列:{}+{}排序, 顺序：{}+{}，错误:{}".format(
            columns, hancolumns, ascending, hanascending, err))
    return [True, ""]


def dfnewcolumn(specialdflist: list, tableindex: list,
                columnname: str = '新列', dtype: str = 'str', columnfillvalue: str = '') -> list:
    """
    把df按指定索引列排序
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columnname:              str: 新增列的列名
    :param dtype:                   str: 新增列的类型
    :param columnfillvalue:         str： 新增列的填充值
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        if columnname in [None, ""]:
            columnname = "新列名"
        df[columnname] = [columnfillvalue for _ in range(df.shape[0])]
        df[columnname] = df[columnname].astype(dtype)
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新列类型
        specialdf.refreshcolumntype(columnnamelist=[columnname])
    except Exception as err:
        return errprintwithname("按列名:{}及列类型{}, 列填充值：{}，错误:{}".format(columnname, dtype, columnfillvalue, err))
    return [True, ""]


def dfwidthwisejoint(specialdflist: list, tableindex: list, positionlist: list) -> list:
    """
    横向拼接多个df，并重建索引
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个

    :param positionlist:    list： [新生成的表号列表]
    :return:                [bool True:成功，False:失败, errstr]
    """
    # global specialdflist
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("合并df列表新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))
    # 当tableindex为空时，默认拼接所有表格
    if tableindex in [[], None]:
        tableindex = [x for x in range(len(specialdflist))]
    try:
        dfcurrentlist = [specialdflist[x].dfdata for x in tableindex]
        specialdflist.append(SpecialDf(dfdata=pd.concat(dfcurrentlist, axis=1)))
    except Exception as err:
        return errprintwithname("合并df列表插入位置：{}，错误:{}".format(positionlist, err))
    return [True, ""]


def dflongitudinaljoint(specialdflist: list, tableindex: list, positionlist: list) -> list:
    """
    纵向拼接多个df，并重建索引
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:    list： [新生成的表号列表]
    :return:                [bool True:成功，False:失败, errstr]
    """
    # global specialdflist
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("合并df列表新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))
    # 当tableindex为空时，默认拼接所有表格
    if tableindex in [[], None]:
        tableindex = [x for x in range(len(specialdflist))]
    try:
        dfcurrentlist = [specialdflist[x].dfdata for x in tableindex]
        specialdflist.append(SpecialDf(dfdata=pd.concat(dfcurrentlist, ignore_index=True)))
    except Exception as err:
        return errprintwithname("合并df列表插入位置：{}，错误:{}".format(positionlist, err))
    return [True, ""]


def dfmerge(specialdflist: list, tableindex: list, positionlist: list, how: str = 'outer') -> list:
    """
    用merge方式融合两张df表格
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list:   生成的表格的位置list
    :param how:                     str： join的方式字符串，用于merge函数
    :return:                        [bool True:成功，False:失败, errstr]
    """
    # global specialdflist
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("合并df列表新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    try:
        dfcurrentlist = [specialdflist[x].dfdata for x in tableindex]
        specialdflist.append(SpecialDf(dfdata=pd.merge(dfcurrentlist[0], dfcurrentlist[1], how=how)))
    except Exception as err:
        return errprintwithname("合并df列表新表序号：{}，错误:{}".format(positionlist, err))
    return [True, ""]


def dfpivot(specialdflist: list, tableindex: list, positionlist: list, index: list = None,
            columns: list = None, values: list = None, ) -> list:
    """
    表格行列转置产生新表, columns values index采用新的columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param index:                   list: 指定的索引列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param columns:                 list: 指定的说明列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param values:                  list: 指定的转置值列的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]

    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("转置df行列新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    myindex = [df.columns[x] for x in index[0]] if index is not None else None
    mycolumns = [df.columns[x] for x in columns[0]] if columns is not None else None
    myvalues = [df.columns[x] for x in values[0]] if values is not None else None
    try:
        dfresult = df.pivot(index=myindex,
                            columns=mycolumns,
                            values=myvalues)
        # 将索引列变成首列
        dfresult.insert(0, "/".join(myindex), dfresult.index)
        # 将行index变成序号
        specialdflist.append(SpecialDf(dfdata=dfresult.reset_index(drop=True)))

    except Exception as err:
        return errprintwithname("转置df行列新表序号：{}，index:{}, columns:{}, values:{} 错误:{}".format(
            positionlist, index, columns, values, err))
    return [True, ""]


def dfmelt(specialdflist: list, tableindex: list, positionlist: list, id_vars=None,
           value_vars=None, var_name=None, value_name="value") -> list:
    """
    表格行列转置产生新表
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param id_vars:                 list: 指定的索引列序号列表 [c1, c2, ....]
    :param value_vars:              list: 指点的转置值列的列序号列表 [v1, v2, ....]
    :param var_name:                str:  生成的说明列的列名
    :param value_name:              str:  生成的值列的列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("列转置行df列新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    try:
        specialdflist.append(SpecialDf(dfdata=df.melt(id_vars=df.columns[id_vars[0]].tolist(),
                                                      value_vars=df.columns[value_vars[0]].tolist(),
                                                      var_name=var_name, value_name=value_name)))
    except Exception as err:
        return errprintwithname("行列转置df新表序号：{}，id_vars:{}, value_vars:{}, var_name:{} value_name:{}错误:{}".format(
            positionlist, id_vars, value_vars, var_name, value_name, err))
    return [True, ""]


def dfswapposition(specialdflist: list, tableindex: list) -> list:
    """
    用merge方式融合两张df表格
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :return:                        [bool True:成功，False:失败, errstr]
    """
    # global specialdflist
    # 插入的位置不在specialdflist表尾
    if len(tableindex) != 2:
        return errprintwithname("交换df列表{}位置不成功, 输入表格数不符合要求".format(len(tableindex)))

    if tableindex[0] >= len(specialdflist) or tableindex[1] >= len(specialdflist):
        return errprintwithname("交换df列表{}位置不成功, 输入表格不在表格清单{}".format(tableindex, len(specialdflist)))

    try:
        specialdflist[tableindex[0]], specialdflist[tableindex[1]] = \
            specialdflist[tableindex[1]], specialdflist[tableindex[0]]
    except Exception as err:
        return errprintwithname("交换df列表{}位置错误:{}".format(tableindex, err))
        
    return [True, ""]


def dfdroptables(specialdflist: list, tableindex: list) -> list:
    """
    删除指定的表格
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if len(tableindex) == 0:
        return [True, ""]

    if False in [0 <= x < len(specialdflist) for x in tableindex]:
        return errprintwithname("dfdroptables 删除列表{}位置超出现有范围{}！".format(tableindex, [0, len(specialdflist)]))
    try:
        # 将dfcurrentlist 转换为set减少同名项，然后list按从大到小排序，specialdflist中pop，避免先弹出小序号影响后续删除序号
        templist = list(set(tableindex))
        templist.sort(reverse=True)
        for i in templist:
            specialdflist.pop(i)
    except Exception as err:
        return errprintwithname("删除表{} :{}".format(tableindex, err))
    return [True, ""]


def dfblockcopy(specialdflist: list, tableindex: list, targetblock: list, sourceblock: list) -> list:
    """
    交换表格位置
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param targetblock:             list: [[startrow, startcolumn],[endrow, endcolumn]]
    :param sourceblock:             list: [[startrow, startcolumn],[endrow, endcolumn]]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    # 判断是表内拷贝还是表间拷贝
    if len(tableindex) == 1:
        sourcedf = targetdf = specialdflist[tableindex[0]].dfdata
    else:
        sourcedf = specialdflist[tableindex[0]].dfdata
        targetdf = specialdflist[tableindex[1]].dfdata

    # 结束点坐标+1， 因为iloc时坐标右侧半封闭
    sourcestartrow, sourcestartcolumn = sourceblock[0][0], sourceblock[0][1]
    sourceendrow, sourceendcolumn = sourceblock[1][0] + 1, sourceblock[1][1] + 1
    targetstartrow, targetstartcolumn = targetblock[0][0], targetblock[0][1]
    targetendrow, targetendcolumn = targetblock[1][0] + 1, targetblock[1][1] + 1

    if sourceendrow == 0 or sourceendrow > sourcedf.shape[0]:  # 填写的是-1， 表示最大行数
        sourceendrow = sourcedf.shape[0]
    if targetendrow == 0 or targetendrow > targetdf.shape[0]:
        targetendrow = targetdf.shape[0]

    # 区域大小是否一致
    # 区域是否单点
    targetheight, targetwidth = (targetendrow - targetstartrow), (targetendcolumn - targetstartcolumn)
    sourceheight, sourcewidth = (sourceendrow - sourcestartrow), (sourceendcolumn - sourcestartcolumn)

    if targetwidth < sourcewidth or targetheight < sourceheight:
        return errprintwithname("目标区域：{}比源区域：{}范围小！".format(targetblock, sourceblock))

    # 判读区域的合法性（边长是否有负值）
    # if targetwidth < 0 or targetheight < 0 or sourceheight < 0 or sourcewidth < 0:
    validflag = [False if x <= 0 else True for x in [targetheight, targetwidth, sourceheight, sourcewidth]]
    if False in validflag:
        return errprintwithname("目标区域：{}和源区域：{} 不合法（部分边长为负值）！".format(targetblock, sourceblock))

    try:
        targetdf.iloc[targetstartrow: targetendrow, targetstartcolumn: targetendcolumn] = \
            sourcedf.iloc[sourcestartrow: sourceendrow, sourcestartcolumn: sourceendcolumn]
        # 刷新列类型
        specialdflist[tableindex[1]].refreshcolumntype(
            columnnamelist=targetdf.columns[targetstartcolumn: targetendcolumn])
    except Exception as err:
        return errprintwithname("目标区域：{}和源区域：{} 失败：{}！".format(targetblock, sourceblock, err))

    return [True, ""]


def dfspecailcolumncopy(specialdflist: list, tableindex: list, targetblock: list,
                        sourceblock: list, indexcolumn: list = None) -> list:
    """
    列拷贝，可以根据是否有索引列采用不同方式拷贝
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param targetblock:             list: [[startrow, startcolumn],[endrow, endcolumn]]
    :param sourceblock:             list: [[startrow, startcolumn],[endrow, endcolumn]]
                                    targetblock 和 sourceblock的列数必须都是1， 行数源数据行数不能大于目标行数
    :param indexcolumn:              list: [源表索引列, 目的表索引列]
    :return:                        [bool True:成功，False:失败, errstr]
    """
    # 如果没有indexcolumn，直接调用dfblockcopy
    dfcurrentspeciallist = [specialdflist[x] for x in tableindex]
    if indexcolumn is None:
        return dfblockcopy(specialdflist, tableindex, targetblock, sourceblock)
    indexnum = len(indexcolumn)
    if indexnum == 0:
        return dfblockcopy(specialdflist, tableindex, targetblock, sourceblock)

    if indexnum != 0 and indexnum != 2:
        return errprintwithname("索引列{}不合法！".format(indexcolumn))

    if indexnum == 2:
        sourceindex, targetindex = indexcolumn[0], indexcolumn[1]
    else:
        sourceindex, targetindex = None, None

    # 判断是表内拷贝还是表间拷贝
    if len(dfcurrentspeciallist) == 1:
        sourcedf = targetdf = dfcurrentspeciallist[0].dfdata
    else:
        sourcedf = dfcurrentspeciallist[0].dfdata
        targetdf = dfcurrentspeciallist[1].dfdata

    # 结束点坐标+1， 因为iloc时坐标右侧半封闭
    sourcestartrow, sourcestartcolumn = sourceblock[0][0], sourceblock[0][1]
    sourceendrow, sourceendcolumn = sourceblock[1][0] + 1, sourceblock[1][1] + 1
    targetstartrow, targetstartcolumn = targetblock[0][0], targetblock[0][1]
    targetendrow, targetendcolumn = targetblock[1][0] + 1, targetblock[1][1] + 1

    if sourceendrow == 0 or sourceendrow > sourcedf.shape[0]:  # 填写的是-1， 表示最大行数
        sourceendrow = sourcedf.shape[0]
    if targetendrow == 0 or targetendrow > targetdf.shape[1]:
        targetendrow = targetdf.shape[0]

    # 区域大小是否一致
    # 区域是否单点
    targetheight, targetwidth = (targetendrow - targetstartrow), (targetendcolumn - targetstartcolumn)
    sourceheight, sourcewidth = (sourceendrow - sourcestartrow), (sourceendcolumn - sourcestartcolumn)

    # 判断是否都是单列
    if targetwidth != 1 or sourcewidth != 1:
        return errprintwithname("源数据列数{} 或目的数据列数{}不为1！".format(sourcewidth, targetwidth))

    # 判读区域的合法性（边长是否有负值）
    # if targetwidth < 0 or targetheight < 0 or sourceheight < 0 or sourcewidth < 0:
    validflag = [False if x <= 0 else True for x in [targetheight, targetwidth, sourceheight, sourcewidth]]
    if False in validflag:
        return errprintwithname("目标区域：{}和源区域：{} 不合法（部分边长为负值）！".format(targetblock, sourceblock))

    # 获取源数据的相关df
    try:
        sourcefilldf = sourcedf.drop(
            sourcedf.columns[[x for x in list(range(len(sourcedf.columns)))
                              if x not in [sourceindex, sourcestartcolumn]]], axis=1)
        # 单独目标索引列df，以及list
        targetfilldf = targetdf.drop(
            targetdf.columns[[x for x in list(range(len(targetdf.columns))) if x not in [targetindex]]], axis=1)
        # targetfilldf = targetdf.iloc[:, targetindex]
        targetindexlist = targetfilldf.iloc[:, 0].to_list()
    except Exception as err:
        return errprintwithname("获取源或者目标数据索引及数据列错误:{}".format(err))

    # 根据索引号和起止行号范围进行赋值(在指定的开始、结束行号范围内）
    try:
        for i in range(sourcestartrow, sourceendrow):
            targetrow, sourcevalue = targetindexlist.index(sourcefilldf.iloc[i, 0]), sourcefilldf.iloc[i, 1]
            # 在目标开始、结束行号内
            if targetstartrow <= targetrow <= targetendrow:
                targetdf.iloc[targetrow, targetstartcolumn] = sourcevalue
    except Exception as err:
        return errprintwithname("数据拷贝过程中错误:{}".format(err))

    return [True, ""]


def dfcolumnindexfill(specialdflist: list, tableindex: list,
                      targetcolumn: list, sourcecolumn: list = None,
                      targetindex: list = None, sourceindex: list = None, fillvalue=None) -> list:
    """
    列值索引查询填充，targetcolumn, sourcecolumn, targetindex,sourceindex 修改为column新格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param targetcolumn:             list: [[column],[columnname]] , 需要填充的列，list的第0项
    :param sourcecolumn:             list: [[column],[columnname]], 数据源表中的数值列
    :param targetindex:              list: 目标区域的索引列，可以多列[[索引列1， 索引列2.....],[索引列名1， 索引列名2.....]]
    :param sourceindex:              list: 源数据区域的索引列，可以多列[[索引列1， 索引列2.....],[索引列名1， 索引列名2.....]]，
                                        必须和目标列索引数及顺序一致，
    :param fillvalue:                list: 查不到索引数据时用来填充的数值
    :return:                         [bool True:成功，False:失败, errstr]

    """
    # 检查参数合法性
    if targetindex is None or sourceindex is None:
        return errprintwithname("目标索引列{} 或者源数据索引列{}不合法！".format(targetindex, sourceindex))
    if len(targetindex[0]) != len(sourceindex[0]) or len(targetindex[0]) == 0 or len(sourceindex[0]) == 0:
        return errprintwithname("目标索引列{} 或者源数据索引列{}列数不一致！".format(targetindex, sourceindex))

    # 判断是表内拷贝还是表间拷贝
    sourcedf = specialdflist[tableindex[0]].dfdata
    targetdf = specialdflist[tableindex[-1]].dfdata

    # 判断目标列是否存在，不存在则创建目标列，并且修改目标列的索引号
    for index, columnid in enumerate(targetcolumn[0]):
        if columnid >= targetdf.shape[1]:
            # 当前targetcolumnid不存在，需要创建
            newcolumnid = targetdf.shape[1]
            newcolumnname = None
            tempname = sourcedf.columns[sourcecolumn[0][index]]
            # 如果源数据列名在目标表格列中已经存在，则更改列名，否则使用源数据列名
            if tempname not in targetdf.columns:
                newcolumnname = tempname
            else:
                for i in range(100):
                    if tempname + str(i) not in targetdf.columns:
                        newcolumnname = tempname + str(i)
                        break
            # 使用1-100作为后缀创建新列都失败，返回失败
            if newcolumnname is None:
                return errprintwithname("目标列序号{}不存在，重新创建不成功！".format(columnid))

            print("dfcolumnindexfill的目标列序号{}不存在，重新创建为列名{}：列序号{}".format(
                columnid, newcolumnid, newcolumnname))
            dfnewcolumn(specialdflist, tableindex[-1:], columnname=newcolumnname)
            targetcolumn[0][index] = newcolumnid

    # 获取源数据的相关df
    try:
        # 将源数据索引值、value转换成字典格式
        # 使用多列index的值的list转换为str后作为字典项
        sourcedict = dict(zip(
            [str(sourcedf[sourcedf.columns[sourceindex[0]]].iloc[x].tolist()) for x in range(sourcedf.shape[0])],
            [sourcedf[sourcedf.columns[sourcecolumn[0]]].iloc[x].tolist() for x in range(sourcedf.shape[0])]))

        # 逐行根据索引填写值
        # 抽取目标的索引列
        fillvaluelist = [fillvalue for _ in range(len(targetcolumn[0]))]
        # 使用listmap，提高整体执行效率
        targetdf[targetdf.columns[targetcolumn[0]]] = list(map(
            lambda x: sourcedict.get(str(x.tolist()), fillvaluelist),
            targetdf[targetdf.columns[targetindex[0]]].values))
        # 刷新目标列类型
        specialdflist[tableindex[-1]].refreshcolumntype(columnnamelist=targetdf.columns[targetcolumn[0]])
    except Exception as err:
        return errprintwithname("获取源或者目标数据索引及数据列错误:{}".format(err))

    return [True, ""]


def dfaddserialno(specialdflist: list, tableindex: list,
                  columnname: str = '序号', startnumber: int = 0) -> list:
    """
    为列举的df分别插入序号列
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param columnname:              str : 新序号列的名称
    :param startnumber:             int:  第一行开始的序号
    :return:                        [bool True:成功，False:失败, errstr]
    """
    # 根据索引号和起止行号范围进行赋值(在指定的开始、结束行号范围内）
    try:
        for x in tableindex:
            specialdf = specialdflist[x]
            df = specialdf.dfdata
            df.insert(loc=0, column=columnname, value=[x + startnumber for x in range(df.shape[0])])
            specialdf.headings = list(df)
            specialdf.colnamelist = specialdf.headings
            specialdf.colnameusedlist = [""] + specialdf.headings
            specialdf.colnamefilterlist = [""] + specialdf.headings
            # 刷新目标列类型
            specialdf.refreshcolumntype(columnnamelist=[columnname])
    except Exception as err:
        return errprintwithname("插入序号列过程中错误:{}".format(err))

    return [True, ""]


def dflabelfilter(specialdflist: list, tableindex: list, positionlist: list,
                  items=None, like=None, regex=None, axis=None) -> list:
    """
    用dataframe filter(行列标签)方式筛选df
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param items:                   对行/列按序号进行筛选
    :param like:                    用like进行筛选
    :param regex:                   regex表示用正则进行匹配
    :param axis:                    axis=0表示对行操作，axis=1表示对列操作
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("过滤表形成新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    try:
        specialdflist.append(SpecialDf(dfdata=df.filter(items=None if items in [None, []] else items,
                                                        like=None if like in [None, "", "None", "none"] else like,
                                                        regex=None if regex in [None, "", "None", "none"] else regex,
                                                        axis=axis)))
    except Exception as err:
        return errprintwithname("item:{} like:{} regex:{} axis(0：行过滤，1：列过滤:{}过滤df列表新表序号：{}，错误:{}".format(
            items, like, regex, axis, positionlist, err))

    return [True, ""]


AGG_FUNCTION_LIST = ["sum", "mean", "max", "min", "count"]
# AGG_FUNCTION_LIST = ["sum", "mean", "max", "min", "count", 'std', 'median', 'abs', 'prod', 'cumsum', 'cumprod']


def dfgroupbyagg(specialdflist: list, tableindex: list,
                 indexcolumns: list = None, aggcolumnlist: list = None,
                 aggfunclist: (list, str) = None, columnnamelist: list = None) -> list:
    """
    根据源df和指定的groupby列、生成求和列，在该组所有行均填写该值, 采用新columns格式，可以同时对多个列分别使用不同的聚合函数，生成各自的列
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]], 为None表示全表聚合，不分组
    :param aggcolumnlist:           list: [[被聚合的列序号]，[被聚合的列名]]，None表示所有列
    :param aggfunclist:             list|str: 聚合的具体算法”sum“，”mean“，“np.min", "lambda x: x.mean().....
                                    aggfunc全部用str传入，可以是匿名函数或则内置聚合函数，np.min等等，用eval转化为实际函数
                                    str: 全部使用相同聚合函数
                                    list时[1]个数必须与aggcolumnlist[1]一致
    :param columnnamelist:          list:  str: 求和列生产的新列名
                                    当None，时，系统自动根据"被聚合列名_聚合函数“生成新列名
                                    list时需要和aggcolumnlist[1]个数一致
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 根据默认参数形成初始参数
    # 聚合算法合法性检查
    paramflag = True
    # 如果在内置函数列表中，可以直接串调用
    if aggfunclist is None or aggcolumnlist is None:
        paramflag = False
    else:
        myaggcolumnlist = aggcolumnlist[0]
        # 参数为单个函数时所有的agg列使用相同的函数
        myaggfunclist = [aggfunclist for _ in range(len(myaggcolumnlist))] \
            if isinstance(aggfunclist, str) else aggfunclist
        # 输入name为空时，自动形成相关列名
        if len(myaggfunclist) != len(myaggcolumnlist):
            paramflag = False
        else:
            # 形成输出列名
            mycolumnnamelist = [
                "".join([aggcolumnlist[1][x], "_", str(myaggfunclist[x])]) for x in range(len(myaggfunclist))] \
                if columnnamelist is None else columnnamelist
            if len(mycolumnnamelist) != len(myaggcolumnlist):
                paramflag = False

    if paramflag is False:
        return errprintwithname("聚合列{}与聚合{}方案、输出列{}数目个数不一致！".format(
            aggcolumnlist, aggfunclist, columnnamelist))

    # 检查聚合函数合法性， 内置函数直接跳过，非内置用eval+try检查
    # 对于非内置函数
    for currentaggfunc in myaggfunclist:
        if currentaggfunc not in AGG_FUNCTION_LIST:
            try:
                eval(currentaggfunc)
            except Exception as err:
                return errprintwithname("对列{}按索引{}分组聚合{}方案不合法：{}".format(
                    aggcolumnlist, indexcolumns, columnnamelist, currentaggfunc, err))

    try:
        # 组合agg字典
        aggdict = dict(zip(df.columns[myaggcolumnlist], myaggfunclist))

        # if indexcolumns is None or len(indexcolumns[0]) == 0:            # 全表聚合
        if indexcolumns in [None, [], [[], []], [[]]]:  # 全表聚合
            currentresult = df.agg(aggdict)
            # 获取groupby当前的index组
            # 该group的index用来给df的对应行赋值
            df.loc[:, mycolumnnamelist] = [currentresult for _ in range(df.shape[0])]
        else:                               # 多列索引groupby再聚合
            groups = df.groupby(df.columns[indexcolumns[0]].to_list())
            # 每组单独求sum然后给改组所有行赋该sum值
            for groupitem in groups:
                # 获取当前group的aggfunclist， aggfunclist全部用str传入，
                # 如果是内置函数，直接调用串，否则可以是匿名函数或则内置聚合函数，np.min等等，用"eval转化为实际函数
                currentresult = groupitem[1].agg(aggdict)
                # 获取groupby当前的index组
                # 该group的index用来给df的对应行赋值
                df.loc[groupitem[1].index, mycolumnnamelist] = [currentresult for _ in range(groupitem[1].shape[0])]
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=mycolumnnamelist)
    except Exception as err:
        return errprintwithname("对列{}按索引{}分组聚合{}到列：{} 失败：{}".format(
            aggcolumnlist, indexcolumns, aggfunclist, columnnamelist, err))

    return [True, ""]


def dfgroupbysum(specialdflist: list, tableindex: list,
                 indexcolumns: list, sumcolumn=None, columnname: str = "新列") -> list:
    """
    根据源df和指定的groupby列、生成求和列，在该组所有行均填写该值, indexcolumns, sumcolumn采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]]
    :param sumcolumn:               list: [[被求和的列序号]，[被求和的列序名]]
    :param columnname:              str: 求和列生产的新列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if sumcolumn is None:
        sumcolumn = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 首先分组
        # 首先给该列赋空值，保证该列存在，并且获取该列的列序号
        df[columnname] = [0 for _ in range(df.shape[0])]
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        groups = df.groupby(df.columns[indexcolumns[0]].to_list())
        # 每组单独求sum然后给改组所有行赋该sum值
        for groupitem in groups:
            # 获取当前group的sum
            currentsum = groupitem[1][df.columns[sumcolumn[0][0]]].sum()
            # 获取groupby当前的index组
            # 该group的index用来给df的对应行赋值
            df.loc[groupitem[1].index, columnname] = [currentsum for _ in range(groupitem[1].shape[0])]
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=[columnname])
    except Exception as err:
        return errprintwithname("对列{}按索引{}分组求和到列：{} 失败：{}".format(sumcolumn, indexcolumns, columnname, err))

    return [True, ""]


def dfgroupbycumsum(specialdflist: list, tableindex: list,
                    indexcolumns: list, sumcolumn=None, columnname: str = "新列") -> list:
    """
    根据源df和指定的groupby列、生成累加求和列, indexcolumns, sumcolumn 采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]]
    :param sumcolumn:               list: [[被累积求和的列序号],[被累积求和的列名]]
    :param columnname:              str: 累积求和列生产的新列名
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if sumcolumn is None:
        sumcolumn = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        df[columnname] = df.groupby(df.columns[indexcolumns[0]].to_list())[df.columns[sumcolumn[0][0]]].cumsum()
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=[columnname])
    except Exception as err:
        return errprintwithname("对列{}按索引{}分组累积求和到列：{} 失败：{}".format(sumcolumn, indexcolumns, columnname, err))

    return [True, ""]


def dfgroupbycountmark(specialdflist: list, tableindex: list,
                       indexcolumns: list, markcolumns: list) -> list:
    """
    根据源df和指定的groupby列、将每个group包含的行数放置到每行的columnname列中，  indexcolumns, markcolumns 采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]]
    :param markcolumns:             list： [[mark列的序号]，[mark列名]]只有第一个有效
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 根据indexcolumn建立groupby对象
        groups = df.groupby(df.columns[indexcolumns[0]].to_list())
        # 获取每个group
        for groupitem in groups:
            # 获取当前分组的行数
            currentcount = groupitem[1].shape[0]
            # 对当前分组中的每一行对应的计数mark列赋值
            for index in groupitem[1].index:
                df.iloc[index, markcolumns[0][0]] = currentcount
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=df.columns[markcolumns[0][:1]])
    except Exception as err:
        return errprintwithname("对列{}分组打计数标签失败{}".format(indexcolumns, err))

    return [True, ""]


def dfgroupbyserialnomark(specialdflist: list, tableindex: list,
                          indexcolumns: list, markcolumns: list) -> list:
    """
    根据源df和指定的groupby列、将每个group在总group数中的序号放置到每行的columnname列中, indexcolumns, markcolumns采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]]
    :param markcolumns:             list： [[mark列的序号],[mark列名]]，只有第一个有效
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 根据indexcolumn建立groupby对象
        groups = df.groupby(df.columns[indexcolumns[0]].to_list())
        # 获取每个group
        for serialno, groupitem in enumerate(groups):
            # 对当前分组中的每一行对应的序号mark列赋值
            for index in groupitem[1].index:
                df.iloc[index, markcolumns[0][0]] = serialno
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=df.columns[markcolumns[0][:1]])
    except Exception as err:
        return errprintwithname("对列{}分组打序号到列{}标签失败{}".format(indexcolumns, markcolumns[0], err))

    return [True, ""]


def dfgroupbyinnerserialnomark(specialdflist: list, tableindex: list,
                               indexcolumns: list, markcolumns: list) -> list:
    """
    根据源df和指定的groupby列、将每个index在group内的顺序数放置到每行的columnname列中,indexcolumns, markcolumns采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:            list： [[groupby的索引列],[groupby的索引列名]]
    :param markcolumns:             list： [[mark列的序号]，[mark列名]]只有第一个有效
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 根据indexcolumn建立groupby对象
        groups = df.groupby(df.columns[indexcolumns[0]].to_list())
        # 获取每个group
        for groupitem in groups:
            # 对当前分组中的每一行对应的序号mark列赋值
            for serialno, index in enumerate(groupitem[1].index):
                df.iloc[index, markcolumns[0][0]] = serialno + 1
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=df.columns[markcolumns[0][:1]])
    except Exception as err:
        return errprintwithname("对列{}分组打内部序号到列{}标签失败{}".format(indexcolumns, markcolumns[0], err))

    return [True, ""]


def dfgroupbyrowcombinemark(specialdflist: list, tableindex: list,
                            indexcolumns: list, combinecolumns: list = None,
                            columnname: str = "新列", seperate='\n', nanreplace='-') -> list:
    """
    根据indexcolumn分组，在分组内合并指定列combinecolumns的所有行，填写到该groupby内的每行的指定columnname列
    indexcolumns , combinecolumns 采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param indexcolumns:        list: groupby的索引列序号、列名列表
    :param combinecolumns:      list: 合并涉及的str类型列序号、列名列表，仅第一个有效
    :param columnname:          生成的合并列名
    :param seperate:            str:  合并行之间的分隔符
    :param nanreplace:          str:  nan的替换符号
    :return:                    [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    try:
        # 首先将目标列全部清空，如果不存在会自动创建
        df[columnname] = ["" for _ in range(df.shape[0])]
        specialdf.headings = list(df)
        specialdf.colnamelist = specialdf.headings
        specialdf.colnameusedlist = [""] + specialdf.headings
        specialdf.colnamefilterlist = [""] + specialdf.headings
        # 首先分组
        groups = df.groupby(df.columns[indexcolumns[0]].to_list())
        # 每组单独求combine值然后给该组所有行赋该合并内容值
        for groupitem in groups:
            # 获取当前group的指定列combine
            # 以下行转列然后用列str.cat多列的方式（实际只生成了一个元素的列表，因为combinecolumns只使用第一个元素
            currentcombine = groupitem[1].iloc[0, combinecolumns[0][:1]].str.cat(
                groupitem[1].iloc[1:, combinecolumns[0][:1]].T, sep=seperate, na_rep=nanreplace).values[0]
            # 获取groupby当前的index组
            # 该group的index用来给df的对应行赋值
            df.loc[groupitem[1].index, columnname] = [currentcombine for _ in range(groupitem[1].shape[0])]
        # 刷新新增列类型
        specialdf.refreshcolumntype(columnnamelist=[columnname])
    except Exception as err:
        return errprintwithname("索引列{}合并列{}输出列{} 错误：{}".format(
            indexcolumns, combinecolumns, columnname, err))

    return [True, ""]


def dfgroupbysplit(specialdflist: list, tableindex: list, positionlist: list,
                   columns: list, groupbyvalues: list, reindex: bool = False) -> list:
    """
    根据源df和指定的groupby列、groupby n个值列表，将df分拆为n个新表放入到specialdflist中,columns采用新columns格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list:   生成的表格的位置列表，数目必须与groupby 值list一致，且必须连续
    :param columns:                 list： [[新生成的表号标记列序号列表],[新生成的表号标记列列名列表]]
    :param groupbyvalues:           list: [groupby get生成子表时用到的值list], 数目必须与positionlist数目一致
    :param reindex:                 bool: True, False
    :return:                        [bool True:成功，False:失败, errstr]
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if len(positionlist) != len(groupbyvalues):
        return errprintwithname("分组形成新表输入分组值个数和新表个数不一致：{}".format(positionlist, groupbyvalues))

    if positionlist[0] != len(specialdflist):
        return errprintwithname("分组拆分表形成新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    # value = None
    # 创建临时空表用于不存在的分组值
    try:
        # tempdf = pd.DataFrame(data=None, columns=df.columns)      # 该方案无法实现dtypes同时设置
        # 创建空df时需要复制df的columns和dtype，dtype使用np.array，整体用dict创建,用
        tempdf = pd.DataFrame(dict(zip(df.columns.to_list(), [np.array([], dtype=x) for x in df.dtypes])))
        groups = df.groupby(df.columns[columns[0][0]])
    except Exception as err:
        return errprintwithname("对列{}按值{}分组形成新表：{}失败:{}".format(columns, groupbyvalues, positionlist, err))

    # 对每个指定分组值抽取表格
    for value in groupbyvalues:
        try:
            if reindex:
                specialdflist.append(SpecialDf(dfdata=groups.get_group(value).reset_index(drop=True)))
            else:
                specialdflist.append(SpecialDf(dfdata=groups.get_group(value)))
        except Exception as err:
            if err.args[0] == value:
                # 获取该value的group失败，直接使用空表填充
                print("dfgroupbysplit 对列{}按值{}分组形成新表,值{}使用空表".format(columns, groupbyvalues, value))
                specialdflist.append(SpecialDf(dfdata=copy.deepcopy(tempdf)))
                continue
            else:
                return errprintwithname("对列{}按值{}分组形成新表：{}失败:{}".format(
                    columns, groupbyvalues, positionlist, err))

    return [True, ""]


dfquerycomparemark = [">", ">=", "<", "<=", "==", "!="]
dfqueryspecialmarkdict = {"contains": "str.contains", "isin": "isin"}
dfqueryconnectmark = ["|", "&"]
dfqueryaddtionmark = ['groupby']
dfquerystartendmark = ["(", ")"]
DF_QUERY_FLAG_STRING = "QUERY"
DF_QUERY_FLAG = 0
DF_QUERY_COLUMN1 = 1
DF_QUERY_COMPARE_MARK = 2
DF_QUERY_VALUE = 3
DF_QUERY_COLUMN2 = 4
DF_QUERY_CONNECT_MARK = 5
DF_QUERY_STARTMARK = 0
DF_QUERY_ENDMARK = 1

LEFT_QUERY_PARSE_DICT = {0: "[", 1: "", 2: "", 3: "", 4: "", 5: "", 6: ""}
RIGHT_QUERY_PARSE_DICT = {0: "", 1: "", 2: "", 3: "", 4: "", 5: "]", 6: ""}
NORMAL_QUERY_INIT_DICT = {0: "[", 1: "", 2: "", 3: "", 4: "", 5: "]", 6: ""}

NORMAL_QUERY_INIT_STRING_DICT = {
    "==": "(df['{0}']=={2})",
    ">=": "(df['{0}']>={2})",
    ">": "(df['{0}']>{2})",
    "<=": "(df['{0}']<={2})",
    "<": "(df['{0}']<{2})",
    "!=": "(df['{0}']!={2})",
    "contains": "(df['{0}'].str.contains('{2}'))",
    "isin": "(df['{0}'].isin({2}))"
}

COLUMNS_QUERY_INIT_STRING_DICT = {
    "==": "(df['{0}']==df['{3}'])",
    ">=": "(df['{0}']>=df['{3}'])",
    ">": "(df['{0}']>df['{3}'])",
    "<=": "(df['{0}']<=df['{3}'])",
    "<": "(df['{0}']<df['{3}'])",
    "!=": "(df['{0}']!=df['{3}'])",
    "contains": "",
    "isin": ""
}

QUERY_FILTER_NEED_TO_STR_EVAL = ['isin']


def querycommandlist2dflist(commandlist: list, resultdictlist=None, embededflag=False) -> list:
    """
    递归将筛选条件list转换为查询dict list以便用df表格边界
    :param commandlist:                     本次处理的条件list
    :param resultdictlist:                  返回的结果list
    :param embededflag:                     bool: 是否有嵌套的标志
    :return:                                [True, list] or [False, str(errstring)]
    """
    #  初步判断格式合法性
    if resultdictlist is None:
        resultdictlist = []
    if not isinstance(commandlist, list):
        return errprintwithname("查询条件{}格式不合法（非list）！".format(commandlist))

    # 缓冲该级的查询组合
    currentquerycombinelist = []
    # 1、获取当list中的item，
    for items in commandlist:
        # 如果item是当前list是否是QUERY元素在第一个的查询list（叶子节点）
        if isinstance(items, list):
            if items[0] == DF_QUERY_FLAG_STRING:
                # 使用默认dict，赋值相应的元素项
                tempdict = copy.deepcopy(NORMAL_QUERY_INIT_DICT)
                tempdict[1], tempdict[2], tempdict[3], tempdict[4], tempdict[6] = \
                    [x.strip() if isinstance(x, str) else x for x in items[1:6]]
                # 如果“isin”, 需要把isin对应的list转换成str
                if tempdict[2] in QUERY_FILTER_NEED_TO_STR_EVAL:
                    tempdict[3] = str(tempdict[3])
                # 将结果压入结果list
                currentquerycombinelist.append(tempdict)
            else:
                # 2.不是查询条件的叶子项，需要递归拆解子项
                # 2.1、递归之前，先将前面已经形成的查询条件压入结果list
                # 确定当前级有多少可以压入左括号的查询组合项，如果只有一个，将组合括号弹出，并且不加入右括号
                if embededflag:  # 存在递归嵌套，压入数据右侧括号
                    resultdictlist.append(copy.deepcopy(LEFT_QUERY_PARSE_DICT))
                    resultdictlist += currentquerycombinelist
                    resultdictlist.append(copy.deepcopy(RIGHT_QUERY_PARSE_DICT))
                else:  # 不存在嵌套，直接压入查询条件
                    resultdictlist += currentquerycombinelist
                # 清理掉当前已经压入的数据
                currentquerycombinelist = []
                # 2.2、递归当前查询条件
                querycommandlist2dflist(items, resultdictlist=resultdictlist, embededflag=True)
        else:
            return errprintwithname("查询条件子项{}格式不合法（非list）！".format(items))

    # 确定当前级有多少可以压入左括号的查询组合项，如果只有一个，将组合括号弹出，并且不加入右括号
    if embededflag:  # 存在递归嵌套，压入数据右侧括号
        resultdictlist.append(copy.deepcopy(LEFT_QUERY_PARSE_DICT))
        resultdictlist += currentquerycombinelist
        resultdictlist.append(copy.deepcopy(RIGHT_QUERY_PARSE_DICT))
    else:  # 不存在嵌套，直接压入查询条件
        resultdictlist += currentquerycombinelist

    return [True, resultdictlist]


def querycommnadlist2functionstring(commandlist: list) -> list:
    """
    递归将筛选条件list转换为查询字符串
    :param commandlist:                     本次处理的条件list
    :return:                                [True, list] or [False, str(errstring)]
    """
    #  初步判断格式合法性
    resultdictlist = []
    result = querycommandlist2dflist(commandlist, resultdictlist=resultdictlist)
    if result[0] is False:
        return errprintwithname("查询条件{}格式不合法: {}！".format(commandlist, result[1]))

    # 逐行dictlist翻译为函数字符串
    # 可能出现的格式：
    # 1、LEFT_QUERY_PARSE_DICT = {0: "[", 1: "", 2: "", 3: "", 4: "", 5: "", 6: ""}, 左括号
    # 2、RIGHT_QUERY_PARSE_DICT = {0: "", 1: "", 2: "", 3: "", 4: "", 5: "]", 6: ""}， 右括号
    # 3、NORMAL_QUERY_INIT_DICT = {0: "[", 1: "", 2: "", 3: "", 4: "", 5: "]", 6: ""}， 查询条件
    #  3.1、{0: "[", 1: "列名1", 2: "比较符号", 3: "比较值", 4: "", 5: "]", 6: "连接符"}
    #  3.2、{0: "[", 1: "列名1", 2: "比较符号", 3: "", 4: "列名2", 5: "]", 6: "连接符"}
    # 【注】当连接符的下一行为2（右括号），需要将连接符放到右括号的后面，
    # 因此做两轮扫描，第一轮找到所有连接符号，并单独成行放到相应位置，
    # 组内连接符在两个条件之间单独成行，组间连接符外溢到组的右括号行之后单独成行
    # 1、第一轮逐项处理list，
    newlist = []
    # 通过堆栈记录左括号、右括号和连接符之间的关系， 每个querylist带的连接符单独成行放到相应位置
    for itemdict in resultdictlist:
        # 根据字典获取查询行list
        currentrow = [x.strip() if isinstance(x, str) else x for x in itemdict.values()]
        # 判断当前row的类别
        if currentrow == ["[", "", "", "", "", "", ""]:     # 左括号，直接处理
            newlist.append("(")
        elif currentrow == ["", "", "", "", "", "]", ""]:   # 右括号，将上一个元素（如果是连接符）弹出放到本括号后面
            if newlist[-1] in dfqueryconnectmark:
                tempmark = newlist.pop()
                newlist.append(")")
                newlist.append(tempmark)
            else:
                newlist.append(")")
        elif currentrow == ["", "", "", "", "", "", ""]:        # 全空行，跳过
            continue
        else:                                                   # 查询条件行，拆分条件和组合连接符为两行
            newlist.append(currentrow[1:5])                     # 先压入查询条件
            if currentrow[6] in dfqueryconnectmark:             # 如果连接符在合法值范围，压入连接符
                newlist.append(currentrow[6])
    # 2、第二轮逐项处理查询条件，转换为命令语句
    resultstring = ""
    for item in newlist:
        if isinstance(item, str):           # "(", ")", "|", "&", 非查询条件list
            resultstring += item
        elif isinstance(item, list):        # 查询list ["列名1", "条件符号（>, contains等）,"比较值”, "比较列名2"]
            # 当比较值不为空时，使用比较值，否则使用比较列
            if item[2] != "":
                resultstring += NORMAL_QUERY_INIT_STRING_DICT[item[1]].format(item[0], item[1], item[2], item[3])
            else:
                resultstring += COLUMNS_QUERY_INIT_STRING_DICT[item[1]].format(item[0], item[1], item[2], item[3])
        else:
            return errprintwithname("命令格式{}及对应结果{}不正确，无法转换".format(commandlist, newlist))

    return [True, resultstring]


def dfquery(specialdflist: list, tableindex: list, positionlist: list,
            querylist=None, inplace=False, reindex=True) -> list:
    """
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param querylist:               list： 查询条件的列表 [[比较列名1， 比较符号， 比较的值， 比较列名2， 下一条件连接符],
                                                        [比较列名1， 比较符号， 比较的值， 比较列名2， 下一条件连接符],
                                                        .....,
                                                        ],
                                                        查询条件可以递归嵌套，
                                    [['a', '>', 1 , "", "|"], ['b', '>', 2 , "", "&"]], ['c', '>', 2 , "", ""]

    :param positionlist:                inplace 为False时，生成的新表的表号
    :param inplace:                 bool: 替换现有表，True，生成新表：False
    :param reindex:                 bool: 结果表格重建index
    :return:                        [bool True:成功，False:失败, errstr]
    """
    if querylist is None:
        querylist = []
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 没有查询条件, 相当于全表复制
    if len(querylist) == 0:
        specialdflist.append(SpecialDf(dfdata=copy.deepcopy(df)))
        return [True, ""]

    # 最后一条记录的连接符应该为空，因为查询条件可以递归嵌套，因此需要使用递归确认最后一个命令的连接符号为空
    def findlastcondition(commandlist: list) -> str:
        # 先找到list的最后一项
        nextlist = commandlist[-1]
        if isinstance(nextlist, list):
            nextlist = findlastcondition(nextlist)
        else:  # 如果最后不是list，返回该项
            pass
        return nextlist

    # 找到命令的最后一条
    findlast = findlastcondition(querylist)
    # if findlast[DF_QUERY_CONNECT_MARK].strip() != "":
    if findlast.strip() != "":
        return errprintwithname("查询条件{}最后一条不应存在连接符: {}".format(querylist, findlast))

    # [["a", ">=", 17, "", "&"],  ["b", "==", "长兴团队","",""]]
    # [比较列名1， 比较符号， 比较的值， 比较列名2， 下一条件连接符], 比较值和比较列名2只能二选一
    result = querycommnadlist2functionstring(querylist)

    if result[0] is False:
        return result
    # 新增插入的位置不在specialdflist表尾
    if inplace is False and positionlist[0] != len(specialdflist):
        return errprintwithname("过滤表形成新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    querystr = result[1]
    print("dfquery 查询条件:{}".format(querystr))
    totalcmd = "df[{}]".format(querystr)
    print("dfquery command: {}".format(totalcmd))
    try:
        if inplace:
            specialdflist[tableindex[0]] = SpecialDf(dfdata=eval(totalcmd))
            if reindex:
                specialdflist[tableindex[0]].dfreindex()
        else:
            specialdflist.append(SpecialDf(dfdata=eval(totalcmd)))
            if reindex:
                specialdflist[-1].dfreindex()
    except Exception as err:
        return errprintwithname("执行{} 错误：{}".format(totalcmd, err))

    return [True, ""]


COMPARE_SAME = 0
COMPARE_DF1_UNIQE = 1
COMPARE_DF2_UNIQE = 2
COMPARE_DIFFERENCE = 3


def dfcompare(specialdflist: list, tableindex: list, positionlist: list,
              indexcolumn: list = None, markcolumnname: str = "差异标记") -> list:
    """
    交换表格位置 indexcolumn采用新column格式
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param indexcolumn:             list: [[比较时使用的索引列序号列表]，[比较时使用的索引列名列表]]
    :param markcolumnname:          str: 标记差异的列，0, 相同 1 表一特有，2 表二特有，3,1-2有差异，
    :return:                        [bool True:成功|False:失败, errstr，[df1独有列信息], [df2独有列信息], [df1、df2差异列信息]]
    """
    dfcurrentlist = [specialdflist[x].dfdata for x in tableindex]
    if len(dfcurrentlist) != 2:
        return errprintwithname("所指定的对比表格不是两个：{}".format(dfcurrentlist))

    if positionlist[0] != len(specialdflist):
        return errprintwithname("比较df列表新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    df1, df2 = dfcurrentlist[0], dfcurrentlist[1]
    if len(df1.columns) != len(df2.columns):
        return errprintwithname("所指定的对比的两个表格{}格式不一致，无法比较！".format(dfcurrentlist))

    # 记录表格列数
    colnumber = df1.shape[1]
    if markcolumnname.strip() == "":
        markcolumnname = "差异标记"

    # 判断索引方式
    myindexcol = [] if indexcolumn is None else indexcolumn[0]
    myindexcolname = [df1.columns[x] for x in indexcolumn[0]] if indexcolumn is not None else None
    indexcolnumber = len(myindexcol)
    # 比较两组数据差异
    try:
        if indexcolnumber == 0:
            dfcompareresult = datacompy.Compare(df1, df2, on_index=True)
        else:
            dfcompareresult = datacompy.Compare(df1, df2, join_columns=myindexcolname)
    except Exception as err:
        return errprintwithname("按索引列{}的对比两个表格{}错误：{}！".format(indexcolumn, dfcurrentlist, err))

    # 获取差异比对True，False矩阵 （每个单元格的差异用True表示相同，False表示不同，
    # matrix 并排放置了df1所有列，+ df2所有列（不包含索引列） + _merge标志列 + 匹配（True|False）列（不包含索引列）
    dfintersect = dfcompareresult.intersect_rows
    # df1独有行
    df1uniqrows = dfcompareresult.df1_unq_rows
    # df1独有行index列表
    df1uniqindex = df1uniqrows.index.tolist()
    # df2独有行
    df2uniqrows = dfcompareresult.df2_unq_rows
    dfmatchmatrix = dfintersect.iloc[:, colnumber * 2 - indexcolnumber + 1:]
    # 获取差异的行（df1, df2的列数相同，数据列是df1_01, df2_01, 逐列对比放置）
    dfmismatch = dfcompareresult.all_mismatch()

    # 当 myindexcol不为空时，需要处理dfmatchmatrix、dfmismatch，以保证数据格式和on_index=True时一致，方便后续处理
    # 补足matrix中缺少的索引列，方便后续程序计算处理，索引列，可以认为对比结果都是“True”
    for items in myindexcol:
        dfmatchmatrix.insert(items, str(items), True)

    # 补足差异行的索引空缺列，方便后面计算处理
    # 先从差异匹配df的开头获取索引列暂存
    dfindex = dfmismatch.iloc[:, : indexcolnumber]
    # 删除差异匹配开头的索引列
    # dfmismatch.drop(dfmismatch.columns[[list(range(indexcolnumber))]], axis=1, inplace=True)
    dfmismatch.drop(dfmismatch.columns[list(range(indexcolnumber))], axis=1, inplace=True)
    # 在差异匹配df的相应列位置插入索引列（插入2份，分别对应df1,df2, 这种方式适应index=True和False两种情况
    i = 0
    for items in myindexcol:
        # 插入表一对应的该列
        dfmismatch.insert(items * 2, str(items) + "_df1", "1")
        # 给该列赋值，利用前面暂存的索引列值
        dfmismatch[str(items) + "_df1"] = dfindex.iloc[:, i]
        # 插入表二对应的该列
        dfmismatch.insert(items * 2 + 1, str(items) + "_df2", "2")
        # 给该列赋值，利用前面暂存的索引列值
        dfmismatch[str(items) + "_df2"] = dfindex.iloc[:, i]
        i += 1

    # 深度复制第一个df到dfresult，作为后续输出的基础
    dfresult = copy.deepcopy(df1)

    # 当on_index = False, join_column 有值时，需要做以下处理
    df1uniqlist = df1uniqindex if indexcolnumber > 0 else []

    # 2、把差异列插入到dfresult将行号记录到dfdifflist
    # 逐行把数据不同的行插入到第一张表中，并且标记不同值的单元为红色
    # 逐行处理有差异的列, 将差异行从并列放置转成并排放置（colnumber*2 列 转成column列）
    dfdifflist = []
    insertedrownumber = 0  # 用于记录已经插入的行数，因为每插入一行，下一个插入行需要根据前面已经插入的行数下移
    for index, mismatchrow in dfmismatch.iterrows():
        # 确定要插入的行
        insert_row = insertedrownumber + index + 1
        dfdifflist.append(insert_row - 1)           # 原表该对应的行也需要标记
        dfdifflist.append(insert_row)
        insertedrownumber += 1
        # 将df2该行到df1对应行的下一行，并将两行的行号添加到dfdifflist
        # 在dfreuslt中insert_row插入该行，拆开dfresult再concat上半部、插入行、下半部
        # df = pd.concat([df.iloc[:i], line, df.iloc[i+1:]]).reset_index(drop=True)
        insertrowdf = pd.DataFrame([mismatchrow[x] for x in range(1, colnumber * 2, 2)]).T
        insertrowdf.columns = dfresult.columns
        dfresult = pd.concat(
            [dfresult.iloc[:insert_row],
             # mismathcrow共 2*columns个元素，偶数下标为df1数值，奇数下标为df2数值
             insertrowdf,
             dfresult.iloc[insert_row:]]
        ).reset_index(drop=True)
        # 调整df1uniqlist中受影响的行索引
        for i, value in enumerate(df1uniqlist):
            if value >= insert_row:
                df1uniqlist[i] = value + 1
    # 当on_index = False, join_column 有值时，需要做以下处理
    # 把df2独有行追加到dfresult的末尾，并记录相应的index作为df2uniqlist
    df2uniqlist = []
    if indexcolnumber > 0:
        # 1、df1中独有的行，后续标注蓝色，不会影响行号, 对于df1, df1uniqlist，不存在于共同df的作为返回项交给外部壳进行显示处理，标注蓝色）
        # 2、把df2中存在但是df1不存在的行追加到dfresult尾，依然不会影响后续需要插入的差异行号
        for index, row in df2uniqrows.iterrows():
            df2uniqlist.append(dfresult.shape[0])  # 记录当前插入的index号作为后续标注的依据， 标注黄色
            dfresult.loc[dfresult.shape[0]] = row

    # 重建dfresult的索引
    dfresult.reset_index(inplace=True, drop=True)
    # 根据记录增加标记列
    # 首先全部标记相同
    dfresult[markcolumnname] = COMPARE_SAME
    # df1uniqlist,  # df1独有行序号
    # df2uniqlist,  # df2独有行序号
    # dfdifflist
    markindex = list(dfresult).index(markcolumnname)
    dfresult.iloc[df1uniqlist, markindex] = COMPARE_DF1_UNIQE
    dfresult.iloc[df2uniqlist, markindex] = COMPARE_DF2_UNIQE
    dfresult.iloc[dfdifflist, markindex] = COMPARE_DIFFERENCE
    specialdflist.append(SpecialDf(dfdata=dfresult))
    # 获取输出结果表(界面GUI表格颜色信息保存到SpecialDf的rowstyleinfo
    resultspecialdf = specialdflist[-1]

    for colorid, item in enumerate([df1uniqlist, df2uniqlist, dfdifflist]):
        for row in item:
            # 如果没有该行的信息，则创建该行
            resultspecialdf.rowstyleinfo[row][ROWSTYLE_ROWCOLOR] = specialcolordefineforeditable[colorid] \
                if TAB_GROUP_EDITABLE else specialcolordefineforuneditable[colorid]

    # # 文件保存的的条件格式style， TODO 为了测试df.style的测试功能，后续归一化到行列无模板带格式输出命令
    # colormap = specialcolordefineforeditable if TAB_GROUP_EDITABLE else specialcolordefineforuneditable
    # df1color = "color:{}; background-color: {}".format(colormap[0][0], colormap[0][1])
    # df2color = "color:{}; background-color: {}".format(colormap[1][0], colormap[1][1])
    # diffcolor = "color:{}; background-color: {}".format(colormap[2][0], colormap[2][1])
    # nocolor = ""
    # # def color_row(row, number_of_columns):
    # #     color = df1color if row[markcolumnname] == COMPARE_DF1_UNIQE \
    # #         else df2color if row[markcolumnname] == COMPARE_DF2_UNIQE \
    # #         else diffcolor if row[markcolumnname] == COMPARE_DIFFERENCE \
    # #         else nocolor
    # #     return [color] * number_of_columns
    # #
    # # df_style = dfresult.style.apply(color_row, number_of_columns=len(dfresult.columns), axis=1)
    # # df_style.to_excel("d:/temp/colorresult.xlsx")
    return [True, ""]  # 新表，需要后续处理的表格


def dfpivottable(specialdflist: list, tableindex: list,
                 positionlist: list, values=None, index=None, columns=None,
                 aggfunc='mean', fill_value=None, margins=False, dropna=True,
                 margins_name='All', observed=False, sort=True) -> list:
    """
    封装透视表pd.pivot_table
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param positionlist:            list： [新生成的表号列表]
    :param columns:                 list: 指定的说明列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param index:                   list: 指定的索引列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param values:                  list: 指定的转置值列的列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param aggfunc:                 使用的函数，默认是均值
    :param fill_value:
    :param margins:                 是否显示总计
    :param dropna:                  缺失值填充
    :param margins_name:            总计显示为All
    :param observed:
    :param sort:                    排序功能  版本1.3.0才有
    :return:
    """
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    # 插入的位置不在specialdflist表尾
    if positionlist[0] != len(specialdflist):
        return errprintwithname("数据透视新表序号：{}，与实际位置{}不符".format(positionlist, len(specialdflist)))

    myindex = [df.columns[x] for x in index[0]] if index is not None else None
    mycolumns = [df.columns[x] for x in columns[0]] if columns is not None else None
    myvalues = [df.columns[x] for x in values[0]] if values is not None else None

    try:
        dfresult = pd.pivot_table(
            df,
            values=myvalues,
            index=myindex,
            columns=mycolumns,
            aggfunc=aggfunc,
            fill_value=fill_value,
            margins=margins,
            dropna=dropna,
            margins_name=margins_name,
            observed=observed,
            sort=sort)
        # 将索引列变成首列
        dfresult.insert(0, "/".join(myindex), dfresult.index)
        # 将行index变成序号
        specialdflist.append(SpecialDf(dfdata=dfresult.reset_index(drop=True)))
    except Exception as err:
        return errprintwithname("创建透视表失败：{}".format(err))

    return [True, ""]


def dfplotshowdf(
        specialdflist: list, tableindex: list, plottype="", indexcolumn=None, valuecolumns=None, title="",
        xlabel="", ylabel="", showgrid=True, showlegend=True, showstacked=False, showwidth=DEF_PLOT_WIDTH,
        showheight=DEF_PLOT_HEIGHT, showdirect=False,  subplots=False, sharex=False, sharey=False,
        use_index=True, logx=False, logy=False, secondary_y=False, mark_right=True,
        savefigurefile=False, figeurefilename=None) -> list:
    """
    plot显示图表数据
    :param specialdflist:           list: SepcialDf 当前系统处理的所有df列表
    :param tableindex:              list: 相对于specialdflist 列表函数要操作的specialdflist中的元素的下标列表[l,[m, n.....]], 元素从0到多个
    :param plottype:                str: 图表类型
    :param indexcolumn:             list: 指定的索引列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param valuecolumns:            list: 指定的值列列序号.列名列表 [[num1, num2, ....],[colname1, colname2,......]
    :param title:                   str: 图标标题
    :param xlabel:                  str: 图标x轴说明
    :param ylabel:                  str: 图标y轴说明
    :param showgrid:                bool：显示网格
    :param showlegend:              bool：显示legend
    :param showstacked:             bool：是否堆叠显示
    :param showwidth:               int： 图表宽度（英寸）
    :param showheight:              int： 图标高度（英寸）
    :param showdirect:              bool：是否自动显示图表
    :param subplots:                bool: 是否显示子图
    :param sharex:                  bool: 如果subplots=True, 如果有子图，子图共x轴刻度。
    :param sharey:                  bool: 如果subplots=True, y轴被隐藏.
    :param use_index:               bool: 默认 True 是否使用索引作为x轴的刻度
    :param logx:                    bool: default False 设置x轴刻度是否取对数
    :param logy:                    bool: default False 设置y轴刻度是否取对数
    :param secondary_y:             bool or sequence default False 置第二个y轴（右y轴）
    :param mark_right :             bool, default True, secondy_y 有值时，mark的位置是否右侧
    :param savefigurefile:          bool：是否直接保存结果图像文件
    :param figeurefilename:         bool：保存结果图像的文件名称
    :return:
    """
    PLOTDATADF = "_plot_data_df_"
    PLOTTYPE = '_plot_type_'
    DATAINDEXCOLUMN = '_data_index_columns_'
    VALUECOLUMNS = '_value_columns_'
    SHOWGRID = '_show_grid_'
    SHOWLEGEND = '_show_legend_'
    SHOWSTACKED = '_show_stacked_'
    FIGURETITLE = '_figure_title_'
    XLABELNAME = '_x_label_name_'
    YLABELNAME = '_y_label_name_'
    FIGUREWIDTH = '_figure_width_'
    FIGUREHEIGHT = '_figure_height_'
    SHOWSUBPLOTS = '_show_subplots_'
    PLOTSHAREX = '_plot_sharex_'
    PLOTSHAREY = '_plot_sharey_'
    PLOTUSEINDEX = '_plot_use_index_'
    PLOTLOGX = '_plot_logx_'
    PLOTLOGY = '_plot_log_'
    PLOTSECONDARYY = '_plot_secondary_y_'
    PLOTMARKRIGHT = '_plot_mark_right_'
    # 表格的编辑页面
    currentdfnumber = len(specialdflist)
    if currentdfnumber == 0 or len(tableindex) < 1:
        return errprintwithname("当前无数据表格")
    plottypelist = [x.value for x in PlotType]
    chooseyesnolist = ["是", "否"]
    specialdf = specialdflist[tableindex[0]]
    df = specialdf.dfdata
    originalcolumns = list(df.columns)
    # values column的初始值需要去掉已经选择的索引列
    originalvaluescolumns = [x for x in originalcolumns if x not in indexcolumn[1]]

    plotlayout = [
        [Sg.HSeparator()],
        [
            [
                Sg.Col([[
                    Sg.Col([
                        [Sg.Text("选择图形显示表号".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo([x for x in range(currentdfnumber)], k=PLOTDATADF, pad=(5, 5),
                                       default_value=tableindex[0], enable_events=True, font=DEF_FT_L,
                                       size=(15, 1), readonly=True)],
                        [Sg.Text("选择图形方式".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(plottypelist, k=PLOTTYPE, pad=(5, 5), default_value=plottype,
                                       font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("图表标题".ljust(2), font=DEF_FT_L)],
                        [Sg.Input(title, k=FIGURETITLE, pad=(5, 5), enable_events=False, font=DEF_FT_L, size=(16, 1))],
                        [Sg.Text("X轴名称".ljust(2), font=DEF_FT_L)],
                        [Sg.Input(xlabel, k=XLABELNAME, pad=(5, 5), enable_events=False, font=DEF_FT_L, size=(16, 1))],
                        [Sg.Text("Y轴名称".ljust(2), font=DEF_FT_L)],
                        [Sg.Input(ylabel, k=YLABELNAME, pad=(5, 5), enable_events=False, font=DEF_FT_L, size=(16, 1))],
                        [Sg.Text("图标宽度(英寸)".ljust(2), font=DEF_FT_L, visible=True)],
                        [Sg.Input(str(showwidth), k=FIGUREWIDTH, pad=(5, 5), enable_events=False,
                                  visible=True, font=DEF_FT_L, size=(16, 1))],
                        [Sg.Text("图标高度(英寸)".ljust(2), font=DEF_FT_L, visible=True)],
                        [Sg.Input(str(showheight), k=FIGUREHEIGHT, pad=(5, 5), enable_events=False,
                                  visible=True, font=DEF_FT_L, size=(16, 1))],
                        [Sg.Text("显示网格".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=SHOWGRID, pad=(5, 5),
                                       default_value="是" if showgrid else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("显示图例".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=SHOWLEGEND, pad=(5, 5),
                                       default_value="是" if showlegend else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                    ]),
                    Sg.Col([
                        [Sg.Text("是否堆叠".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=SHOWSTACKED, pad=(5, 5),
                                       default_value="是" if showstacked else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("显示子图".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=SHOWSUBPLOTS, pad=(5, 5),
                                       default_value="是" if subplots else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("共享x轴".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTSHAREX, pad=(5, 5),
                                       default_value="是" if sharex else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("共享y轴".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTSHAREY, pad=(5, 5),
                                       default_value="是" if sharey else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("使用索引为X轴刻度".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTUSEINDEX, pad=(5, 5),
                                       default_value="是" if use_index else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("x轴取对数".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTLOGX, pad=(5, 5),
                                       default_value="是" if logx else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("y轴取对数".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTLOGY, pad=(5, 5),
                                       default_value="是" if logy else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("显示第二y轴".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTSECONDARYY, pad=(5, 5),
                                       default_value="是" if secondary_y else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                        [Sg.Text("第二y轴刻度靠右".ljust(2), font=DEF_FT_L)],
                        [Sg.InputCombo(chooseyesnolist, k=PLOTMARKRIGHT, pad=(5, 5),
                                       default_value="是" if mark_right else "否",
                                       enable_events=False, font=DEF_FT_L, size=(15, 1), readonly=True)],
                    ])],
                    [Sg.Text("选择数据索引列".ljust(2), font=DEF_FT_L)],
                    [Sg.InputCombo(originalcolumns, k=DATAINDEXCOLUMN, pad=(5, 5), font=DEF_FT_L, size=(36, 1),
                                   default_value=indexcolumn[1][0] if len(indexcolumn[1]) > 0 else [],
                                   readonly=True, enable_events=True)],
                    [Sg.Text("选择数据列".ljust(2), font=DEF_FT_L)],
                    [Sg.Listbox(originalvaluescolumns, k=VALUECOLUMNS,
                                default_values=valuecolumns[1],
                                highlight_background_color=HIGHLIST_SELECTED_COLOR,
                                highlight_text_color='white',
                                text_color=INPUT_TEXT_COLOR,
                                select_mode=Sg.LISTBOX_SELECT_MODE_MULTIPLE,
                                enable_events=False,
                                tooltip="选择数据列",
                                pad=DEF_PD_L, size=(36, 15))],
                    [
                        Sg.Button('确定', k='_ok_'), Sg.Button('保存', k='_save_'), Sg.Button('关闭', k='_close_')
                    ]
                ]),
                Sg.Col([
                    [Sg.Canvas(key='-TOOLBAR-')],
                    [Sg.Canvas(key='-CANVAS-')]
                ])
            ]
        ],
    ]

    plotwindow = Sg.Window(
        '数据图形显示',
        plotlayout,
        finalize=True,
        element_justification='center',
        font=DEF_FT_L,
        icon=LOCAL_SYSTEM_ICON,
        use_ttk_buttons=True,
        return_keyboard_events=True,  # 把mouse事件暂时由主窗口屏蔽
        resizable=True,
        modal=True,
        keep_on_top=False
    )
    fig_canvas_agg = None
    dfplot = False
    # 设置了直接启动显示
    if showdirect:
        # event, values = plotwindow.read()
        plotwindow.write_event_value('_ok_', {})
    # 等待界面输入
    while True:
        event, values = plotwindow.read()
        # --- Process buttons --- #
        if event in (Sg.WIN_CLOSED, '_close_'):
            break
        elif event == "_save_":
            filetemp = Sg.popup_get_file('选择文件',
                                         save_as=True,
                                         file_types=(('png文件', '*.png'), ),
                                         initial_folder="d:/temp/",
                                         default_path="figure.png",



                                         history_setting_filename="d:/temp/figure.png",
                                         keep_on_top=True, modal=True,
                                         no_window=True, icon=LOCAL_SYSTEM_ICON)
            print(filetemp)
            if dfplot:
                dfplot.save_figure(filetemp)
            continue
        elif event in [PLOTDATADF, DATAINDEXCOLUMN]:
            try:
                currentspecialdf = int(values[PLOTDATADF])
                currentindexcolumn = values[DATAINDEXCOLUMN]
            except:
                Sg.popup_error("表格参数不合规！", modal=True, keep_on_top=True)
                continue
            currentdfcolumnlist = specialdflist[currentspecialdf].dfdata.columns.tolist()
            # 填写
            if event == PLOTDATADF:
                # 表格选择变化，刷新对应表格的索引列以及数值列
                plotwindow[DATAINDEXCOLUMN].update(values=currentdfcolumnlist)
            if currentindexcolumn != "":
                # valuecolumns 需要清除掉已选的indexcolumn
                currentvaluecolumnlist = specialdflist[currentspecialdf].dfdata.columns.tolist()
                currentvaluecolumnlist.remove(currentindexcolumn)
                plotwindow[VALUECOLUMNS].update(values=currentvaluecolumnlist)
            plotwindow.refresh()
            continue

        elif event in ['_ok_']:
            # 获取各个参数，调用画图
            # 调用绘图函数
            if fig_canvas_agg is not None:
                canvas_clear(fig_canvas_agg)
            try:
                currentspecialdf = int(values[PLOTDATADF])
                currentplottype = values[PLOTTYPE]
                currentcolumns = specialdflist[currentspecialdf].dfdata.columns.tolist()
                currentindexcolumn = values[DATAINDEXCOLUMN]
                currentindexcolumnid = currentcolumns.index(currentindexcolumn)
                currentvaluecolumns = values[VALUECOLUMNS]
                currentvaluecolumnsid = [currentcolumns.index(x) for x in currentvaluecolumns]
                currenttitle = values[FIGURETITLE].strip()
                currentxlabel = values[XLABELNAME].strip()
                currentylabel = values[YLABELNAME].strip()
                currentgrid = values[SHOWGRID]
                currentlegend = values[SHOWLEGEND]
                currentstacked = values[SHOWSTACKED]
                currentwidth = int(values[FIGUREWIDTH])
                currentheight = int(values[FIGUREHEIGHT])
                currentsubplots = values[SHOWSUBPLOTS]
                currentsharex = values[PLOTSHAREX]
                currentsharey = values[PLOTSHAREY]
                currentuseindex = values[PLOTUSEINDEX]
                currentlogx = values[PLOTLOGX]
                currentlogy = values[PLOTLOGY]
                currentsecondaryy = values[PLOTSECONDARYY]
                currentmarkright = values[PLOTMARKRIGHT]
            except:
                Sg.popup_error("参数不合规！", modal=True, keep_on_top=True)
                continue
            if len(currentvaluecolumns) == 0:
                Sg.popup_error("未选择数据列！", modal=True, keep_on_top=True)
                continue
            dfplot = DfPlot(
                dfdata=specialdflist[currentspecialdf].dfdata,
                plottype=currentplottype,
                valuecolumns=[currentvaluecolumnsid, currentvaluecolumns],
                indexcolumn=[[currentindexcolumnid], [currentindexcolumn]],
                title=currenttitle,
                xlabel=currentxlabel if currentxlabel != "" else currentindexcolumn,
                ylabel=currentylabel,
                figsize=(currentwidth, currentheight),
                grid=False if currentgrid in ["否", ""] else True,
                legend=False if currentlegend == "否" else True,
                stacked=False if currentstacked in ["否", ""] else True,
                showflag=False,
                subplots=False if currentsubplots == "否" else True,
                sharex=False if currentsharex == "否" else True,
                sharey=False if currentsharey == "否" else True,
                use_index=False if currentuseindex == "否" else True,
                logx=False if currentlogx == "否" else True,
                logy=False if currentlogy == "否" else True,
                secondary_y=False if currentsecondaryy == "否" else True,
                mark_right=False if currentmarkright == "否" else True,
            )
            fig_canvas_agg = draw_figure(dfplot.get_figure(),
                                         plotwindow['-CANVAS-'].TKCanvas,
                                         plotwindow['-TOOLBAR-'].TKCanvas)

            continue

    # 关闭窗体
    plotwindow.close()

    return [True, ""]


def is_mem_virtual_file_name(object_file_name: str) -> bool:
    """
    判断文件名是否为内存虚拟对象
    :param object_file_name:
    :return:
    """
    return (object_file_name.endswith(".{}".format(RAM_FILE_SUFFIX)) and
            object_file_name.startswith(RAM_FILE_PREFIX))


def is_mem_virtual_file_exist(object_file_name: str) -> bool:
    """
    判断内存虚拟文件是否存在
    :param object_file_name:
    :return:
    """
    if not is_mem_virtual_file_name(object_file_name):
        return False
    try:
        # 尝试转化为实际对象，成功表示存在，失败表示不存在
        eval(object_file_name[4: 0 - (len(RAM_FILE_SUFFIX) + 1)])
        return True
    except Exception as err:
        print("读取内存虚拟对象{}失败r: {}".format(object_file_name, err))
        return False


def save_object_to_virtual_file(object_file_name: str, originaldata: object) -> list:
    """
    将指定的object指针保存到虚拟的文件对象中
    :param object_file_name:    str： 后缀必须是.object
    :param originaldata:        需要保存的对象指针
    :return:                    True， 成功
    """
    if not is_mem_virtual_file_name(object_file_name):
        return errprintwithname("{} 内存文件名不合规，前缀不是：{}或后缀不是：{}".format(
            object_file_name, RAM_FILE_PREFIX,RAM_FILE_SUFFIX))
    try:
        exec("global {0}\n{0} = originaldata".format(object_file_name[4: 0 - (len(RAM_FILE_SUFFIX) + 1)]))
    except Exception as err:
        return [False, "保存内存虚拟对象文件{0} 失败: {1}".format(object_file_name, err)]

    return [True, ""]


def load_object_from_virtual_file(object_file_name: str, useorignalmemory=True) -> object:
    """
    从虚拟的对象存储文件获取对象指针
    :param object_file_name:
    :param useorignalmemory:            bool： 使用原内存对象，不进行deepcopy
    :return:
    """
    if not is_mem_virtual_file_name(object_file_name):
        print("{} 内存文件名不合规，前缀不是：{}或后缀不是：{}".format(object_file_name, RAM_FILE_PREFIX, RAM_FILE_SUFFIX))
        return None
    try:
        tempobject = eval(object_file_name[4: 0 - (len(RAM_FILE_SUFFIX) + 1)])
    except Exception as err:
        print("读取内存虚拟对象{}失败r: {}".format(object_file_name, err))
        return None
    # 返回对象的完整拷贝，不能使用指针，否则操作实在原对象上进行
    return tempobject if useorignalmemory else copy.deepcopy(tempobject)

