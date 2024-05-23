"""Python library extending yEd functionality through programmatic interface to graphs.

Currently supports...
* Building of yEd graph objects from scratch
* Reading of yEd graph files
* Management of yEd graph data in Excel (for simplified addition / management of data)
* Opening and basic control of yEd application

"""

# import asyncio
import io
import os
import platform
import re
import subprocess
import sys
import xml.etree.ElementTree as ET
from shutil import which
from time import sleep
from tkinter import messagebox as msg
from typing import Any, List, Optional
from xml.dom import minidom

import openpyxl as pyxl
import psutil

# import pygetwindow as gw

# Enumerated parameters / Constants
PROGRAM_NAME = "yEd.exe"

# Testing related triggers
testing = False
local_testing = None

LINE_TYPES = [
    "line",
    "dashed",
    "dotted",
    "dashed_dotted",
]

FONT_STYLES = [
    "plain",
    "bold",
    "italic",
    "bolditalic",
]

HORIZONTAL_ALIGNMENTS = [
    "left",
    "center",
    "right",
]

VERTICAL_ALIGNMENTS = [
    "top",
    "center",
    "bottom",
]

CUSTOM_PROPERTY_SCOPES = [
    "node",
    "edge",
]  # TODO: DOES THIS NEED GROUP?

CUSTOM_PROPERTY_TYPES = [
    "string",
    "int",
    "double",
    "boolean",
]


def checkValue(
    parameter_name: str,
    value: Any,
    validValues: Optional[List[str]] = None,
) -> None:
    """Check whether given inputs
    (e.g. Shape, Arrow type, Line type, etc.)
    are valid (within existing enumerated options).
    If not valid - returns error message for input."""

    if validValues:
        if value not in validValues:
            raise ValueError(f"{parameter_name} '{value}' is not supported. Use: '{', '.join(validValues)}'")


class File:
    """Object to check and act on yEd files / filepaths (or the excel files during bulk data management)."""

    def __init__(self, file_name_or_path=None):
        self.DEFAULT_FILE_NAME = "temp"
        self.EXTENSION = ".graphml"
        self.dir = self.path_validate(file_name_or_path)
        self.basename = self.base_name_validate(file_name_or_path)
        self.fullpath = os.path.join(self.dir, self.basename)
        self.window_search_name = self.basename + " - yEd"
        self.file_exists = os.path.isfile(self.fullpath)

    def full_path_validate(self):
        self.file_exists = os.path.isfile(self.fullpath)

    def path_validate(self, temp_name_or_path=None):
        """Validate if the file was initialized with valid path - returning the same path - if not valid, return working directory as default path."""
        path = os.getcwd()
        if temp_name_or_path:
            path = os.path.dirname(temp_name_or_path)
            if not os.path.exists(path):
                path = os.getcwd()
        return os.path.realpath(path)

    def base_name_validate(self, temp_name_or_path=None):
        """Validate / Build valid file name with GraphML extension"""
        temp_name = ""
        if temp_name_or_path:
            temp_name = os.path.basename(temp_name_or_path)
        temp_name = temp_name or f"{self.DEFAULT_FILE_NAME}"
        if not temp_name.endswith(self.EXTENSION) and not temp_name.endswith(".xlsx"):
            temp_name += self.EXTENSION
        return temp_name

    def open_with_yed(self, force=False):
        """Method to open GraphML file directly with yEd application (must be installed and on path)."""
        print("opening file with yed...")
        open_yed_file(self, force)
        return get_yed_pid()


class Label:
    """Generic Label Class for nodes / edges in yEd"""

    graphML_tagName = None

    def __init__(
        self,
        text="",
        height="18.1328125",
        width=None,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        horizontalTextPosition="center",
        underlined_text="false",
        text_color="#000000",
        icon_text_gap="4",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        border_color=None,
        background_color=None,
        has_background_color="false",
    ):
        # make class abstract
        if type(self) is Label:
            raise Exception("Label is an abstract class and cannot be instantiated directly")

        self._text = text

        # Initialize dictionary for parameters
        self._params = {}
        self.updateParam("horizontalTextPosition", horizontal_text_position, HORIZONTAL_ALIGNMENTS)
        self.updateParam("verticalTextPosition", vertical_text_position, VERTICAL_ALIGNMENTS)
        self.updateParam("alignment", alignment, HORIZONTAL_ALIGNMENTS)
        self.updateParam("fontStyle", font_style, FONT_STYLES)

        # TODO: Implement range checks
        self.updateParam("fontFamily", font_family)
        self.updateParam("iconTextGap", icon_text_gap)
        self.updateParam("fontSize", font_size)
        self.updateParam("textColor", text_color)
        self.updateParam("visible", visible.lower(), ["true", "false"])
        self.updateParam("underlinedText", underlined_text.lower(), ["true", "false"])
        if background_color:
            has_background_color = "true"
        self.updateParam("hasBackgroundColor", has_background_color.lower(), ["true", "false"])
        self.updateParam("width", width)
        self.updateParam("height", height)
        self.updateParam("borderColor", border_color)
        self.updateParam("backgroundColor", background_color)

    def updateParam(
        self,
        parameter_name,
        value,
        validValues=None,
    ):
        if value is None:
            return False
        checkValue(parameter_name, value, validValues)

        self._params[parameter_name] = value
        return True

    def addSubElement(self, shape):
        label = ET.SubElement(shape, self.graphML_tagName, **self._params)
        label.text = self._text


class NodeLabel(Label):
    """Node specific label"""

    VALIDMODELPARAMS = {
        "internal": ["t", "b", "c", "l", "r", "tl", "tr", "bl", "br"],
        "corners": ["nw", "ne", "sw", "se"],
        "sandwich": ["n", "s"],
        "sides": ["n", "e", "s", "w"],
        "eight_pos": ["n", "e", "s", "w", "nw", "ne", "sw", "se"],
    }

    graphML_tagName = "y:NodeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="internal",
        border_color=None,
        background_color=None,
        model_position="c",
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, NodeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, NodeLabel.VALIDMODELPARAMS[model_name])


class EdgeLabel(Label):
    """Edge specific label"""

    VALIDMODELPARAMS = {
        "two_pos": ["head", "tail"],
        "centered": ["center"],
        "six_pos": ["shead", "thead", "head", "stail", "ttail", "tail"],
        "three_center": ["center", "scentr", "tcentr"],
        "center_slider": None,
        "side_slider": None,
    }

    graphML_tagName = "y:EdgeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="centered",
        model_position="center",
        border_color=None,
        background_color=None,
        preferred_placement=None,
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, EdgeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, EdgeLabel.VALIDMODELPARAMS[model_name])
        self.updateParam("preferredPlacement", preferred_placement)


class CustomPropertyDefinition:
    """Custom properties which can be added to yEd objects / graph as a whole"""

    def __init__(
        self,
        scope,
        name,
        property_type,
        default_value,
    ):
        """
        scope: [node|edge]
        name: name of the custom property
        property_type: [string|boolean|int|double]
                        boolean: Java keywords [true|false]
        default_value: any above datatype represented as a string
        """
        self.scope = scope
        self.name = name
        self.property_type = property_type
        self.default_value = default_value
        self.id = "%s_%s" % (self.scope, self.name)

    def convert_to_xml(self):
        custom_prop_key = ET.Element("key", id=self.id)
        custom_prop_key.set("for", self.scope)
        custom_prop_key.set("attr.name", self.name)
        custom_prop_key.set("attr.type", self.property_type)

        return custom_prop_key


class Group:
    """yEd Group Object (Visual Container of Nodes / Edges / also can recursively act as Node)"""

    VALID_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        group_id,
        parent_graph,
        label=None,
        label_alignment="center",
        shape="rectangle",
        closed="false",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.label = label
        if label is None:
            self.label = group_id

        self.parent = None
        self.group_id = group_id
        self.nodes = {}
        self.groups = {}
        self.parent_graph = parent_graph
        self.edges = {}
        self.num_edges = 0

        # node shape
        checkValue("shape", shape, Group.VALID_SHAPES)
        self.shape = shape

        self.closed = closed

        # label formatting options
        self.font_family = font_family
        self.underlined_text = underlined_text

        checkValue("font_style", font_style, FONT_STYLES)
        self.font_style = font_style
        self.font_size = font_size

        checkValue("label_alignment", label_alignment, HORIZONTAL_ALIGNMENTS)
        self.label_alignment = label_alignment

        self.fill = fill
        self.transparent = transparent

        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_node(self, node_name, **kwargs):
        """Adding node within Group"""
        if node_name in self.parent_graph.existing_entities:
            raise RuntimeWarning("Node %s already exists" % node_name)

        node = Node(node_name, **kwargs)
        node.parent = self
        self.nodes[node_name] = node
        self.parent_graph.existing_entities[node_name] = node
        return node

    def add_group(self, group_id, **kwargs):
        """Adding a group within current group object (and same parent graph)."""
        if group_id in self.parent_graph.existing_entities:
            raise RuntimeWarning("Group %s already exists" % group_id)

        group = Group(group_id, self.parent_graph, **kwargs)
        group.parent = self
        self.groups[group_id] = group
        self.parent_graph.existing_entities[group_id] = group
        return group

    def add_edge(self, node1_id, node2_id, **kwargs):
        """Adds edge - input: node names, not actual node objects"""

        node1 = self.parent_graph.existing_entities.get(node1_id) or self.add_node(node1_id)

        node2 = self.parent_graph.existing_entities.get(node2_id) or self.add_node(node2_id)

        # http://graphml.graphdrawing.org/primer/graphml-primer.html#Nested
        # The edges between two nodes in a nested graph have to be declared in a graph,
        # which is an ancestor of both nodes in the hierarchy.

        if not (self.is_ancestor(node1) and self.is_ancestor(node2)):
            raise RuntimeWarning("Group %s is not ancestor of both %s and %s" % (self.group_id, node1_id, node2_id))

        self.parent_graph.num_edges += 1
        kwargs["edge_id"] = str(self.parent_graph.num_edges)
        edge = Edge(node1_id, node2_id, **kwargs)
        edge.parent = self
        self.edges[edge.edge_id] = edge
        return edge

    def remove_node(self, node_name) -> None:
        """Remove/Delete a node from a group"""
        if node_name not in self.nodes:
            raise RuntimeWarning("Node %s doesn't exist" % node_name)
        del self.nodes[node_name]
        del self.parent_graph.existing_entities[node_name]

    def remove_group(self, group_id) -> None:
        """Removes a group from within current group object (and same parent graph)."""
        if group_id not in self.groups:
            raise RuntimeWarning("Group %s doesn't exist" % group_id)
        group = self.groups[group_id]

        # reroute dependents
        for node in group.nodes.values():
            node.parent = self  # reassign parent: node side
            self.nodes[node.node_name] = node  # reassign parent: group side

        for edge in group.edges.values():
            edge.parent = self  # reassign parent: edge side
            self.edges[edge.edge_id] = edge  # reassign parent: group side

        for group in group.groups.values():
            # edge.parent = self  #reassign parent: edge side
            group.parent = self
            self.groups[group.group_id] = group  # reassign parent: group side

        # remove records
        del self.groups[group_id]
        del self.parent_graph.existing_entities[group_id]

    def remove_edge(self, edge_id):
        """Removing edge from group  - uses node names not node objects."""
        if not self.edges[edge_id]:
            raise RuntimeWarning("Edge %s does not exist under group %s" % (edge_id, self.group_id))

        del self.edges[edge_id]
        # self.num_edges -= 1
        self.parent_graph.num_edges -= 1

    def is_ancestor(self, node):
        """Check for possible nesting conflict of this id usage"""
        return node.parent is not None and (node.parent is self or self.is_ancestor(node.parent))

    def convert_to_xml(self):
        """Converting graph object to graphml xml object"""

        node = ET.Element("node", id=self.group_id)
        node.set("yfiles.foldertype", "group")
        data = ET.SubElement(node, "data", key="data_node")

        # node for group
        pabn = ET.SubElement(data, "y:ProxyAutoBoundsNode")
        r = ET.SubElement(pabn, "y:Realizers", active="0")
        group_node = ET.SubElement(r, "y:GroupNode")

        if self.geom:
            ET.SubElement(group_node, "y:Geometry", **self.geom)

        ET.SubElement(group_node, "y:Fill", color=self.fill, transparent=self.transparent)

        ET.SubElement(
            group_node,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        label = ET.SubElement(
            group_node,
            "y:NodeLabel",
            modelName="internal",
            modelPosition="t",
            fontFamily=self.font_family,
            fontSize=self.font_size,
            underlinedText=self.underlined_text,
            fontStyle=self.font_style,
            alignment=self.label_alignment,
        )
        label.text = self.label

        ET.SubElement(group_node, "y:Shape", type=self.shape)

        ET.SubElement(group_node, "y:State", closed=self.closed)

        graph = ET.SubElement(node, "graph", edgedefault="directed", id=self.group_id)

        if self.url:
            url_node = ET.SubElement(node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(node, "data", key="description_node")
            description_node.text = self.description

        for node_id in self.nodes:
            n = self.nodes[node_id].convert_to_xml()
            graph.append(n)

        for group_id in self.groups:
            n = self.groups[group_id].convert_to_xml()
            graph.append(n)

        for edge_id in self.edges:
            e = self.edges[edge_id].convert_to_xml()
            graph.append(e)

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return node
        # ProxyAutoBoundsNode crap just draws bar at top of group


class Node:
    """yEd Node object"""

    custom_properties_defs = {}

    VALID_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        node_name,
        label=None,
        label_alignment="center",
        shape="rectangle",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        shape_fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        node_type="ShapeNode",
        UML=False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.list_of_labels = []  # initialize list of labels
        if label:
            self.add_label(
                label,
                alignment=label_alignment,
                font_family=font_family,
                underlined_text=underlined_text,
                font_style=font_style,
                font_size=font_size,
            )
        else:
            self.add_label(
                node_name,
                alignment=label_alignment,
                font_family=font_family,
                underlined_text=underlined_text,
                font_style=font_style,
                font_size=font_size,
            )

        self.node_name = node_name

        self.node_type = node_type
        self.UML = UML

        self.parent = None

        # node shape
        checkValue("shape", shape, Node.VALID_SHAPES)
        self.shape = shape

        # shape fill
        self.shape_fill = shape_fill
        self.transparent = transparent

        # border options
        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        # geometry
        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs):
        """Adding node label"""
        self.list_of_labels.append(NodeLabel(label_text, **kwargs))
        return self

    def convert_to_xml(self):
        """Converting node object to xml object"""

        node = ET.Element("node", id=str(self.node_name))
        data = ET.SubElement(node, "data", key="data_node")
        shape = ET.SubElement(data, "y:" + self.node_type)

        if self.geom:
            ET.SubElement(shape, "y:Geometry", **self.geom)
        # <y:Geometry height="30.0" width="30.0" x="475.0" y="727.0"/>

        ET.SubElement(shape, "y:Fill", color=self.shape_fill, transparent=self.transparent)

        ET.SubElement(
            shape,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        for label in self.list_of_labels:
            label.addSubElement(shape)

        ET.SubElement(shape, "y:Shape", type=self.shape)

        if self.UML:
            UML = ET.SubElement(shape, "y:UML")

            attributes = ET.SubElement(UML, "y:AttributeLabel", type=self.shape)
            attributes.text = self.UML["attributes"]

            methods = ET.SubElement(UML, "y:MethodLabel", type=self.shape)
            methods.text = self.UML["methods"]

            stereotype = self.UML["stereotype"] if "stereotype" in self.UML else ""
            UML.set("stereotype", stereotype)

        if self.url:
            url_node = ET.SubElement(node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(node, "data", key="description_node")
            description_node.text = self.description

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return node

    @classmethod
    def set_custom_properties_defs(cls, custom_property):
        cls.custom_properties_defs[custom_property.name] = custom_property


class Edge:
    """yEd Edge - connecting Nodes or Groups"""

    custom_properties_defs = {}

    ARROW_TYPES = [
        "none",
        "standard",
        "white_delta",
        "diamond",
        "white_diamond",
        "short",
        "plain",
        "concave",
        "convex",
        "circle",
        "transparent_circle",
        "dash",
        "skewed_dash",
        "t_shape",
        "crows_foot_one_mandatory",
        "crows_foot_many_mandatory",
        "crows_foot_many_optional",
        "crows_foot_one",
        "crows_foot_many",
        "crows_foot_optional",
    ]

    def __init__(
        self,
        node1,
        node2,
        label=None,
        arrowhead="standard",
        arrowfoot="none",
        color="#000000",
        line_type="line",
        width="1.0",
        edge_id="",
        label_background_color="",
        label_border_color="",
        source_label=None,
        target_label=None,
        custom_properties=None,
        description="",
        url="",
    ):
        self.node1 = node1
        self.node2 = node2

        self.list_of_labels = []  # initialize list of labels

        if label:
            self.add_label(
                label,
                border_color=label_border_color,
                background_color=label_background_color,
            )

        if not edge_id:
            edge_id = "%s_%s" % (node1, node2)

        self.edge_id = str(edge_id)

        if source_label is not None:
            self.add_label(
                source_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        if target_label is not None:
            self.add_label(
                source_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        checkValue("arrowhead", arrowhead, Edge.ARROW_TYPES)
        self.arrowhead = arrowhead

        checkValue("arrowfoot", arrowfoot, Edge.ARROW_TYPES)
        self.arrowfoot = arrowfoot

        checkValue("line_type", line_type, LINE_TYPES)
        self.line_type = line_type

        self.color = color
        self.width = width

        self.description = description
        self.url = url

        self.parent = None

        # Handle Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Edge.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs):
        """Adding edge label"""
        self.list_of_labels.append(EdgeLabel(label_text, **kwargs))
        # Enable method chaining
        return self

    def convert_to_xml(self):
        """Converting edge object to xml object"""

        edge = ET.Element("edge", id=str(self.edge_id), source=str(self.node1), target=str(self.node2))

        data = ET.SubElement(edge, "data", key="data_edge")
        pl = ET.SubElement(data, "y:PolyLineEdge")

        ET.SubElement(pl, "y:Arrows", source=self.arrowfoot, target=self.arrowhead)
        ET.SubElement(pl, "y:LineStyle", color=self.color, type=self.line_type, width=self.width)

        for label in self.list_of_labels:
            label.addSubElement(pl)

        if self.url:
            url_edge = ET.SubElement(edge, "data", key="url_edge")
            url_edge.text = self.url

        if self.description:
            description_edge = ET.SubElement(edge, "data", key="description_edge")
            description_edge.text = self.description

        # Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            edge_custom_prop = ET.SubElement(edge, "data", key=definition.id)
            edge_custom_prop.text = getattr(self, name)

        return edge

    @classmethod
    def set_custom_properties_defs(cls, custom_property):
        cls.custom_properties_defs[custom_property.name] = custom_property


class GraphStats:
    """Object to query and carry complete structure of current (recursive) graph objects and relationships."""

    def __init__(self, graph_or_input_node):
        self.all_nodes = dict()
        self.all_groups = dict()
        self.all_edges = dict()
        self.recursive_id_extract(graph_or_input_node)

    def recursive_id_extract(self, graph_or_input_node):
        """Gather complete structure of current (recursive) graph objects and relationships."""
        sub_nodes = graph_or_input_node.nodes.values()
        sub_groups = graph_or_input_node.groups.values()
        sub_edges = graph_or_input_node.edges.values()

        for node in sub_nodes:
            id = node.node_name
            self.all_nodes[id] = node

        for group in sub_groups:
            id = group.group_id
            self.all_groups[id] = group
            self.recursive_id_extract(group)

        for edge in sub_edges:
            id = edge.edge_id
            self.all_edges[id] = edge


class ExcelManager:
    """Object to handle interfacing of graphs with Excel in bulk data management operations."""

    def __init__(self):
        # Template operations ==============
        self.WB_TYPES = ["obj_and_hierarchy", "object_data", "relations"]
        self.TEMP_EXCEL_SHEET = File("yedextended_temp.xlsx").fullpath
        self.OBJECTS_WS_NAME = "Objects_and_Groups"
        self.RELATIONS_WS_NAME = "Relations"

        # Graph operations ================
        self.graph = Graph()
        self.original_stats = None
        self.original_id_to_label = dict()
        self.original_id_to_obj = dict()
        self.obj_keys = list()
        self.obj_values = list()
        self.dup_objects = list()

    def kill_excel(self):
        os.system('taskkill /f /im "excel.exe"')

    def excel_type_verification(self, type):
        """Check if the given type is valid for Excel operations."""
        self.type = type or self.WB_TYPES[0]  # default
        if self.type not in self.WB_TYPES:
            raise RuntimeWarning("Invalid Excel type. Use: %s" % ", ".join(self.WB_TYPES))

    def create_excel_template(self, type):
        """Generate excel wb per template for that wb type."""

        self.excel_type_verification(type)

        if os.path.isfile(self.TEMP_EXCEL_SHEET):
            os.remove(self.TEMP_EXCEL_SHEET)

        # create workbook
        excel_wb = pyxl.Workbook()

        # Inserting /organizing sheets
        excel_ws = excel_wb.active

        objects_ws = excel_wb.create_sheet(self.OBJECTS_WS_NAME, 0)
        objects_ws.cell(
            row=1, column=1, value="FORMAT -> LABEL | ID (INDENTATION OF INFO MEANS BELONGING TO GROUP ABOVE.)"
        )

        if self.type == "relations":
            relations_ws = excel_wb.create_sheet(self.RELATIONS_WS_NAME, 1)
            relations_ws.cell(row=1, column=1, value="FORMAT -> NODE1 | NODE2 | EDGE_LABEL | EDGE_ID ")

        if excel_ws:
            excel_wb.remove(excel_ws)  # removing default sheet
        self.kill_excel()
        excel_wb.save(self.TEMP_EXCEL_SHEET)
        excel_wb.close()

    def open_close_excel(*args, **kwargs):
        """Provide wrapper for opening / saving / closing excel."""
        save = kwargs.get("save", False)

        def decorator(func):
            def wrapper_func(self, *args, **kwargs):
                in_mem_file = None
                excel_data = kwargs.get("excel_data", None)
                type = kwargs.get("type", None)
                self.excel_type_verification(type)

                if save:
                    self.create_excel_template(type)
                    sleep(0.5)
                    with open(self.TEMP_EXCEL_SHEET, "rb") as f:
                        in_mem_file = io.BytesIO(f.read())

                else:
                    # In case we are given a file path or in-memory file, for excel to graph
                    if excel_data:
                        if isinstance(excel_data, io.BytesIO):
                            in_mem_file = excel_data
                        elif isinstance(excel_data, str):
                            with open(excel_data, "rb") as f:
                                in_mem_file = io.BytesIO(f.read())
                    # Primary use case - from template for graph to excel
                    else:
                        with open(self.TEMP_EXCEL_SHEET, "rb") as f:
                            in_mem_file = io.BytesIO(f.read())

                        # if not os.path.isfile(self.TEMP_EXCEL_SHEET):
                        #     self.create_excel_template(type)
                        #     sleep(1)
                        # with open(self.TEMP_EXCEL_SHEET, "rb") as f:
                        #     in_mem_file = io.BytesIO(f.read())

                if not in_mem_file:
                    raise RuntimeWarning("No excel data found to open.")
                self.excel_wb = pyxl.load_workbook(in_mem_file)
                self.objects_ws = self.excel_wb[self.OBJECTS_WS_NAME]
                if self.type == "relations":
                    self.relations_ws = self.excel_wb[self.RELATIONS_WS_NAME]
                func(self, *args, **kwargs)
                if save:
                    self.excel_wb.save(self.TEMP_EXCEL_SHEET)
                self.kill_excel()

            return wrapper_func

        return decorator

    @open_close_excel(save=True)
    def graph_to_excel_conversion(self, type=None, graph=None) -> None:
        """Converting graph object to excel sheet format for bulk data management operations."""
        if graph:
            self.graph = graph

        self.original_stats = self.graph.gather_graph_stats()
        row = 2

        def graph_object_extract_to_excel(self, input_node, indent_level):
            nonlocal row

            sub_nodes = input_node.nodes
            sub_groups = input_node.groups

            for node in sub_nodes.values():
                id = node.node_name or ""
                labels = getattr(node, "list_of_labels")
                # desc = node.get(description, "")
                # url = node.get(url, "")
                if labels:
                    label = labels[0]._text  # ASSUMPTION: that this would make sense for most graphs
                else:
                    label = ""

                self.original_id_to_label[id] = label
                self.original_id_to_obj[id] = node

                # posting to excel
                self.objects_ws.cell(row=row, column=indent_level, value=label)
                self.objects_ws.cell(row=row, column=indent_level + 1, value=id)
                row += 1

            for group in sub_groups.values():
                id = group.group_id or ""
                label = getattr(group, "label", "")

                # posting to excel
                self.objects_ws.cell(row=row, column=indent_level, value=label)
                self.objects_ws.cell(row=row, column=indent_level + 1, value=id)
                row += 1

                self.original_id_to_label[id] = label
                self.original_id_to_obj[id] = group

                # Recursive extraction - adapting indent
                graph_object_extract_to_excel(self, group, indent_level=indent_level + 1)

        def relations_extract_to_excel(self, input_node):
            nonlocal row

            sub_groups = input_node.groups
            sub_edges = input_node.edges

            for edge in sub_edges.values():
                id = edge.edge_id or ""
                node1 = getattr(edge, "node1")
                node2 = getattr(edge, "node2")
                labels = getattr(edge, "list_of_labels")
                if labels:
                    label = labels[0]._text  # ASSUMPTION: that this would make sense for most graphs
                else:
                    label = ""

                def deduplicate_obj(id):
                    name = self.original_id_to_label[id]
                    if name in self.dup_objects:
                        return id + "##" + name
                    else:
                        return name

                node1name = deduplicate_obj(node1)
                node2name = deduplicate_obj(node2)

                group_name = ""
                if isinstance(input_node, Group):
                    group_name = deduplicate_obj(input_node.group_id)

                # post to excel
                self.relations_ws.cell(row=row, column=col, value=node1name)
                self.relations_ws.cell(row=row, column=col + 1, value=node2name)
                self.relations_ws.cell(row=row, column=col + 2, value=label)
                self.relations_ws.cell(row=row, column=col + 3, value=id)
                self.relations_ws.cell(row=row, column=col + 4, value=group_name)
                row += 1

            for group in sub_groups.values():
                relations_extract_to_excel(self, group)

        if self.type == "obj_and_hierarchy" or self.type == "relations":
            # self.objects_ws = self.excel_wb[self.OBJECTS_WS_NAME]
            graph_object_extract_to_excel(self, self.graph, indent_level=1)

        if self.type == "relations":
            # self.relations_ws = self.excel_wb[self.RELATIONS_WS_NAME]
            row = 2
            col = 1
            self.obj_keys = list(self.original_id_to_label.keys())
            self.obj_values = list(self.original_id_to_label.values())
            self.dup_objects = list(filter(lambda x: True if self.obj_values.count(x) > 1 else False, self.obj_values))

            relations_extract_to_excel(self, self.graph)

    @open_close_excel(save=False)
    def excel_to_graph_conversion(self, type=None, excel_data=None):
        self.excel_type_verification(type)

        if not self.original_stats:
            self.original_stats = Graph().gather_graph_stats()

        if self.type == "obj_and_hierarchy":
            # read the data back into structure - making changes
            # read excel for data / get some stats

            obj_data = list(self.objects_ws.values)
            obj_data.pop(0)  # remove header

            # identifying indents in excel (marker for groupings)
            indent: list[int] = []
            for row in obj_data:
                none_i = 0
                for val in row:
                    if val == None:
                        none_i += 1
                    else:
                        break  # per row for
                indent.append(none_i)

            # identifying groups
            num_items = len(obj_data)
            group_identifiers = list(map(lambda x: 1 if indent[x + 1] > indent[x] else 0, range(0, num_items - 1)))
            group_identifiers.append(0)  # small limitation - deepest or last cannot be group

            # identifying ownership
            owner = dict()
            indent_and_group_ident = list(zip(indent, group_identifiers))
            for i in range(0, num_items):
                owner[i] = None
                if i == 0:
                    continue
                curr_indent = indent[i]
                for j, (indent_i, is_group) in enumerate(reversed(indent_and_group_ident[:i])):
                    if indent_i == curr_indent - 1 and is_group == 1:
                        owner[i] = i - (j + 1)
                        break

            # of format: label|id (optional)

            # identifying last used Id # TODO: NEEDS REFACTORING - ID COUNTING in yEd IS PER OWNER
            ids = sorted(self.original_id_to_label)
            if ids:
                last_id = ids[-1]
                last_id_num = re.match(r".*?(\d+)", last_id)
                if last_id_num and last_id_num.group(1):
                    last_id = int(last_id_num.group(1))
            else:
                last_id = 0

            objects = list()
            # Building / Modifying objects
            all_bulk_mod_ids = set()
            all_curr_obj_ids = set(self.original_id_to_obj.keys())
            for i, (nr_indent, gr_i, obj_row) in enumerate(zip(indent, group_identifiers, obj_data)):
                # possibilities:
                #     completely new node / group
                #     changed label - same node and structure
                #     changed structuring - groups vs nodes
                # changed structuring from node <-> group
                #     changed owner
                #     changed owning
                #     there can be multiple groups at same level
                #     negative - when in this mode - dont have ability to detect  / correct now problem relationships
                # deleted nodes (id is no longer existing)

                # Extracting label and id
                label = obj_row[nr_indent]
                id = None
                try:
                    id = obj_row[nr_indent + 1]
                except Exception as e:
                    print(f"Node missing Id: {e}. Id will be assigned.")
                else:
                    all_bulk_mod_ids.add(id)

                # Checks
                id_given = id is not None
                id_exists = id is not None and id in self.original_id_to_label
                is_group = gr_i == 1
                is_node = gr_i == 0
                is_group_owned = owner[i] is not None

                # giving an id if not one assigned
                if not id_given:
                    if is_group:
                        id = "g" + str(last_id + 1)
                    if is_node:
                        id = "n" + str(last_id + 1)
                    last_id += 1

                if is_group_owned:
                    owner_inst = objects[owner[i]]

                # This is a new object
                if not id_exists:
                    if is_group:
                        if is_group_owned:
                            objects.append(owner_inst.add_group(group_id=id, label=label))
                        else:
                            objects.append(self.graph.add_group(group_id=id, label=label))

                    elif is_node:
                        if is_group_owned:
                            objects.append(owner_inst.add_node(node_name=id, label=label))
                        else:
                            objects.append(self.graph.add_node(node_name=id, label=label))

                    else:
                        raise NotImplementedError("This state not implemented")

                elif id_exists:
                    # changed name
                    existing_obj = self.original_id_to_obj[id]

                    # changed structuring from node <-> group

                    # changed owner

                    # changed owning

                    objects.append(existing_obj)

                    pass

            # Deleted objects
            all_deleted_obj_ids = all_curr_obj_ids.difference(all_bulk_mod_ids)
            for obj_id in all_deleted_obj_ids:
                obj: Group | Node = self.original_id_to_obj[obj_id]
                parent = obj.parent or self.graph
                if isinstance(obj, Group):  # group
                    # find all immediate dependents and connect them to owner
                    parent.remove_group(obj_id)

                elif isinstance(obj, Node):  # node
                    parent.remove_node(obj_id)

        elif self.type == "relations":
            # Columns
            # node1name
            # node2name
            # label
            # id
            self.relations_ws = self.excel_wb[self.RELATIONS_WS_NAME]
            relations_data = self.relations_ws.values
            row_length = None
            edge_ids_after_mod = set()

            count = 0
            for row in relations_data:
                if count == 0:
                    row_length = len(row)
                    count += 1
                    continue
                node1_name = None
                node2_name = None
                edge_label = None
                edge_id = None

                if row_length == 4:
                    node1_name, node2_name, edge_label, edge_id = row
                elif row_length == 3:
                    node1_name, node2_name, edge_label = row
                elif row_length == 2:
                    node1_name, node2_name = row

                edge_id = str(edge_id)

                two_node_id_check = node1_name is not None and node2_name is not None
                if not two_node_id_check:
                    continue

                id_assigned = edge_id is not None
                label_check = edge_label is not None

                id_exist = False
                if id_assigned:
                    id_exist = str(edge_id) in self.original_stats.all_edges.keys()

                node1_found = node1_name in self.obj_values
                node1_dedup_req = node1_name in self.dup_objects
                if node1_found and not node1_dedup_req:
                    node1_id = self.obj_keys[self.obj_values.index(node1_name)]
                # TODO: ADD LOGIC FOR node1_found and node1_dedup_req

                node2_found = node2_name in self.obj_values
                node2_dedup_req = node2_name in self.dup_objects
                if node2_found and not node2_dedup_req:
                    node2_id = self.obj_keys[self.obj_values.index(node2_name)]
                # TODO: ADD LOGIC FOR node2_found and node2_dedup_req

                nodes_found = False
                if node1_id and node2_id:
                    nodes_found = True

                # modify
                if id_exist and nodes_found:
                    # make changes - but only if nodes are found
                    edge = self.original_stats.all_edges[edge_id]
                    edge.node1_id = node1_id
                    edge.node2_id = node2_id
                    edge.label = edge_label

                    # keep track of edges that were existing and still existing
                    edge_ids_after_mod.add(edge.edge_id)

                elif not nodes_found:
                    if not node1_found:
                        raise NameError(f"Node {node1_id} not found.")
                    if not node2_found:
                        raise NameError(f"Node {node2_id} not found.")

                else:  # new / id not existing
                    edge_init_dict = dict()
                    edge_init_dict["node1_id"] = node1_id
                    edge_init_dict["node2_id"] = node2_id
                    edge_init_dict["label"] = edge_label
                    edge_init_dict = {key: value for (key, value) in edge_init_dict.items() if value is not None}
                    self.graph.add_edge(**edge_init_dict)

            # Deleting edges that have been deleted
            all_bulk_edge_ids = set(list(self.original_stats.all_edges.keys()))
            all_deleted_edge_ids = all_bulk_edge_ids.difference(edge_ids_after_mod)
            for del_edge_id in all_deleted_edge_ids:
                edge_obj: Edge = self.original_stats.all_edges[del_edge_id]
                parent = edge_obj.parent or self.graph
                parent.remove_edge(del_edge_id)

        elif self.type == "object_data":  # TODO: Implement management of deeper data - url, description, formatting
            raise NotImplementedError("This functionality is not yet implemented.")

        # Run Checks to avoid problems of manual data management
        self.graph.run_graph_rules()

    def give_user_chance_to_modify(self):
        """Open Excel for user to modify data."""
        # kill excel process
        self.kill_excel()

        if not testing:
            os.startfile(self.TEMP_EXCEL_SHEET)

            user_response = msg.askokcancel(
                title="yEd Bulk Data Management - Async", message="To process changes, save workbook and press ok."
            )

            if not user_response:
                print("User cancelled operation.")
                exit()

    def bulk_data_management(self, type=None, graph=None, excel_data=None):
        """Process of converting to excel for manual operations, allows user to modify and then convert back to graph."""
        self.graph_to_excel_conversion(type=type, graph=graph)
        self.give_user_chance_to_modify()
        self.excel_to_graph_conversion(type=type, excel_data=excel_data)

        return self


class Graph:
    """Graph structure - carries yEd graph information"""

    def __init__(self, directed="directed", graph_id="G"):
        self.nodes = {}
        self.edges = {}
        self.num_edges = 0

        self.directed = directed
        self.graph_id = graph_id
        self.existing_entities = dict()

        self.groups = {}

        self.custom_properties = []

        self.graphml: ET.Element

    # Addition of items ============================
    def add_node(self, node_name, **kwargs):
        """Adding defined node to graph"""
        if node_name in self.existing_entities:
            raise RuntimeWarning("Node %s already exists" % node_name)

        node = Node(node_name, **kwargs)
        self.nodes[node_name] = node
        self.existing_entities[node_name] = node
        return node

    def add_edge(self, node1_id, node2_id, **kwargs):
        """Adding edge to graph - uses node names not node objects."""

        # Ensuring nodes are existing at time of edge creation (error avoidance) - dont need assignments
        node1 = self.existing_entities.get(node1_id) or self.add_node(node1_id)
        node2 = self.existing_entities.get(node2_id) or self.add_node(node2_id)

        self.num_edges += 1
        kwargs["edge_id"] = str(self.num_edges)
        edge = Edge(node1_id, node2_id, **kwargs)
        self.edges[edge.edge_id] = edge
        edge.parent = self
        return edge

    def add_group(self, group_id, **kwargs):
        """Adding group to graph"""
        if group_id in self.existing_entities:
            raise RuntimeWarning("Group %s already exists" % group_id)

        group = Group(group_id, self, **kwargs)
        self.groups[group_id] = group
        self.existing_entities[group_id] = group
        return group

    def define_custom_property(self, scope, name, property_type, default_value):
        """Adding custom properties to graph (which makes them available on the contained objects in yEd)"""
        if scope not in CUSTOM_PROPERTY_SCOPES:
            raise RuntimeWarning("Scope %s not recognised" % scope)
        if property_type not in CUSTOM_PROPERTY_TYPES:
            raise RuntimeWarning("Property Type %s not recognised" % property_type)
        if not isinstance(default_value, str):
            raise RuntimeWarning("default_value %s needs to be a string" % default_value)
        custom_property = CustomPropertyDefinition(scope, name, property_type, default_value)
        self.custom_properties.append(custom_property)
        if scope == "node":
            Node.set_custom_properties_defs(custom_property)
        elif scope == "edge":
            Edge.set_custom_properties_defs(custom_property)

    # Removal of items ==============================
    def remove_node(self, node_name) -> None:
        """Remove/Delete a node from graph"""
        if node_name not in self.existing_entities:
            raise RuntimeWarning("Node %s doesn't exist" % node_name)
        del self.nodes[node_name]
        del self.existing_entities[node_name]

    def remove_group(self, group_id):
        """Removes a group from within current group object (and same parent graph)."""
        if group_id not in self.existing_entities:
            raise RuntimeWarning("Group %s doesn't exist" % group_id)

        # reroute dependents
        for node in self.nodes.values():
            node.parent = self  # reassign parent: node side
            self.nodes[node.node_name] = node  # reassign parent: group side

        for edge in self.edges.values():
            edge.parent = self  # reassign parent: edge side
            self.edges[edge.edge_id] = edge  # reassign parent: group/graph side

        for group in self.groups.values():
            group.parent = self
            self.groups[group.group_id] = group  # reassign parent: group side

        # eliminate records
        del self.groups[group_id]
        del self.existing_entities[group_id]

    def remove_edge(self, edge_id):
        """Removing edge from graph - uses node names not node objects."""
        if edge_id not in self.edges:
            raise RuntimeWarning("Edge %s does not exist under graph" % (edge_id))
        del self.edges[edge_id]
        self.num_edges -= 1

    # TODO: ADD FUNCTIONALITY TO REMOVE / MODIFY CUSTOM PROPERTIES

    # Graph functionalities ===========================
    def construct_graphml(self):
        """Creating template graphml xml structure and then placing all graph items into it."""

        # Creating XML structure in Graphml format
        # xml = ET.Element("?xml", version="1.0", encoding="UTF-8", standalone="no")

        graphml = ET.Element("graphml", xmlns="http://graphml.graphdrawing.org/xmlns")
        graphml.set("xmlns:java", "http://www.yworks.com/xml/yfiles-common/1.0/java")
        graphml.set("xmlns:sys", "http://www.yworks.com/xml/yfiles-common/markup/primitives/2.0")
        graphml.set("xmlns:x", "http://www.yworks.com/xml/yfiles-common/markup/2.0")
        graphml.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        graphml.set("xmlns:y", "http://www.yworks.com/xml/graphml")
        graphml.set("xmlns:yed", "http://www.yworks.com/xml/yed/3")
        graphml.set(
            "xsi:schemaLocation",
            "http://graphml.graphdrawing.org/xmlns http://www.yworks.com/xml/schema/graphml/1.1/ygraphml.xsd",
        )

        # Adding some implementation specific keys for identifying urls, descriptions
        node_key = ET.SubElement(graphml, "key", id="data_node")
        node_key.set("for", "node")
        node_key.set("yfiles.type", "nodegraphics")

        # Definition: url for Node
        node_key = ET.SubElement(graphml, "key", id="url_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Node
        node_key = ET.SubElement(graphml, "key", id="description_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: url for Edge
        node_key = ET.SubElement(graphml, "key", id="url_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Edge
        node_key = ET.SubElement(graphml, "key", id="description_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: Custom Properties for Nodes and Edges
        for prop in self.custom_properties:
            graphml.append(prop.convert_to_xml())

        edge_key = ET.SubElement(graphml, "key", id="data_edge")
        edge_key.set("for", "edge")
        edge_key.set("yfiles.type", "edgegraphics")

        # Graph node containing actual objects
        graph = ET.SubElement(graphml, "graph", edgedefault=self.directed, id=self.graph_id)

        # Convert python graph objects into xml structure
        for node in self.nodes.values():
            graph.append(node.convert_to_xml())

        for node in self.groups.values():
            graph.append(node.convert_to_xml())

        for edge in self.edges.values():
            graph.append(edge.convert_to_xml())

        self.graphml = graphml

    def persist_graph(self, file=None, pretty_print=False, overwrite=False) -> File:
        """Convert graphml object->xml tree->graphml file.
        Temporary naming used if not given.
        """

        graph_file = File(file)

        if graph_file.file_exists and not overwrite:
            raise FileExistsError

        self.construct_graphml()

        if pretty_print:
            raw_str = ET.tostring(self.graphml)
            pretty_str = minidom.parseString(raw_str).toprettyxml()
            with open(graph_file.fullpath, "w") as f:
                f.write(pretty_str)
        else:
            tree = ET.ElementTree(self.graphml)
            tree.write(graph_file.fullpath)  # Uses internal method to XML Etree

        # recheck the file as existing or not
        graph_file.full_path_validate()
        print("persisting graph to file:", graph_file.fullpath)
        return graph_file

    def stringify_graph(self):
        """Returns Stringified version of graph in graphml format"""
        self.construct_graphml()
        # Py2/3 sigh.
        if sys.version_info.major < 3:
            return ET.tostring(self.graphml, encoding="UTF-8")
        else:
            return ET.tostring(self.graphml, encoding="UTF-8").decode()

    def from_existing_graph(self, file: str | File):
        """Parse GraphML xml of existing/stored graph file into python Graph structure."""

        # Manage file input ==============================
        if isinstance(file, File):
            graph_file = file
        else:
            graph_file = File(file)
        if not graph_file.file_exists:
            raise FileNotFoundError

        # Simplify input into string ==============================
        graph_str = xml_to_simple_string(graph_file.fullpath)

        # Begin XML parsing
        root = ET.fromstring(graph_str)

        # Extract off key information==============================
        all_keys = root.findall("key")
        key_dict = dict()
        for a_key in all_keys:
            sub_key_dict = dict()

            key_id = a_key.attrib.get("id")
            # sub_key_dict["label"] = a_key.attrib.get("for")
            sub_key_dict["attr"] = a_key.attrib.get("attr.name", None)
            # sub_key_dict["label"] = a_key.attrib.get("type")
            key_dict[key_id] = sub_key_dict

        # Get major graph node
        graph_root = root.find("graph")

        # get major graph info
        graph_dir = graph_root.get("edgedefault")
        graph_id = graph_root.get("id")

        # instantiate graph object
        new_graph = Graph(directed=graph_dir, graph_id=graph_id)

        # Parse graph

        def is_group_node(node):
            return "foldertype" in node.attrib

        def process_node(parent, input_node):
            # Get sub nodes of this node (group or graph)
            current_level_nodes = input_node.findall("node")
            current_level_edges = input_node.findall("edge")

            for node in current_level_nodes:
                # normal nodes
                if not is_group_node(node):
                    node_init_dict = dict()

                    # <node id="n1">
                    node_init_dict["node_name"] = node.attrib.get("id", None)

                    data_nodes = node.findall("data")
                    info_node = None
                    for data_node in data_nodes:
                        info_node = data_node.find("GenericNode") or data_node.find("ShapeNode")
                        if info_node is not None:
                            node_init_dict["node_type"] = info_node.tag

                            node_label = info_node.find("NodeLabel")
                            if node_label is not None:
                                node_init_dict["label"] = node_label.text

                                # TODO: PORT REST OF NODELABEL

                            # <Fill color="#FFCC00" transparent="false" />
                            fill = info_node.find("Fill")
                            if fill is not None:
                                node_init_dict["shape_fill"] = fill.get("color")
                                node_init_dict["transparent"] = fill.get("transparent")

                            # <BorderStyle color="#000000" type="line" width="1.0" />
                            border_style = info_node.find("BorderStyle")
                            if border_style is not None:
                                node_init_dict["border_color"] = border_style.get("color")
                                node_init_dict["border_type"] = border_style.get("type")
                                node_init_dict["border_width"] = border_style.get("width")

                            # <Shape type="rectangle" />
                            shape_sub = info_node.find("Shape")
                            if shape_sub is not None:
                                node_init_dict["shape"] = shape_sub.get("type")

                            uml = info_node.find("UML")
                            if uml is not None:
                                node_init_dict["shape"] = uml.get("AttributeLabel")
                            # TODO: THERE IS FURTHER DETAIL TO PARSE HERE under uml
                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    node_init_dict[info_type] = info
                    # Removing empty items
                    node_init_dict = {key: value for (key, value) in node_init_dict.items() if value is not None}
                    # create node
                    parent.add_node(**node_init_dict)

                # group nodes
                # <node id="n2" yfiles.foldertype="group">
                elif is_group_node(node):
                    group_init_dict = dict()

                    # <node id="n1">
                    group_init_dict["group_id"] = node.attrib.get("id", None)

                    # Actual Group Data ===================================
                    data_nodes = node.findall("data")
                    for data_node in data_nodes:
                        proxy = data_node.find("ProxyAutoBoundsNode")
                        if proxy is not None:
                            realizer = proxy.find("Realizers")

                            group_nodes = realizer.findall("GroupNode")

                            for group_node in group_nodes:
                                geom_node = group_node.find("Geometry")
                                if geom_node is not None:
                                    group_init_dict["height"] = geom_node.attrib.get("height", None)
                                    group_init_dict["width"] = geom_node.attrib.get("width", None)
                                    group_init_dict["x"] = geom_node.attrib.get("x", None)
                                    group_init_dict["y"] = geom_node.attrib.get("y", None)

                                fill_node = group_node.find("Fill")
                                if fill_node is not None:
                                    group_init_dict["fill"] = fill_node.attrib.get("color", None)
                                    group_init_dict["transparent"] = fill_node.attrib.get("transparent", None)

                                borderstyle_node = group_node.find("BorderStyle")
                                if borderstyle_node is not None:
                                    group_init_dict["border_color"] = borderstyle_node.attrib.get("color", None)
                                    group_init_dict["border_type"] = borderstyle_node.attrib.get("type", None)
                                    group_init_dict["border_width"] = borderstyle_node.attrib.get("width", None)

                                nodelabel_node = group_node.find("NodeLabel")
                                if nodelabel_node is not None:
                                    group_init_dict["label"] = nodelabel_node.text
                                    group_init_dict["font_family"] = nodelabel_node.attrib.get("fontFamily", None)
                                    group_init_dict["font_size"] = nodelabel_node.attrib.get("fontSize", None)
                                    group_init_dict["underlined_text"] = nodelabel_node.attrib.get(
                                        "underlinedText", None
                                    )
                                    group_init_dict["font_style"] = nodelabel_node.attrib.get("fontStyle", None)
                                    group_init_dict["label_alignment"] = nodelabel_node.attrib.get("alignment", None)

                                group_shape_node = group_node.find("Shape")
                                if group_shape_node is not None:
                                    group_init_dict["shape"] = group_shape_node.attrib.get("type", None)

                                group_state_node = group_node.find("State")
                                if group_state_node is not None:
                                    group_init_dict["closed"] = group_state_node.attrib.get("closed", None)
                                    # group_init_dict["aaa"] = group_state_node.attrib.get("closedHeight",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("closedWidth",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("innerGraphDisplayEnabled",None)

                                break

                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    group_init_dict[info_type] = info

                    # Group - Graph node
                    sub_graph_node = node.find("graph")

                    # Removing empty items
                    group_init_dict = {key: value for (key, value) in group_init_dict.items() if value is not None}

                    # Creating new group
                    new_group = parent.add_group(**group_init_dict)

                    # Recursive processing
                    if sub_graph_node is not None:
                        process_node(parent=new_group, input_node=sub_graph_node)

                # unknown node type
                else:
                    raise NotImplementedError

            # edges then establish connections
            for edge_node in current_level_edges:
                edge_init_dict = dict()

                # <node id="n1">
                edge_init_dict["edge_id"] = edge_node.attrib.get("id", None)
                edge_init_dict["node1_id"] = edge_node.attrib.get("source", None)
                edge_init_dict["node2_id"] = edge_node.attrib.get("target", None)

                # <data key="d5">
                data_nodes = edge_node.findall("data")
                for data_node in data_nodes:
                    polylineedge = data_node.find("PolyLineEdge")

                    if polylineedge is not None:
                        # TODO: ADD POSITION MANAGEMENT
                        # path_node = polylineedge.find("Path")
                        # if path_node:
                        #   edge_init_dict["label"] = path_node.attrib.get("sx")
                        #   edge_init_dict["label"] = path_node.attrib.get("sy")
                        #   edge_init_dict["label"] = path_node.attrib.get("tx")
                        #   edge_init_dict["label"] = path_node.attrib.get("ty")

                        linestyle_node = polylineedge.find("LineStyle")
                        if linestyle_node is not None:
                            edge_init_dict["color"] = linestyle_node.attrib.get("color", None)
                            edge_init_dict["line_type"] = linestyle_node.attrib.get("type", None)
                            edge_init_dict["width"] = linestyle_node.attrib.get("width", None)

                        arrows_node = polylineedge.find("Arrows")
                        if arrows_node is not None:
                            edge_init_dict["arrowfoot"] = arrows_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = arrows_node.attrib.get("target", None)

                        edgelabel_node = polylineedge.find("EdgeLabel")
                        if edgelabel_node is not None:
                            edge_init_dict["label"] = edgelabel_node.text
                            edge_init_dict["arrowfoot"] = edgelabel_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = edgelabel_node.attrib.get("target", None)

                    else:
                        info = data_node.text
                        if info is not None:
                            info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                            info = re.sub(r"\]\]>", "", info)  # unneeded schema

                            the_key = data_node.attrib.get("key")

                            info_type = key_dict[the_key]["attr"]
                            if info_type in ["url", "description"]:
                                edge_init_dict[info_type] = info

                # bendstyle_node = polylineedge.find("BendStyle")
                # edge_init_dict["smoothed"] = linestyle_node.attrib.get("smoothed") # TODO: ADD THIS

                # TODO:
                #   CUSTOM PROPERTIES

                # Removing empty items
                edge_init_dict = {key: value for (key, value) in edge_init_dict.items() if value is not None}
                parent.add_edge(**edge_init_dict)

        process_node(parent=new_graph, input_node=graph_root)

        return new_graph

    def manage_graph_data_in_excel(self, type=None):
        """Port graph data into Excel in several formats for easy and bulk creation and management.  Then ports back into python graph structure."""
        return ExcelManager().bulk_data_management(graph=self, type=type)

    def gather_graph_stats(self) -> GraphStats:
        """Creating current Graph Stats for the current graph"""
        return GraphStats(self)

    def run_graph_rules(self, correct=None):
        """Check a few graph items that are most likely to fail following manual data management."""
        if correct is None:  #  ("auto", "manual")
            correct = "auto"

        stats = self.gather_graph_stats()

        def stranded_edges_check(self, graph_stats, correct):
            """Check for edges with no longer valid nodes (these will prevent yEd from opening the file).  Correct them automatically or manually."""
            stranded_edges = set()
            for edge_id, edge in graph_stats.all_edges.items():
                node1_exist = edge.node1 in graph_stats.all_nodes.keys() or edge.node1 in graph_stats.all_groups.keys()
                node2_exist = edge.node2 in graph_stats.all_nodes.keys() or edge.node2 in graph_stats.all_groups.keys()
                stranded_edge = not node1_exist or not node2_exist
                at_least_one_edge = node1_exist or node2_exist
                if stranded_edge:
                    stranded_edges.add(edge)

            if correct == "auto":
                for edge in stranded_edges:
                    edge.parent.remove_edge(edge.edge_id)

            elif correct == "manual":
                # Excel - run relations and highlight edges with issues?
                raise NotImplementedError("Manual correction of stranded edges is not yet implemented.")

            # offer review or update edges
            return stranded_edges

        stranded_edges = stranded_edges_check(self, stats, correct)


# App related functions ===========================
def is_yed_findable():
    """Find yEd exe path locally"""
    path = which(PROGRAM_NAME) or None
    yed_found_bool = path is not None
    if not yed_found_bool:
        msg.showerror(
            title="Could not find yEd application.",
            message="Please install / add yEd to path.",
        )
    return yed_found_bool


def get_yed_pid():
    """Return process id for yEd application or None if not running"""
    yed_pid = None
    yed_process = get_yed_process()
    if yed_process:
        yed_pid = yed_process.pid
    return yed_pid


def get_yed_process():
    """Return process object for yEd application, if there is one running."""
    process = None
    for process_iter in psutil.process_iter(["name"]):
        if process_iter.info["name"] == PROGRAM_NAME:
            process = process_iter
            break
    return process


def is_yed_open() -> bool:
    """Check whether yEd is currently open - windows, linux, os, etc."""
    return get_yed_pid() is not None


def start_subprocess(command):
    try:
        # Start the subprocess
        process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        pid = process.pid
        # print(f"Started process with PID: {pid}")
        return process
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
    return None


def open_yed_file(file: File, force=False):
    """Opens yed file - also will start yed if not open. Returns process or None."""
    if not file.file_exists:
        return None

    if force:
        print("Act on yEd message box...")
        answer = msg.askokcancel(title="Force yEd App Close", message="Are you ok to force yEd closure?")
        if not answer:
            print("Exiting program")
            exit()

        kill_yed()

    os.startfile(file.fullpath)

    if force:
        sleep(4)
        if get_yed_pid() is None:
            os.startfile(file.fullpath)

    return get_yed_process()


def start_yed(wait=False):
    """Starts yEd program (if installed and on path and not already open). Returns process or None.

    Wait - if True, will wait for yEd to close before returning.
    """
    process = None
    if is_yed_findable():
        if not is_yed_open():
            if wait is False:
                command = [
                    PROGRAM_NAME,
                ]
                process = start_subprocess(command)

                return process or None
            else:
                subprocess.run(PROGRAM_NAME)
                return None


def kill_yed() -> None:
    """Ends yEd program (if installed and on path and open)."""
    yed_process = get_yed_process()
    if yed_process:
        yed_process.kill()


# Utilities =======================================
def xml_to_simple_string(file_path) -> str:
    """Takes GraphML xml in string format and reduces complexity of the string for simpler parsing (without loss of any significant information).  Returns simplified string."""
    graph_str = ""
    try:
        with open(file_path, "r") as graph_file:
            graph_str = graph_file.read()

    except FileNotFoundError:
        print(f"Error, file not found: {file_path}")
        raise FileNotFoundError(f"Error, file not found: {file_path}")
    else:
        # Preprocessing of file for ease of parsing
        graph_str = graph_str.replace("\n", " ")  # line returns
        graph_str = graph_str.replace("\r", " ")  # line returns
        graph_str = graph_str.replace("\t", " ")  # tabs
        graph_str = re.sub("<graphml .*?>", "<graphml>", graph_str)  # unneeded schema
        graph_str = graph_str.replace("> <", "><")  # empty text
        graph_str = graph_str.replace("y:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("xml:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("yfiles.", "")  # unneeded namespace prefix
        graph_str = re.sub(" {1,}", " ", graph_str)  # reducing redundant spaces

    return graph_str
