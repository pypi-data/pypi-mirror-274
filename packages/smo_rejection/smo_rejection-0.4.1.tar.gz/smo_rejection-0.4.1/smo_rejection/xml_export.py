import openpyxl
from openpyxl.styles import Font, Alignment
from .utils import calculate_mean_served_requests

def export_to_excel(results, filename):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for i, result in enumerate(results, start=1):
        worksheet.merge_cells(f'A{1 + (i - 1) * (len(result.request_times) + 3)}:I{1 + (i - 1) * (len(result.request_times) + 3)}')
        header_cell = worksheet.cell(row=1 + (i - 1) * (len(result.request_times) + 3), column=1, value=f"---------------------------- Симуляция номер: {i} ----------------------------")
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        headers = ["Индекс", "Случайное число", "МЕЖ", "Время в очереди", "Сервер 1", "Сервер 2", "Сервер 3", "Обслужено", "Отказов"]
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=2 + (i - 1) * (len(result.request_times) + 3), column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row, entry in enumerate(result.request_times, start=3 + (i - 1) * (len(result.request_times) + 3)):
            worksheet.cell(row=row, column=1, value=entry["index"])
            worksheet.cell(row=row, column=2, value=entry["rand_value"])
            worksheet.cell(row=row, column=3, value=entry["iba"])
            worksheet.cell(row=row, column=4, value=entry["app_time"])
            for j in range(1, 4):
                worksheet.cell(row=row, column=4 + j, value=entry[f"server_{j}"])
            worksheet.cell(row=row, column=8, value=entry["Обслужено"])
            worksheet.cell(row=row, column=9, value=entry["Отказов"])

        served_count_cell = worksheet.cell(row=len(result.request_times) + 3 + (i - 1) * (len(result.request_times) + 3), column=1, value=f"Количество исполненных заявок: {result.served_requests}")
        rejected_count_cell = worksheet.cell(row=len(result.request_times) + 4 + (i - 1) * (len(result.request_times) + 3), column=1, value=f"Количество отказов: {result.rejected_requests}")

    mean_served_cell = worksheet.cell(row=worksheet.max_row + 1, column=1, value=f"В качестве оценки искомого математического ожидания a – числа обслуженных заявок примем выборочную среднюю: a = {calculate_mean_served_requests(results)}")

    workbook.save(filename)