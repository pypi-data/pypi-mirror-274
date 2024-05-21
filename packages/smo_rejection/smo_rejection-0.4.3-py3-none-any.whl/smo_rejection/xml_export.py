import openpyxl
from openpyxl.styles import Font, Alignment
from .utils import calculate_mean_served_requests

def export_to_excel(results, filename):
    """
    Экспортирует результаты симуляций в файл Excel.

    ### Параметры:
    * `results (list)` - Список результатов симуляций, где каждый элемент содержит данные о заявках.
    * `filename (str)` - Имя файла для сохранения Excel-документа.

    ### Описание:
    Функция создает новую рабочую книгу Excel, записывает в неё результаты симуляций, включая заголовки и данные для каждой заявки,
    а также итоговые значения для каждой симуляции и среднее количество обслуженных заявок.

    """
    # Создание новой рабочей книги и активного листа
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    current_row = 1  # Начальная строка

    for i, result in enumerate(results, start=1):
        # Запись заголовка для каждой симуляции
        header_text = f"---------------------------- Симуляция номер: {i} ----------------------------"
        header_cell = worksheet.cell(row=current_row, column=1, value=header_text)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center')
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        
        current_row += 1  # Перемещение на следующую строку для заголовков колонок

        # Запись заголовков колонок
        headers = ["Индекс", "Случайное число", "МЕЖ", "Время в очереди", "Сервер 1", "Сервер 2", "Сервер 3", "Обслужено", "Отказов"]
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        current_row += 1  # Перемещение на следующую строку для данных

        # Запись данных для каждой записи в симуляции
        for entry in result.request_times:
            worksheet.cell(row=current_row, column=1, value=entry["index"])
            worksheet.cell(row=current_row, column=2, value=entry["rand_value"])
            worksheet.cell(row=current_row, column=3, value=entry["iba"])
            worksheet.cell(row=current_row, column=4, value=entry["app_time"])
            for j in range(1, 4):
                worksheet.cell(row=current_row, column=4 + j, value=entry[f"server_{j}"])
            worksheet.cell(row=current_row, column=8, value=entry["Обслужено"])
            worksheet.cell(row=current_row, column=9, value=entry["Отказов"])
            current_row += 1  # Перемещение на следующую строку для следующей записи

        # Запись итогов для каждой симуляции
        worksheet.cell(row=current_row, column=1, value=f"Количество исполненных заявок: {result.served_requests}")
        current_row += 1
        worksheet.cell(row=current_row, column=1, value=f"Количество отказов: {result.rejected_requests}")
        current_row += 2  # Добавление пустой строки для разделения симуляций

    # Запись среднего количества обслуженных заявок в конце документа
    mean_served_value = calculate_mean_served_requests(results)
    mean_served_cell = worksheet.cell(row=current_row, column=1, value=f"В качестве оценки искомого математического ожидания a – числа обслуженных заявок примем выборочную среднюю: a = {mean_served_value}")

    # Попытка сохранить рабочую книгу
    try:
        workbook.save(filename)
    except PermissionError:
        # Если файл уже открыт или доступ запрещен, сохраняем под альтернативным именем
        alternative_filename = "alternative_" + filename
        workbook.save(alternative_filename)
