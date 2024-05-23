import base64
import hashlib
import json
import os
import random
import re
import socket
import uuid
import zipfile
import datetime


def hashids_encode(ids: str, **kwargs):
    """
    hashids 加密
    :param ids:
    :param kwargs:
    :return:
    """
    from hashids import Hashids
    hashids = Hashids(**kwargs)
    return hashids.encode(ids)


def hashids_decode(string: str, **kwargs):
    """
    hashids 解密
    :param string:
    :param kwargs:
    :return:
    """
    from hashids import Hashids
    hashids = Hashids(**kwargs)
    return hashids.decode(string)


def get_mac_address():
    """
    获取 本机 mac 地址
    :return:
    """
    mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
    return ":".join([mac[e:e + 2] for e in range(0, 11, 2)])


def get_host_name():
    """
    获取主机名
    :return:
    """
    return socket.getfqdn(socket.gethostname())


def get_host_ip():
    """
    获取IP地址
    :return:
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()
    return ip


def get_client_ip(request):
    """
    获取请求的ip
    :param request:
    :return:
    """
    headers = request.headers
    addr = headers.get('http_x_forwarded_for', None)
    if not addr:
        addr = headers.get('http_client_ip', None)
    if not addr:
        addr = headers.get('remote_addr', None)
    if not addr:
        addr = headers.get('remote-host', None)
    if not addr:
        addr = headers.get('x-real-ip', None)
    if not addr:
        addr = headers.get('x-forwarded-for', None)
    if not addr:
        addr = headers.get('x-forwarded-for', None)
    return addr or None


def md5_bytes(string: bytes):
    """
    md5 字符串
    :param string:
    :return:
    """
    md5hash = hashlib.md5(string)
    return md5hash.hexdigest()


def md5_string(string: str, encoding="utf8"):
    """
    md5 字符串
    :param string:
    :param encoding:
    :return:
    """
    return md5_bytes(string.encode(encoding))


def md5_file(file):
    """
    md5 文件
    :param file:
    :return:
    """
    m = hashlib.md5()
    with open(file, 'rb') as obj:
        while True:
            data = obj.read(4096)
            if not data:
                break
            m.update(data)

    return m.hexdigest()


def str_bytes(string: str, **kwargs):
    """
    字符串转字节
    :param string:
    :param kwargs:
    :return:
    """
    return bytes(string, **kwargs)


def bytes_str(b: bytes, **kwargs):
    """
    字节转字符串
    :param b:
    :param kwargs:
    :return:
    """
    return str(b, **kwargs)


def plural(word: str):
    """
    字符串转复制
    :param word:
    :return:
    """
    if word.endswith('y'):
        return word[:-1] + "ies"
    elif word[-1] in 'sx' or word[-2:] in ["sh", "ch"]:
        return word + "es"
    elif word.endswith('an'):
        return word[:-2] + "en"
    else:
        return word + "s"


def write_file(path, content, mode="wb", **kwargs):
    """
    写入文件
    :param path:
    :param content:
    :param mode:
    :param kwargs:
    :return:
    """
    dir = os.path.dirname(path)
    if not os.path.exists(dir):
        os.makedirs(dir)
    if not os.path.isdir(path):
        with open(path, mode, **kwargs) as f:
            f.write(content)
        return True
    else:
        return False


def zip_files(files, zip_name):
    """
    压缩文件
    """
    if not os.path.exists(os.path.dirname(zip_name)):  # os模块判断并创建
        os.makedirs(os.path.dirname(zip_name))
    zip = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
    for file in files:
        zip.write(file)
    zip.close()
    return True


def count(arr: list):
    """计算数据元素总和"""
    total = 0
    for num in arr:
        total += num
    return total


def probability_extract(collect: list, key: str = "probability", type=None):
    """
    概率抽取
    :param collect: 抽取list
    :param key: 概率 键
    :param type: 概率 键
    :return:
    """

    def get(ob, key, default):
        if type == "dict":
            return ob.get(key, default)
        else:
            return getattr(ob, key, default)

    res = None
    sum = count([get(ob, key, 0) for ob in collect])
    if sum > 0:
        randint = random.randint(1, sum)
        for ob in collect:
            probability = get(ob, key, 0)
            if randint > probability:
                randint -= probability
            else:
                res = ob
                break
        return res
    return None


def get_file_content(path: str, mode="rb", **kwargs):
    """获取文件内容"""
    if os.path.isfile(path):
        with open(path, mode, **kwargs) as f:
            return f.read()
    return bytes()


def file_to_base64(path: str):
    """文件转base64"""
    import filetype
    content = get_file_content(path)
    base = base64.b64encode(content).decode()
    type = filetype.guess(path)
    return "data:%s;base64,%s" % (type.mime, base)


def filter_content(content: str):
    """过滤内容"""
    return re.sub(r'\s+', ' ', content)  # 去掉换行 和 多余空格


def convert(one_string, space_character="_"):
    """
    字符串转成驼峰
    :param one_string: 输入的字符串
    :param space_character: 字符串的间隔符，以其做为分隔标志
    :return:
    """
    string_list = str(one_string).split(space_character)  # 将字符串转化为list
    first = string_list[0].lower()
    others = string_list[1:]
    others_capital = [word.capitalize() for word in others]  # str.capitalize():将字符串的首字母转化为大写
    others_capital[0:0] = [first]
    hump_string = ''.join(others_capital)  # 将list组合成为字符串，中间无连接符。
    return hump_string


def convert_dict_key(data: dict, space_character="_"):
    """
    字典键转驼峰
    :param data:
    :param space_character:
    :return:
    """
    return {convert(key, space_character=space_character): value for key, value in data.items()}


def convert_key(data, space_character="_"):
    """
    字典键转驼峰
    :param data:
    :param space_character:
    :return:
    """
    if type(data) is dict:
        return {convert(key, space_character=space_character): convert_key(value) if type(value) is list or type(value) is dict else value for key, value in data.items()}
    elif type(data) is list:
        return [convert_key(datum) if type(datum) is dict else datum for datum in data]
    return data


def aes_encrypt(data, key="0000000000000000", iv="0000000000000000"):
    """
    AES的ECB模式加密方法
    :param data:被加密字符串（明文）
    :param key: 密钥
    :param iv: 密斯偏移量
    :return:密文
    """
    from Crypto.Cipher import AES
    BLOCK_SIZE = 16  # Bytes
    pad = lambda s: s + (BLOCK_SIZE - len(s) % BLOCK_SIZE) * chr(BLOCK_SIZE - len(s) % BLOCK_SIZE)
    key = key.encode('utf8')
    data = pad(data)
    cipher = AES.new(key, AES.MODE_EAX, iv.encode('utf8'))
    # 加密后得到的是bytes类型的数据，使用Base64进行编码,返回byte字符串
    result = cipher.encrypt(data.encode('utf8'))
    return base64.b64encode(result).decode('utf8')


def aes_decrypt(data, key="0000000000000000", iv="0000000000000000"):
    """
    :param data: 加密后的数据（密文）
    :param key: 密钥
    :param iv: 密斯偏移量
    :return:明文
    """
    from Crypto.Cipher import AES
    unpad = lambda s: s[:-ord(s[len(s) - 1:])]
    key = key.encode('utf8')
    data = base64.b64decode(data)
    cipher = AES.new(key, AES.MODE_EAX, iv.encode('utf8'))
    return unpad(cipher.decrypt(data)).decode('utf8')


def filter_None_key(data):
    """
    过滤值为None的键
    :param data:
    :return:
    """
    if type(data) is dict:
        return {key: value for key, value in data.items() if value is not None}
    elif type(data) is list:
        return [filter_None_key(datum) for datum in data]
    return data


def get_attr(object, name: str, default=None):
    """
    获取对象属性 多层 以.分割
    """
    _name = name.split('.')
    for name in _name:
        if type(object) is dict:
            object = object.get(name, default)
        else:
            object = getattr(object, name, default)
    if object is None:
        object = ""
    if object is True:
        object = 1
    if object is False:
        object = 0
    if type(object) is dict:
        object = json.dumps(object)
    return object


def export_file(sheet_name: str, export_name: str, col_items: dict = {}, db_list: list = [], is_header=True):
    """
    导出文件
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    is_header and ws.append([datum for datum in col_items.values()])
    # [ws.append([str(get_attr(db_obj, key, "")) for key in col_items]) for db_obj in db_list]
    for db_obj in db_list:
        if type(db_obj) is list:
            ws.append(db_obj)
        else:
            _list = []
            for key in col_items:
                _list.append(get_attr(db_obj, key, ""))
            ws.append(_list)

    if not os.path.exists(os.path.dirname(export_name)):
        os.makedirs(os.path.dirname(export_name))
    return wb.save(export_name)


def export_files(export_name: str, sheet_data: list):
    """
    导出文件
    """
    from openpyxl import Workbook
    wb = Workbook()
    for key, sheet in enumerate(sheet_data):
        ws = wb.create_sheet(sheet.get("sheet_name"), key)
        header_label = sheet.get("col_items", {})
        if sheet.get("is_label", True):
            ws.append([datum for datum in header_label.values()])
        for db_obj in sheet.get("db_list", []):
            if type(db_obj) is list:
                ws.append(db_obj)
            else:
                _list = []
                for key in header_label.keys():
                    _list.append(get_attr(db_obj, key, ""))
                ws.append(_list)

        if not os.path.exists(os.path.dirname(export_name)):
            os.makedirs(os.path.dirname(export_name))
    return wb.save(export_name)


def export_file_to_dict(sheet_name: str, export_name: str, col_items: dict, db_list: list):
    """
    导出文件
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([datum for datum in col_items.values()])
    [ws.append([str(db_obj[key] if key in db_obj else "") for key in col_items]) for db_obj in db_list]
    if not os.path.exists(os.path.dirname(export_name)):
        os.makedirs(os.path.dirname(export_name))
    return wb.save(export_name)


def import_file(file: str, sheet: int = 0) -> list:
    """读取excel文件内容"""
    from openpyxl import load_workbook
    wb = load_workbook(file)
    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    # 第一个表格的名称
    sheet_first = sheets[sheet]
    # 获取特定的worksheet
    ws = wb[sheet_first]
    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    content = []
    # 迭代所有的行
    for row in rows:
        # line = []
        # for col in row:
        #     line.append((type(col.value) ===str ? col.value.strip(): col.value ))
        line = [(col.value.strip() if type(col.value) == str else col.value) for col in row]
        content.append(line)
    return content


def export_demo(sheet_name: str, export_name: str, col_items: dict, db_list: list):
    """https://www.cnblogs.com/MDD-Blog/p/14187228.html"""
    from openpyxl import Workbook
    wb = Workbook()
    # ws = wb.active
    # ws.title = sheet_name
    ws = wb.create_sheet(sheet_name, 0)
    ws.append([datum for datum in col_items.values()])
    [ws.append([str(get_attr(db_obj, key, "")) for key in col_items]) for db_obj in db_list]
    if not os.path.exists(os.path.dirname(export_name)):
        os.makedirs(os.path.dirname(export_name))
    return wb.save(export_name)


def string_replace(content: str, str_dict: dict) -> str:
    """
    字典替换
    :param content:
    :param str_dict:
    :return:
    """
    for item, value in str_dict.items():
        content = content.replace(item, value)
    return content


def string_re_replace(content: str, str_dict: dict) -> str:
    """
    字典替换
    :param content:
    :param str_dict:
    :return:
    """
    for item, value in str_dict.items():
        content = re.sub(item, value, content)
    return content


def get_date_of_last_month(form="%Y-%m-%d"):
    """
    获取上月开始结束日期
    :param form 返回值显示格式
    :return: str，date tuple
    """
    today = datetime.date.today()
    year = today.year
    month = today.month
    if month == 1:
        begin_of_last_month = datetime.date(year - 1, 12, 1).strftime(form)
    else:
        begin_of_last_month = datetime.date(year, month - 1, 1).strftime(form)
    end_of_last_month = (datetime.date(year, month, 1) + datetime.timedelta(-1)).strftime(form)
    return begin_of_last_month, end_of_last_month


def get_qrcode(data, path, version=1, box_size=10, border=5, img_fill_color="black", img_back_color="white"):
    """
    生成二维码
    Args:
        data:
        path:
        version:
        box_size:
        border:
        img_fill_color:
        img_back_color:

    Returns:

    """
    from qrcode import QRCode
    # 创建QRCode对象
    qr = QRCode(version=version, box_size=box_size, border=border)
    # 设置二维码数据
    qr.add_data(data)
    # 填充数据并生成二维码
    qr.make(fit=True)
    img = qr.make_image(fill_color=img_fill_color, back_color=img_back_color)
    # 保存二维码图片
    img.save(path)
    return True


def clean_icode(text, restr=''):
    """
    过滤表情
    Args:
        text:
        restr:

    Returns:

    """
    try:
        # co = re.compile(u'['u'\U0001F300-\U0001F64F' u'\U0001F680-\U0001F6FF' u'\u2600-\u2B55]+')
        co = re.compile(u'['u'\U0001F300-\U0001F64F' u'\U0001F680-\U0001F6FF' u'\u2600-\u2B55' u'\U00010000-\U0010ffff]+')
    except re.error:
        co = re.compile(u'('u'\ud83c[\udf00-\udfff]|'u'\ud83d[\udc00-\ude4f\ude80-\udeff]|'u'[\u2600-\u2B55])+')
    return co.sub(restr, text)
