import os
import sys
from openpyxl.utils.cell import get_column_letter
import tomllib
from tkinter import filedialog
from fnschool import *
from fnschool.canteen.path import *
from fnschool.canteen.food import *
from fnschool.canteen.spreadsheet.base import *
from openpyxl.worksheet.datavalidation import DataValidation


class Purchasing(SpreadsheetBase):
    def __init__(self, bill):
        super().__init__(bill)
        self._path = None
        self.food_name_cols = ["商品名称", "食材名称", "食品名称"]
        self.food_name_col_name = None
        self.unit_name_cols = ["订货单位", "食材单位", "订购单位", "计量单位"]
        self.unit_name_col_name = None
        self.total_price_cols = ["总价", "折前金额", "折后金额", "总金额"]
        self.total_price_col_name = None
        self.xdate_cols = ["送货日期", "检查日期", "清点日期", "x日期", "日期"]
        self.xdate_col_name = None
        self.purchaser_name_cols = [
            "客户名称",
            "购买者",
            "购买者名称",
            "顾客名称",
            "下单单位名",
            "购入单位名",
        ]
        self.purchaser_name_col_name = None
        self.count_cols = ["总数", "数量", "下单数量", "订货数量", "发货数量"]
        self.count_col_name = None
        self.abandoned_cols = [
            "不计",
            "是不计",
            "未入库",
            "非入库",
            "不需入库",
            "是非入库",
        ]
        self.abandoned_col_name = None
        self.inventory_cols = [
            "盘存",
            "存余",
            "结余",
            "是结余",
            "剩余",
            "是剩余",
            "是盘存",
        ]
        self.food_class_col_name = "食材大类"
        self.inventory_col_name = None

    @property
    def food_classes(self):
        food_classes = self.bill.food_classes
        return food_classes

    def food_name_like(self, name, like):
        not_likes = None
        if "!" in like:
            like = like.split("!")
            not_likes = like[1:]
            like = like[0]

        result = None
        like_value = like.replace("*", "")
        if like.startswith("*") and not like.endswith("*"):
            result = name.endswith(like_value)
        elif like.endswith("*") and not like.startswith("*"):
            result = name.startswith(like_value)
        elif not "*" in like:
            result = like_value == name
        elif like.startswith("*") and like.endswith("*"):
            result = like_value in name

        if not_likes:
            result = result and not any(
                [self.food_name_like(name, nl) for nl in not_likes]
            )
        return result
        pass

    def get_food_class(self, name):
        food_classes = self.food_classes
        for fclass, name_likes in food_classes.items():
            for name_like in name_likes:
                if self.food_name_like(name, name_like):
                    return fclass
        return "蔬菜类"

    def set_col_names(self, columns):
        columns = list(columns)
        for column_name in columns:
            if column_name in self.food_name_cols:
                self.food_name_col_name = column_name
            if column_name in self.unit_name_cols:
                self.unit_name_col_name = column_name
            if column_name in self.count_cols:
                self.count_col_name = column_name
            if column_name in self.total_price_cols:
                self.total_price_col_name = column_name
            if column_name in self.abandoned_cols:
                self.abandoned_col_name = column_name
            if column_name in self.inventory_cols:
                self.inventory_col_name = column_name
            if column_name in self.purchaser_name_cols:
                self.purchaser_name_col_name = column_name
            if column_name in self.xdate_cols:
                self.xdate_col_name = column_name

        for col_name, col_names in [
            (self.food_name_col_name, self.food_name_cols),
            (self.unit_name_col_name, self.unit_name_cols),
            (self.count_col_name, self.count_cols),
            (self.total_price_col_name, self.total_price_cols),
            (self.xdate_col_name, self.xdate_cols),
            (self.purchaser_name_col_name, self.purchaser_name_cols),
        ]:
            if not col_name:
                print_error(
                    _("There should be column ({0}), please add it.").format(
                        "|".join(col_names)
                    )
                )
                exit()

    @property
    def path(self):
        if not self._path:
            print_info(
                _(
                    "{0} need a purchasing list file, "
                    + "and it's file type should be '.xlsx'. "
                    + "The column names of it:"
                ).format(app_name)
                + _(
                    ""
                    + "\n\tcolumn   type    example"
                    + "\n\t送货日期 Text    2024-03-01"
                    + "\n\t食材名称 Text    香菜"
                    + "\n\t数量     Number  20"
                    + "\n\t计量单位 Text    斤"
                    + "\n\t总价     Number  20.0"
                    + "\n\t购买者   Text    "
                    + "\n\t是不计   Text    y"
                    + "\n\t是结余   Text    y"
                )
            )
            print_info(_("Please select a purchasing file."))
            filetypes = ((_("Spreadsheet Files"), "*.xlsx"),)

            filename = filedialog.askopenfilename(
                title=_("Please select the purchasing file"),
                initialdir=(Path.home() / "Downloads").as_posix(),
                filetypes=filetypes,
            )
            if filename is None or filename == ():
                print_warning(_("No file was selected, exit."))
                exit()
                return None
            print_info(
                _('Purchasing list file "{0}" has been selected.').format(
                    filename
                )
            )
            self._path = filename
        return self._path

    def update_fclass(self):
        wb = load_workbook(self.path)
        sheet = wb.active
        headers = [
            h
            for h in [
                sheet.cell(1, ci).value for ci in range(1, sheet.max_column + 1)
            ]
            if h
        ]
        if self.food_class_col_name in headers:
            wb.close
            return

        merged_ranges = list(sheet.merged_cells.ranges)
        for cell_group in merged_ranges:
            sheet.unmerge_cells(str(cell_group))

        food_name_col_index = -1
        for h in headers:
            food_name_col_index += 1
            if h in self.food_name_cols:
                break
        if food_name_col_index < 0:
            print_error(_("Unable to find food name column, exitt."))
            exit()
        food_class_col_index = food_name_col_index + 1 + 1
        food_class_col_letter = get_column_letter(food_class_col_index)

        food_class_list_dv = DataValidation(
            type="list",
            formula1='"'
            + ",".join(["蔬菜类"] + list(self.food_classes.keys()))
            + '"',
        )
        sheet.add_data_validation(food_class_list_dv)

        sheet.insert_cols(food_class_col_index, 1)
        sheet.cell(1, food_class_col_index, self.food_class_col_name)
        food_len = 0
        for row_index in range(2, sheet.max_row + 1):
            food_name = sheet.cell(row_index, food_name_col_index + 1).value
            if not food_name:
                break
            sheet.cell(
                row_index, food_class_col_index, self.get_food_class(food_name)
            )
            food_class_list_dv.add(sheet.cell(row_index, food_class_col_index))
            food_len += 1
        wb.save(self.path)
        wb.close()
        print_info(
            _(
                'Column "{0}" has been updated, '
                + "please verify/modify it. "
                + "Feel free to open new issue if some "
                + "food with the wrong class ({1}). "
                + "(Press any key to check it)"
            ).format(self.food_class_col_name, get_new_issue_url())
        )
        input(">_ ")
        open_file(self.path)
        print_info(_("Ok, I checked it, it's ok. (Press any key to continue)"))
        input(">_ ")
        pass

    def read_pfoods(self):
        self.update_fclass()
        foods = pd.read_excel(self.path)
        self.set_col_names(foods.columns)
        _foods = []
        for __, food in foods.iterrows():
            _food = Food(
                self.bill,
                name=food[self.food_name_col_name],
                unit_name=food[self.unit_name_col_name],
                count=food[self.count_col_name],
                total_price=food[self.total_price_col_name],
                xdate=food[self.xdate_col_name],
                purchaser=food[self.purchaser_name_col_name],
                fclass=food[self.food_class_col_name],
            )
            if self.abandoned_col_name:
                _food.is_abandoned = not pd.isna(food[self.abandoned_col_name])
            if self.inventory_col_name:
                _food.is_inventory = not pd.isna(food[self.inventory_col_name])
            _foods.append(_food)

        foods = _foods
        foods = sorted(foods, key=lambda f: f.xdate)
        self.bill.foods = foods
        self.spreadsheet.preconsuming.pre_consume_foods()

        return foods
        pass
