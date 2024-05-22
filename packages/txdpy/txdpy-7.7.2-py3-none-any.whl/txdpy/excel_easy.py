import os.path
from .pytmysql import PyMySQL
import xlrd,xlsxwriter,zipfile
from .bk_179 import optstr,prvadepl
from file_ls import ls
from tqdm import tqdm

def gen_excel(data:list,save_path,header=None,is_optstr=True):
    """
    :param data:列表数据
    :param save_path:保存excel路径
    :param header:表头
    """
    if header:
        data = list(data)
        data.insert(0,header)
    workbook = xlsxwriter.Workbook(save_path if save_path.endswith('.xlsx') else f'{save_path}.xlsx')
    sheet = workbook.add_worksheet()
    for row_num, row_data in enumerate(tqdm(data,desc=f'{os.path.basename(save_path)}')):
        for col_num, col_value in enumerate(row_data):
            if col_value:
                if is_optstr:
                    col_value = optstr(col_value)
            if type(col_value) not in [int, float] and col_value is not None:
                col_value = str(col_value)
            sheet.write(row_num, col_num, col_value)
    workbook.close()

#读取execl数据
def read_excel(file_path,sheet_index=0):
    """
    :param file_path:文件路径或路径下第一个文件
    :param sheet_index:工作表索引默认第一个
    return:表格数据
    """
    if not file_path.endswith('.xlsx') and not file_path.endswith('.xls'):
        file_paths=ls(file_path)
        for path in file_paths:
            if path.endswith('.xlsx') or path.endswith('.xls'):
                file_path=path
                break

    datas = []
    if file_path.endswith('.xlsx'):
        from openpyxl import load_workbook
        try:
            workbook = load_workbook(filename=file_path,read_only=True)
            worksheet = workbook.worksheets[sheet_index]
            for row in worksheet.iter_rows():
                datas.append([prvadepl(cell.value) if isinstance(cell.value, float) else cell.value for cell in row])
        except zipfile.BadZipFile:
            from txdpy import webptablesl
            with open(file_path, 'r', encoding='utf-8') as f:
                datas = webptablesl(f.read(), '//table')
    elif file_path.endswith('.xls'):
        import xlrd
        workbook = xlrd.open_workbook(file_path)
        sheet_by_index = workbook.sheet_by_index(sheet_index)
        for row_idx in range(sheet_by_index.nrows):
            datas.append([prvadepl(cell) if isinstance(cell, float) else cell for cell in sheet_by_index.row_values(row_idx)])

    i = [i for i, l in enumerate(datas[0]) if not l and datas[0][i - 1]]

    data = []
    if i:
        i = i[-1]
        for l in datas:
            if any(l): data.append(l[:i])
    else:
        for l in datas:
            if any(l): data.append(l)
    return data