# coding:utf-8

import json, pymysql, re, xlrd,xlsxwriter,os
from typing import Union,List
from loguru import logger
from sshtunnel import SSHTunnelForwarder
from tqdm import tqdm
import zipfile

class ReadData:
    """
    读取mysql和.xlsx文件数据，提供一些方便的方法
    """
    def __init__(self,xlsx_mysql_sql:Union[str,List[List]],needful_field:List[str]=None,select_sql=None,xlsx_sheet_index=0,replace_th=True):
        """
        :param name:二维列表数据或者.xlsx文件路径或者mysql数据库数据表名称信息关键字或者mysql查询语句  比如：[['姓名','年齡'],['张三',26]]/内蒙古计划.xlsx/陕西计划、院校数据/select * from bk_school
        :param needful_field:需要保留的数据字段,只支持中文,表头字段名称会被(强制)自动转为中文表头名称,sql语句查询有筛选字段时,不支持保留字段  比如：['院校名称','专业名称']
        :param xlsx_sheet_index:.xlsx文件的索引默认第一个  比如：0
        :param select_sql:查询数据库数据表名时筛选部分的sql语句  比如：year=2022 and school_name like '%北%大学'
        :param replace_th:是否替换表头字段为中文，默认为True
        """
        self.data=[]#所有数据以二维列表存放
        self.replace_th=replace_th
        with open('c:/mysql_config.json', 'r', encoding='utf-8') as f:
            self.mysql_config=json.load(f)
        self.connect_mysql()  # 连接mysql数据库

        data_info = xlsx_mysql_sql
        if type(data_info) == list:
            if all(isinstance(sub_lst, list) for sub_lst in data_info):
                self.data = data_info
                self.data[0] = [str(f) for f in self.data[0]]  # 数据第一行必须作为列索引且为字段名称类型字符串
            else:
                raise ValueError('数据必须是二维列表')
        elif data_info.endswith('.xlsx') or data_info.endswith('.xls'):  # 判断是否为.xlsx文件读取数据
            self.read_excel(data_info, xlsx_sheet_index)

        else:
            self.read_mysql(data_info, select_sql)  # 当做查询数据库数据

        # 替换表头字段名称为中文
        self.data[0] = self.replace_th_name(self.data[0])

        if len(self.data) == 0:
            raise ValueError('表格数据为空，此类(ReadData)读取数据第一行必须用作为索引')

        self.column_index = {field: i for i, field in enumerate(self.data[0])}  # 创建列索引

        if needful_field:  # 判断是否有需要保留的字段，按保留的字段保留数据
            self.reserve_needful_field(needful_field)

        # self.abnormal_inspection(list(self.column_index.keys()))  # 异常数据检查

        for k, v in self.column_index.items():
            try:
                exec(f'self.{k}={v}')
            except:
                pass

        self.len_row = len(self.data)  # 行数
        self.len_col = len(self.data[0])  # 列数

    def connect_mysql(self):
        self.db = pymysql.connect(host=self.mysql_config["1"]['host'], port=3306, user='root',
                                  password=self.mysql_config["1"]['password'],
                                  database=self.mysql_config["1"]['database'])
        self.cursor = self.db.cursor()
        self.cursor.execute("""select * from 公司表名称对应关系""")
        self.table_dict = {k: v for k, v in (v[0].split(':') for v in self.cursor.fetchall())}

    def select_mysql_data(self, sql, table_name=None):
        if not table_name: table_name = extract_table_name(sql)
        if table_name in self.table_dict.values():
            if 'information_schema.COLUMNS' in sql:
                sql = sql.replace('需要替换的数据库名称', self.mysql_config["2"]['db'])
            server = SSHTunnelForwarder(
                ssh_address_or_host=eval(self.mysql_config["2"]['ssh_address_or_host']),
                ssh_username='root',
                ssh_password=self.mysql_config["2"]['ssh_password'],
                remote_bind_address=eval(self.mysql_config["2"]['remote_bind_address'])
            )
            # 启动隧道服务
            server.start()

            db = pymysql.connect(host='127.0.0.1', port=server.local_bind_port, user=self.mysql_config["2"]['user'],
                                 password=self.mysql_config["2"]['passwd'], database=self.mysql_config["2"]['db'])
            cursor = db.cursor()
            cursor.execute(sql)
            data = cursor.fetchall()
            db.close()
            cursor.close()
            server.close()
            return data
        if 'information_schema.COLUMNS' in sql:
            sql = sql.replace('需要替换的数据库名称', self.mysql_config["1"]['database'])
        self.cursor.execute(sql)
        return self.cursor.fetchall()

    def get_th(self, table_name):  # 获取表头字段
        table_name = self.table_dict.get(table_name, table_name)
        table_header = self.select_mysql_data(f"""SELECT COLUMN_NAME   
                                FROM information_schema.COLUMNS
                                WHERE TABLE_SCHEMA = '需要替换的数据库名称' AND TABLE_NAME = '{table_name}'   
                                ORDER BY ORDINAL_POSITION ASC;""", table_name)
        return [k[0] for k in table_header]

    def replace_th_name(self, lit): # 替换表头名称
        if not self.replace_th: return lit
        th_dict = {k: v for k, v in (v[0].split(':') for v in self.select_mysql_data("""select * from 表头字段对应名称"""))}
        new_name = []
        for name in lit:
            new_name.append(th_dict.get(name, name))
        return new_name

    def read_mysql(self, name, select_sql):
        if name[:6].lower() == 'select':
            table_name = extract_table_name(name)
            self.table_name = table_name
            table_name = self.table_dict.get(table_name, table_name)
            sql = re.sub(r'(?i)(FROM\s+)[^\s]+', r'\1' + f'`{table_name}`', name, 1)
            logger.info(sql)
            table_th = re.search('select\s(.+?)\sfrom', name, re.IGNORECASE).group(1)
            if table_th == '*':
                self.data.append(self.replace_th_name(self.get_th(table_name)))
            else:
                self.data.append(
                    self.replace_th_name(table_th.replace('`', '').replace('"', '').replace("'", '').split(',')))
            self.data += [list(v) for v in self.select_mysql_data(sql)]
        else:
            self.table_name = name
            table_name = self.table_dict.get(name, name)
            self.data.append(self.replace_th_name(self.get_th(table_name)))
            sql = f"""select * from `{table_name}`{' where ' + select_sql if select_sql else ''}"""
            logger.info(sql)
            self.data += [list(v) for v in self.select_mysql_data(sql)]

    # 读取execl数据
    def read_excel(self, file_path, sheet_index=0):
        """
        :param file_path:文件路径或路径下第一个文件
        :param sheet_index:工作表索引默认第一个
        return:表格数据
        """
        if not file_path.endswith('.xlsx') and not file_path.endswith('.xls'):
            file_paths = ls(file_path)
            for path in file_paths:
                if path.endswith('.xlsx') or path.endswith('.xls'):
                    file_path = path
                    break

        datas = []
        if file_path.endswith('.xlsx'):
            from openpyxl import load_workbook
            try:
                workbook = load_workbook(filename=file_path, read_only=True)
                worksheet = workbook.worksheets[sheet_index]
                for row in worksheet.iter_rows():
                    datas.append(
                        [prvadepl(cell.value) if isinstance(cell.value, float) else cell.value for cell in row])
            except zipfile.BadZipFile:
                from txdpy import webptablesl
                with open(file_path, 'r', encoding='utf-8') as f:
                    datas = webptablesl(f.read(), '//table')
        elif file_path.endswith('.xls'):
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet_by_index = workbook.sheet_by_index(sheet_index)
            for row_idx in range(sheet_by_index.nrows):
                datas.append([prvadepl(cell) if isinstance(cell, float) else cell for cell in
                              sheet_by_index.row_values(row_idx)])
        self.data = [row for row in datas if any(row)]
        self.data[0] = [str(f) for f in self.data[0]]  # 数据第一行必须作为列索引且为字段名称类型字符串

    # def abnormal_inspection(self, header):
    #     """
    #     数据异常检查
    #     """
        # if '' in header:
        #     raise ValueError('表头存在空字符串')
        # for char in ',', '-', '、':
        #     if char in ''.join(header):
        #         raise ValueError(f'表头名称不允许出现“{char}”')
        # for v in header:
        #     if is_num(v):
        #         raise ValueError('表头名称不允许出现纯自然数')
        # if len(header) != len(set(header)):
        #     raise ValueError('表头字段不允许出现出重复名称')

    def reserve_needful_field(self, needful_field):
        """
        保留数据所需字段
        """
        data = []
        for row in self.data:
            col = []
            for field in needful_field:
                col.append(row[self.column_index[field]])
            data.append(col)
        self.data = data
        self.column_index = {field: i for i, field in enumerate(self.data[0])}  # 创建列索引

    def sel_cl(self, row_mark: str = None, column_mark: str = None, da_ts=1):
        """
        筛选行和列
        :param row_mark:行索引只支持自然数 比如'1,10-20'
        :param column_mark:列索引支持自然数和字段名 比如'1,10-20','1,院校名称,专业代码-最低分,14'
        :param da_ts:默认为1横向显示数据，2为纵向显示数据
        """

        def diiini(i):
            if i == '':
                i = self.len_col
            elif not is_num(i):
                i = self.column_index.get(i)
                if not i:
                    raise ValueError(f'传入索引错误{i}')
            return int(i)

        def index_split(mark, typ):
            rcis = []
            for i1 in mark.split(','):
                i2 = i1.split('-')
                if len(i2) == 1:
                    rcis.append(diiini(i2[0]) if typ == 'col' else int(i2[0]))
                else:
                    [rcis.append(i) for i in range((diiini(i2[0]) if typ == 'col' else int(i2[0] or 0)), (
                        diiini(i2[1]) if typ == 'col' else int(i2[1] or self.len_row)) + 1 or 0)]
            return sorted(list(set(rcis)))

        ris = index_split(row_mark, 'row') if row_mark else range(self.len_row + 1)
        cis = index_split(column_mark, 'col') if column_mark else range(self.len_col + 1)

        if da_ts == 1:
            for ri in ris:
                if ri < self.len_row - 1:
                    row_data = []
                    for ci in cis:
                        if ci < self.len_col:
                            row_data.append(self.data[ri + 1][ci])
                    yield row_data
        else:
            for ci in cis:
                if ci < self.len_col:
                    row_data = []
                    for ri in ris:
                        if ri < self.len_row - 1:
                            row_data.append(self.data[ri + 1][ci])
                    yield row_data

    def group(self, *args):
        data = {}
        for row in self.data[1:]:
            key = ','.join([str(row[self.column_index[arg]]) for arg in args])
            if key in data:
                data[key].append(row)
            else:
                data[key] = [row]
        return data

    def save(self, fp=None):
        if not fp:
            fp = self.table_name
        workbook = xlsxwriter.Workbook(fp if fp.endswith('.xlsx') else f'{fp}.xlsx')
        sheet = workbook.add_worksheet()
        for row_num, row_data in enumerate(tqdm(self.data, desc=f'{os.path.basename(fp)}')):
            for col_num, col_value in enumerate(row_data):
                sheet.write(row_num, col_num, col_value)
        workbook.close()

    def __str__(self):
        return '你很牛逼，你很棒！'

    def __del__(self):
        if self.db:
            try:
                self.db.close()
                self.cursor.close()
            except:
                pass

# 提取数据表名称
extract_table_name = lambda sql: re.search('from\s([^\s]+)', sql, re.IGNORECASE).group(1).split('.')[-1].strip('`')
# 判断是否为纯数字
is_num = lambda s: type(s) == int or re.search('^([0-9]+)$', str(s))