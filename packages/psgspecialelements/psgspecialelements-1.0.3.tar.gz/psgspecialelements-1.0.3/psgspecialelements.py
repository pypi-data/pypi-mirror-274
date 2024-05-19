# -*- coding: utf-8 -*-
# !/usr/bin/env python
import ast

import PySimpleGUI as Sg
import pandas as pd
import numpy as np
import copy
import os
from openpyxl import load_workbook
import inspect
from dataclasses import dataclass

SYSTEM_ICON = './config/img/table_arrow9_transparent_64.ico'

INPUTFILE_SRTING_COLUMNS = '-string-columns'
OUTPUTFILE_RENAME_KEY = '-output'

# 表格tab中的最大阈值，表格数及每张表格的最大列数
MAX_TABLE_NUMBER = 15  # 最大配置，大于18运行错误
MAX_TABLE_SHOW_COLUMNS = 20


colornames = {
    'aliceblue':            '#F0F8FF',
    'antiquewhite':         '#FAEBD7',
    'aqua':                 '#00FFFF',
    'aquamarine':           '#7FFFD4',
    'azure':                '#F0FFFF',
    'beige':                '#F5F5DC',
    'bisque':               '#FFE4C4',
    'black':                '#000000',
    'blanchedalmond':       '#FFEBCD',
    'blue':                 '#0000FF',
    'blueviolet':           '#8A2BE2',
    'brown':                '#A52A2A',
    'burlywood':            '#DEB887',
    'cadetblue':            '#5F9EA0',
    'chartreuse':           '#7FFF00',
    'chocolate':            '#D2691E',
    'coral':                '#FF7F50',
    'cornflowerblue':       '#6495ED',
    'cornsilk':             '#FFF8DC',
    'crimson':              '#DC143C',
    'cyan':                 '#00FFFF',
    'darkblue':             '#00008B',
    'darkcyan':             '#008B8B',
    'darkgoldenrod':        '#B8860B',
    'darkgray':             '#A9A9A9',
    'darkgreen':            '#006400',
    'darkkhaki':            '#BDB76B',
    'darkmagenta':          '#8B008B',
    'darkolivegreen':       '#556B2F',
    'darkorange':           '#FF8C00',
    'darkorchid':           '#9932CC',
    'darkred':              '#8B0000',
    'darksalmon':           '#E9967A',
    'darkseagreen':         '#8FBC8F',
    'darkslateblue':        '#483D8B',
    'darkslategray':        '#2F4F4F',
    'darkturquoise':        '#00CED1',
    'darkviolet':           '#9400D3',
    'deeppink':             '#FF1493',
    'deepskyblue':          '#00BFFF',
    'dimgray':              '#696969',
    'dodgerblue':           '#1E90FF',
    'firebrick':            '#B22222',
    'floralwhite':          '#FFFAF0',
    'forestgreen':          '#228B22',
    'fuchsia':              '#FF00FF',
    'gainsboro':            '#DCDCDC',
    'ghostwhite':           '#F8F8FF',
    'gold':                 '#FFD700',
    'goldenrod':            '#DAA520',
    'gray':                 '#808080',
    'green':                '#008000',
    'greenyellow':          '#ADFF2F',
    'honeydew':             '#F0FFF0',
    'hotpink':              '#FF69B4',
    'indianred':            '#CD5C5C',
    'indigo':               '#4B0082',
    'ivory':                '#FFFFF0',
    'khaki':                '#F0E68C',
    'lavender':             '#E6E6FA',
    'lavenderblush':        '#FFF0F5',
    'lawngreen':            '#7CFC00',
    'lemonchiffon':         '#FFFACD',
    'lightblue':            '#ADD8E6',
    'lightcoral':           '#F08080',
    'lightcyan':            '#E0FFFF',
    'lightgoldenrodyellow': '#FAFAD2',
    'lightgreen':           '#90EE90',
    'lightgray':            '#D3D3D3',
    'lightpink':            '#FFB6C1',
    'lightsalmon':          '#FFA07A',
    'lightseagreen':        '#20B2AA',
    'lightskyblue':         '#87CEFA',
    'lightslategray':       '#778899',
    'lightsteelblue':       '#B0C4DE',
    'lightyellow':          '#FFFFE0',
    'lime':                 '#00FF00',
    'limegreen':            '#32CD32',
    'linen':                '#FAF0E6',
    'magenta':              '#FF00FF',
    'maroon':               '#800000',
    'mediumaquamarine':     '#66CDAA',
    'mediumblue':           '#0000CD',
    'mediumorchid':         '#BA55D3',
    'mediumpurple':         '#9370DB',
    'mediumseagreen':       '#3CB371',
    'mediumslateblue':      '#7B68EE',
    'mediumspringgreen':    '#00FA9A',
    'mediumturquoise':      '#48D1CC',
    'mediumvioletred':      '#C71585',
    'midnightblue':         '#191970',
    'mintcream':            '#F5FFFA',
    'mistyrose':            '#FFE4E1',
    'moccasin':             '#FFE4B5',
    'navajowhite':          '#FFDEAD',
    'navy':                 '#000080',
    'oldlace':              '#FDF5E6',
    'olive':                '#808000',
    'olivedrab':            '#6B8E23',
    'orange':               '#FFA500',
    'orangered':            '#FF4500',
    'orchid':               '#DA70D6',
    'palegoldenrod':        '#EEE8AA',
    'palegreen':            '#98FB98',
    'paleturquoise':        '#AFEEEE',
    'palevioletred':        '#DB7093',
    'papayawhip':           '#FFEFD5',
    'peachpuff':            '#FFDAB9',
    'peru':                 '#CD853F',
    'pink':                 '#FFC0CB',
    'plum':                 '#DDA0DD',
    'powderblue':           '#B0E0E6',
    'purple':               '#800080',
    'red':                  '#FF0000',
    'rosybrown':            '#BC8F8F',
    'royalblue':            '#4169E1',
    'saddlebrown':          '#8B4513',
    'salmon':               '#FA8072',
    'sandybrown':           '#FAA460',
    'seagreen':             '#2E8B57',
    'seashell':             '#FFF5EE',
    'sienna':               '#A0522D',
    'silver':               '#C0C0C0',
    'skyblue':              '#87CEEB',
    'slateblue':            '#6A5ACD',
    'slategray':            '#708090',
    'snow':                 '#FFFAFA',
    'springgreen':          '#00FF7F',
    'steelblue':            '#4682B4',
    'tan':                  '#D2B48C',
    'teal':                 '#008080',
    'thistle':              '#D8BFD8',
    'tomato':               '#FF6347',
    'turquoise':            '#40E0D0',
    'violet':               '#EE82EE',
    'wheat':                '#F5DEB3',
    'white':                '#FFFFFF',
    'whitesmoke':           '#F5F5F5',
    'yellow':               '#FFFF00',
    'yellowgreen':          '#9ACD32'}

# 当单元readonly时，无法修改背景色，需要分别设置不同颜色方案
specialcolordefineforuneditable = [("red", "white"), ("green", "white"), ("blue", "white")]
specialcolordefineforeditable = [("white", "red"), ("white", "green"), ("white", "blue")]
TAB_GROUP_EDITABLE = True


def columntypebyname(df: pd.DataFrame, columnnamelist: list) -> list:
    """
    将行值list按对应的列值类型进行转换
    :param df:                  dataframe
    :param columnnamelist:        list： 要处理的列名列表
    :return:                    list： 处理后的列类型 list
    """
    # 首先获取主类型
    typenamelist = [df[x].dtypes.name for x in columnnamelist]
    # 根据第一行数据将其中的"object"类型更换为更细的子类型
    # 需要找到该列第一个不为np.nan的值确定本列类型，完全无值或者全部空值使用'float64'
    # 首先获取相应列的空值矩阵df，不做全表检测，减少计算量
    dfnulljudge = df[columnnamelist].isnull()
    # 指定类的单元格值类型分布df， 后续可以用此细细化禁用第一个非空值判断类型的方案
    dftypejudge = df[columnnamelist].applymap(type)
    # 0、首先排除无数据和所有列均为非object的情况
    if df.shape[0] == 0 or 'object' not in typenamelist:
        typesubnamelist = typenamelist
    else:
        # 主类型是object并且有数据时
        typesubnamelist = []
        for x in zip(columnnamelist, typenamelist):
            typeset = set(dftypejudge[x[0]])
            # 1、非object，直接使用
            if x[1] != "object":
                typesubnamelist.append(x[1])
            # 2、其次全列nan则默认为float64，之后的else情况一定有非nan
            elif False not in dfnulljudge[x[0]].tolist():
                typesubnamelist.append("float64")
            # 3、全列类型set 后只有一种，直接用str(set.pop())
            elif len(typeset) == 1:
                typesubnamelist.append(str(typeset.pop()).replace('class ', ""))
            # 4、全列两个set，列中有nan， 其中nan个数和非nan个数与float个数一致，使用该类别
            elif len(typeset) == 2 and True in dfnulljudge[x[0]].tolist() \
                    and dfnulljudge[x[0]].tolist().count(True) == dftypejudge[x[0]].tolist().count(type(np.nan)):
                # 移除nan对应的类型
                typeset.remove(type(np.nan))
                typesubnamelist.append(str(typeset.pop()).replace('class ', ""))
            # 完全没有类型， 使用主类型
            elif len(typeset) == 0:
                typesubnamelist.append(x[1])
            # 6、本列不止一个非nan类型，全部通过set放到类型中,忽略其中的“str"简化为“s“
            else:
                typesubnamelist.append(
                    "-".join([str(x).replace('class ', "") if "str" not in str(x) else "<s>" for x in typeset]))
                # typesubnamelist.append(str(type(df.loc[df.index[dfnulljudge[x[0]].tolist().index(False)], x[0]])))

    return typesubnamelist


# 根据函数返回值，如果是错误则修饰器打印带函数名称的错误信息
def errprintwithtunctionname(func):
    """
    修饰函数，用于对df处理函数的返回值添加函数名信息
    首先执行函数，根据返回值[True, ....] 和 [False, errstr]
    :param func:                    # 修饰的函数
    :return:                        [True, ....] | [False, 修饰函数名 + errstr]
    """

    def run(*argc, **argv):
        result = func(*argc, **argv)
        if result[0] is True:
            return result
        else:
            return [False, "\n=========== 函数{}执行错误 ===========\n{}".format(func.__name__, result[1])]

    return run


# 用于输出错误信息， 并返回给主函数
def errprintwithname(errstr: str):
    """
    返回带函数名称的错误信息
    :param errstr:                      str：错误信息
    :return:
    """
    # print(inspect.stack()[1].function)
    # print(type(inspect.stack()[1].function))
    return [False, "\n=========== 函数{}执行错误 ===========\n{}".format(inspect.stack()[1].function, errstr)]


# 定义显示数据类型和内部数据类型的关系
VALIE_DATA_TYPE = {
    "字符串": "str",
    "浮点": "float",  # 等价于 float64
    "浮点(32位)": "float32",
    "浮点(64位)": "float64",
    "整型(32位)": "int32",
    "整型": "int64",
    "日期时间": "datetime",
    "日期": "date",
    "列表": "list",
    "python非字符串对象": "str2object"
}

# 快速提示消息显示的时间
AUTO_CLOSE_DURATION = 0.5
# GUI中表格处理编辑、新增模式时用到的定义
EDIT_METHOD = 1
ADD_METHOD = 2
# GUI中用到的默认字体
deffont = ("宋体", 15)
DEF_FT = ('黑体', 10)
SPECIAL_TABLE_DEF_FT = ('黑体', 10)
MAX_SPECIAL_TABLE_ROWS = 16
MAX_SPECIAL_TABLE_COLUMNS = 10
SPECIAL_TABLE_CELL_INPUT = 1
SPECIAL_TABLE_CELL_COMBOX = 2
SPECIAL_TABLE_CELL_TEXT = 3
DEF_CELL_SIZE = (10, 1)
DEF_WINDOW_SIZE = (1378, 390)
DEF_PAD = (1, 1)
DEF_BD = 0

# for i in range(1, 100):
#     w = (mainwindow[(tablekey, -1, -1)].char_width_in_pixels(("黑体", i), 'w'))
#     h = (mainwindow[(tablekey, -1, -1)].char_height_in_pixels(("黑体", i)))
#     print("{} : [{}, {}],".format(i, w, h))

FONT_SIZE_DICT = {
    1: [1, 2], 2: [2, 3], 3: [2, 4], 4: [3, 5], 5: [4, 7], 6: [4, 8], 7: [5, 9], 8: [6, 11], 9: [6, 12],
    10: [7, 13], 11: [8, 15], 12: [8, 16], 13: [9, 17], 14: [10, 19], 15: [10, 20], 16: [11, 21], 17: [12, 23],
    18: [12, 24], 19: [13, 25], 20: [14, 27], 21: [14, 28], 22: [15, 29], 23: [16, 31], 24: [16, 33], 25: [17, 33],
    26: [18, 35], 27: [18, 36], 28: [19, 37], 29: [20, 39], 30: [20, 40], 31: [21, 41], 32: [22, 43], 33: [22, 44],
    34: [23, 45], 35: [24, 47], 36: [24, 48], 37: [25, 49], 38: [26, 51], 39: [26, 52], 40: [27, 53], 41: [28, 55],
    42: [28, 56], 43: [29, 57], 44: [30, 59], 45: [30, 60], 46: [31, 61], 47: [32, 63], 48: [32, 64], 49: [33, 65],
    50: [34, 67], 51: [34, 68], 52: [35, 69], 53: [36, 71], 54: [36, 72], 55: [37, 73], 56: [38, 75], 57: [38, 76],
    58: [39, 77], 59: [40, 79], 60: [40, 80], 61: [41, 81], 62: [42, 83], 63: [42, 84], 64: [43, 85], 65: [44, 87],
    66: [44, 88], 67: [45, 89], 68: [46, 91], 69: [46, 92], 70: [47, 93], 71: [48, 95], 72: [48, 97], 73: [49, 97],
    74: [50, 99], 75: [50, 100], 76: [51, 101], 77: [52, 103], 78: [52, 104], 79: [53, 105], 80: [54, 107],
    81: [54, 108], 82: [55, 109], 83: [56, 111], 84: [56, 112], 85: [57, 113], 86: [58, 115], 87: [58, 116],
    88: [59, 117], 89: [60, 119], 90: [60, 120], 91: [61, 121], 92: [62, 123], 93: [62, 124], 94: [63, 125],
    95: [64, 127], 96: [64, 128], 97: [65, 129], 98: [66, 131], 99: [66, 132],
}

# ICON IMAGE Base64 file
icondict = {
    "HOUSE64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAF50lEQVRIiYWVX2wc1RWHf+ece+/szu7a47Vjx+s42CRA/hASAQFCEgcTgkjAjVryQFNXtJUqFfJQqe0DbZ+KKvEcVU1VpAYa+idSq1IKFFTVgUBccKAJSYkViC2TxCZZx2uv7V3Wu56Z24fZNU4aykhXGmnune9+v3N0L/AlDzEDAC/JZPDS/v1bsod++7M9u3cnAUCJ0Jetl//3kYnIWiuu54W/ePKJrV3DIwcnXnn1a11bu+KX6+r6Bs+eDYmIAFw7EIvFKJlM8hcCmBnWWhZjwj88/fS9D50bfqH/9ZfaBsq5ibaPPtmx6/7ulmE38erQuXOWKRJREv3fAojH45xKpei6ACKCtZabMpnw+R8/1dV95Ohf33y7LzW8LTWf2FTvDQ5dydW9eaqrZ3v30nwm8974TPHb8VjdrkKhsEk75sEg8I+JSCAi/wtYiCWdDn/5rccf2nni5AvH3u93L25vDNdvu8Fb1d7K0/WhPjdemHTfOrl16+13ZG7rufv+W9p574ab0tuD0PJYNv9cMpm0nufJVYCFWOLx8I8//MEDO//17sHj/Ucbzj/aMX/nfcu9zuYMnHgSbU0xKTSTHhotzKijH9x6g5nVD3x9nfPIfTerDz8afea9wcvvl8tlmpqaCtXiWMIw5KZly8Jf9e7d0f27w38ZmPrUXnx8bXn5inpv5FIdLs1YGH8KFeXZ1kTFyGNO6sIrF/P5F4+3FGdLvPknXwVMLA0ATU1N3NLSEhV5IZbGxvDArp27H/7HPw+dmByT7N5bg7VbOrxsVuF5vxctG7+BN05fwgdrfk7rVRY3t8xJsDQu2aLvF45+rFS+RBdSDX9/++TQO77vU6EwGwozk7WWxHXDw729PY/0HXn2dPZC4tPvbvRX3NPhtTUtQ25iBqpcwio3j/riEO5p9XFj+RQSDR7S6ZSybUpPTPnFXN+gWellMNnZ+efzo6NBZmmrklq3HNqz5ys7f3/4T/+hEmef3OyvvKvDW+K1QZTG5VwJL8tuxFd349hYgA+XPIq73AtI6RmIU2/TqQTplQmaKFGucuTf63esXr1uMpPpGzhxYla8pia7/95Nj+3pe+PgGVWxk9/bHLRv7PAaU60gHYMii9x0gPrOTdiyKgFz5WPcvmYV1pcHAKqAdIy0E0d9IiZ6uauuVChXev2dO+7u7Owotbe/RU/19Gx4ZnTsxbPDg61jP314rvW2ZfUNiWYQKwAWREC5UIQjAsfRoPIsyCSB8gxKbhrWAhYAgTA3N4Wx8fHKmd8M5KXvTPPaffsOSEtb21wq5mSGNjevuGXHusYGt4XYuCCSCEIKM8U55D+bQ75YQd5nTBXnkPcVtIlBm1h1LkPrpHUNK789Redn1fFxN31IvdzfP/038PefaNsg23R8nziuZRICRa3r+wGe/fVhTI1nobWCDUMABD+0+OZ3enHnxnWoVCogEIjFBkWhlTfeVHxtNf1o/4Hn3lVB4HMQhEEIzivtQMSAWQOwYCIEoY+gOINEZRocEmAtCEChAlT8EErFEAQEIgKRgJWGk6ifDwOaBAAFWzsiWEQ0SEw1/8iAQkY8ZsBJBZKoLgwAcxaiTDRf7OcAMWBisgglAtQIQAhisDgQqRowQUKBUQw3rhYKL2QRIASzgigHEmABQJ/fALYKWHSKgqIdiAEQgplBwnCMQrMxoGp0IMK8nQexBosDFiwyuPr8VFfhiEDVmCIhBgnBKIWkdgBWMBzik4KDXOUzKJFFEQFECqAvANQcWAxYG8BWDXyCoxW8pAFV76c1MYsEEcAGrAw4iADMGrQAoGsBkbqIA2GnGpFAhGG0IOkQQARrAaMY0yUBiQJLDCKIDLjWIMH1DagWkXIAG4JYQAI4WuC5GiCBBaAZSDgqqolyQP4iA2ZY68Pa8HoRMZgNRMwCgNlCaY2GlAsihrWAVoRUwYJZAWwgEkYGYmqFtlqbawC1biWORu2dGT40ZoK4BTMsABUQKmGZ3Gjb1TVR7o4Tw8jISHDy1OkyAPwXWfQkSWcWg6cAAAAASUVORK5CYII=',
    "TIMER64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAGgElEQVRIiaVVbUxb1xl+3nOvr++1Y2MbzJeB8Bk+CklDU7VTPjrIkoW00xoqRZVSlmrafk7VplWVqv3Ypkn5k1Vo035V2fajCsqqJo1SVU3TINKxJRVhNDVgMJhQwBiwjT+xjX3vOftjGO02qdPO73Oe933O+zzvI+H/OD0njvc/3X34BZMkP/e95/s6ykpdzsDjxWUAfOeO9L8AEhEAQNU0nP5O7/etFkv2+s1bQxuRyL2tTGaipbmps9xdVvF48cvFnTfsm4IzxsAYIwBQVbNHU9WGRDpzu+9sX++rFy9emPXPce+078O6mtp6ImjfmIEkSUREEuechBASAG5WlKbNzc18taeGXjj7/DsNDfU/PvPdU+2Li0vDDx+OP7udL0zqup77rwyKnTIAMAxDEJHh8Xh4U1OTYbfbkclmlrs6n7D9YOCVN00muWV+zo/llZWDNpvN2d52IEJEhR0s+evgRMSEEADAy8vL5XPn+g/Z7LZT3Ye7KzWLxTQx8Y9EKpn6m9vlUGempy+oFgs2o1FUVHl+k4zHPBWVFVld19O7eF8DhxCCqqqqxKVLl851P/XU64uBwLfWQ6vCMHTSdR2ZbBbEJCEr5g3f1GRFIZ9PWCzalGEY1+/d+3Tc558bISISxS53Z8AYIyEE+vv7Sy5fvvzLUpfrrU9HRvZ75xaQZiqEtRS0zwVDsSCTzVE8GrZwbtD+/fXBjXDkV29f+ePQ4cPdoWPHjr4sSZIWCoVWiIq6K1ZEVVWVGBoa+q0kST+7du0vhrX2AD3Te4a1tjVDcAOFbQMWu4KtWAbzvknhfziK0GKAuBCfEdFPjh49+nNNNZ+Px2IP3rk61Dc8PByX/vU7JAYHB3/oLCm5dO3au6Lt5IvU92I/M/M8woksgutRJDJZRDZiyORycDhc1Nb9LOWzaawuBjyqaj4X24wemp70yi6nazYajY1MTk1GWVExoqenp+TIkSOv//3+fXI0d9FzvSdZIhKBN7CMx0vLYCYFFus+GHoe8fAaTKoGa4kNTx7rRXPbE3xmZtady20/0CyWH733/s2Xb31wy78jUwKA4ydOnJ7xTbdtZgo4dqqPsolNTExOIZPLora+AZIQSG6E4HA44Kmrh2pWkI3HQQCePv5t7nS5IJlM3o8/Gb4yPDwcy2azBACMc47a2lp0dnb2htfX4PDUi+aWOkzN+iGbNcRWHuPDP/8Bqeg6XGVlyCRjcJTYkQyvYXl+BnbbPjS0dkgHDz2J0dHR09PT03WSJBlCCNphwIUQ5vz2dlVqK4tKTw0yGQ5buQfNHV04+dIFqIoZ77/9FoKBGVRX10CRJVRVV6O+sQmMG2AQKC0rAxFpQgjJMAwUVbrrVlNma0vLGwY0VRHzU58jvLQAGYCJEQZ++gZqGw7gxpXfQ1NMMDGCqpiQikWxODuN6NoqJNkEs6Jw7Nmku06WZXkbRClwA8Lg1HSwG654GmZFgQQOkS/g1dfeQDYVh8QAmQQkAloOtIAZjVBkBv8X40il07IQghUNu8uACSEKhYK+QIJjc20VigTwQhb6dgYyI0gkoMgM5eXlUBjBxAgobCO/lYJJYpBJiGg4DKvVGtI0LSmE2F3tEhFRMpkU7R0d3GKxvpJOJ5nDXY2FmUlkUwlUVlZCNZnAwMEEh2IiWFUZM94vsB5cBoFjK5U0blx/T3I4HO+mUqkbkUhEYoxxIQQkxpgQQsBqtX7Z0NjYsxZcqdcsFv7MybO0z2rF8twsSkrsKLFbYVUlZJJJBGamUVdbi9b2dtitmhj+5GPp0eeP4sFg8M3x8fEVxhjjnItdmRIR3blzh3u93l87HY7w2Mhttu73Gno2DX07A0WWEFwIwDfxCDIjyIwQj4bBuMHHx8bERx/dhtvt/l0wGLxf9JWxmyd7YyAUCi00NTenIcTZiQejrMxZond1HxFlZU6KhFYRXQ+hs7MDddVVopDPG38dGWZDV68yIrq5srLy2tjYmAFgd8BfWdfFyTO73c4HBgZe0jRt0O/317S2tomOzi7a39gIu82G2GYUG2shMen1ks/nM5xO5+DS0tIv7t69myviiT1NfzUPGGPgnJPD4RDnz5/v4JxfjEYjZ6wWa51JUSxmRWEFXc+l0+lIPp//LBAI/CmRSIwEg8FtXdf3xsB/LrCXiaqqvLS0FDU1NRWqqnatra2V53I5pbS0NOp2u+eXlpZmfT4fL25i/Bty8fwTRd0OV+xMEysAAAAASUVORK5CYII=',
    "CLOSE64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAE30lEQVRIiZ2VXYgdRRqGn6+quvucM/85iRoTNevMBJFEWY0GFQTBC1HBlaz/jMpoFFfXBdmFvdiLvRIEFRHFGBXMjUQhF/6Bol6sSNaIruCNir/R/Dlx5iRzck736e6qby/6JDlx9CIWFN10Ue/7vW+9X7XcDn8bryWPL2vERkNQQPj9Q72K7F3s7Hxb9bZ98L0bj91jt1y23kxNTxIEGUQ/aTYR6WW9cud/Prx01zf7/7FP5EHXHG7Y6bVTpBPLMSegCWKEEMKvkihgjEWDP+FbEjxTa1bjv9l/CsIKF3ypHhUDSFGACCKC956iKKjV6/hfkCjgUNK0TW1oCA3h+EJk8UUBYFCsQaSyRajArUWLnEONcTrT68nTLtZaEKmmMTiUlsREGy9HO0dgcL1y6lgtZrAsEYFexhwxq2buYfru+1mcOo+828UYg4rgUH7OSkY3zbDq1lkaV1yFP9TqEyy18jiBCMF7DjYmOOu+hxifnCSKItZuvp/F6fPJ05TEwE+dHhN33MfpGy4iFAVjf7qF8etvBV9y1IilBApGIMt6TExOM372JKqKqhLFMdOz93Jk6jx+bHVoztzLyj9eiHqP2Gq7O3UlGAuq1RwYDlUwhoChMdSAz3ZxaEeD8T/fBggaAnGtxpqZWdKFBSbOPLMCCQGJItJPdrHw4lOYRgNsBM6dSCDGErIuodtGkhoyPEr68U5svcbI1ZsQY0CV2vAw9ZGRKjEiSBTR/fQjDm9/AddcjqoSul182kYHVDhJauRffUH7wD7ilatxzVOwI6PM7XiJLO2x4rob0CgGVTSEKigidD94j/ltW9Dg0b0/4BfmyQ8ewKUdWLZ6wCIB9SXFXJvQ+hLkc6QeEznHf199jY1rpjh1w0ZUFTGm7z18/tSj2Hffor5shKLdhhJCADMcw7IlKRIkAqkJRIa4LPl6d5c/PPJkBd5vpArcArD+ue101l1Md08bFxuIBUlOyOUggUIAVIl94Kv5wKqtz7L+7r/0bRHEmApcFbwnHhljw6tv0b3kEtK5gDWmj/GbfQAWZbdaztjyPOfP3oN6D8GDCO133uDAvx9CyxKsRX1JMjbBBa+8Rnbl5RSpR35RfXUGfVLnYGFBcTfdwLo77yLkPYy14CLa773JngfuoNy7QOh2WPnw09WVkufUm8s598G/s+eT9wmBJZ1m+sVTFNBc4Wi8vJ3v//kAJk7AOhbf3MGezTfjWwuYCcv8s1s58K+/okWOxDGdjz5g7+YZtKRSoL+igCp5FKVntGk48sTTzDWb1C+4mB833wgETD2CELBjEfNbtyAjo4xdcz27N11L6B5GGoZQhN+26KiSoII9LebnJx9BkggzNIQkyfEdItiRQGvbM7S2bQHJMGN1NO8ds2dQhBORYBCjAFEE1kFSw0QxuAiTJCAGce64vz4gviTkOTJcErIMMRbyDIxg7bHTFnc47clcmpdj43VkeBRJEkytgdTqSL2OiRMkSRDroH9t4EtCUaBZhmYpIUurZ9pFfVnuX+w62xfjeq3D3/6vbifXrT1XkzgWdREmipA4RlwMUYRY21cg/X+lJ5gSbIHGOVovCHmOCSX7DrbMx599icIhVI2cA5c5mC1gbGnITm4oqAOr0PoOXs9g51HAGiITyCDByXDp4KuiaoESmP8/YC0Y5GajmEsAAAAASUVORK5CYII=',
    "PSG64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAMAAADXqc3KAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAMAUExURQAAABlbkiVhkyZikyFjmShjkyhlly1mli5nlylnmS5olyppnC5qmi5rmzBpmDBpmTFqmDFqmTFqmjNrmzFrnDNsmzJsnDJtnTFunjNvnzRsmTRtmzVunTVunjVvnzZwnjlxny5toDFvozJwoTNzpjVwoDRzpDRzqDd4rThyojp0ozl1pTt2pjt3pz11ozt4qDl4qTp6rTl6rzx4qD55qTx5qj97qz17rT58rD98rT9+rz5+sEB3pEF9rkB+r0F+sEF/sT2AtD6AtT6Btk+Aqk2CrEOBs0GBtEKCtEGCtUSBskWBs0SCtEWDtUaEtkWFt0aGt0KFuUCEukOGu0eFuEWHu0eJvEiGuUiHukiIuUiIukmJu0uKvEyKvFCAp1CCqlODrVKCrl2Kr1OEs1KGsFWGsFaIsVaPvFqKsl6NsmKNsWeTuG6YvHadvlySwV6UxF+YxW+bxW6dxnKewHGex3SdwHSfx3egwXGizHmgwHqkxnyjxHio0f/RMf7QMv/TMf/SMv3SNf7UOv7UO//UPP/UPf/UPv/VP//WPP/XPv/VQ//WQv7XQ//WRP/XRf/WR//YRv/YSP/YSf/YSv/ZS//aS//ZTf/aTP7aTf7aTv7bT//cT//bUP/cUP7cUv/cU/7cVf/eVf/fVv/eV/7dWP/eWP/eWf/fWv/fW//fYvzcaf/hW//gXP/gXv/gX//hYP/hYf/hY//iYP/jY//gZf/iZP7iZf/jZv/iZ//lZv/jaf/kav7ka//maf/ma//kbf/lbv7mbP/mbv/mb//id//mcP/ncv/nc//ld//ndv/meP/ocf/ocv/oc//odP/odf/odv/peP/pff/qfY+11ZSzz5G41qC81aW/1P/jgf/qiv/qjv7qoMnZ5szb587d6eDm2+fo1+7v3e/x3vXw1fHx3gAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJblQd8AAAEAdFJOU////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////wBT9wclAAAACXBIWXMAABcQAAAXEAEYYRHbAAAAGHRFWHRTb2Z0d2FyZQBwYWludC5uZXQgNC4xLjFjKpxLAAABzUlEQVQoU2P4jwNAJZIEuLJA9M2u/iNgAaiEELOgIFPc//+r6puaalvAQmAJdg4pPj4h5oT/2+raWtqaGmAS/PxC/IJAGdbEB7saW5pb20AyQAkFNiFhUSEgkGZNfjizua29u70XJJHr8j+eV0RGVkRMTJb56u2mvt7eSR0gCT2gPsbMGzbi8hJyPDl3OidPnDRlwhagRHbG/zTXe5WSqqqqmpzXb/VMmz5jztSVIDtSWFLvl3Jrampq8ZY8WThj1tx586ZCXFV9t1xRR1tbR6Lw0f6ZC+YvWDAb6tz/xUom+rrGymWPD8xaunjZ0oUgMZBEsYqZqampWsnTY/PWLF+xZhFIHCRRpW5raWFhUPT/3IJ1a9euW/H//5oTYAlDezs7Kwvv//+XbN6wcev6//+3/z8FltDwcrC3N8/7v3rHtu07Nv3/vxVo0CWQhJGPm5ubdf7/TXt279699//JnTA70j38fH19wv//33b00OGj+w6fPXz5KMRVTiH+/gHuFf//7zl+5szZs2fO7YPo+H/FOSIyPMqz5v//g+dAMocvQCX+XwsMjYmNdgSy9p0/d/bgRZAYWOL//4LgoDAwY+++02AaJoEJcEj8/w8A4UqG4COjF7gAAAAASUVORK5CYII=',
    "CPU64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAFi0lEQVRIiZ1WS2wbRRj+/5l92O6u7WycTW03xLEd6uBUjVO1hfbSByLQEwhxoIhbCwcuvApUiEOFhAQVEiAOvK4gqoojSEhtgUNApaJKG5qkdWpa1JfcZL3teu3sZmeGQ9ZVGhIJ8Z/+2Vn93//45psBWMMqlcoXxWLxEACAaZpju3btOkkIoZRSsnv37hOmaY4BABSLxUOVSuXzteJIq32klMqyLBuqqhoAAKqqGpTSpKIoSUQESZK6OnuKohiKohiUUpkxtvivWCvWBABEOp1+sr+//5V169ZtnJub+6FUKh3Rdf3hVqv1l6Zp5Ww2ezASifQ0Go3fhoaGjsZisYdardaM4zjTiEgAQHQC4j0HkQghAAC4oiiJRCKxBQBIs9m8oOt6iRASa7VaVwEAYrFYP+e85TjOpXg8PiyE4LZtn/F93wYAgogghOD3AYS+UFW1q1AovGoYxp4wGxIEgS2EEIQQCQCAcx7gkslCCB8AwLbt07Va7SPP8xqdWPdmIElSxDTNfZZlncrn828MDg6+VavVPvF9fy4Wi/X19fUdWJHMfSaEYJlMZgwRpVqtdtQwjD31ev2HIAgWpJGRkS8VRTEMw9g9OTm5v7u7+9GpqamXq9XqxwAAmzZt+oBzjpzzYC0QIQRDRJpIJLanUqmdw8PDX1mW9ZPv+5bkOM5FVVVTiURia1i24rruDQCAUqn09sDAwCHGGEdEadnwlgOJZT5BRMIYc5rN5iXP8+ax0y9N04qc84Vt27aduHjx4uuEED46Ovo95xxEOH1ExKWEhQh9DPe4JEl0fn7+14mJiecQUWo2m7MAgNQ0zb3d3d3bhoaGjrTb7Wld1x/p6uoa2bBhw4uyLGsAEFBKKSIi51xQSjFcIiICIQRDAhDXdWue502Vy+X3hRALqqr2SoODg2/KsmzE4/GNlNJ1nPOF9evXPxYEAbiue7lWq72rKIphmub+GzdufBeNRg1d14cZYx4hhBJClFQqNRbOQlBKo8lkcms+n48vLi5a0vj4+OOKoiTT6fQzjuNcJYRIQRCALMswOzv7LSEk0tPT85TjOBeCIKi12+1rtm3/ruv6FgDgAMB7e3vHgiAAQgh1HOfquXPnXr958+Zx3/dtshopltp7nyEiUtd1rxuG8URfX99B13Un2+32rKIo3ZzztRgMdOfOnT/mcrkX+vv79zcajVOapm3XNC3HGINoNNpnWdZJz/P+TiQSOzRNK6bT6WcjkUh/q9WaQUTIZrMHEFEjhECz2fzL9/2ZkZGRz0zT3JfNZp+WqtXq+5FIJJXL5V5kjLVDdgDnnMVisYFyufxVSFHgnO9gjDFElIvF4jth34ExxgCAIiIyxtq2bZ+5cuXK5wsLC3NSvV4/BQDCsqw/hBBBLpeTO+WF/KdhC0TIHAoAIggCjogYMnjpEBAi27Z96ezZsy90aCoVCoXXVFVNZbPZ/TMzMy9xzr1ljSdhYLHicN0DCkFYWKFnGMamUqn06fXr17/xPG9e0nV9Y6jnWqiAPCydrTm5laxY+pcCABdCcEqprmnag4qiWNLExMTBZWI3Ho/Hd2Qymb1CCBpm+V8AQJZluHPnzum5ubnx8+fPH+iI3apync/nX04mk9vDXihCiMX/K9drXTjJZDK5FRHJ3bt3/9R1/cH/e+Esb0FnkKK3t3ff5s2bv+7p6Rm7devWsXK5/GGhUDjsOM5kNBp9oFKpfKNp2kC9Xv9xdHT0eCaTed513fPhlYmd4CsBOiDQarVmu7q6KpZl/XLt2rVjQggvHo8PTE9PH242m1PpdPrRy5cvf3L79u2fo9GoyRi7U61W3wsDL5fv1V8VjLFF3/ct3/ctAADP86wgCBq+7zcAABljtud5FgCA7/uWLMvWai8KAIB/ACsf4Gh+DNwbAAAAAElFTkSuQmCC',
    "CAMERA64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAF4ElEQVRIiaWVS2xcVxnHf+fc58ydh2fGnthO7Dixa6etQtQKaChKqUCVCqFSN92yQEKqVFGExKZSJcQCZcWSTReIFXQDCKkIVEAqURVQCImdtLVJmiaxM+PXjOc99965557DYkwSVJAqcY7O5nv99T/6f98n+Kznwo3ngDeBFWD+gd2whuDrvHF6+7+l2Z8ZIDW/zbjWxPe+WOb8Yp52pPnp1SZ/WO+ewZO/Ac7+XwC+awU/Olfl22cmqGTGaVOBjRCC36+2nvlfeYILN5ZJzR+fms3ML5dcImU+HaUMx8quev3zFXsqYzFKDUIILAHvbQ146+9NtkLFZlcxiFKAHSzxHd44/Y7gwo29Y3ln6s1nJzl/Mk87TsdFH8URgiQ1tKKURD90SAFZR5IauLIz5OpOzOV6yL1OzCglQXLeBsovLxepBh6rjQRlDAIwBsQjGA/LCkAghAEDw3jsXS5n+cJsnienBvzyozYf7g1tjfixDUKUAhfhOHRHGsTD4kYYHsCYMRMBYARGHBoPg402tLopC6UMXzuhOYiUqPWTeZtGJH/2bo23HUmSaowxaGMwevyEZSEtMEahVESaJqg0QZsUKS1s4eC5rjpSzo/OPj6TOTWbE6V8hsVSloOwZ9sME2rDIVgCXAfPd/F9Fy/j4Gd8wiim1WhCMmC+5DI3nWO2nCHrWvRCxVZzyGZzaH24uevebw7155Zm5BMnyuJYJce1nUFgozUyq/EyFkHgUCxmKRZy5PMBtitp7TdYCFyerBb50mOTHK8UOFrxCDzohrDVHLJR74iLG7v2pY0Dc+n6phnEmsJUTiCFtFEpaIHRAq0hVYYkSYnjhP3dFtlRh3MnJnjhzHGCwGV/YFjdUShtsKWk6Gd5ZiXLyekCxyo18c7lOmsb9/VEuyK1NMImScBYCAHSgEBgDMRRiBn0OLtU5htPz5FKh19ca3G5NqDWSYhVim9bLFZ8nl8IeHE5z4un5+jHWrx7dUfubO1ru1KQNumYwaE4xnI0mngQslTx+crKFLbj8Ku1Jr9b26G/t0vY65MmCbHrsD5RotOZRqA5f2qCc49N8sl2n96dgUm6obYZKTAWCDlWuACDQcUjVhYnWKoGfNCIeW9jj+7uPvNZw5mlaaZyLlutkGu1AbWtXf7sWzy3kOfUdIGVuYJYr4dWchBKySgBYxjfsbQFgB6xXA2YLVjsdSI+rnexdcJCtcBctcj0ZIGF6QnmygFJFPHPWodeOOJE2eXkVB5HGpF2htJmNAITfKpjPVswmXUpWJDEMckgwi5kiaXH7XaK3U1RGpTrI72Ubm+IVoqSC5XAxRGg+xH2vxmIwy9CCIQQZD0fNZ4GeCikUBi/yN2BxI00UkBqIFQWTjaL0+3iCYMyoAHPcRCJujlmoFMwGmM0JlVobeN4LvuDEX0Fk77FkbxFbAxtbSMeTClBqlMskbBQsCh6klYIrWGCLYSxBf+QKI1IYtJ4QNTt0Nnbp7G9Tb8/pNaO6UaaJ6ZzPDuXp9/cR+sUy/dxggLS8xjFEUnngK8ul5gtOjSHCfVmSNweKmHSS4d9YEBIkBJpW1jSJjEWdzsj1uo9vjyf5/svLIJWXLx1j/pGCCMNnsXCdMA3n57h1eeXsG2LD3ZD7tR6tG83Wkl78GubUdLHdXPC87AMOJ6Pm3HRlsv20HClFlINXE7NFPnhS4+zXmuz140YqRTfsThWDjh1tMREwWe1lXDl1gG3rtWS3t3GK+bGa3UbpUK9dZAVOV861RJexiOTy+BnPRLf4ZPU5i97isiyeGqmyOmjxf9cdkAtgvdrA/56q8WlP91M7l+99630xmsXAWxS/ZJZr/9cWdZygpQjy0JmfUzGIbFdhhIanuFuxWXtSMDxis/RskfgWnQjxeZBzN1GxM16T6/+7U5//f2PXx1d/+7bj64nWP7JCsa8heFhQ4jDpkg1Xs5jZrHK/Mo01fkKlaNF/KzLsBfR3O7QrHdU4/7B1u3VrR9E11+/9yjDfwGSndm1qwVxegAAAABJRU5ErkJggg==',
    "CHECKMARK64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAD2ElEQVRIibWVe4hUVRzHP+feuTOz82gfrqwu5gM0zTQpS1yfoGQPyc3IjOwlYRElPcDICEP6Q6i/IvojCCIpTA0tSlSqJSvbTLZE0dbWxz4nXWdnxrk7M/feOY/+sHyEa7pt57/z45zP53x/B84R/B/jzeNLgeXADDHE4IeBV2Ihpo6LK1rzNqEhAW9oGw18UhelYeWUGHFL0tx5lqPZyBAI3mi9o8YRm16cWTlsVFLQVwjY2+4SSI2W+j8KXj+ybmwytO69xjo7Wyqzr8sldbaE60ksIdBlhTVo+KuHXppY5azftKzeNsbQkurntOuTKQQobYhFbCgPNsGaA5NDWm94ZlYV7fmAX3pcenIlTucDAqlJRENUxcJQLgwiwfMtYcpq4503JMJjq8M0d+XpyBRJnfUpBpJwyKYqFqbCcSCQg0gQyCeq4qHp90yr5Pd0kY6+ImnXJ1CaeDhEdSJCTSJKzLEHLXhu4oQEuWKZ79uzZAoX2hKPhOn+I6DtuEdfLriC4NE9L4CYhzEP8dH84Hz9kT0NBHLqvMlJmo5nyBQDylITj4RwM5rmw70orcEA0AL8Q/DgN8OBr/DltL8q64G1F52+obomwr6US7boE0hNhRPiVIdHx7H+EvA2sJ0tC3/+e8uFS27c/SS+7ElGrGkbnp5EfV0UArmGxt0Lzq/x5YzKWocz/T4FXyGEINvj0XE410QgJ7Fl4dqL4ecS3PVlJYgdllKzx04ZxqolY8h4mkm315JPl+z+nP8Bd++4hZ2LM/hyuokLCr7Eti28TJnOA5ndGLOUnYtLl+u2YMHnJ4BxY2bWsWj2SA72eoBBG4PnBvy2qwvpq81gVjhJp1Q7q9axLIFVMqSaz3ytfLWEpsbLwgFs6pc1o/R9+e7+eK9joSMWvjR4gSLA4FSGKLS7UyirUmRkbJFTG0VI6N17+oR0/bl8d/+A8HMJAG7bPB7BTmGL8TVz64mMiKGNQSuN0hqvq59CS59Kzq2zo8MrcH/s1V6qMIf9y5uvBL8gALj54xpgG5aYH589klB9BdoYjDY0XJ9k9HURPj2aRZ/ycL/tfouDK17+N/ilAoAbP6wAsRGLB8INI7BGJUAYLGEhLAtLCApfnDymc95NtD4eDMC8ZNiXzNKfSdLbt5K8N6o68nNMwoHqKCAwlkVwKI06ln2MtpWtVwMHBnjspHyNQO1Xe7pRbTmUEchCGbk/laKsdl0tfGBB51OKQM0hUD/ppk7kkTTy11NQku/TuUpdi+DKn/7wdyuAHzDcii0Uykwg/ezJoRMAVL9TCWwFjpJdvfpa4AB/Akx4zQw8GDagAAAAAElFTkSuQmCC',
    "COOKBOOK64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABQAAAAYCAIAAAB1KUohAAAACXBIWXMAAAsSAAALEgHS3X78AAADMUlEQVQ4jY1UQUgCWxS9782M5hQzle3NTRGEQdBipCKKECoEg4JaJAqCmyBsYauEaNOqjQtXLSKINuFiKMiwIMtF1iLK2kWRZLUoDRxo5s37ixfSt4+/u7xzz51z7r3nwc7Ojt1uBwCEEPwtWKXL5eIvLy+HhoYIIYIgCIJAKa0Do5RyHPfy8nJycnJ1dcUjhJxOZygUOj09zefzFoulDp5SihCKRqPLy8vJZBI4jgOAo6Mjj8cDABjj/6WdTqdDoRAAfJeyFn8MQohhGADAY4xFUSyVSpIkAYBpmgih+soRQmxm2GazbW5u7u7ujoyMKIrCmP+ePMdxv9nhSqXi8/lmZmb29vay2Syrs1gs8EM/QogQQgipBWOMOzs7397eWlpabDYbAMiyHAwGu7u7mQTWzu/3R6PRxsZG+HERvNVqjcVix8fHfX19Nzc3T09PHo+HUjo1NVUulx8fHwFgbW0tEolQSguFwtbWVpU/rlQqs7Ozc3NzqqrmcjmXy9Xe3m61WgcGBubn5wGgo6NjYWEBAEql0t3dHQBUx8ljjNva2orFYnNzM8/zBwcHFoslGo329/cXCgUA6OnpwRh/fHwsLS3lcjm2qm9wQ0NDPB7f398fHBx8eHjIZrOqqhaLRUmSwuFwPB53OBw+ny+dTn9+ftYujed5AEilUhMTE9U9saTX66WUJhKJmv0dHh4Gg0FgF4YxJoQwANNjGIaiKLFYbHp62ul0Li4umqb5H5crSVIymQwEAolEwu12s6SiKNfX15OTkwDgcDguLi4ikUgVUv0zCIJgs9lUVWWlrP3q6qrf72dfAaCrq2tjY0OW5RowTynVNM1qteq6XqW9srJiGAZCSNd1hNDt7W04HGZm+NeFiaKYTCa3t7fHx8fdbjez+9fXV7UR87Cu66Zp1oI1TQsEAl6vN51Os9smhCCEfpbWmMw0TZbBpmm+v7+3traWy2VKKdP825I/M7Isi6IIAFxTU9P6+nomk+nt7X19fX1+fsYY1/ez0+k8Pz+/v7/nMMblcnl4eDifz5+dnWmaVgfGolQq2e32sbGx7wcok8mMjo7C396wVCpFKSWE/ANWXYLwO0+V8wAAAABJRU5ErkJggg==',
    "DOWNLOAD64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAYAAADgdz34AAAACXBIWXMAAAsSAAALEgHS3X78AAAEl0lEQVRIia2WS2hdZRDHf/Odc+7NvbZpYky1Gk2rEbRa26pt8UWLitYHIqigCD426kIQxCfWhYu68U0RBBERXCq+WrXoorYIFWpbEY0lpk0NTWKtuUlzn+d834yLm5smqQspzurMcM785j+c+eYTTtUcESAzvhIAm/+azHacA4FodszAVNFTrUPm+Q6iecmUyJkEna5OiDCCXCtvJ2cnV+Ep+9R7/VIfQhmeryKeySywok+SSMMKMwqAihDXPIcPDDMURUQCgiPWjJCck6x87ZXXV3cXu3XTO5tkYOvAHbnIfZipTpghLdAMIEngi1cXvtlzwfrHpG0x5ismzsvo0E9D9z7z++M799s2EcSm67OUkAs5cpbzkkoMtPtAAzdXQ9zqjHkt1Ol5SHofx0KWYRUxrdiS3FlLtzz51wd7+v2OQl7qHnPtorUXS3ZxPRUKUT5x4mTDWu559LbCNS+9X9v025Duc4KoYdMAA7A4Mk92EMp/JFIZwR/rx9dL1teVdC2/Qe8yzQg+pS0JvLUzx3hjioPVQamGGlcu47KNq6qrPj+fsd+GeAEYA2SmRQiCNSJKP1Ad3IVaG0nnlWRxKqkkVlYxJxGZwhmFIo34U/fh0Hv4v6YYrY+ihYtkorDUNj+298GPvzv6ZRrkMzA/oyCXh9rEMOOHfiLfcx+5zhXkOnppswxEpJHVxdTjs0CycDHy9XcMlwc5a0E3EoTconOls/dyBsb6lYRLY4m/9T6blDgi8oHw3rPx83fesubl4oVPWFvXBUKoQzqB92Xitpite77n/k/epaN7AZO1CTIROtZ14fJC6ccS9ndGUhRLK0Eum1h2YGpH5eFfD47sjluzcFo+f+vp655F03alNhZhASMjloA1qtzedzab125kiw2QLhHaQ0zIFM2MztUdkBcqx1Lp+0o59NGRP49OVQs0Z3d6nEyMUMP8OGgVtAJaA19CagP4xn4e6DPuPhox1V9HTRFr/h9mRmWkwbJtGSsHK4xXq4cQGQDCDABM0ClEy6DlJiA9DLV90BgktirFzhrPXX0mT6Y9lAaqkAhRItRKGT3bjetTYd2aYM7JYcwm5wwaAP44hDyQYukokg5jliICZoFIoNjZ4Ol1HdhueOPgCLlFjt7twvo63HwztGuipml20lEBBlrGfBXzR5BsDGjOPBrAAkJKRKBwuuepNUXyP5/HN7tKXFGvcuMGY/3qhAO/NLCTJ7kFmIT0OPgjmAhiYKYIASFgGoCUyAILu+o8ckng0jSwsF1YuzxP0hYwm3tizwIIpKPQOIY4BXUYCiiYYWSIKYYHMoRAV1fKTddFxJKQOA/mmW9zFWRjoCmYw6R1lrcg2kxgAfCIeRxKMa+YBSw0Vc7fOScAZuAnMXWYE8yaIUFBDFSbS8sCgscsayZWD3jMAmhT7b8CnDPIeZw6RGTOLmwWFRALMA3BZvkamoBcwM3Zh7MA9Yb5I3v/YKoKTlr9sROKZVrlTGDWsylmkMTGxCQ4h0ObGaT1aRJzHsbtwJJmWSet0/9kIpB69gPbgersJA4oMm/pn6JlQI1/uWX87/YP06p9rkZQnAYAAAAASUVORK5CYII=',
    "GITHUB64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAMAAADXqc3KAAAABGdBTUEAALGPC/xhBQAAAwBQTFRFAAAADAwMDQ0NAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAhyjGAAAAQB0Uk5T////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////AFP3ByUAAAAJcEhZcwAADdUAAA3VAT3WWPEAAAAYdEVYdFNvZnR3YXJlAHBhaW50Lm5ldCA0LjEuMWMqnEsAAABzSURBVChTbYxRFoAgDMPQ+98Z1zbIeJqPbU3RMRfDECqyGpjMg6ivT6NBbKTw5WySq0jKt/sHrXiJ8PwpAAVIgQGkwABSYAApMIAUGEAalFmK9UJ24dC1i7qdj6IO5F+xnxfLu0jS0c7kqxd3Dk+JY8/5AKFrLuM7mfCAAAAAAElFTkSuQmCC',
    "RUN64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAMAAADXqc3KAAAABGdBTUEAALGPC/xhBQAAAwBQTFRFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAszD0iAAAAQB0Uk5T////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////AFP3ByUAAAAJcEhZcwAADdUAAA3VAT3WWPEAAAAYdEVYdFNvZnR3YXJlAHBhaW50Lm5ldCA0LjEuMWMqnEsAAABqSURBVChTpY5JDsAwCMTy/09TMGvFpVF9aAZPRHpkcXC7OIodPg0uCjPq+MwCrWRGKkiIvLyTqzw3aqoI73eqUNAoXBXlg4zudxF+NONfPIVvbSZPgww5oW0Vz8T4Lgbt/xbjia+rahR5AEYEg4vdzh2JAAAAAElFTkSuQmCC',
    "STORAGE64": 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAMAAADXqc3KAAAABGdBTUEAALGPC/xhBQAAAwBQTFRFAAAABwcHDQ0NDg4ODw8PFxcXGRkZGhoaGxsbHh4eIyMjJSUlJiYmJycnKCgoMTExMjIyNTU1NjY2Nzc3AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAouNksgAAAQB0Uk5T////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////AFP3ByUAAAAJcEhZcwAADdQAAA3UAe+RuhUAAAAYdEVYdFNvZnR3YXJlAHBhaW50Lm5ldCA0LjEuMWMqnEsAAAC5SURBVChTfZLbDsMgDEPpbb3TDv7/W7PYuAztYUeqhO2QAGowkXIMIeYkaSU4QsNBi4GcyhNINpTglmq4GWSphvy/ldkuLXZ4HmAxy3NmFJaA4guKGCwsjClfV05+fWdhYBtFw+amB292aygW3M7fsPTwjmadZkCvHEtWaAYTViBqVwgTA3tJVnB6D/xhaimItDhjMBvlhtFsaIafnEtOaAY/twAw/eslK70CbX8obUvgJNw9Jv0+Zh8D4s5+VAm/LwAAAABJRU5ErkJggg==',
    "LITTLEBEAR64": 'data:image/png;base64, iVBORw0KGgoAAAANSUhEUgAAABgAAAAYCAMAAADXqc3KAAADAFBMVEUAAAAAADMAAGYAAJkAAMwAAP8AKwAAKzMAK2YAK5kAK8wAK/8AVQAAVTMAVWYAVZkAVcwAVf8AgAAAgDMAgGYAgJkAgMwAgP8AqgAAqjMAqmYAqpkAqswAqv8A1QAA1TMA1WYA1ZkA1cwA1f8A/wAA/zMA/2YA/5kA/8wA//8zAAAzADMzAGYzAJkzAMwzAP8zKwAzKzMzK2YzK5kzK8wzK/8zVQAzVTMzVWYzVZkzVcwzVf8zgAAzgDMzgGYzgJkzgMwzgP8zqgAzqjMzqmYzqpkzqswzqv8z1QAz1TMz1WYz1Zkz1cwz1f8z/wAz/zMz/2Yz/5kz/8wz//9mAABmADNmAGZmAJlmAMxmAP9mKwBmKzNmK2ZmK5lmK8xmK/9mVQBmVTNmVWZmVZlmVcxmVf9mgABmgDNmgGZmgJlmgMxmgP9mqgBmqjNmqmZmqplmqsxmqv9m1QBm1TNm1WZm1Zlm1cxm1f9m/wBm/zNm/2Zm/5lm/8xm//+ZAACZADOZAGaZAJmZAMyZAP+ZKwCZKzOZK2aZK5mZK8yZK/+ZVQCZVTOZVWaZVZmZVcyZVf+ZgACZgDOZgGaZgJmZgMyZgP+ZqgCZqjOZqmaZqpmZqsyZqv+Z1QCZ1TOZ1WaZ1ZmZ1cyZ1f+Z/wCZ/zOZ/2aZ/5mZ/8yZ///MAADMADPMAGbMAJnMAMzMAP/MKwDMKzPMK2bMK5nMK8zMK//MVQDMVTPMVWbMVZnMVczMVf/MgADMgDPMgGbMgJnMgMzMgP/MqgDMqjPMqmbMqpnMqszMqv/M1QDM1TPM1WbM1ZnM1czM1f/M/wDM/zPM/2bM/5nM/8zM////AAD/ADP/AGb/AJn/AMz/AP//KwD/KzP/K2b/K5n/K8z/K///VQD/VTP/VWb/VZn/Vcz/Vf//gAD/gDP/gGb/gJn/gMz/gP//qgD/qjP/qmb/qpn/qsz/qv//1QD/1TP/1Wb/1Zn/1cz/1f///wD//zP//2b//5n//8z///8AAAAAAAAAAAAAAADZ9vIoAAAA/XRSTlP///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////8A9k80AwAAAYtJREFUeJxdkr1y5CAMx93lmsuzWG7AXd7DFaSx3ey9R1zJV4C3yT7INmivWNysHyRXsQ1qTjgfkzkxwwz6SX+QRMWfli5+Wb9OXO17Xu1jOw6t/w8kNSzjuJyOlrcvkDN7vQyH4rmBjd8yDs+/f5Vz7jucJ6BPMAzDK2/b3QLOwRFZULgD2668eehmct6FcPa9hSxgvevBA/pAIZDbl6kLeGgaC0QiEojI09U7gsuNq0HrxhU/XUnE3iZjiEzkCjRYCST6myQBqXeecKpjxXKBiAQ0EExtUqRIdDYkdUR0iAGRYoyYMtIbvkCuNl4MTJ24bWLq6hqzaEJf8TarmYBiQgWYc2IBc38XqX6myZAEJk652NVJIRVnOJPtgK7FJyQRln5V3NeedPNj9nQrKEqvLBTwU2vdtq0GkCdlr2xjGVcB8DSOT8c/GprmARcFuo0cewHqdDqu68YRQHIP66BKs8s8bvvUpVCt7eH0umhm0ALi8gGYN3UBgGZjpaVX8X3M+y9Y7PsHmp/jPzIEhv9pBFozAAAAAElFTkSuQmCC'
}

def isinnerrightclickmenu(currentwindow, currentevent) -> bool:
    """
    判断事件串是否为
    :param currentwindow： 当前事件的window
    :param currentevent:   当前读取的事件
    :return:               bool True：是innerrightmenuclick event， False:不是
    """
    # 如果事件不是字符串，直接返回不成功
    if not isinstance(currentevent, str):
        return False

    # 找寻字符串中代表右键菜单事件的关键字
    returnindex = currentevent.find(Sg.MENU_KEY_SEPARATOR)
    # 如果找不到，返回不成功
    if returnindex < 0:
        return False
    else:
        try:
            # 从字符串事件中恢复右键菜单元组事件并回写到主窗口
            currentwindow.write_event_value(eval(currentevent[returnindex + len(Sg.MENU_KEY_SEPARATOR):]), None)
        except Exception as err:
            # 回写失败则返回不成功
            print("反向回送表格内部右键事件{}失败：{}".format(currentevent, err))
            return False
        print("反向回送表格内部右键事件{}".format(currentevent))
        return True


class SgKeyGenerate(Sg.Text):
    """
    用于给自定义类Sg类产生key的隐藏实用类， 方便系统使用windows[tablekey].eventprocess(event, window, values) 及 update
    初始化
    SgKeyGenerate(eventprocessfunct, updatefunc, "", k=tablekey, visible=False)
    使用:
    if isinstance(event, tuple):
        windows[event[0]].eventprocess(event, window, values)
        windows[event[0]].update(.....)
    所有事件的处理，全部通过update方式传给类内部进行处理
    """

    def __init__(self, updatefunc, getfunc, *args, **kwargs):
        self.updatefunc = updatefunc
        self.getfunc = getfunc
        super().__init__(*args, **kwargs)

    # 用于设置类的各种参数
    def update(self, *args, **kwargs):
        # print("hidingkey udpate:{}, {}".format(args, kwargs))
        return self.updatefunc(*args, **kwargs)

    # 用于获取类的各种属性
    def get(self, **kwargs):
        # print("hidingkey get:type: {} args:{} kwargs:{}".format(attribute, args, kwargs))
        return self.getfunc(**kwargs)


class SpecialCombox(Sg.Combo):
    """
    自定义Combo用于过滤update text_color， background_color, 因为Sg.Combo没有此option
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    # 用于设置类的各种参数
    def update(self, *args, **kwargs):
        # 从combo 参数中去掉 text_color 和background_color, 因为combo不支持，后续可以改写相关方法，支持相关color的设置
        newkwargs = kwargs
        tempkwargs = {}

        # PysimpleGUI.py line: 15409
        # combostyle = ttk.Style()
        # if element.TextColor not in (None, COLOR_SYSTEM_DEFAULT):
        #     combostyle.configure(style_name, foreground=element.TextColor)
        #     combostyle.configure(style_name, selectbackground=element.TextColor)
        #     combostyle.configure(style_name, insertcolor=element.TextColor)
        # if element.BackgroundColor not in (None, COLOR_SYSTEM_DEFAULT):
        #     combostyle.configure(style_name, selectforeground=element.BackgroundColor)
        #     combostyle.map(style_name, fieldbackground=[('readonly', element.BackgroundColor)])
        #     combostyle.configure(style_name, fieldbackground=element.BackgroundColor)
        #
        # if element.button_arrow_color not in (None, COLOR_SYSTEM_DEFAULT):
        #     combostyle.configure(style_name, arrowcolor=element.button_arrow_color)
        # if element.button_background_color not in (None, COLOR_SYSTEM_DEFAULT):
        #     combostyle.configure(style_name, background=element.button_background_color)

        # 获取传入的“text_color"
        tempresult1 = newkwargs.get("text_color", None)
        # 如果存在就截获选项并调用configure进行设置
        if tempresult1 is not None:
            tempkwargs["foreground"] = tempresult1
            # 从参数中去掉"text_color"
            newkwargs.pop("text_color")
        # 获取传入的“background_color"
        tempresult2 = newkwargs.get("background_color", None)
        # 如果存在就截获选项并调用configure进行设置
        if tempresult2 is not None:
            tempkwargs["fieldbackground"] = tempresult2
            # 从参数中去掉"text_color"
            newkwargs.pop("background_color")
        if tempkwargs != {}:
            # self.TKCombo.configure(**tempkwargs) 此方法只能设置textcolor, 需要是用ttk底层的configure
            # 首先获取ttk.Style的configure需要的stylename
            styletype = self.TKCombo.configure("style")
            stylename = styletype[4]
            # print("stlye get: {}".format(styletype))
            # print("stlye name : {}".format(stylename))
            # 然后创建combostyle实例
            combostyle = Sg.ttk.Style()
            # 使用实例处理各个要设置的参数
            combostyle.configure(stylename, **tempkwargs)
            # combostyle.configure(stylename, foreground="red")
            # combostyle.configure(stylename, selectbackground="red")
            # combostyle.configure(stylename, insertcolor="red")
            # combostyle.configure(stylename, selectforeground="red")
            # combostyle.configure(stylename, fieldbackground="red")
            # combostyle.configure(stylename, arrowcolor="red")
            # combostyle.configure(stylename, background="red")
        # print("Special combox preprocess :{}".format(tempkwargs))
        super().update(*args, **newkwargs)


SPECIAL_TABLE_GET_MAXROWNUMBER = 'get_max_row_number'
SPECIAL_TABLE_GET_MAXCOLUMNNUMBER = 'get_max_column_number'
SPECIAL_TABLE_GET_CURRENTROW = 'get_current_row'
SPECIAL_TABLE_GET_CURRENTROW_CONTENT = "get_current_row_content"
SPECIAL_TABLE_GET_SPECIALEVENTLIST = 'get_specail_eventlist'
SPECIAL_TABLE_GET_RESULT_DICT = 'get_result_dict'
SPECIAL_TABLE_GET_EVENTINFLUENCEROW = 'event_influnce_row'
SPECIAL_TABLE_GET_SELECTED_ROWCOLUMN = 'get_selected_row_column'
SPECIAL_TABLE_GET_SELECTED_ROWCONTENT = 'get_selected_row_content'
SPECIAL_TABLE_GET_COLUMNN_NAME = 'get_column_name'

SPECIAL_TABLE_SET_ROWCOLOR = 'set_row_color'
SPECIAL_TABLE_SET_CELLVALUE = 'set_cell_value'
SPECIAL_TABLE_SET_ROWVALUE = 'set_row_value'
SPECIAL_TABLE_SET_CURRENTROW = 'set_current_row'
SPECIAL_TABLE_SET_WHOLEVALUE = 'set_whole_value'
SPECIAL_TABLE_SET_BIND = 'set_bind_event'
SPECIAL_TABLE_TUPLE_EVENTPROCESS = 'tuple_event_process'
SPECIAL_TABLE_SET_MOUSEWHEEL = 'set_mouse_wheel'
SPECIAL_TABLE_SET_ADD_NEWROW = 'set_new_row'
SPECIAL_TABLE_SET_SAVETOFILE = 'save_to_file'
SPECIAL_TABLE_SET_LOADFROMFILE = 'load_from_file'

SPECIAL_TABLE_GROUP_REFRESH = '_table_group_refresh_'  # 子类使用的命令，刷新整个group
SPECIAL_TABLE_GROUP_TRIGGER_TAB = '_table_group_tigger_tab_'

SPECIAL_TABLE_CHANGE_EVENT_TRIGGER = 'data_change_event_trigger'

SPECIAL_TABLE_SELECT_MODE_NONE = Sg.TABLE_SELECT_MODE_NONE  # 不能选择
SPECIAL_TABLE_SELECT_MODE_BROWSE = Sg.TABLE_SELECT_MODE_BROWSE  # 只能单选，但是再点击取消
SPECIAL_TABLE_SELECT_MODE_EXTENDED = Sg.TABLE_SELECT_MODE_EXTENDED  # 可以多选
SPECIAL_TABLE_LISTBOX_SELECT_MODE_SINGLE = Sg.LISTBOX_SELECT_MODE_SINGLE  # 模仿list选择，增加一个选项(点击后不取消）

MOUSEWHEEL_UP = "MouseWheel:Up"
MOUSEWHEEL_DOWN = "MouseWheel:Down"

ROWSTYLE_ROWCOLOR = 'rowcolor'
COLSTYLE_DTYPE = 'dtype'


class SpecialDf(object):
    """
    将dataframe数据进行各种df操作后返回，
    """

    def __init__(self,
                 dfdata: pd.DataFrame = None,
                 headings: list = None,
                 ):
        """
        :param dfdata:                原始数据，pd格式
        :type dfdata:                 dataframe格式
        :param headings:              数据列名
        :type headings:               list; [列名1， 列名2......]
        """
        # 用于存储输出给显示的list
        self.headings = headings
        # 用于存储完整的数据
        self.dfdata = None
        # 用于存储处理后的df，与tablevalue一致
        self.dffilter = None
        self.currentrow = None
        self.CURRENTCOLINDEX = 0
        self.STARTROWINDEX = 1
        self.ENDROWINDEX = 2
        self.LIKEINDEX = 3
        self.currentfilter = [None, None, None, None]
        # 数据来源，文件或者数据库表信息，导入、导出时做记录
        self.dataimportsource = ""
        self.dataexporttarget = ""

        # 初始化数据
        if dfdata is not None:
            self.dfdata = dfdata
            if headings is None:
                self.headings = list(self.dfdata)
            self.dffilter = self.dfdata
        else:
            self.dfdata = pd.DataFrame([], columns=headings)  # 出入df为None
        # 为了df显示，需要保存dfcolumns，然后使用headings
        self.dfcolumns = list(self.dfdata)
        self.dfdata.columns = self.headings
        self.text_color = "black"
        self.text_background_color = "white"
        self.head_text_color = "white"
        self.head_background_color = "dark blue"
        self.input_text_color = "blue"
        self.input_text_background_color = "white"
        self.selected_text_color = "dark gray"
        self.selected_background_color = "blue"
        self.combox_button_background_color = "blue"
        self.combox_button_arrow_color = "white"
        self.cell_editable = True
        self.defaultcolor = (self.input_text_color, self.input_text_background_color) if self.cell_editable \
            else (self.text_color, self.text_background_color)

        # 以下部分如果在GroupTab中切换保存到SpecicalTable层级，可以放回SpecialTable，
        # 主要是考虑SpecialGroupTable备份不同表格的动态数据，所以迁移到此
        self.showstartpoint = [0, 0]
        self.selectedrawrow = set()
        self.currentshowrow = set()    
        self.selectedrawcolumn = set()
        self.currentshowcolumn = set()
        self.currentrowpage = 0
        self.currentcolumnpage = 0
        self.colnamelist = self.headings
        self.colnameusedlist = [""] + self.headings
        self.colnamefilterlist = [""] + self.headings
        self.visible_column_map = [True for _ in range(len(self.headings))]
        self.column_edit_element = []
        self.infilterstatus = False
        self.filter_startrow = ""
        self.filter_endrow = ""
        self.filter_filter_column = ""
        self.filter_filter_like = ""
        self.filter_set_value_column = ""
        self.filter_set_column_used_value = ""
        self.v_slider_position = 0
        self.h_slider_position = 0
        self.indexcol = []
        self.col_value_scale = None
        self.visible_column_map = []
        self.rowstyleinfo = None
        self.columnstyleinfo = None

        # 用于保存每行的文本和背景颜色，填写到行格式信息字典中
        self.dfbuildrowstyleinfo()
        self.dfbuildcolumnstyleinfo()

    def dfbuildrowstyleinfo(self):
        # 用于保存每行的文本和背景颜色，填写到行格式信息字典中，只保存有非默认颜色的行信息，从每行都有的list改变为行号字典方式，
        # "row":{ROWSTYLE_ROWCOLOR: ("color", "backgroundcolor")}
        # self.rowstyleinfo = {}
        # [{ROWSTYLE_ROWCOLOR: ("color", "backgroundcolor")}, {}, ....]
        if self.dfdata is not None:
            self.rowstyleinfo = [{} for _ in range(self.dfdata.shape[0])]
        else:
            self.rowstyleinfo = None

    def refreshcolumntype(self, columnnamelist=None):
        """
        根据列名清单刷列类型
        :param columnnamelist:
        :return:                与列名匹配的列类型list
        """
        if self.dfdata is not None:
            currentcolumnsname = list(self.dfdata) if columnnamelist is None else columnnamelist
            currenttypelist = columntypebyname(self.dfdata, currentcolumnsname)
            for i, columname in enumerate(currentcolumnsname):
                self.columnstyleinfo[columname] = {COLSTYLE_DTYPE: currenttypelist[i]}
            return

    def getcolumnstype(self, columnnamelist=None) -> list:
        """
        获取列名对应的列类型
        :param columnnamelist:              列名称列表
        :return:                            对应列名称的列类型列表
        """
        if self.dfdata is not None:
            currentcolumnsname = list(self.dfdata) if columnnamelist is None else columnnamelist
            typelist = [self.columnstyleinfo[x].get(COLSTYLE_DTYPE, 'float64') for x in currentcolumnsname]
        else:
            typelist = []
        return typelist

    def dfbuildcolumnstyleinfo(self):
        # 用于保存每列的格式信息，首先保存类型信息
        if self.dfdata is not None:
            self.columnstyleinfo = {x: {} for x in list(self.dfdata)}
            self.refreshcolumntype(columnnamelist=None)
        else:
            self.columnstyleinfo = None

    def filterdf2table(self):
        return self.dffilter.values.tolist()

    def dfsetfilter(self, columnname=None, startrow=None, endrow=None, like=None):
        # if columnname is not None or startrow is not None or endrow is not None or like is not None:
        if [columnname, startrow, endrow, like] not in [[None, None, None, None]]:
            # 外部输入过滤条件 保存
            self.currentfilter[self.CURRENTCOLINDEX] = columnname
            self.currentfilter[self.STARTROWINDEX] = startrow
            self.currentfilter[self.ENDROWINDEX] = endrow
            self.currentfilter[self.LIKEINDEX] = like

        querystr = ""
        if self.currentfilter == [None, None, None, None]:
            # 无过滤条件
            self.dffilter = self.dfdata
        else:
            # 有过滤条件
            if (self.currentfilter[self.CURRENTCOLINDEX] is not None and
                    self.currentfilter[self.LIKEINDEX] is not None):
                querystr = " {}.str.contains('{}') ".format(self.currentfilter[self.CURRENTCOLINDEX],
                                                            self.currentfilter[self.LIKEINDEX])
            if (self.currentfilter[self.STARTROWINDEX] is not None and
                    self.currentfilter[self.ENDROWINDEX] is not None):
                querystr += " and " if querystr != "" else ""
                querystr += " index >= {} and index <= {} ".format(self.currentfilter[self.STARTROWINDEX],
                                                                   self.currentfilter[self.ENDROWINDEX])
        # print(querystr)
        setflag = False
        if querystr != "":
            self.dffilter = self.dfdata.query(querystr)
            setflag = True
        self.currentrow = None
        return setflag

    def dffilterclear(self):
        self.currentfilter = [None, None, None, None]
        self.dffilter = self.dfdata
        self.currentrow = None
        self.infilterstatus = False

    def dfbatsetcolumnvalue(self, columnname, value=None):
        # 对过滤项指定列进行批量赋值
        optiontemp = pd.get_option('mode.chained_assignment')
        # 暂时关闭告警
        pd.set_option('mode.chained_assignment', None)
        self.dffilter.loc[:, columnname] = value
        # 恢复告警
        pd.set_option('mode.chained_assignment', optiontemp)
        filterindex = 0
        self.currentrow = None

        # 将过滤表的处理结果一致化到原始数据df，注意filter的index和iloc的下标不同，
        for index in self.dffilter.index:
            # print(index)
            try:
                self.dfdata.iloc[index, :] = self.dffilter.iloc[filterindex, :]
                filterindex += 1
            except Exception as err:
                print("df copy from index: {}  to index :{} error:{}".format(filterindex, index, err))

    def setcurrentrow(self, tablerownumber):
        if tablerownumber is None:
            self.currentrow = None
        else:
            self.currentrow = self.dfdata.index[tablerownumber]

    def dfdeleterow(self):
        if self.currentrow is not None:
            # 首先删除原始表中的该项
            self.dfdata.drop(self.currentrow, axis=0, inplace=True)
            # 删除该行的rowstyleinfo
            self.rowstyleinfo.pop(self.currentrow)
            # 删除后重新排序
            self.dfreindex()

    def dfaddrow(self, rowvalue: list):
        # 判断数据量是否合法，
        if len(rowvalue) == len(self.headings):
            # 合法则在原始数据末尾添加
            # self.dfdata = self.dfdata.append(pd.Series(rowvalue, self.headings), ignore_index=True)
            # valueseries = pd.Series(rowvalue, self.headings)
            # self.dfdata = pd.concat([self.dfdata, valueseries], axis=1)
            # or
            self.dfdata.loc[len(self.dfdata.index)] = rowvalue
            # 增加该行的rowstyleinfo
            self.rowstyleinfo.append({})
            # 用当前过滤器重新过滤形成过滤数据和过滤列表
            self.dfreindex()

    def dfmoverow(self, targetrownumber):
        # 判断数据量是否合法，
        if self.currentrow is not None:
            # 获取当前行的值
            movevalueseries = self.dfdata.iloc[self.currentrow, :]
            # 获取当前行的styleinfo, 并删除
            temprowstyleinfo = self.rowstyleinfo.pop(self.currentrow)
            # movevalue = movevalueseries.tolist() 使用series 和 list都可以？
            # 删除当前行
            self.dfdata.drop(self.currentrow, axis=0, inplace=True)
            # 当前行值插入到指定行，该操作影响了df的columns，需要复原
            self.dfdata = pd.DataFrame(np.insert(self.dfdata.values, targetrownumber,
                                                 values=movevalueseries, axis=0))
            # 将当前行rowstyleinfo插入到目标行位置
            self.rowstyleinfo.insert(targetrownumber, temprowstyleinfo)
            self.dfdata.columns = self.headings
            self.dfreindex()

    def dfgetrowvalue(self):
        if self.currentrow is not None:
            return self.dfdata.iloc[self.currentrow, :]
        else:
            return None

    def dfsetrowvalue(self, rowvalue: list):
        # 判断数据量是否合法，
        if len(rowvalue) == len(self.headings):
            # 合法则在原始数据末尾添加
            # valueseries = pd.Series(rowvalue, self.headings)
            self.dfdata.iloc[self.currentrow, :] = rowvalue
            # 用当前过滤器重新过滤形成过滤数据和过滤列表
            self.dfsetfilter()
            self.dfreindex()

    def dfsetcellvalue(self, location: tuple, cellvalue: str):
        """
        设置单元格的值
        :param location:      元组 ， （x, y） 赋值的单元格位置
        :param cellvalue:     值
        :return:
        """
        # 判断坐标是否合法
        if location[0] <= self.dfdata.shape[0] and location[1] <= self.dfdata.shape[1]:
            self.dfdata.iloc[location[0], location[1]] = cellvalue
            # 用当前过滤器重新过滤形成过滤数据和过滤列表
            self.dfsetfilter()
            self.dfreindex()

    def dfsetwholevalue(self, dfdata: pd.DataFrame = None, headings: list = None):
        if dfdata is None and headings is None:
            print("dfsetwholevalue 输入数据均为空值！")
            return
        self.dfdata = dfdata if dfdata is not None else pd.DataFrame([], columns=headings)
        self.dfcolumns = list(self.dfdata)
        if headings is None:
            self.headings = list(self.dfdata)
        else:
            self.headings = headings
        self.dfdata.columns = self.headings
        self.dffilter = self.dfdata
        self.currentfilter = [None, None, None, None]
        self.dfsetfilter()
        self.dfreindex()
        self.dfbuildcolumnstyleinfo()
        self.dfbuildrowstyleinfo()

    def dfsetcolumnname(self, column: list, columnname: list):
        """
        设置列名
        :param column:
        :param columnname:
        :return:
        """
        for i, columnid in enumerate(column):
            # 修改heading
            self.headings[columnid] = columnname[i]

        self.dfdata.columns = self.headings
        self.dffilter.columns = self.headings

    def dfreindex(self):
        self.dfdata.reset_index(drop=True, inplace=True)
        # 用当前过滤器重新过滤形成过滤数据和过滤列表
        self.dfsetfilter()
        self.currentrow = None

    def getrowoutputstyle(self, rowcontent: pd.Series, rowid: int) -> list:
        """
        根据行号，将行styleinfo转换为实际的输出文件条件格式style
        :param rowcontent:          本行的内容，, 暂时未使用
        :param rowid:               int, row 行号
        :return: str
        """
        try:
            color = self.rowstyleinfo[rowid][ROWSTYLE_ROWCOLOR]
            # 在gui方式下，颜色可以有light，在excel输出时没有，系统默认替换为black，因此需要提前去掉light
            return ["color:{}; background-color: {}".format(
                color[0].replace("light ", ""), color[1].replace("light ", ""))] * self.dfdata.shape[1]
        except Exception as err:
            # print("行：{}无色彩信息：{}".format(rowid, err))
            return [""] * self.dfdata.shape[1]

    def getrowcolor(self, row: int) -> tuple:
        """
        获取指定行的颜色配置，没有时使用默认配置
        :param row:             int ： 行号
        :return:                tuple     color
        """
        try:
            color = self.rowstyleinfo[row][ROWSTYLE_ROWCOLOR]
        except Exception as err:
            # print("行：{}无色彩信息：{}".format(row, err))
            color = self.defaultcolor
        return color

    @property
    def maxrownumber(self):
        return self.dfdata.shape[0]

    @property
    def maxcolnumber(self):
        return self.dfdata.shape[1]

    def getmaxrownumber(self):
        return self.dfdata.shape[0]

    def getmaxcolnumber(self):
        return self.dfdata.shape[1]


class SpecialTable(object):
    def __init__(self, tablekey, specialdf: SpecialDf = None, dfshowdata: pd.DataFrame = None,
                 tabledatadictlist: list = None,  itemeditfunc=None, itemeditfuncargsdic: dict = {},
                 itemeditusecolvaluescale=False, indexcol=None, visible_column_map=None,
                 def_col_width=DEF_CELL_SIZE[0], auto_size_columns=True, max_col_width=20, cell_justification='r',
                 col_widths=None, size=DEF_WINDOW_SIZE, cell_size=DEF_CELL_SIZE, scrollable=False,
                 col_value_scale: list = None, filter_visible=False, edit_button_visible=True,
                 filter_set_value_visible=False, border_width=DEF_BD, use_inner_right_click_menu=False,
                 move_row_button_visible=True, visible=True, headings=None, right_click_menu=None,
                 vertical_scroll_only=True, expand_x=False, expand_y=False, select_mode=None, font=SPECIAL_TABLE_DEF_FT,
                 pad=DEF_PAD, max_rows=MAX_SPECIAL_TABLE_ROWS, max_columns=MAX_SPECIAL_TABLE_COLUMNS,
                 text_color='black', text_background_color='white', head_text_color='white', show_cell_tooltips=True,
                 head_background_color='dark blue', input_text_color='blue', input_text_background_color='white',
                 selected_text_color='dark gray', selected_background_color='blue', combox_button_arrow_color="blue",
                 combox_button_background_color="white", column_cell_edit_type_map=None,
                 v_slider_resolution: float = 0.1, h_slider_resolution: float = 0.1, slider_every_row_column=True,
                 v_slider_show=True, h_slider_show=True, relief=Sg.RELIEF_FLAT, cell_editable=True,
                 slider_relief=Sg.RELIEF_FLAT, disable_slider_number_display=False, datachangeeventtrigger=False,
                 showcolnamewithid=False, showcolumnname=True, display_row_numbers=True,
                 enable_events=False, enable_click_events=False, importsource=""):

        """
        表格编辑框类
        :param tablekey:                Used with window.find_element and with return values to uniquely identify
                                        this element to uniquely identify this element
        :type tablekey:                 str | int | tuple | object
        :param specialdf:               初始化时优先使用specialdf，其次使用dfshowdata + headings，
                                        再次使用tabledatadictlist + headings
        :param dfshowdata:               pd.DataFrame, 用于显示的数据
        :param tabledatadictlist:        dict格式的数据，dfshowdata数据为None时，使用tabledatadict的数据
        :param itemeditfunc:             单条记录编辑时用到的界面处理函数，输入参数需要用字典方式带入。以适应可变参数
        :param itemeditfuncargsdic:      dict:  itemeditfunc函数的参数，row编辑函数，
        :param itemeditusecolvaluescale:    bool: True, itemeditfunc 需要将col_value_scale传递过去，
        :param indexcol:                  索引列序号（从0开始,目前只支持单列索引， None为没有），
                                          如果有索引列，会根据索引列判断是否有重复值并强制提醒
                                          None|int|[int,int...]
        :param visible_column_map:        是否显示该栏，
        :type visible_column_map:         List[bool]
        :param def_col_width:              默认列宽
        :param auto_size_columns:          列宽根据初始化自适应
        :param max_col_width：              最大列宽
        :param col_widths:                该栏显示宽度
        :type col_widths:                  List[int]
        :param size:                        tuple(int,int) , 整个lay的尺寸
        :param col_value_scale:            list , None or [[], [], None, ......]
                                           None: 没有任何限制，全部列都可以修改
                                          （二维list， 对应每列的取值范围，用于列快速填充）
                                           [None]:该列不允许批量编辑
                                           []该列任意取值范围
                                           [a,b,c]不为空，该列取值限制范围
        :param use_inner_right_click_menu: bool: True, 显示并内部右键菜单
        :param filter_visible:             bool: True, 显示筛选功能，False：关闭筛选功能
        :param filter_set_value_visible:   bool: True, 显示列赋值功能，False：关闭列赋值功能
        :param edit_button_visible:        bool: True, 各种编辑按钮，False：关闭各种编辑按钮，
        :param move_row_button_visible:    bool: True, 各种移动行按钮，False：关闭各种移动行按钮，
        :param right_click_menu:           A list of lists of Menu items to show when this element is right clicked.
        :type right_click_menu:            List[List[ List[str] | str ]]
        :param vertical_scroll_only:       if True only the vertical scrollbar will be visible
        :param select_mode:                Valid values include:
                                            TABLE_SELECT_MODE_NONE          不能选择
                                            TABLE_SELECT_MODE_BROWSE        只能单选
                                            TABLE_SELECT_MODE_EXTENDED      可以多选
                                            TABLE_LISTBOX_SELECT_MODE_SINGLE      增加一个选项 (点击后不取消）
        :param font:                        窗口中使用的字体大小
        :param pad:                         窗口中元素间的pad
        :param max_rows:                   最大表行数
        :param max_columns:                最大表列数
        :param expand_x:                   横向自动扩展
        :param expand_y:                   纵向自动扩展
        :param show_cell_tooltips:         cell方式下显示tooltips
        :param text_color:                 表格文本颜色
        :param text_background_color:      表格文本背景色
        :param border_width:                 表格边界线宽度
        :param head_text_color:             标题及序号文本颜色
        :param head_background_color:       标题及序号背景色
        :param selected_text_color:         选中文本颜色
        :param selected_background_color:   选中文本背景色
        :param combox_button_arrow_color:   cell为combox时，下拉箭头的颜色
        :param combox_button_background_color:  cell为combox时，下拉箭头的背景色
        :param input_text_color:            可编辑框的文本颜色
        :param input_text_background_color:  可编辑框的背景颜色
        :param column_cell_edit_type_map:       None, 所有单元格使用默认Sg.Input,
                                            list: 长度和表格数据的列数一致（该种方式下，所有列在一个max_column 页内显示，
                                            即表格页不要横行左右翻转，
                                            SPECIAL_TABLE_CELL_INPUT, SPECIAL_TABLE_CELL_COMBOX, SPECIAL_TABLE_CELL_TEXT
        :param relief:                      cell的relief
        :param cell_editable:               表格单元是否可以直接编辑
        :param slider_relief:               slider relief  style.Use constants -
                        RELIEF_RAISED  RELIEF_SUNKEN  RELIEF_FLAT RELIEF_RIDGE RELIEF_GROOVE RELIEF_SOLID
        :param v_slider_resolution:         v  slider分辨率，默认0.1
        :param h_slider_resolution:         h slider 分辨率， 默认0.1
        :param slider_every_row_column:    是否可以逐行、逐列滚动
        :param v_slider_show:               bool, 是否显示垂直滚动条
        :param h_slider_show:               bool, 是否显示水平滚动条
        :param disable_slider_number_display:          bool: True 不显示， False：显示
        :param datachangeeventtrigger:      bool: 数据修改后是否向主窗口发送数据变化事件
        :param showcolnamewithid:           bool: True, 显示列号， False：不显示列号
        :param showcolumnname:              bool: True, 显示列名， False：显示列名
        :param display_row_numbers:                bool: True, 显示序号列， False：不显示
        :param enable_events:           Turns on the element specific events. Table events happen when row is clicked
        :type enable_events:            (bool)
        :param enable_click_events:     (NO USE) Turns on the element click events that will give you (row, col)
                                        click data when the table is clicked
        :type enable_click_events:      (bool)
        :param importsource:            str: 用于记录数据来源文件名（主要用于TabGroup的tab tips）
        :return:                           用于Sg.window的页面layout list
        """
        self.max_rows = max_rows
        self.max_columns = max_columns
        self.datachangeeventtrigger = datachangeeventtrigger
        self.use_inner_right_click_menu = use_inner_right_click_menu
        self.enable_click_events = enable_click_events
        self.enable_events = enable_events
        self.dataimportsource = importsource
        if itemeditfunc is None:
            self.itemeditfunc = self.itemeditgeneral
            self.itemeditfuncargsdic = {}
            self.itemeditusecolvaluescale = False
        else:
            self.itemeditfunc = itemeditfunc
            self.itemeditfuncargsdic = itemeditfuncargsdic
            self.itemeditusecolvaluescale = itemeditusecolvaluescale

        # 优先使用 specialtable
        # 本地化记录Specialdf
        self.specialdf = specialdf
        if self.specialdf is not None:
            if headings is not None:
                self.specialdf.headings = headings
                # self.specialdf.tabledatadictlist = dfshowdata.to_dict(orient='records')
        # 输入数据是df，只有dfshowdata和headings有效，其他全部从df和headings推算
        elif dfshowdata is not None:
            headings = headings if headings is not None else list(dfshowdata)
        # 数据输入格式是字典list，不是dfdat，使用字典格式相关的参数以及字段说明字典
        elif tabledatadictlist is not None and tabledatadictlist != []:
            headings = headings if headings is not None else list(tabledatadictlist[0])
            # 如果有数据 , dict 的list理论上可以直接转换成df，但是有顺序问题，
            # dfshowdata = pd.DataFrame(self.tabledatadictlist)         # 不采用此方式，因为转换后数据顺序有问题
            dfshowdata = pd.DataFrame([list(item.values()) for item in tabledatadictlist],
                                      columns=list(tabledatadictlist[0]))
        else:  # specialdf , dfshowdata 和 tabledatadictlist 都为空，只是为了创建空表，此时headings必须有值，
            # 对dfshowdata使用空df
            dfshowdata = pd.DataFrame([], columns=headings)
        if specialdf is None:
            self.specialdf = SpecialDf(dfdata=dfshowdata, headings=headings)
        # 以上完成后，形成self.specialdf 及 self.specialdf.headings ，！！！
        # 在Specialdf初始化时已经执行，此处不应该重复执行, 除非dfcolumns未赋值
        if self.specialdf.dfcolumns is None:
            self.specialdf.dfcolumns = list(self.specialdf.dfdata)
            self.specialdf.columns = self.specialdf.headings
        # 每项的合法范围，df方式下如果没有输入就用df生成默认
        if visible_column_map is None:
            self.specialdf.visible_column_map = [True for _ in range(len(dfshowdata.columns))]
        else:
            self.specialdf.visible_column_map = visible_column_map
        # 对于不足最大列数的部分，进行补足
        self.specialdf.visible_column_map += [True for _ in range(len(self.specialdf.visible_column_map),
                                                                  self.max_columns)]
        # 从values获取对应的字典keys名，便于save时获取结果dict
        self.specialdf.indexcol = indexcol
        self.specialdf.col_value_scale = col_value_scale
        self.filter_visible = filter_visible
        self.filter_set_value_visible = filter_set_value_visible
        self.edit_button_visible = edit_button_visible
        self.move_row_button_visible = move_row_button_visible
        self.tablekey = tablekey
        self.scrollable = scrollable
        self.visible = visible
        self.def_col_width = def_col_width
        self.auto_size_columns = auto_size_columns
        self.max_col_width = max_col_width
        self.showcolnamewithid = showcolnamewithid
        self.showcolumnname = showcolumnname
        self.display_row_numbers = display_row_numbers
        if col_widths is None:
            self.col_widths = [cell_size[0] for _ in range(len(self.specialdf.headings))]
        else:
            collen = len(col_widths)
            headlen = len(self.specialdf.headings)
            if collen == headlen:
                self.col_widths = col_widths
            else:
                if collen > headlen:
                    self.col_widths = col_widths[:headlen]
                else:
                    self.col_widths = col_widths + [cell_size[0] for _ in range(headlen - collen)]

        self.vertical_scroll_only = vertical_scroll_only
        self._tablelayout = None
        self.expand_x = expand_x
        self.expand_y = expand_y
        self.border_width = border_width
        self.font = font
        self.pad = pad
        self.lineheight = FONT_SIZE_DICT[self.font[1]][1]
        tablelendiff = {
            # filter_visible, filter_set_value_visible, edit_button_visible,  move_row_button_visible,
            (True, True, True, True): (6.5 + 3.5) * self.lineheight,
            (True, True, True, False): (6.5 + 3.5) * self.lineheight,
            (True, True, False, True): (6.5 + 3.5) * self.lineheight,
            (True, True, False, False): (6.5 + 0) * self.lineheight,
            (True, False, True, True): (4 + 3.5) * self.lineheight,
            (True, False, True, False): (4 + 3.5) * self.lineheight,
            (True, False, False, True): (4 + 3.5) * self.lineheight,
            (True, False, False, False): (4 + 0) * self.lineheight,
            (False, True, True, True): (0 + 3.5) * self.lineheight,
            (False, True, True, False): (0 + 3.5) * self.lineheight,
            (False, True, False, False): (0 + 0) * self.lineheight,
            (False, False, True, True): (0 + 3.5) * self.lineheight,
            (False, False, True, False): (0 + 3.5) * self.lineheight,
            (False, False, False, True): (0 + 3.5) * self.lineheight,
            (False, False, False, False): 10,
        }
        self.size = size
        self.window_size = self.size[0], self.size[1] - int(tablelendiff[(
            self.filter_visible, self.filter_set_value_visible,
            self.edit_button_visible, self.move_row_button_visible)]),
        self.cell_size = cell_size
        self.relief = relief
        self.slider_relief = slider_relief
        self.slider_every_row_column = slider_every_row_column
        if self.slider_every_row_column:
            self.v_slider_resolution = 1 / self.max_rows
            self.h_slider_resolution = 1 / self.max_columns
        else:
            self.v_slider_resolution = v_slider_resolution
            self.h_slider_resolution = h_slider_resolution
        self.v_slider_show = v_slider_show
        self.h_slider_show = h_slider_show
        self.cell_justification = cell_justification
        self.show_cell_tooltips = show_cell_tooltips
        self.disable_slider_number_display = disable_slider_number_display
        self.lay_col_size = \
            (self.window_size[0] - (45 if self.disable_slider_number_display else 66),
             self.window_size[1] - (1 if self.disable_slider_number_display else 10))
        self.table_col_size = (self.lay_col_size[0] - 6, self.lay_col_size[1] - 20)
        self.datalay_size = (self.table_col_size[0] - 15, self.table_col_size[1] - 15)
        # H_SLIDER_SIZE = (int(DEF_CELL_SIZE[0] * MAX_SPECIAL_TABLE_COLUMNS * 1.2) - 6, 10)
        # V_SLIDER_SIZE = (int(DEF_CELL_SIZE[1] * (MAX_SPECIAL_TABLE_ROWS + 1) * 1.2), 10)
        # slider的size是为字符个数而不是pixel，所以需要根据font进行换算
        self.h_slider_size = \
            (int(self.lay_col_size[0] / FONT_SIZE_DICT[self.font[1]][0]), 10)
        self.v_slider_size = \
            (int(self.lay_col_size[1] / FONT_SIZE_DICT[self.font[1]][1] / 1.4), 10)
        self.specialdf.text_color = text_color
        self.specialdf.text_background_color = text_background_color
        self.specialdf.head_text_color = head_text_color
        self.specialdf.head_background_color = head_background_color
        self.specialdf.selected_text_color = selected_text_color
        self.specialdf.selected_background_color = selected_background_color
        self.specialdf.input_text_background_color = input_text_background_color
        self.specialdf.input_text_color = input_text_color
        self.specialdf.combox_button_background_color = combox_button_background_color
        self.specialdf.combox_button_arrow_color = combox_button_arrow_color
        self.specialdf.cell_editable = cell_editable
        self.specialdf.defaultcolor = (self.specialdf.input_text_color, self.specialdf.input_text_background_color) \
            if self.specialdf.cell_editable else (self.specialdf.text_color, self.specialdf.text_background_color)

        if column_cell_edit_type_map is None:
            self.specialdf.column_cell_edit_type_map = [SPECIAL_TABLE_CELL_INPUT
                                                        if self.specialdf.cell_editable else SPECIAL_TABLE_CELL_TEXT
                                                        for _ in range(len(self.specialdf.headings))]
        else:
            # 如果非编辑方式，原来的INPUT换成TEXT
            self.specialdf.column_cell_edit_type_map = [
                SPECIAL_TABLE_CELL_TEXT if not self.specialdf.cell_editable and x == SPECIAL_TABLE_CELL_INPUT else x
                for x in column_cell_edit_type_map]

        celltypedict = {
            SPECIAL_TABLE_CELL_COMBOX: SpecialCombox,
            SPECIAL_TABLE_CELL_INPUT: Sg.Input,
            SPECIAL_TABLE_CELL_TEXT: Sg.Text
        }
        # 形成针对各个列的element
        if self.specialdf.column_cell_edit_type_map is not None and self.specialdf.col_value_scale is not None:
            self.specialdf.column_edit_element = []
            for item, celltype in enumerate(self.specialdf.column_cell_edit_type_map):
                elementtype = celltypedict.get(celltype, None)
                if elementtype is SPECIAL_TABLE_CELL_COMBOX:
                    # combox模式下，要确定有多选的list且内容不能为空，否则根据cell_editable 确认input或者text
                    if self.specialdf.col_value_scale[item] is not None:
                        if len(self.specialdf.col_value_scale[item]) > 0:
                            self.specialdf.column_edit_element.append(elementtype)
                        else:
                            self.specialdf.column_edit_element.append(
                                Sg.Input if self.specialdf.cell_editable else Sg.Text)
                    else:
                        self.specialdf.column_edit_element.append(
                            Sg.Input if self.specialdf.cell_editable else Sg.Text)
                else:
                    self.specialdf.column_edit_element.append(elementtype)
        else:
            self.specialdf.column_edit_element = [Sg.Input if self.specialdf.cell_editable
                                                  else Sg.Text for _ in range(len(self.specialdf.headings))]
        # 对于不足显示最大列的部分，补足Sg.input
        for _ in range(len(self.specialdf.column_edit_element), self.max_columns):
            self.specialdf.column_edit_element.append(Sg.Input)

        # # 用于保存每行的文本和背景颜色，填写到行格式信息字典中
        self.buildrowstyleinfo()
        self.buildcolumnstyleinfo()

        # combox只显示可以编辑的列名
        self.select_mode = select_mode if select_mode is not None else SPECIAL_TABLE_SELECT_MODE_NONE
        if self.specialdf.col_value_scale is None:  # 全部列都可以修改
            for i in range(len(self.specialdf.headings)):
                # 排除不显示的列
                if self.specialdf.visible_column_map is not None:  # 当该项会在表格显示
                    if self.specialdf.visible_column_map[i]:
                        # 当该项会在表格显示
                        self.specialdf.colnameusedlist.append(self.specialdf.headings[i])
                        self.specialdf.colnamefilterlist.append(self.specialdf.headings[i])
                else:
                    self.specialdf.colnameusedlist = [""] + self.specialdf.headings
                    self.specialdf.colnamefilterlist = [""] + self.specialdf.headings
        else:
            for i in range(len(self.specialdf.headings)):
                # 排除不允许批量编辑的列和不显示的列
                if self.specialdf.visible_column_map is not None:  # 当该项会在表格显示
                    # 过滤列不用考虑是否可以编辑
                    self.specialdf.colnamefilterlist.append(self.specialdf.headings[i])
                    if self.specialdf.col_value_scale[i] is not None:  # 当该项允许被批量编辑
                        if self.specialdf.visible_column_map[i]:
                            self.specialdf.colnameusedlist.append(self.specialdf.headings[i])
        self.specialdf.currentrow = None
        self.tablelay = None
        # mousewheel事件是否已经绑定的标志
        self.mousewheelbind = False
        # 组成函数处理字典
        self.eventprocessdic = {}
        # 列批量复制相关elements的key及事件
        # 对于数据表格改变后发送给主窗口的触发事件，如果又回到本类，说明主窗口没有处理，本类也不做处理
        self.TABLE_CHANGE_EVENT_TRIGGER = (self.tablekey, SPECIAL_TABLE_CHANGE_EVENT_TRIGGER)
        self.eventprocessdic[self.TABLE_CHANGE_EVENT_TRIGGER] = self.donothing
        self.COLUMNFILTER = (self.tablekey, '_colunmfilter_')
        self.eventprocessdic[self.COLUMNFILTER] = self.donothing
        self.ROW_START = (self.tablekey, '_row_start_')
        self.eventprocessdic[self.ROW_START] = self.updaterowscale
        self.ROW_END = (self.tablekey, '_row_end_')
        self.eventprocessdic[self.ROW_END] = self.updaterowscale
        self.COL_BAT_VALUE = (self.tablekey, '_col_bat_value_')
        self.eventprocessdic[self.COL_BAT_VALUE] = self.donothing
        self.COL_BAT_OK = (self.tablekey, '_col_bat_ok_')
        self.eventprocessdic[self.COL_BAT_OK] = self.columnbatsetvalue
        # 表格附加按钮的相关事件及处理函数
        self.EDITROW = (self.tablekey, '_edit_')
        self.eventprocessdic[self.EDITROW] = self.editrow
        self.ADDROW = (self.tablekey, '_add_')
        self.eventprocessdic[self.ADDROW] = self.addrow
        self.DELETEROW = (self.tablekey, '_del_')
        self.eventprocessdic[self.DELETEROW] = self.deleterow
        self.CLEARALL = (self.tablekey, '_clear_')
        self.eventprocessdic[self.CLEARALL] = self.clearall
        self.MOVEROWUP = (self.tablekey, '_up_')
        self.eventprocessdic[self.MOVEROWUP] = self.moverowup
        self.MOVEROWDOWN = (self.tablekey, '_down_')
        self.eventprocessdic[self.MOVEROWDOWN] = self.moverowdown
        self.MOVEROWTOTOP = (self.tablekey, '_top_')
        self.eventprocessdic[self.MOVEROWTOTOP] = self.moverowtotop
        self.MOVEROWTOBOTTOM = (self.tablekey, '_bottom_')
        self.eventprocessdic[self.MOVEROWTOBOTTOM] = self.moverowtobottom
        self.MOVEROWTOCERTAIN = (self.tablekey, '_moveto_')
        self.eventprocessdic[self.MOVEROWTOCERTAIN] = self.moverowtocertain
        self.TARGETROWNUMBER = (self.tablekey, '_torownumber_')
        self.eventprocessdic[self.TARGETROWNUMBER] = self.moverowtocertain
        self.FILTERCONDITIONOK = (self.tablekey, '_fillterok_')
        self.eventprocessdic[self.FILTERCONDITIONOK] = self.setfilter
        self.FILTERCONDITIONCANCEL = (self.tablekey, '_filltercancel_')
        self.eventprocessdic[self.FILTERCONDITIONCANCEL] = self.filterclear
        self.FILTERFRAME = (self.tablekey, '_filterframe_')
        self.FILTERSETVALUEFRAME = (self.tablekey, '_filtersetvalueframe_')
        self.CONTENTLIKE = (self.tablekey, '_contentlike_')
        self.eventprocessdic[self.CONTENTLIKE] = self.donothing
        self.COLCHOOSE = (self.tablekey, '_col_choose_')
        self.eventprocessdic[self.COLCHOOSE] = self.colnameselected
        self.V_SLIDER_KEY = (self.tablekey, '_v_slider_')
        self.eventprocessdic[self.V_SLIDER_KEY] = self.sliderclicked
        self.H_SLIDER_KEY = (self.tablekey, '_h_slider_')
        self.eventprocessdic[self.H_SLIDER_KEY] = self.sliderclicked
        self.EDIT_ROW_FRAME = (self.tablekey, '_edit_row_frame_')
        self.MOVE_ROW_FRAME = (self.tablekey, '_move_row_frame_')
        self.getattrdict = {
            SPECIAL_TABLE_GET_MAXROWNUMBER: self.specialdf.getmaxrownumber,
            SPECIAL_TABLE_GET_MAXCOLUMNNUMBER: self.specialdf.getmaxcolnumber,
            SPECIAL_TABLE_GET_SPECIALEVENTLIST: self.getspecialeventlist,
            SPECIAL_TABLE_GET_CURRENTROW: self.getcurrentrow,
            SPECIAL_TABLE_GET_CURRENTROW_CONTENT: self.getcurrentrowcontent,
            SPECIAL_TABLE_GET_RESULT_DICT: self.getresultdic,
            SPECIAL_TABLE_GET_EVENTINFLUENCEROW: self.getcommandinfluncerow,
            SPECIAL_TABLE_GET_SELECTED_ROWCOLUMN: self.getselectedrowcolumn,
            SPECIAL_TABLE_GET_SELECTED_ROWCONTENT: self.getselectedrowcontent,
            SPECIAL_TABLE_GET_COLUMNN_NAME: self.getcolumnname
        }

        # 用于记录非表格单元的事件及右键菜单的字典
        self.specialeventdict = {
            self.EDITROW: " 编  辑 ",
            self.ADDROW: " 增  加 ",
            self.DELETEROW: " 删  除 ",
            self.CLEARALL: "全部清除",
            self.MOVEROWUP: "记录上移",
            self.MOVEROWDOWN: "记录下移",
            self.MOVEROWTOTOP: "移至表首",
            self.MOVEROWTOBOTTOM: "移至表尾",
            self.MOVEROWTOCERTAIN: "移至行号"  # 该项对于right_click_menu 将清除
        }

        # 需外部特殊处理的事件list, 记录非表格单元的事件
        self.specialeventlist = list(self.specialeventdict)
        # 定义内部inner_right_click_menu, 只加入主要的编辑按钮，不考虑移动到（因为需要定义目标行号）
        # 单个窗口多个相同的menu处理的底层思路
        # https://www.pysimplegui.org/en/latest/#right-click-menus
        #
        # Keys for Menus
        # Beginning in version 3.17 you can add a key to your menu entries.
        # The key value will be removed prior to be inserted into the menu.
        # When you receive Menu events, the entire menu entry, including the key is returned.
        # A key is indicated by adding :: after a menu entry, followed by the key.
        # To add the key _MY_KEY_ to the Special menu entry, the code would be:
        # ['&Edit', ['Paste', ['Special::_MY_KEY_', 'Normal',], 'Undo'],]
        # If you want to change the characters that indicate a key follows from '::' to something else,
        # change the variable MENU_KEY_SEPARATOR
        # 根据此特性，将SpecialTable的默认right_click_menu关联相关编辑Button，
        # 如：（"删除"-->("删除::" + str(self.DELETEROW))， window事件处理中，先拆分为元组事件，然后再返还给SpecialTable类
        # 相当于self.specialeventlist 去掉最后一个事件
        self.inner_right_click_menu = [["", [
                                                "{}{}{}".format(value, Sg.MENU_KEY_SEPARATOR, key) for key, value in
                                                self.specialeventdict.items()
                                            ][0: -1]]]
        # 用于判断鼠标进去的Frame
        self.TABLE_FRAME = (self.tablekey, '_table_frame_')
        self.MOUSE_IN = 'mouse in'
        self.MOUSE_LEAVE = 'mouse leave'
        self.MOUSE_WHEEL = 'mouse wheel'
        self.MOUSE_CLICK = 'mouse click'
        self.mousein = False
        # 用于记录所有的类事件
        self.eventlist = [a for a in self.eventprocessdic.keys()]
        # 追加二维显示表的所有事件
        self.tablecellkeylist = \
            [(self.tablekey, x, y) for y in range(-1, self.max_columns)
             for x in range(-2, self.max_rows)]
        self.eventlist += self.tablecellkeylist
        # 优先获取及使用外部定义的right_click_menu
        self.right_click_menu = right_click_menu
        # 当外部没有定义right_click_menu并且允许使用内部默认right_click_menu时，改写right_click_menu
        if self.use_inner_right_click_menu and self.right_click_menu is None:
            self.right_click_menu = self.inner_right_click_menu
        if isinstance(self.right_click_menu, list):
            menulistlen = len(self.right_click_menu)
            self.row_right_click_menu = self.right_click_menu[0] if menulistlen >= 1 else None
            self.column_right_click_menu = self.right_click_menu[1] if menulistlen >= 2 else self.row_right_click_menu
            self.tabledata_right_click_menu = self.right_click_menu[2] \
                if menulistlen >= 3 else self.row_right_click_menu
        else:
            self.row_right_click_menu, self.column_right_click_menu, self.tabledata_right_click_menu = None, None, None

        # self.v_slider_range = (int(self.specialdf.dfdata.shape[0] / self.max_rows) - 1 +
        #                        (1 if (self.specialdf.dfdata.shape[0] % self.max_rows) > 0 else 0), 0),
        # self.h_slider_range = (0, int(self.specialdf.dfdata.shape[1] / self.max_columns) - 1 +
        #                        (1 if (self.specialdf.dfdata.shape[1] % self.max_columns) > 0 else 0)),

        # 形成开始、结束行号的combolist
        rowcombolist = [""] + [str(a) for a in range(0, self.specialdf.dfdata.shape[0])]

        self._tablelayout = []
        # 先添加过滤选项
        if self.filter_visible:
            self._tablelayout.append(
                [
                    Sg.Col([[Sg.Frame(
                        "按列值筛选" if self.filter_set_value_visible else '按列值筛选:', [[
                            Sg.Col([[Sg.Frame("开始行号", [
                                [Sg.Combo(rowcombolist, readonly=True, k=self.ROW_START, enable_events=True,
                                          pad=(10, 3), size=(8, 1))]],
                                              border_width=0,
                                              title_location=Sg.TITLE_LOCATION_LEFT
                                              if not self.filter_set_value_visible else Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[Sg.Frame("结束行号", [
                                [Sg.Combo(rowcombolist, readonly=True, k=self.ROW_END, enable_events=True,
                                          pad=(10, 3), size=(8, 1))]],
                                              border_width=0,
                                              title_location=Sg.TITLE_LOCATION_LEFT
                                              if not self.filter_set_value_visible else Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[Sg.Frame("筛选列名", [
                                [Sg.Combo(self.specialdf.colnamefilterlist, readonly=True, k=self.COLUMNFILTER,
                                          enable_events=False, pad=(10, 3), size=(20, 1))]],
                                              border_width=0,
                                              title_location=Sg.TITLE_LOCATION_LEFT
                                              if not self.filter_set_value_visible else Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[Sg.Frame("包含内容", [
                                [Sg.Input("", readonly=False, k=self.CONTENTLIKE,
                                          enable_events=False, pad=(10, 3), size=(20, 1))]],
                                              border_width=0,
                                              title_location=Sg.TITLE_LOCATION_LEFT
                                              if not self.filter_set_value_visible else Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[
                                Sg.Button('筛选确认', k=self.FILTERCONDITIONOK, pad=(3, 1)),
                                Sg.Button('筛选取消', k=self.FILTERCONDITIONCANCEL, disabled=True, pad=(3, 1))]])
                        ]], pad=(10, 3), visible=self.filter_visible, k=self.FILTERFRAME,
                        border_width=None if self.filter_set_value_visible else 0,
                        # title_color='blue',
                        title_color='white' if self.filter_set_value_visible else 'blue',
                        title_location=Sg.TITLE_LOCATION_LEFT
                        if not self.filter_set_value_visible else Sg.TITLE_LOCATION_TOP)]], justification='c'),
                    Sg.Col([[Sg.Frame(
                        "列批量赋值", [[
                            Sg.Col([[Sg.Frame(
                                "赋值列名", [[Sg.Combo(self.specialdf.colnameusedlist, readonly=True, k=self.COLCHOOSE,
                                                   enable_events=True, pad=(10, 3), size=(20, 1))]],
                                border_width=0, title_location=Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[Sg.Frame(
                                "赋值内容",
                                [[Sg.Combo([], readonly=False, k=self.COL_BAT_VALUE, pad=(1, 1), size=(20, 1))]],
                                border_width=0, title_location=Sg.TITLE_LOCATION_TOP)]]),
                            Sg.Col([[Sg.Button(' 赋值 ', k=self.COL_BAT_OK, pad=(10, 3))]])
                        ]], pad=(10, 3), visible=self.filter_set_value_visible and self.filter_visible,
                        title_color='white', k=self.FILTERSETVALUEFRAME,
                        title_location=Sg.TITLE_LOCATION_TOP)]], justification='c')]),

        # 添加表格, 不同方式下加入的layout不同
        self._tablelayout.append(self.elementtablelay())
        # 用来产生tablekey的隐藏element
        layhiding = [Sg.Col([[SgKeyGenerate(self.update, self.get, "", k=self.tablekey, visible=False)]],
                            visible=False, justification='center', element_justification='center')]
        self._tablelayout.append(layhiding)
        # 添加编辑按钮
        if self.edit_button_visible or self.move_row_button_visible:
            self._tablelayout.append([
                Sg.Col([[
                    Sg.Frame("", [[
                        Sg.Button(self.specialeventdict[self.EDITROW], k=self.EDITROW, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.ADDROW], k=self.ADDROW, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.DELETEROW], k=self.DELETEROW, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.CLEARALL], k=self.CLEARALL, pad=(1, 1)),
                    ]], pad=(1, 1), visible=self.edit_button_visible, k=self.EDIT_ROW_FRAME),
                    Sg.Frame("", [[
                        Sg.Button(self.specialeventdict[self.MOVEROWUP], k=self.MOVEROWUP, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.MOVEROWDOWN], k=self.MOVEROWDOWN, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.MOVEROWTOTOP], k=self.MOVEROWTOTOP, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.MOVEROWTOBOTTOM], k=self.MOVEROWTOBOTTOM, pad=(1, 1)),
                        Sg.Button(self.specialeventdict[self.MOVEROWTOCERTAIN], k=self.MOVEROWTOCERTAIN, pad=(1, 1)),
                        Sg.Input('0', size=(5, 0), justification='r', k=self.TARGETROWNUMBER, pad=(1, 1))
                    ]], pad=(1, 1), visible=self.move_row_button_visible, k=self.MOVE_ROW_FRAME)]],
                    justification='center', visible=self.edit_button_visible or self.move_row_button_visible)
            ])

    def usedfdatainit(self, dfshowdata):
        # 只是用dfdata初始化，无法知道取值范围
        self.specialdf.col_value_scale = None
        # 过滤用的变量
        self.specialdf.colnameusedlist = [""] + self.specialdf.headings
        self.specialdf.colnamefilterlist = [""] + self.specialdf.headings
        self.specialdf.infilterstatus = False

    def elementtablelay(self):  # 返回element 拼接table的layout
        # 初始化所有的layout
        # 获取显示区域的数据(超出df范围的区域不会返回dfresult, 利用此特性不用单独计算当前是否到最后页)
        dfresult = self.specialdf.dfdata.iloc[
                   self.specialdf.showstartpoint[0]: self.specialdf.showstartpoint[0] + self.max_rows,
                   self.specialdf.showstartpoint[1]: self.specialdf.showstartpoint[1] + self.max_columns
                   ]
        # 生成空白的完整填充表
        dfnull = pd.DataFrame(["" for _ in range(self.max_columns)] for _ in range(self.max_rows))

        # 形成完整的显示数据（有数据部分会覆盖"", 无数据部分依然是""）
        dfnull.iloc[0:dfresult.shape[0], 0:dfresult.shape[1]] = dfresult
        # 为 使用elements方式准备
        tempcolwidths = [self.cell_size[0]] + self.col_widths
        # 列序号行，根据需要再拼接
        laycolumnnumber = [[
            Sg.T(
                "[{}]".format(s) if i > 0 and s != "" else "列序号" if i == 0 else "",
                tooltip="[{}]".format(s) if self.show_cell_tooltips else None,
                k=(self.tablekey, -2, i - 1),  # headings 加上了序号列，左上角从 （-2， -1#）开始
                justification='c',
                enable_events=True,
                # 序号列-1根据是否显示序号判断
                visible=(self.specialdf.visible_column_map[i - 1] and self.showcolnamewithid)
                if i > 0 else (self.display_row_numbers and self.showcolnamewithid),
                font=self.font,
                border_width=1 if self.border_width == 0 else self.border_width,
                relief=Sg.RELIEF_RIDGE,
                background_color=self.specialdf.head_background_color,
                text_color=self.specialdf.head_text_color,
                # size=self.cell_size,
                size=(tempcolwidths[i], 1) if i < len(tempcolwidths) else (self.cell_size[0]),
                right_click_menu=self.column_right_click_menu,
                # expand_x=False,
                # expand_y=False,
                pad=self.pad) for i, s in enumerate(
                (['列序号'] + [str(x) for x in range(len(self.specialdf.headings))] +
                 ['' for _ in range(len(self.specialdf.headings) + 1, self.max_columns + 1)])[:self.max_columns + 1])
        ]]
        layheading = [[
            Sg.T(
                s,
                tooltip=("[{}]".format(i - 1) if i > 0 else "") + str(s) if self.show_cell_tooltips else None,
                k=(self.tablekey, -1, i - 1),  # headings 加上了序号列，左上角从 （-1， -1#）开始
                justification='c',
                enable_events=True,
                # 序号列-1根据是否显示序号判断
                visible=(self.specialdf.visible_column_map[i - 1] and self.showcolumnname)
                if i > 0 else (self.display_row_numbers and self.showcolumnname),
                font=self.font,
                border_width=1 if self.border_width == 0 else self.border_width,
                relief=Sg.RELIEF_RIDGE,
                background_color=self.specialdf.head_background_color,
                text_color=self.specialdf.head_text_color,
                # size=self.cell_size,
                size=(tempcolwidths[i], 1) if i < len(tempcolwidths) else (self.cell_size[0]),
                right_click_menu=self.column_right_click_menu,
                # expand_x=False,
                # expand_y=False,
                pad=self.pad) for i, s in enumerate(
                (['列  名'] + self.specialdf.headings +
                 ['' for _ in range(len(self.specialdf.headings) + 1, self.max_columns + 1)])[:self.max_columns + 1])
        ]]

        laydata = []
        for r in range(self.max_rows):
            # 行序号列
            temprownumber = [Sg.T(
                r,
                tooltip=r if self.show_cell_tooltips else None,
                k=(self.tablekey, r, -1),
                # size=self.cell_size,
                size=(tempcolwidths[0], 1),
                pad=self.pad,
                font=self.font,
                border_width=1 if self.border_width == 0 else self.border_width,
                relief=Sg.RELIEF_RIDGE,
                enable_events=True,
                justification='center',
                background_color=self.specialdf.head_background_color,
                expand_x=False,
                expand_y=False,
                right_click_menu=self.row_right_click_menu,
                visible=self.display_row_numbers,
                text_color=self.specialdf.head_text_color)]
            temprowdata = []
            for c in range(self.max_columns):
                # 对于combox、text、input类型，第一个输入参数不同，另外combox需要额外的default_value参数，
                if self.specialdf.column_edit_element[c] == SpecialCombox:
                    firstparam = self.specialdf.col_value_scale[c]
                    kwargsdict = {
                        "default_value": dfnull.iloc[self.specialdf.showstartpoint[0] + r, 
                                                     self.specialdf.showstartpoint[1] + c],
                        # "disabled": not self.cell_editable,
                        # "readonly": not self.cell_editable,
                        "button_arrow_color": self.specialdf.combox_button_arrow_color,
                        "button_background_color": self.specialdf.combox_button_background_color,
                        "size": (tempcolwidths[c + 1] - 2, 1)
                        if c + 1 < len(tempcolwidths) else (self.cell_size[0] - 2, self.cell_size[1])
                    }
                else:
                    # Input， text共用部分
                    firstparam = "{}".format(
                        dfnull.iloc[self.specialdf.showstartpoint[0] + r, self.specialdf.showstartpoint[1] + c])
                    kwargsdict = {
                        "justification": self.cell_justification,
                        "right_click_menu": self.tabledata_right_click_menu,
                        "border_width": self.border_width,
                        "size": (tempcolwidths[c + 1], 1) if c + 1 < len(tempcolwidths) else self.cell_size
                    }

                # input 可以设置disable,readonly， use_readonly_for_disable, Text没有该选项。
                # if self.specialdf.column_edit_element[c] == Sg.Input:
                #     kwargsdict["disabled"] = not self.cell_editable
                #     kwargsdict["readonly"] = not self.cell_editable
                #     kwargsdict["use_readonly_for_disable"] = True
                #     kwargsdict["disabled_readonly_text_color"] = \
                #         self.input_text_color if self.cell_editable else self.text_color,

                temprowdata += [self.specialdf.column_edit_element[c](
                    firstparam,
                    # enable_events=self.cell_editable or self.enable_events,
                    enable_events=True,
                    visible=self.specialdf.visible_column_map[c],
                    tooltip=
                    dfnull.iloc[self.specialdf.showstartpoint[0] + r, self.specialdf.showstartpoint[1] + c]
                    if self.show_cell_tooltips else None,
                    text_color=self.specialdf.input_text_color if self.specialdf.cell_editable 
                    else self.specialdf.text_color,
                    background_color=self.specialdf.input_text_background_color
                    if self.specialdf.cell_editable else self.specialdf.text_background_color,
                    font=self.font,
                    k=(self.tablekey, r, c),
                    expand_x=False, expand_y=False,
                    pad=self.pad,
                    **kwargsdict
                )]
            # 组织一行的数据，给laydata1添加一行
            laydata.append(temprownumber + temprowdata)

        dftablelay = [Sg.Col(
            [[Sg.Frame("", laycolumnnumber + layheading + laydata,
                       pad=(0, 0),
                       k=self.TABLE_FRAME)]],
            visible=True, vertical_alignment='top', pad=(0, 0),
            expand_x=False,
            right_click_menu=self.tabledata_right_click_menu,
            # expand_y=True,
            scrollable=self.scrollable, size=self.table_col_size, justification='center')]

        # hr = (0, int(self.specialdf.dfdata.shape[1] / self.max_columns) - 1 +
        #                                  (1 if (self.specialdf.dfdata.shape[1] % self.max_columns) > 0 else 0))
        self.h_slider_range = (0, int(self.specialdf.dfdata.shape[1] / self.max_columns) - 1 +
                               (1 if (self.specialdf.dfdata.shape[1] % self.max_columns) > 0 else 0))

        # vr = (int(self.specialdf.dfdata.shape[0] / self.max_rows) - 1 +
        #                                  (1 if (self.specialdf.dfdata.shape[0] % self.max_rows) > 0 else 0), 0)
        self.v_slider_range = (int(self.specialdf.dfdata.shape[0] / self.max_rows) - 1 +
                               (1 if (self.specialdf.dfdata.shape[0] % self.max_rows) > 0 else 0), 0)
        totallay = [
            Sg.Col(
                [
                    dftablelay,
                    [
                        Sg.Slider(orientation='h',
                                  k=self.H_SLIDER_KEY,
                                  resolution=self.h_slider_resolution,
                                  range=self.h_slider_range,
                                  border_width=2,
                                  relief=self.slider_relief,
                                  disable_number_display=self.disable_slider_number_display,
                                  size=self.h_slider_size,
                                  tooltip="左右翻页",
                                  visible=self.h_slider_show,
                                  enable_events=True,
                                  pad=(0, 0),
                                  # expand_x=True
                                  )
                    ]
                ],
                visible=True,
                # vertical_alignment='bottom',
                pad=(0, 0),
                expand_x=False,
                # expand_y=True,
                # scrollable=False,
                size=self.lay_col_size,
                justification='center'
            ),
            Sg.Col(
                [
                    [
                        Sg.Slider(orientation='v',
                                  border_width=2,
                                  # relief=Sg.RELIEF_FLAT,
                                  enable_events=True,
                                  default_value=0,
                                  k=self.V_SLIDER_KEY,
                                  resolution=self.v_slider_resolution,
                                  relief=self.slider_relief,
                                  range=self.v_slider_range,
                                  size=self.v_slider_size,
                                  pad=(0, 0),
                                  visible=self.v_slider_show,
                                  tooltip="上下翻页",
                                  disable_number_display=self.disable_slider_number_display,
                                  expand_x=True
                                  )
                    ]
                ], pad=(0, 0), size=self.lay_col_size, expand_x=False, justification='left'
            )
        ]

        return [Sg.Col([totallay], justification='center', element_justification='center')]

    def recoverspecialdf(self, newspecialdf: SpecialDf = None):
        if newspecialdf is None:
            return
        self.specialdf = newspecialdf

    @property
    def layout(self):
        # 如果_tablelayout已经创建（调用过),直接返回该值
        if self._tablelayout is not None:
            return self._tablelayout

    def eventprocess(self, eventstr: tuple, window, values):
        """
        事件处理函数，从主windows送达的事件以及windows的values 清单
        :param eventstr:                事件
        :param window:                  主window
        :param values:                  windwos的 values
        :return:
        """
        # TABLE 被点击后，会连续输出两个事件，一个是元组，格式如下，包含了点击的（行号，列号），行号-1表示点击的是heading, -2点击列序号
        # TABLE CLICKED Event has value in format ('-TABLE=', '+CLICKED+', (row,col))
        # 第二个事件是表格本身key事件
        # 表格点击事件
        if isinstance(eventstr, tuple):
            # print("Special Table event :{}".format(eventstr))
            # 首先查看bind是否已经完成，没有则做一次
            if not self.mousewheelbind:
                self.setbind(window)
                self.mousewheelbind = True
                print("{}, binding mouse event.".format(eventstr))
            # bind 事件， 查找是否鼠标滚轮事件
            if eventstr[0][0] == self.tablekey and eventstr[1] == self.MOUSE_WHEEL:
                # userbindevent = window.find_element(eventstr[0]).user_bind_event
                userbindevent = window[eventstr[0]].user_bind_event
                # print("mousewheel event :{}".format(userbindevent))
                if userbindevent.delta > 0:
                    mouseevent = MOUSEWHEEL_UP
                    # 获取滚轮滚动的格子数
                    mousetimes = int(userbindevent.delta / 120)
                elif userbindevent.delta < 0:
                    mouseevent = MOUSEWHEEL_DOWN
                    # 获取滚轮滚动的格子数
                    mousetimes = int((-userbindevent.delta) / 120)
                else:
                    print("unknow mousewheel event :{}".format(userbindevent))
                    return
                for _ in range(mousetimes):
                    self.mousewheelupdown(currentwindow=window, currentvalues=values, mousewheelupdown=mouseevent)
                return

        if len(eventstr) == 3 and eventstr[0] == self.tablekey:
            self.resetselected(eventstr, window, values)

        elif eventstr == self.MOVEROWTOCERTAIN:
            # 最后事件需要指定移动到的行数，其他不需要参数
            # 获取目标行序号
            torownumber = values[self.TARGETROWNUMBER]
            self.moverowtocertain(eventstr, window, values, torownumber)

        elif eventstr == self.COL_BAT_OK:
            # 列批量赋值
            self.columnbatsetvalue(window, values[self.COLCHOOSE], values[self.COL_BAT_VALUE])

        elif eventstr == self.FILTERCONDITIONOK:
            # 设置过滤值
            templikestr = str(values[self.CONTENTLIKE]).strip()
            columnname = str(values[self.COLUMNFILTER]).strip()
            startrow = str(values[self.ROW_START]).strip()
            endrow = str(values[self.ROW_END]).strip()
            self.setfilter(currentwindow=window, columnname=columnname, startrow=startrow,
                           endrow=endrow, like=templikestr)
        else:
            self.eventprocessdic[eventstr](eventstr, window, values)
        return

    def addrow(self, currentevent, currentwindow, currentvalues, newrowcontent=None):
        if self.itemeditfunc is None:
            return False
        if self.specialdf.indexcol is not None:
            # 需要判断索引项下是否有重复值 如果需要，把当前编辑方式、列序号、列内容传递给行编辑页面
            self.itemeditfuncargsdic["method"] = ADD_METHOD
            self.itemeditfuncargsdic["current_row"] = self.specialdf.currentrow
            # indexcol输入时没有考虑行序号，在父类中加入了序号
            self.itemeditfuncargsdic["index_col_value"] = \
                self.specialdf.dfdata.iloc[:, self.specialdf.indexcol].values.tolist() \
                    if self.specialdf.dfdata.shape[0] > 0 else []
        if self.itemeditusecolvaluescale is True:
            self.itemeditfuncargsdic["col_value_scale"] = self.specialdf.col_value_scale
        # 首先判断是否外部调用addrow而不是类内部的增加按钮
        if self.specialdf.currentrow is not None and newrowcontent in [None, []]:
            # 行记录编辑函数的第一个参数固定是menuitem，其他可变
            # indexcol输入时没有考虑行序号，在父类中加入了序号, 需要去掉
            temprowcontent = self.specialdf.dfdata.iloc[self.specialdf.currentrow, :]
        elif newrowcontent not in [None, []]:
            # 判断行内容是否合法
            if len(newrowcontent) != len(self.specialdf.headings) and self.specialdf.currentrow is None:
                print("添加的行内容{}列数{}和需要列数{}不符！".format(
                    newrowcontent, len(newrowcontent), len(self.specialdf.headings)))
                return False
            temprowcontent = newrowcontent
        else:
            # 行记录编辑函数的第一个参数固定是menuitem，其他可变
            # 传递["","".....]改为传递series
            # temprowcontent = ['' for _ in range(len(self.columninfodic))]
            temprowcontent = pd.Series(data=['' for _ in range(len(self.specialdf.headings))],
                                       index=self.specialdf.headings)
        ret = self.itemeditfunc(
            temprowcontent,
            **self.itemeditfuncargsdic)
        if ret is not None:
            # 修改后台缓冲记录，增加该行内容
            self.specialdf.dfaddrow(ret)
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
            return True
        else:
            return False

    def deleterow(self, currentevent, currentwindow, currentvalues):
        if self.specialdf.currentrow is not None:
            # 弹出确认窗口
            poprow = self.specialdf.currentrow
            if Sg.popup_ok_cancel("删除记录项:{}\n{}".format(
                    self.specialdf.currentrow, self.specialdf.dfdata.iloc[self.specialdf.currentrow, :]),
                    keep_on_top=True, modal=True, icon=SYSTEM_ICON) == "OK":
                self.specialdf.dfdeleterow()
                # # 刷新显示, 设置数据改变标志
                self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:  # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！", keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def clearall(self, currentevent, currentwindow, currentvalues):
        if Sg.popup_ok_cancel("确定清除所有记录？？？！！！",
                              keep_on_top=True, modal=True, icon=SYSTEM_ICON) == "OK":
            # 从表尾开始删除，避免currentrow改变
            for row in range(self.specialdf.maxrownumber - 1, -1, -1):
                self.specialdf.setcurrentrow(row)
                self.specialdf.dfdeleterow()
            self.specialdf.currentrow = None
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)

    def editrow(self, currentevent, currentwindow, currentvalues):
        # 判断是否选择了记录
        if self.itemeditfunc is None:
            return
        if self.specialdf.currentrow is not None:
            # 判断是否需要根据索引列保证数据不重复
            if self.specialdf.indexcol is not None:
                # 需要判断索引项下是否有重复值，如果需要，把当前编辑方式、列序号、列内容传递给行编辑页面，必须先判断是否选择了当前编辑行
                self.itemeditfuncargsdic["method"] = EDIT_METHOD
                self.itemeditfuncargsdic["current_row"] = self.specialdf.currentrow
                # indexcol输入时没有考虑行序号，在父类中加入了序号
                self.itemeditfuncargsdic["index_col_value"] = \
                    self.specialdf.dfdata.iloc[:, self.specialdf.indexcol].values.tolist() if self.specialdf.dfdata.shape[0] > 0 else []
            if self.itemeditusecolvaluescale is True:
                self.itemeditfuncargsdic["col_value_scale"] = self.specialdf.col_value_scale
            # 调用编辑子窗口，行记录编辑函数的第一个参数固定是menuitem，其他可变
            ret = self.itemeditfunc(
                # indexcol输入时没有考虑行序号，在父类中加入了序号
                self.specialdf.dfdata.iloc[self.specialdf.currentrow, :],
                **self.itemeditfuncargsdic)
            if ret is not None:
                # 修改后台缓冲记录
                self.specialdf.dfsetrowvalue(ret)
                # # 刷新显示, 设置数据改变标志
                self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！", keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def setrowvalue(self, currentwindow, rownumber, rowcontent):
        """
        外部调用update设置指定行的值
        :param currentwindow:
        :param rownumber:
        :param rowcontent:
        :return:
        """
        # 临时获取当前row，
        tempcurrentrow = self.getcurrentrow()
        # 将当前row设置为需要设置值的row
        self.specialdf.setcurrentrow(rownumber)
        # 设置row值
        self.specialdf.dfsetrowvalue(rowcontent)
        # 恢复当前row
        self.specialdf.setcurrentrow(tempcurrentrow)
        # # 刷新显示, 设置数据改变标志
        self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)

    def setcellvalue(self, currentwindow, location: tuple, cellvalue: str):
        """
        设置单元格的值
        :param currentwindow:       表格所在的窗口
        :param location:      元组 ， （x, y） 赋值的单元格位置
        :param cellvalue:     值
        :return:
        """
        self.specialdf.dfsetcellvalue(location, cellvalue)
        # 对于在显示范围内的row进行页面刷新
        # 获取页面显示的序号
        row = location[0]
        showrowstr = [currentwindow[(self.tablekey, x, -1)].get() for x in range(self.max_rows)]
        showrowlist = [int(x) if x != "" else "" for x in showrowstr]
        # 如果当前赋值行在显示窗内，做相应设置
        if row in showrowlist:
            for column in range(self.max_columns):
                currentwindow[(self.tablekey, showrowlist.index(row), location[1])].update(value=cellvalue)

        # 如果设置了数据修改事件，触发该事件
        if self.datachangeeventtrigger:
            currentwindow.write_event_value((self.tablekey, SPECIAL_TABLE_CHANGE_EVENT_TRIGGER), None)

    def moverowup(self, currentevent, currentwindow, currentvalues):
        # 判断是否选择了记录
        if self.specialdf.currentrow is not None:
            # 判断当前行是否第一行
            if self.specialdf.currentrow == 0:
                Sg.popup_quick_message("当前行{}已是第一行，无法上移！".format(self.specialdf.currentrow),
                                       modal=True, keep_on_top=True,
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       text_color='white', background_color='dark blue')
                return
            # 调用父类移动行方法
            self.specialdf.dfmoverow(self.specialdf.currentrow - 1)
            # 刷新显示
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！", keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def moverowdown(self, currentevent, currentwindow, currentvalues):
        # 判断是否选择了记录
        if self.specialdf.currentrow is not None:
            # 判断当前行是否最后一行
            if self.specialdf.currentrow == (self.specialdf.maxrownumber - 1):
                Sg.popup_quick_message("当前行{}已是最后一行，无法下移！".format(self.specialdf.currentrow),
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
                return
            # 调用父类移动行方法
            self.specialdf.dfmoverow(self.specialdf.currentrow + 1)
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！",
                        keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def moverowtotop(self, currentevent, currentwindow, currentvalues):
        # 判断是否选择了记录
        if self.specialdf.currentrow is not None:
            # 判断当前行是否第一行
            if self.specialdf.currentrow == 0:
                Sg.popup_quick_message("当前行{}已是第一行，无法上移！".format(self.specialdf.currentrow),
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
                return
            # 调用父类移动行方法
            self.specialdf.dfmoverow(0)
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！", keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def moverowtobottom(self, currentevent, currentwindow, currentvalues):
        # 判断是否选择了记录
        if self.specialdf.currentrow is not None:
            # 判断当前行是否最后一行
            if self.specialdf.currentrow == (self.specialdf.maxrownumber - 1):
                Sg.popup_quick_message("当前行{}已是最后一行，无法下移！".format(self.specialdf.currentrow),
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
                return
            # 调用父类移动行方法
            self.specialdf.dfmoverow(self.specialdf.maxrownumber - 1)  # 需要当前df的最大行数
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！",
                        keep_on_top=True, modal=True, icon=SYSTEM_ICON)

    def moverowtocertain(self, currentevent, currentwindow, currentvalues, torownumber: str):
        # 获取目标行号
        try:
            targetrownumber = int(torownumber)
        except Exception as err:
            Sg.popup_quick_message("目标行{}不正确:{}！".format(torownumber, err),
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
            return
        # 判断是否超出范围
        if targetrownumber < 0 or targetrownumber >= self.specialdf.maxrownumber:
            Sg.popup_quick_message("目标行{}超出范围！".format(torownumber),
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
            return
        # 判断是否选择了记录
        if self.specialdf.currentrow is not None:
            # 判断当前行是否目标行
            if self.specialdf.currentrow == targetrownumber:
                Sg.popup_quick_message("当前行{}已经是目标行{}！".format(self.specialdf.currentrow, targetrownumber),
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
                return
            # 调用父类移动行方法
            self.specialdf.dfmoverow(targetrownumber)
            # # 刷新显示, 设置数据改变标志
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False, datachanged=True)
        else:
            # 没有选择行
            Sg.popup_ok("当前没有选择唯一记录项！", modal=True, keep_on_top=True, icon=SYSTEM_ICON)

    def buildrowstyleinfo(self):
        self.specialdf.dfbuildrowstyleinfo()

    def buildcolumnstyleinfo(self):
        self.specialdf.dfbuildcolumnstyleinfo()

    def setrowcolor(self, tablekey, currentwindow, row: int, color: tuple):
        """
        设置行的text，background颜色
        :param tablekey:            操作的表格key
        :param currentwindow:       表格所在的窗口
        :param row:                 行号,绝对行号
        :param color:               元组 ， （文本颜色， 背景颜色）
        :return:
        """
        # 更新保存的颜色
        self.specialdf.rowstyleinfo[row][ROWSTYLE_ROWCOLOR] = color
        # 对于在显示范围内的row进行页面刷新
        # 获取页面显示的序号
        showrowstr = [currentwindow[(self.tablekey, x, -1)].get() for x in range(self.max_rows)]
        showrowlist = [int(x) if x != "" else "" for x in showrowstr]
        # 当操作行在当前显示页面时，进行相应操作
        if row in showrowlist:
            for column in range(self.max_columns):
                currentwindow[(tablekey, showrowlist.index(row), column)].update(
                    text_color=color[0], background_color=color[1])

    def recoverrowcolor(self, tablekey, currentwindow, row: int):
        """
        恢复行的text，background原颜色
        :param tablekey:            操作的表格key
        :param currentwindow:       表格所在的窗口
        :param row:                 行号，页面显示
        :return:
        """
        # 当前相对行号转变为绝对行号
        rawrowstr = currentwindow[(self.tablekey, row, -1)].get()
        if rawrowstr == '':
            return
        rawrow = int(rawrowstr)
        # 更新保存的颜色
        color = self.specialdf.getrowcolor(rawrow)
        # 对于在显示范围内的row进行页面刷新
        for column in range(self.max_columns):
            currentwindow[(tablekey, row, column)].update(text_color=color[0], background_color=color[1])

    def recovercolumncolor(self, tablekey, currentwindow, column: int):
        """
        恢复行的text，background原颜色
        :param tablekey:            操作的表格key
        :param currentwindow:       表格所在的窗口
        :param column:              页面相对列号
        :return:
        """
        # 获取当前页面所有行
        # 获取页面显示的序号
        showrowstr = [currentwindow[(self.tablekey, x, -1)].get() for x in range(self.max_rows)]
        # 获取每行对应的颜色， 当页面有行无数据时，列选择会造成越界错误。不足showrowlist 不足max_rows部分利用"" 返回默认color
        currentcolumncolor = list(map(self.specialdf.getrowcolor, 
                                      [ast.literal_eval(x) if x != "" else "" for x in showrowstr]))

        # 填写该列每行的颜色
        # 3.2清理可能在当前页反显的列
        for x in range(self.max_rows):
            currentwindow[(tablekey, x, column)].update(text_color=currentcolumncolor[x][0],
                                                        background_color=currentcolumncolor[x][1])

    def setcolumnname(self, windows, column: list, columnname: list) -> bool:
        """
        设置列名称
        :param windows:       表格所在的窗口
        :param column:                  list 用于设置的列序号列表
        :param columnname:              list 用于设置的列名称列表，
        :return:                        bool: True 成功， False 失败
        """
        # 检查输入合法性
        if len(column) != len(columnname):
            print("设置列名指定列{}与列名{}数目不符".format(column, columnname))
            return False

        for x in column:
            if x >= len(self.specialdf.headings):
                print("指定列{}与超出现有列数{}范围".format(x, len(self.specialdf.headings)))
                return False

        # 修改原始数据
        self.specialdf.dfsetcolumnname(column, columnname)
        # 逐项设置列名称
        for i, j in enumerate(column):
            windows[(self.tablekey, -1, j)].update(value=columnname[i])
        return True

    def judgesetcurrentrow(self):
        # 选中项判断完成后，重新设置currentrow，给单选中移动及编辑、删除使用
        # 当前已选所有行是否是1， 如果是1， 则currentrow设置，否则清空
        if len(self.specialdf.selectedrawrow) != 1:
            self.specialdf.setcurrentrow(None)
        else:
            self.specialdf.setcurrentrow([x for x in self.specialdf.selectedrawrow][0])

    def getresultdic(self):
        # 回传整个表格
        # 回传需要修改dfdata.columns为columndict的字典名字, 转换为二维listdict后，再修改回来
        self.specialdf.dfdata.columns = self.specialdf.dfcolumns
        resultdictlist = self.specialdf.dfdata.to_dict(orient='records')
        self.specialdf.dfdata.columns = self.specialdf.headings
        return resultdictlist

    def getresultdf(self):
        return self.specialdf.dfdata

    def geteventlist(self):
        return self.eventlist

    def colnameselected(self, currentevent, currentwindow, currentvalues):
        # 如果当前列的col_value_scale 为None, 将赋值框清空改变为可以任意填写值
        colindex = self.specialdf.headings.index(currentvalues[self.COLCHOOSE])
        if self.specialdf.col_value_scale is not None:
            if len(self.specialdf.col_value_scale[colindex]) != 0:  # None列已经在形成combo时排除,为0表示任意值
                # 如果该列取值有范围，将范围挂接到赋值框，并且把赋值框变成只读
                currentwindow[self.COL_BAT_VALUE].update(values=self.specialdf.col_value_scale[colindex],
                                                         value=self.specialdf.col_value_scale[colindex][0],
                                                         readonly=True)
        else:
            # 如果取值范围是None，关闭只读，清空combx values及values
            currentwindow[self.COL_BAT_VALUE].update(value="", values=[], readonly=False)
        # 更新row的范围信息
        self.updaterowscale(currentevent, currentwindow, currentvalues)
        return

    def updaterowscale(self, currentevent, currentwindow, currentvalues):
        """
        根据最新记录数，更新开始、结束行号的combo
        :param currentevent:        event
        :param currentwindow:      主window
        :param currentvalues:      和主window相关的values
        :return:
        """
        rowcount = self.getrowcount()
        rowspinlist = [""] + [str(a) for a in range(0, rowcount)]
        # 更新开始行号
        value = currentvalues[self.ROW_START]
        currentwindow[self.ROW_START].update(values=rowspinlist, value=value)
        # 更新结束行号
        value = currentvalues[self.ROW_END]
        currentwindow[self.ROW_END].update(values=rowspinlist, value=value)
        return

    def columnbatsetvalue(self, currentwindow, columnname, colvalue: str):
        if columnname.strip() == "":
            Sg.popup_quick_message("请选择赋值的列！", auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
            return

        # 修改数据前，先再次确认
        if (Sg.popup_ok_cancel("提示", "将筛选结果中列<{}>赋值为：{}，请再次确认！".format(columnname, colvalue),
                               modal=True, keep_on_top=True, icon=SYSTEM_ICON) == "OK"):
            # 更新窗口
            self.specialdf.dfbatsetcolumnvalue(columnname=columnname, value=colvalue)
            self.resetshow(currentwindow=currentwindow, dffilterchoose=False)
        return

    def setfilter(self, currentwindow=None, columnname=None, startrow=None, endrow=None, like=None):
        # 判断列名是否合法
        tempcolumnname = None if columnname.strip() == "" else columnname.strip()
        tempstartrow = None if startrow.strip() == "" else startrow.strip()
        tempendrow = None if endrow.strip() == "" else endrow.strip()
        templike = None if like.strip() == "" else like.strip()

        if [tempcolumnname, tempstartrow, tempendrow, templike] == [None, None, None, None]:
            Sg.popup_quick_message("没有设置任何筛选条件！",
                                   modal=True, keep_on_top=True,
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   icon=SYSTEM_ICON, text_color='white', background_color='dark blue')
            return

        if (tempcolumnname is not None and templike is None) or (tempcolumnname is None and templike is not None):
            Sg.popup_quick_message("筛选列及筛选内容不匹配！",
                                   modal=True, keep_on_top=True,
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   icon=SYSTEM_ICON, text_color='white', background_color='dark blue')
            return

        # 判断行号范围
        try:
            if tempstartrow is not None and tempendrow is not None:
                if int(tempstartrow) > int(tempendrow):
                    temp = tempstartrow
                    tempstartrow = tempendrow
                    tempendrow = temp
            elif (tempstartrow is not None and tempendrow is None) or (tempstartrow is None and tempendrow is not None):
                Sg.popup_quick_message("开始和结束行号数据不完整！",
                                       modal=True, keep_on_top=True,
                                       auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                       icon=SYSTEM_ICON, text_color='white', background_color='dark blue')
                return
        except Exception as err:
            Sg.popup_quick_message("开始行号<{}>或结束行号<{}>不正确：{}！".format(startrow, endrow, err),
                                   modal=True, keep_on_top=True,
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   icon=SYSTEM_ICON, text_color='white', background_color='dark blue')
            return

        try:
            if self.specialdf.dfsetfilter(columnname=tempcolumnname, startrow=tempstartrow, endrow=tempendrow,
                                like=templike) is True:
                # 设置成功，修改“筛选取消”按钮
                if currentwindow is not None:
                    currentwindow[self.FILTERCONDITIONCANCEL].update(disabled=False)
                self.specialdf.infilterstatus = True
                # 刷新已经选择的行列
                self.resetselected(currentevent=self.V_SLIDER_KEY,
                                   currentwindow=currentwindow, currentvalues=None)

        except Exception as err:
            Sg.popup_quick_message("表格过滤错误：{}！".format(err),
                                   modal=True, keep_on_top=True,
                                   auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                   icon=SYSTEM_ICON, text_color='white', background_color='dark blue')
            return

        # 设置过滤后，使用过滤df进行显示，所以不能调用setwholevalue，会覆盖原始dfdata
        # # 刷新显示，使用过滤的数据
        # 此处使用过滤，dffilterchoose=True
        self.resetshow(currentwindow=currentwindow, dffilterchoose=True)

    def filterclear(self, currentevent, currentwindow, currentvalues):
        self.specialdf.dffilterclear()
        # 设置成功，修改“筛选取消”按钮
        if currentwindow is not None:
            currentwindow[self.FILTERCONDITIONCANCEL].update(disabled=True)

        # 取消过滤后，使用原始dfdata
        # # 刷新显示
        self.resetshow(currentwindow=currentwindow, dffilterchoose=False)

    def getrowcount(self):
        return self.specialdf.maxrownumber

    def refreshcurrentshow(self, currentwindow, dffilterchoose=False):
        # 获取显示区域的数据(超出df范围的区域不会返回dfresult, 利用此特性不用单独计算当前是否到最后页)
        # 当主动调用了dffileterchoose 或者dffilter ！= dfdata，说明刚产生过过滤
        if dffilterchoose is False and id(self.specialdf.dfdata) == id(self.specialdf.dffilter):
            # 显示dfdata
            dfresult = self.specialdf.dfdata.iloc[
                       self.specialdf.showstartpoint[0]: self.specialdf.showstartpoint[0] + self.max_rows,
                       self.specialdf.showstartpoint[1]: self.specialdf.showstartpoint[1] + self.max_columns
                       ]
            # 刷新滚动条, 该update会产生slider事件！
            self.v_slider_range = (0, int(self.specialdf.dfdata.shape[0] / self.max_rows) - 1 +
                                   (1 if (self.specialdf.dfdata.shape[0] % self.max_rows) > 0 else 0))
            currentwindow[self.V_SLIDER_KEY].update(range=self.v_slider_range)
            self.h_slider_range = (0, int(self.specialdf.dfdata.shape[1] / self.max_columns) - 1 +
                                   (1 if (self.specialdf.dfdata.shape[1] % self.max_columns) > 0 else 0))
            currentwindow[self.H_SLIDER_KEY].update(range=self.h_slider_range)
            # 确定拷贝的行列数，将剩余行序号和列名清空
            copyrows = dfresult.shape[0]
            copycolumns = dfresult.shape[1]
            currentcolor = \
                list(map(self.specialdf.getrowcolor,
                         range(self.specialdf.showstartpoint[0], self.specialdf.showstartpoint[0] + self.max_rows))) + \
                [self.specialdf.defaultcolor for _ in range(self.max_rows - copyrows)]
        else:
            # 显示dffilter
            dfresult = self.specialdf.dffilter.iloc[
                       self.specialdf.showstartpoint[0]: self.specialdf.showstartpoint[0] + self.max_rows,
                       self.specialdf.showstartpoint[1]: self.specialdf.showstartpoint[1] + self.max_columns
                       ]

            # 刷新滚动条, 该update会产生slider事件！
            self.v_slider_range = (0, int(self.specialdf.dffilter.shape[0] / self.max_rows) - 1 +
                                   (1 if (self.specialdf.dffilter.shape[0] % self.max_rows) > 0 else 0))
            currentwindow[self.V_SLIDER_KEY].update(range=self.v_slider_range)
            self.h_slider_range = (0, int(self.specialdf.dffilter.shape[1] / self.max_columns) - 1 +
                                   (1 if (self.specialdf.dffilter.shape[1] % self.max_columns) > 0 else 0))
            currentwindow[self.H_SLIDER_KEY].update(range=self.h_slider_range)
            # 确定拷贝的行列数，将剩余行序号和列名清空
            copyrows = dfresult.shape[0]
            copycolumns = dfresult.shape[1]
            # 获取当前颜色，不足最大显示行部分补充普通颜色，根据是否editable
            currentcolor = \
                list(map(self.specialdf.getrowcolor, self.specialdf.dffilter.index.to_list())) + \
                [self.specialdf.defaultcolor for _ in range(self.max_rows - copyrows)]

        # 生成空白的完整填充表
        dfnull = pd.DataFrame(["" for _ in range(self.max_columns)] for _ in range(self.max_rows))

        # 形成完整的显示数据（有数据部分会覆盖"", 无数据部分依然是""）
        dfnull.iloc[0:dfresult.shape[0], 0:dfresult.shape[1]] = dfresult
        # 刷新完整显示内容
        for x in range(self.max_rows):
            for y in range(self.max_columns):
                currentwindow[(self.tablekey, x, y)].update(value="{}".format(dfnull.iloc[x, y]),
                                                            text_color=currentcolor[x][0],
                                                            background_color=currentcolor[x][1])
                if self.show_cell_tooltips:
                    # 获取当前每个单元格的对象类型tips，只有显示列号时才需要类型展示
                    if self.showcolnamewithid:
                        typetip = "- ({})".format(str(type(dfresult.iloc[x, y])).replace("class ", "")) \
                            if x < dfresult.shape[0] and y < dfresult.shape[1] else ""
                    else:
                        typetip = ""
                    currentwindow[(self.tablekey, x, y)].set_tooltip(
                        "{}{}".format(dfnull.iloc[x, y], typetip) if dfnull.iloc[x, y] != "" else "" + typetip)

        # 要填写的行序号(不能直接计算，而是要根据结果dfresult获取）
        rowindex = dfresult.index.to_list() + ["" for _ in range(copyrows, self.max_rows)]
        # 更新行序号
        for x in range(self.max_rows):
            currentwindow[(self.tablekey, x, -1)].update(value=rowindex[x])
            if self.show_cell_tooltips:
                currentwindow[(self.tablekey, x, -1)].set_tooltip(rowindex[x])
        # 要填写的列名，当要求显示列号+列名时，加上列号
        columnname = [self.specialdf.headings[self.specialdf.showstartpoint[1] + y] for y in range(copycolumns)] + \
                     ["" for _ in range(copycolumns, self.max_columns)]
        # 更新列名 以及列序号
        for y in range(self.max_columns):
            currentwindow[(self.tablekey, -1, y)].update(value=columnname[y])
            currentwindow[(self.tablekey, -2, y)].update(
                value="[{}]".format(self.specialdf.showstartpoint[1] + y) if y >= 0 and columnname[y] != "" else "")
            if self.show_cell_tooltips:
                currentwindow[(self.tablekey, -1, y)].set_tooltip(
                    "[{}]".format(self.specialdf.showstartpoint[1] + y if y >= 0 else "") + str(columnname[y]))
                if self.showcolnamewithid:
                    # -2列序号的tips加上列类型
                    # 获取当前的类型
                    typename = self.specialdf.getcolumnstype([columnname[y]]) \
                        if y >= 0 and columnname[y] != "" else [""]
                    currentwindow[(self.tablekey, -2, y)].set_tooltip(
                        "[{}]-({})".format(self.specialdf.showstartpoint[1] + y, typename[0])
                        if y >= 0 and columnname[y] != "" else "")

    def resetselected(self, currentevent, currentwindow, currentvalues):
        """
        清理当前选中的行或者列：
        :param currentwindow                    操作的窗口
        :param currentevent                     操作的事件
        :param  currentvalues                    当前的窗口values
        :return:
        """
        # 没有选中模式，直接返回
        if self.select_mode == SPECIAL_TABLE_SELECT_MODE_NONE:
            return

        # self.cell_editable , self.enable_events 可以产生该事件，所以对于数据单元要区分：
        # a、self.cell_editable， True ==> 更新df数据
        # b、self.cell_editable, False ==> 从df读原始数覆盖输入单元
        # c、self.enable_events, 产生行点击事件（tablekey, 行号， -1）
        # 行点击事件产生row选择变化， 如果self.enable_events , 产生tablekey 事件，value为当前行content

        # 1、获取当前事件,以及点击行列类型
        if len(currentevent) == 3:  # 当前命令是表格点击,没有翻页，showstartpoint 没有改变
            # 事件
            # 获取当前行号
            if currentevent[1] not in [-1, -2] and currentevent[2] != -1:
                try:
                    changedrawrow = int(currentwindow[(currentevent[0], currentevent[1], -1)].get())
                except Exception as err:
                    print("获取当前行号失败，超出表格范围，不处理")
                    # 清空当前非法输入
                    currentwindow[currentevent].update(value="")
                    Sg.popup_quick_message("获取当前行号失败，超出表格范围，不处理",
                                           modal=True, keep_on_top=True,
                                           auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                           text_color='white', background_color='dark blue')
                    return
                # 如果点击的是数据单元, cell_editable = True 则更新后台df， False ， 则反向用原始数据覆盖显示
                changedrawcolumn = self.specialdf.showstartpoint[1] + currentevent[2]
                # 如果修改在dfdata的合法单元格，修改相关数据、tooltips和dffilter
                if changedrawrow < self.specialdf.dfdata.shape[0] and changedrawcolumn < self.specialdf.dfdata.shape[1]:
                    if self.specialdf.cell_editable:
                        self.specialdf.dfdata.iloc[changedrawrow, changedrawcolumn] = currentvalues[currentevent]
                        # 同步更新tooltip
                        if self.show_cell_tooltips:
                            currentwindow[currentevent].set_tooltip(currentvalues[currentevent])
                        if id(self.specialdf.dfdata) != id(self.specialdf.dffilter):  # 当前是filter有效状态，需要同步更新dffilter
                            if changedrawrow in self.specialdf.dffilter.index.to_list():  # 判断当前行是否在dffliter中
                                # 当前过滤后的df， index、列名未变，但是绝对位置变化，所以使用loc而不是iloc
                                self.specialdf.dffilter.loc[changedrawrow, self.specialdf.dffilter.columns[changedrawcolumn]] = \
                                    currentvalues[currentevent]
                    else:
                        # 反向覆盖输入数据, 获取原始数覆盖显示
                        currentwindow[currentevent].update(self.specialdf.dfdata.iloc[changedrawrow, changedrawcolumn])
                    # 合法单元格，如果enable_events,
                    if self.enable_events:
                        # 如果enable_event, 发送改行序号被点击事件
                        print("send event ({}, {}, -1)".format(self.tablekey, currentevent[1], -1))
                        currentwindow.write_event_value((self.tablekey, currentevent[1], -1), None)
                else:
                    # 清空当前非法输入
                    currentwindow[currentevent].update(value="")
                    Sg.popup_quick_message("当前单元格超出表格范围，输入无效！",
                                           modal=True, keep_on_top=True,
                                           auto_close_duration=AUTO_CLOSE_DURATION, non_blocking=False,
                                           text_color='white', background_color='dark blue')
                return

            # 行点击时，如果enable_events, 发送表格tablekey事件, value为当前行内容
            tempevent = currentevent[0]
            # 如果点击坐标不是(-1, -1)并且行为-2，-1或者列为-1，说明点击了选择行、列
            # clickedtable = False if currentevent[1] == -1 and currentevent[2] == -1 else True
            clickedrow = currentevent[1] if currentevent[1] not in [-1, -2] and currentevent[2] == -1 else None
            clickedcolumn = currentevent[2] if currentevent[1] in [-1, -2] and currentevent[2] != -1 else None
            try:
                clickedrawrow = \
                    int(currentwindow[(currentevent[0], currentevent[1], -1)].get()) if clickedrow is not None else None
            except Exception as err:
                print("选中行号无法转换为int")
                return
            # clickedrawrow = self.specialdf.showstartpoint[0] + currentevent[1] if clickedrow is not None else None
            clickedrawcolumn = self.specialdf.showstartpoint[1] + currentevent[2] if clickedcolumn is not None else None
            if clickedrawrow is not None:
                if clickedrawrow >= self.specialdf.dfdata.shape[0]:
                    print("{}坐标在表格数据范围({}, {})之外，不处理".format(currentevent, self.max_rows, self.max_columns))
                    return
            if clickedrawcolumn is not None:
                if clickedrawcolumn >= self.specialdf.dfdata.shape[1]:
                    print("{}坐标在表格数据范围({}, {})之外，不处理".format(currentevent, self.max_rows, self.max_columns))
                    return
        else:  # 当前没有点击表格，上下翻页
            tempevent, clickedtable, clickedrow, clickedcolumn, clickedrawrow, clickedrawcolumn = \
                currentevent[0], False, None, None, None, None
        """
        如果是BROWSE单选模式，先做清理，再做设置
        如果是EXTEND多选模式，不需要清理raw set，只需要清理 show set, 但是 raw set 需要根据当前点击和showset进行调整
            点中某项时，判断是否需要选中或者去选中
        """
        # 因为随后会再设置，所以清理相对简单，只需要全部做清理，否则挑选出不做清理的内容会相对复杂的逻辑
        # 2、重新梳理不同模式下的set
        # 2.1、多选模式下的新set
        rowselectedflag = False  # 由于标记是否点击了行并选择
        if self.select_mode == SPECIAL_TABLE_SELECT_MODE_EXTENDED:
            if clickedrow is not None:
                rowselectedflag = True
                if clickedrow in self.specialdf.currentshowrow:
                    self.specialdf.selectedrawrow.remove(clickedrawrow)
                else:
                    self.specialdf.selectedrawrow.add(clickedrawrow)

            if clickedcolumn is not None:
                if clickedcolumn in self.specialdf.currentshowcolumn:
                    self.specialdf.selectedrawcolumn.remove(clickedrawcolumn)
                else:
                    self.specialdf.selectedrawcolumn.add(clickedrawcolumn)

        # 2.2、单选模式下的新set
        if self.select_mode in [SPECIAL_TABLE_SELECT_MODE_BROWSE, SPECIAL_TABLE_LISTBOX_SELECT_MODE_SINGLE]:
            # 重新设置单选模式下的set
            if clickedrawrow is not None:
                rowselectedflag = True
                self.specialdf.selectedrawcolumn.clear()
                # 确定当前行是否要加入set
                if clickedrawrow not in self.specialdf.selectedrawrow:
                    # 如果点击的行不是已选行，则放入已选行set中，准备刷新
                    self.specialdf.selectedrawrow.clear()
                    self.specialdf.selectedrawrow.add(clickedrawrow)
                elif self.select_mode == SPECIAL_TABLE_SELECT_MODE_BROWSE:
                    # 如果点击的行为已选行且BROWSE模式，则从已选行set中清理，不再显示
                    self.specialdf.selectedrawrow.clear()

            if clickedrawcolumn is not None:
                self.specialdf.selectedrawrow.clear()
                # 确定当前列是否要加入set
                if clickedrawcolumn not in self.specialdf.selectedrawcolumn:
                    # 如果点击的列不是已选列，则放入已选行set中，准备刷新
                    self.specialdf.selectedrawcolumn.clear()
                    self.specialdf.selectedrawcolumn.add(clickedrawcolumn)
                elif self.select_mode == SPECIAL_TABLE_SELECT_MODE_BROWSE:
                    # 如果点击的列为已选行且BROWSE模式，则从已选行set中清理，不再显示
                    self.specialdf.selectedrawcolumn.clear()

        # 2.3、选中项判断完成后，重新设置currentrow，给单选中移动及编辑、删除使用
        self.judgesetcurrentrow()

        # 3、清理所有当前的反显
        # 3.1、清理可能在当前页选中的行
        for item in self.specialdf.currentshowrow:
            self.recoverrowcolor(self.tablekey, currentwindow, item)
        self.specialdf.currentshowrow.clear()
        # 3.2清理可能在当前页反显的列
        for item in self.specialdf.currentshowcolumn:
            self.recovercolumncolor(self.tablekey, currentwindow, item)
        self.specialdf.currentshowcolumn.clear()

        # 4、根据新set情况重新刷新
        # 4.1、刷新行
        # 获取当前显示页中的行号列表
        currentshowrowstr = [(currentwindow[(currentevent[0], x, -1)].get()) for x in range(self.max_rows)]
        currentshowrow = [int(x) if x != "" else None for x in currentshowrowstr]
        for item in self.specialdf.selectedrawrow:
            # 查看是否需要刷新的行,考虑过滤行引起的特殊性，不能使用showstartpoint[0],必须使用实际的当前页中的行序号
            if item in currentshowrow:
                # 添加到当前显示set
                tempshowrow = currentshowrow.index(item)
                self.specialdf.currentshowrow.add(tempshowrow)
                for y in range(self.max_columns):
                    currentwindow[(tempevent, tempshowrow, y)].update(
                        text_color=self.specialdf.selected_text_color,
                        background_color=self.specialdf.selected_background_color)

        # 4.2、刷新列
        for item in self.specialdf.selectedrawcolumn:
            # 查看是否存在需要刷新的列
            if self.specialdf.showstartpoint[1] <= item < self.specialdf.showstartpoint[1] + self.max_columns:
                # 添加到当前显示set
                self.specialdf.currentshowcolumn.add(item - self.specialdf.showstartpoint[1])
                for x in range(self.max_rows):
                    currentwindow[(tempevent, x, item - self.specialdf.showstartpoint[1])].update(
                        text_color=self.specialdf.selected_text_color,
                        background_color=self.specialdf.selected_background_color)

        # 5、判断是否发送表格tablekey事件, 只有在某些事件时才发送该事件
        if self.enable_events and rowselectedflag:
            currentwindow.write_event_value(self.tablekey, self.getselectedrowcontent())

    def sliderclicked(self, currentevent, currentwindow, currentvalues):
        """
        翻页当前的行或者列：
        :param currentevent                窗口event
        :param currentwindow               窗口
        :param currentvalues               窗口values
        :return:
        """
        # 获取要翻到的页
        if currentevent == self.V_SLIDER_KEY:
            # 将精度改为float，可以逐行上下翻
            # torowpage = int(currentvalues[currentevent])
            torowpage = float(currentvalues[currentevent])
            # print("上下翻页到第{}页".format(torowpage))
            tocolumnpage = self.specialdf.currentcolumnpage
        else:
            torowpage = self.specialdf.currentrowpage
            # 将精度改为float，可以逐列左右翻
            # tocolumnpage = int(currentvalues[currentevent])
            tocolumnpage = float(currentvalues[currentevent])
            # print("左右翻页到第{}页".format(tocolumnpage))

        #   如果依然是原行、列页，直接返回
        if self.specialdf.currentrowpage == torowpage and self.specialdf.currentcolumnpage == tocolumnpage:
            return
        else:
            # 更新起始点
            self.specialdf.showstartpoint = [int(torowpage * self.max_rows), int(tocolumnpage * self.max_columns)]
            self.specialdf.currentrowpage = torowpage
            self.specialdf.currentcolumnpage = tocolumnpage
            self.refreshcurrentshow(currentwindow)
            # 重新刷新选中的行、列情况
            self.resetselected(currentevent, currentwindow, currentvalues)

    def setwholevalue(self, currentwindow, dfdata: pd.DataFrame = None,
                      infodict: dict = None, headings: list = None):
        """
        全部更新数据表和显示
        :param currentwindow:           当前窗口
        :param dfdata:                  新数据
        :param infodict                 字典方式数据
        :param headings                list方式下的列headings信息
        :return:
        """
        if dfdata is not None:
            # 首先更新dflist class数据
            self.specialdf.dfsetwholevalue(dfdata=dfdata, headings=headings)
            # print("设置全部数据:\n{}".format(dfdata))
            self.usedfdatainit(dfdata)  # 设置colnamelist、tableinfodictlist等
        elif infodict is not None and headings is not None:
            # 将命令相关参数信息字典保存到二维list中，对应于界面表格的数据
            if infodict == []:  # 空list必须有columns
                dfdata = pd.DataFrame(infodict, columns=headings)
            else:  # 非空list不能有columns，否则结果不对， 先生成df，再赋值columns
                dfdata = pd.DataFrame(infodict)
            # 首先更新df数据
            self.specialdf.dfsetwholevalue(dfdata=dfdata, headings=headings)
        # 需要更新filter的行数，筛选列，以及可以批量赋值的列
        self.specialdf.colnameusedlist = [""] + self.specialdf.headings
        self.specialdf.colnamefilterlist = [""] + self.specialdf.headings
        rowlist = [""] + [str(a) for a in range(0, self.specialdf.maxrownumber)]
        if self.filter_visible:
            currentwindow[self.ROW_START].update(values=rowlist)
            currentwindow[self.ROW_END].update(values=rowlist)
            currentwindow[self.COLUMNFILTER].update(values=self.specialdf.colnamefilterlist)
            currentwindow[self.COLCHOOSE].update(values=self.specialdf.colnameusedlist)

        # 数据全部更新，需要更新rowcolor
        # 用于保存每行的文本和背景颜色，
        self.buildrowstyleinfo()
        self.buildcolumnstyleinfo()

        # 更新table显示
        self.resetshow(currentwindow=currentwindow, dffilterchoose=False)

    def showfilterresult(self, currentwindow, dffiler: pd.DataFrame):
        if dffiler is None:
            return
        self.resetshow(currentwindow=currentwindow, dffilterchoose=False)

    def resetshow(self, currentwindow, dffilterchoose=False, datachanged=False):
        """
        :param currentwindow               当前窗口
        :param dffilterchoose               当前是否过滤方式
        从0， 0开始，刷新整个显示页面
        :param datachanged                 本次刷新数据是否变化
        :return:
        """
        # 1、初始化所有相关信息
        # 1.1、显示起始坐标， 三个必须全部清零，因为slider的update会触发slider点击事件
        self.specialdf.showstartpoint = [0, 0]
        self.specialdf.currentrowpage = 0
        self.specialdf.currentcolumnpage = 0
        # 1.2、当前选中行列清除
        for item in self.specialdf.currentshowrow:
            for y in range(self.max_columns):
                currentwindow[(self.tablekey, item, y)].update(
                    text_color=self.specialdf.input_text_color
                    if self.specialdf.cell_editable else self.specialdf.text_color,
                    background_color=self.specialdf.input_text_background_color
                    if self.specialdf.cell_editable else self.specialdf.text_background_color)
        self.specialdf.currentshowrow.clear()
        # 1.3、清理可能在当前页反显的列
        for item in self.specialdf.currentshowcolumn:
            for x in range(self.max_rows):
                currentwindow[(self.tablekey, x, item)].update(
                    text_color=self.specialdf.input_text_color
                    if self.specialdf.cell_editable else self.specialdf.text_color,
                    background_color=self.specialdf.input_text_background_color
                    if self.specialdf.cell_editable else self.specialdf.text_background_color)
        self.specialdf.currentshowcolumn.clear()
        # 1.4、全局选中行列清理
        self.specialdf.selectedrawrow.clear()
        self.specialdf.selectedrawcolumn.clear()

        # 2、刷新数据
        self.refreshcurrentshow(currentwindow, dffilterchoose)
        # 3、选中项判断完成后，重新设置currentrow，给单选中移动及编辑、删除使用
        self.judgesetcurrentrow()

        # 4、判断是否要发送datachanged事件
        if self.datachangeeventtrigger and datachanged:
            currentwindow.write_event_value((self.tablekey, SPECIAL_TABLE_CHANGE_EVENT_TRIGGER), None)

    def donothing(self, currentevent, currentwindow, currentvalues):
        print("eventprocess {} do nothing!".format(currentevent))
        return

    def getcolnameusedlist(self):
        return self.specialdf.colnameusedlist

    def getcurrentrow(self):
        return self.specialdf.currentrow

    def getcurrentrowcontent(self):
        if self.specialdf.currentrow is None:
            return None
        else:
            return self.specialdf.dfdata.iloc[self.specialdf.currentrow, :].tolist()

    def getcommandinfluncerow(self, currentevent=None, currentwindow=None, currentvalues=None, executednumber=None):
        """
        根据当前指定的行号，返回当前命令、currrentrow、tonumber结合对比，返回不受影响的行数（不是行号），没有则None
        :param currentevent             当前的事件
        :param currentwindow            当前窗口
        :param currentvalues           当前事件值
        :param executednumber            用于比较的行号
        :return:
        """
        # 如果是CLEARALL，全部会被清空，不受影响的行数是0
        if currentevent == self.CLEARALL:
            return 0
        # 没有选中currentrow
        if self.specialdf.currentrow is None:
            return executednumber

        # ADDROW命令不影响任何已经存在的行
        if currentevent == self.ADDROW:
            return executednumber

        # currentrow >= excutednumber 保持原excutednumber及颜色
        # currentrow < excutednumber  最小currentrow - 1
        if currentevent in (self.EDITROW, self.DELETEROW):
            return self.specialdf.currentrow if self.specialdf.currentrow < executednumber else executednumber

        if currentevent == self.MOVEROWTOTOP:
            tonumber = 0

        if currentevent == self.MOVEROWUP:
            tonumber = self.specialdf.currentrow - 1 if self.specialdf.currentrow - 1 >= 0 else 0

        if currentevent == self.MOVEROWDOWN:
            tonumber = self.specialdf.currentrow + 1 if self.specialdf.currentrow + 1 < self.specialdf.dfdata.shape[0] else self.specialdf.currentrow

        if currentevent == self.MOVEROWTOBOTTOM:
            tonumber = self.specialdf.dfdata.shape[0] - 1 if self.specialdf.dfdata.shape[0] - 1 >= 0 else 0

        if currentevent == self.MOVEROWTOCERTAIN:
            try:
                tonumber = int(currentwindow[self.TARGETROWNUMBER].get())
            except Exception as err:
                print("获取要到达的行数失败：{}".format(err))
                tonumber = self.specialdf.dfdata.shape[0] - 1 if self.specialdf.dfdata.shape[0] - 1 >= 0 else 0

        # excutednumber - 1 <  currentrow、tonumber 中的小者  保持原excutednumber及颜色
        # excutednumber - 1 >=  currentrow、tonumber 中的小者	 执行到currentrow \tonumber中的小者
        if currentevent in (self.MOVEROWTOCERTAIN, self.MOVEROWTOBOTTOM, self.MOVEROWTOTOP,
                            self.MOVEROWDOWN, self.MOVEROWUP):
            #  根据 executednumber 和 currentrow、tonumber的关系判断处理
            comparevalue = self.specialdf.currentrow if self.specialdf.currentrow < tonumber else tonumber
            if executednumber - 1 < comparevalue:
                return executednumber
            else:
                return comparevalue

        return executednumber

    def getselectedrowcolumn(self):
        return [self.specialdf.selectedrawrow, self.specialdf.selectedrawcolumn]

    def getselectedrowcontent(self):
        # 将所选中的所有行内容形成列表返回
        return self.specialdf.dfdata.iloc[list(self.specialdf.selectedrawrow), :].values.tolist()

    def getcolumnname(self):
        return list(self.specialdf.dfdata)

    def getspecialeventlist(self):
        return self.specialeventlist

    def itemeditgeneral(self, rowcontent: list,
                        method: int = None,
                        current_row: int = None,
                        index_col_value: list = None):
        """
        df方式时，如果没有指定的钩子函数，用此通用行编辑函数
        :param rowcontent:              list:当前编辑行的值list
        :param method:                  EDIT_METHOD or ADD_METHOD
        :param current_row:             当前编辑的行号，用于判断是否重复
        :param index_col_value:         索引列的当前列值，用于判断是否重复值
        :return:
        """
        # 当前表的列序号list
        currentcolumnsindex = [x for x in range(len(self.specialdf.headings))]
        # 当前设置的列名list
        currentvaluelist = rowcontent
        defaultkey = "_itemedit_general_"
        # ROW_EDIT_GUI_KEY = "".join([self.tablekey, defaultkey])
        ROW_EDIT_GUI_KEY = (self.tablekey, defaultkey)
        ROW_EDIT_OK = (ROW_EDIT_GUI_KEY, '_ok_')
        ROW_EDIT_CANCEL = (ROW_EDIT_GUI_KEY, '_cancel_')
        # 序列旧名称和新名称的二维列表
        tablevalues = [[self.specialdf.headings[x], currentvaluelist[x]] for x in currentcolumnsindex]
        # 显示的表格头
        tempcellsize = (18, 1)
        tempvaluesize = (56, 1)
        tempargvlist = []
        # 确定使用Input还是Combo作为cell编辑的类型
        if self.specialdf.col_value_scale is not None:
            tempcelltypelist = [Sg.Combo if (isinstance(x, list) and len(x) > 0)
                                else Sg.Input for x in self.specialdf.col_value_scale]
        else:
            tempcelltypelist = [Sg.Input for _ in self.specialdf.headings]

        for i, item in enumerate(tempcelltypelist):
            if item == Sg.Combo:
                # 形成每个编辑单元格需要的差异化参数
                tempargvlist.append(
                    [
                        self.specialdf.col_value_scale[i],
                        {
                            "default_value": tablevalues[i][1],
                            "button_arrow_color": self.specialdf.combox_button_arrow_color,
                            "button_background_color": self.specialdf.combox_button_background_color,
                            "size": (tempvaluesize[0] - 2, tempvaluesize[1])
                        }
                    ]
                )
            else:
                # 形成每个编辑单元格需要的差异化参数
                tempargvlist.append(
                    [
                        tablevalues[i][1],
                        {
                            "justification": "l",
                            # "border_width": self.border_width,
                            "size": tempvaluesize
                        }
                    ]
                )

        headinglayout = [[
            Sg.T(s,
                 k=(ROW_EDIT_GUI_KEY, -1, i - 1),  # headings 加上了序号列，左上角从 （-1， -1）开始
                 justification='center',
                 enable_events=False,
                 font=self.font,
                 border_width=1 if self.border_width == 0 else self.border_width,
                 relief=Sg.RELIEF_RIDGE,
                 background_color=self.specialdf.head_background_color,
                 text_color=self.specialdf.head_text_color,
                 size=tempvaluesize if s == "列值" else tempcellsize,
                 pad=self.pad) for i, s in enumerate(['列序号', '列名', '列值'])
        ]]
        # 显示表格内容,key用元组（ROW_EDIT_GUI_KEY，行, 列），只有第2列可以编辑，后续事件处理只需要判断key元组的[0],[2]即可
        datalayout = [
            [Sg.T(r,
                  k=(ROW_EDIT_GUI_KEY, r, -1),
                  size=tempcellsize,
                  pad=self.pad,
                  font=self.font,
                  enable_events=False,
                  justification='center',
                  relief=Sg.RELIEF_RIDGE,
                  visible=self.specialdf.visible_column_map[r],
                  border_width=1 if self.border_width == 0 else self.border_width,
                  background_color=self.specialdf.head_background_color,
                  text_color=self.specialdf.head_text_color)] +
            [Sg.T(
                tablevalues[r][0],
                justification='right',
                # disabled=True,
                background_color=self.specialdf.text_background_color,
                enable_events=False,
                text_color=self.specialdf.text_color,
                font=self.font,
                relief=self.relief,
                border_width=self.border_width,
                visible=self.specialdf.visible_column_map[r],
                k=(ROW_EDIT_GUI_KEY, r, 0),
                size=tempcellsize,
                pad=self.pad)] +
            [tempcelltypelist[r](
                tempargvlist[r][0],
                **tempargvlist[r][1],
                background_color=self.specialdf.input_text_background_color,
                enable_events=False,
                text_color=self.specialdf.input_text_color,
                font=self.font,
                visible=self.specialdf.visible_column_map[r],
                k=(ROW_EDIT_GUI_KEY, r, 1),
                pad=self.pad)] for r in range(len(tablevalues))
        ]
        taillayout = [
            [Sg.Col([[
                Sg.Frame("",
                         [[Sg.Button('确定', k=ROW_EDIT_OK),
                           Sg.Button('取消', k=ROW_EDIT_CANCEL)]],
                         vertical_alignment='top', element_justification='center', expand_x=True, border_width=0)
            ]], justification='center', expand_x=True)]]

        currentlayout = [[Sg.Col(headinglayout + datalayout, size=(686, 500), expand_x=True, expand_y=False,
                                 justification='center', scrollable=True)]] + taillayout

        # 开启window
        itemwindow = Sg.Window(
            '表格行编辑',
            currentlayout,
            font=SPECIAL_TABLE_DEF_FT,
            icon=SYSTEM_ICON,
            return_keyboard_events=True,
            resizable=True,
            size=(706, 600),
            modal=True,
            # element_justification='center',
            keep_on_top=True,
            finalize=True
        )
        rowresult = None
        while True:
            event, values = itemwindow.read()
            if event in (Sg.WIN_CLOSED, '_close_'):
                rowresult = None
                break

            elif event == ROW_EDIT_OK:
                rowresult = [itemwindow[(ROW_EDIT_GUI_KEY, x, 1)].get() for x in range(len(self.specialdf.headings))]
                # 添加模式下判断index是否重复
                # 首先判断是否有indexcol
                if self.specialdf.indexcol is not None:
                    # 获取当前编辑行indexcol的值
                    currentindexvalue = [rowresult[x] for x in self.specialdf.indexcol] \
                        if isinstance(self.specialdf.indexcol, list) else rowresult[self.specialdf.indexcol]
                    if currentindexvalue in index_col_value:
                        # 获取该值在整个索引列中位置信息
                        cmdindex = index_col_value.index(currentindexvalue)
                        if not (method == EDIT_METHOD and cmdindex == current_row):  # 排除本行编辑的情况
                            # 不是编辑当前行，说明和某行的index重复
                            # 新增内容和原有内容重复
                            valuevalide = False
                            Sg.popup_error("{}和第{}条记录同名！".format(currentindexvalue, cmdindex),
                                           keep_on_top=True, modal=True, icon=SYSTEM_ICON)
                            continue
                # 没有重名情况， 关闭窗口，返回修改或者新增的值
                break

            elif event == ROW_EDIT_CANCEL:
                rowresult = None
                break
            else:
                continue

        itemwindow.close()
        # 返回结果字典
        return rowresult

    def setbind(self, currentwindow=None):
        """
        window初始化后，调用update bind，内部调用该函数，初始化相关的bind事件
        :param currentwindow:
        :return:
        """
        # currentwindow[self.TABLE_FRAME].bind("<Enter>", self.MOUSE_IN, propagate=True)
        # currentwindow[self.TABLE_FRAME].bind("<Leave>", self.MOUSE_LEAVE, propagate=True)
        # self.windowmousedict = windowmousedict
        # currentwindow[self.TABLE_FRAME].bind("<MouseWheel>", self.MOUSE_WHEEL, propagate=True)
        # 将表格中所有单元格绑定mousewheel事件，包括-1行 -1列,
        # print("bind mouse wheel！")
        for row in range(-2, self.max_rows):
            for column in range(-1, self.max_columns):
                currentwindow[(self.tablekey, row, column)].bind("<MouseWheel>", self.MOUSE_WHEEL, propagate=True)

    def mousewheelupdown(self, currentwindow=None, currentvalues=None, mousewheelupdown=None):
        """
        接手窗口的wheelupdown并且进行处理
        :param currentwindow:               当前主窗口
        :param currentvalues:               当前窗口values
        :param mousewheelupdown:            MOUSEWHEEL_UP, MOUSEWHEEL_DOWN
        :return:
        """
        # print("{}:{}".format(self.tablekey, mousewheelupdown))
        # 获取slider的当前值
        value = currentvalues[self.V_SLIDER_KEY]
        rangetemp = [self.v_slider_range[0], self.v_slider_range[1]]
        rangetemp.sort()
        rangemin, rangemax = rangetemp
        # 计算新的value
        newvalue = value
        if mousewheelupdown == MOUSEWHEEL_DOWN:
            newvalue = value + self.v_slider_resolution if value + self.v_slider_resolution < rangemax else rangemax
        elif mousewheelupdown == MOUSEWHEEL_UP:
            newvalue = rangemin if value - self.v_slider_resolution < rangemin else value - self.v_slider_resolution

        # 设置v_slider的新值
        currentwindow[self.V_SLIDER_KEY].update(value=newvalue)
        # 模拟sliderclick事件
        self.sliderclicked(self.V_SLIDER_KEY, currentwindow, currentvalues)

    def loaddatafromfile(self, currentwindow, inputfilename: str = None, headline: int = 0, tailline: int = 0,
                         sheetnumber: int = 0, headname: int = 0) -> bool:
        """
        从文件导入数据到df
        :param currentwindow:                           当前主窗口，为后续刷新显示用
        :param inputfilename:                           str： 文件全路径名
        :param headline:                                int: 导出表首丢弃行数
        :param tailline:                                int： 导出表尾丢弃行数
        :param sheetnumber:                             int： 导出的子表号
        :param headname:                                int: 丢弃表首后，列名所在行，>=0有效
        :return:                                        bool：True 导入成功 False 导入失败
        """
        # 读取指定的文件
        fileresult = os.path.splitext(inputfilename)
        # print(fileresult[-1])
        # 如果是xlsx文件
        # 2、pandas读取csv文件
        # import pandas as pd
        # data = pd.read_csv('path', sep=',', header=0, names=["第一列"，"第二列"，"第三列"]，encoding = 'utf-8')
        # path: 要读取的文件的绝对路径
        # sep: 指定列和列的间隔符，默认sep =‘, ’
        # 若sep =‘’\t",即列与列之间用制表符\t分割，相当于tab——四个空格
        # header: 列名行，默认为0
        # names: 列名命名或重命名
        # encoding: 指定用于unicode文本编码格式
        # 3、pandas读取txt文件
        # read_csv
        # data = pd.read_table('path', sep='\t', header=None, names=['第一列', '第二列', '第三列'])

        myhead = headname if headname >= 0 else None
        if fileresult[-1] == ".xlsx":
            try:
                # # 读取时跳过表头行和表尾行
                readdf = pd.read_excel(inputfilename, header=myhead,
                                       skiprows=headline,
                                       skipfooter=tailline, sheet_name=sheetnumber)
                # head需要转换为str以便后续处理
                readdf.columns = readdf.columns.map(str)
            except Exception as err:
                Sg.popup_error("读取{}失败:{}！".format(inputfilename, err),
                               modal=True, keep_on_top=True, icon='ic_umdos.ico')
                return False
        # TODO 增加read_csv时分隔符设置。
        elif fileresult[-1] == ".csv":
            try:
                # # 读取时跳过表头行和表尾行
                readdf = pd.read_csv(inputfilename, header=myhead,
                                     skiprows=headline,
                                     skipfooter=tailline, encoding='utf-8')
                # head需要转换为str以便后续处理
                readdf.columns = readdf.columns.map(str)
            except Exception as err:
                Sg.popup_error("读取{}失败:{}！".format(inputfilename, err),
                               modal=True, keep_on_top=True, icon='ic_umdos.ico')
                return False

        elif fileresult[-1] == ".txt":
            try:
                # # 读取时跳过表头行和表尾行
                readdf = pd.read_csv(inputfilename, header=myhead,
                                     skiprows=headline, sep=',',
                                     skipfooter=tailline, encoding='utf-8')
                # head需要转换为str以便后续处理
                readdf.columns = readdf.columns.map(str)
            except Exception as err:
                Sg.popup_error("读取{}失败:{}！".format(inputfilename, err),
                               modal=True, keep_on_top=True, icon='ic_umdos.ico')
                return False
        else:
            Sg.popup_error("无法识别后缀的文件:{}！".format(inputfilename),
                           modal=True, keep_on_top=True, icon='ic_umdos.ico')
            return False

        self.setwholevalue(currentwindow=currentwindow, dfdata=readdf)
        Sg.popup_quick_message("导入文件{}完成".format(inputfilename),
                               modal=True, keep_on_top=True, text_color='white', background_color='dark blue')

        return True

    def savedatatofile(self, outputfilename=None, modelfilename=None, index: bool = True,
                       sheetnumber: int = 0, headline: int = 0, tailline: int = 0) -> bool:
        """
        将df内容保存到文件
        :param outputfilename:                  str：输出文件全路径名
        :param modelfilename:                   str：输出模板文件全路径名
        :param index:                           bool: True 输出index， False： 不输出
        :param sheetnumber:                     int: 输入模板的子表号
        :param headline:                        int：输出模板表首跳过行数
        :param tailline:                        int: 输出模板表尾跳过行数
        :return:                                bool： True 导出成功， False 导出失败
        """
        if outputfilename is None:
            return False

        if modelfilename is None:  # 没有模板文件，直接导出到excel
            try:
                self.specialdf.dfdata.to_excel(outputfilename, index=index)
            except Exception as err:
                print("导出数据到文件{}失败：{}！".format(outputfilename, err))
                if guiflag:
                    Sg.popup_quick_message("导出数据到文件{}失败：{}！".format(outputfilename, err),
                                           modal=True, keep_on_top=True, text_color='white',
                                           background_color='dark blue')
                return False
        else:
            # model 方式导出
            return self.savedftomodelexcelfile(outputfilename=outputfilename,
                                               modelfilename=modelfilename,
                                               sheetnumber=sheetnumber,
                                               headline=headline,
                                               tailline=tailline)

        print("导出数据到文件{}子表{}完成！".format(outputfilename, sheetnumber))
        Sg.popup_quick_message("导出数据到文件{}完成！".format(outputfilename),
                               modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
        return True

    def savedftomodelexcelfile(self, outputfilename: str = None,
                               modelfilename: str = None, sheetnumber: int = 0,
                               headline: int = 0, tailline: int = 0) -> bool:
        """
        将df内容保存到文件
        :param outputfilename:                  str：输出文件全路径名
        :param modelfilename:                   str：输出模板文件全路径名
        :param sheetnumber:                     int: 输出到文件的子表号
        :param headline:                        int：输出模板表首跳过行数
        :param tailline:                        int: 输出模板表尾跳过行数
        :return:                                bool： True 导出成功， False 导出失败
        """
        # 打开模板文件，处理结束后保存到目标文件
        global guiflag
        try:
            wb = load_workbook(modelfilename)
            ws = wb.worksheets[sheetnumber]
        except Exception as err:
            print("打开模板文件{}子表{}失败：{}！".format(modelfilename, sheetnumber, err))
            if guiflag:
                Sg.popup_error("打开模板文件{}子表{}失败：{}！！".format(modelfilename, sheetnumber, err),
                               modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
            return False

        # 将结果数据填入模板
        # 逐行\逐列处理，笨办法
        # ws 起始行
        # 旧残留数据行数
        olddatanumber = ws.max_row - headline - tailline
        # 插入行数累计
        insertednumber = 0
        rownumber = headline + 1  # 跳过表头
        # 能够拷贝的列数，小于模板列数时，使用源数据列数
        maxcolnumber = self.specialdf.dfdata.shape[1] if self.specialdf.dfdata.shape[1] < ws.max_column else ws.max_column  # 获取模板列数
        # 是否在处理第一行，从数据第二行开始，复制第一行的格式
        isfirstrow = True
        # 逐行获取源数据信息进行处理
        for index, rowseries in self.specialdf.dfdata.iterrows():
            row = rowseries.values
            # ws 起始列，ws列号从1开始，但是dataframe从0开始
            # 不是第一行时插入新行，将表尾行往下推，复制上一行的各个单元格格式
            # 复制单元格属性
            if isfirstrow is False:
                ws.insert_rows(rownumber)
                for j in range(1, ws.max_column + 1):
                    lastrow = rownumber - 1
                    ws.cell(rownumber, j).number_format = ws.cell(lastrow, j).number_format
                    ws.cell(rownumber, j).font = copy.copy(ws.cell(lastrow, j).font)
                    ws.cell(rownumber, j).alignment = copy.copy(ws.cell(lastrow, j).alignment)
                    ws.cell(rownumber, j).fill = copy.copy(ws.cell(lastrow, j).fill)
                    ws.cell(rownumber, j).border = copy.copy(ws.cell(lastrow, j).border)

            # ws的列数从1开始，但是dataframe从0 开始，所以此处range如此设定
            for colnumber in range(maxcolnumber):
                # ws的列数从1开始，但是dataframe从0 开始，所以此处需要加1
                ws.cell(rownumber, colnumber + 1, row[colnumber])
                # 本列处理完成，列数+1
                colnumber += 1
            # 本行处理完成，行数+1
            rownumber += 1
            insertednumber += 1
            # 翻转标志，保证从第二行开始复制第一行数据单元的格式。
            isfirstrow = False

        # 清理模板文件表头和表尾之间原来残留的数据，为了保持格式完整，先插入新数据，再删除老数据
        deleterownumber = headline + insertednumber + 1
        # 根据要删除的行数，持续删除head后的第一行，直到行尾
        for _ in range(olddatanumber + 1):
            ws.delete_rows(deleterownumber)

        # 保存处理好的数据表到输出目标excel文件
        try:
            wb.save(outputfilename)
        except Exception as err:
            print("输出文件{}错误：{}".format(outputfilename, err))
            if guiflag:
                Sg.popup_error("输出文件{}错误：{}".format(outputfilename, err),
                               modal=True, keep_on_top=True, icon='ic_umdos.ico')
            return False

        print("{}处理完成！".format(outputfilename))
        if guiflag:
            Sg.popup_quick_message("将数据按模板文件{}导出到文件{}完成".format(modelfilename, outputfilename),
                                   modal=True, keep_on_top=True, text_color='white', background_color='dark blue')
        return True

    def update(self, commandtype=None, currentevent=None, currentwindow=None, currentvalues=None,
               dfshowdata: pd.DataFrame = None, infodict: dict = None, rownumber=None, color=None,
               location=None, value=None,windowmousedict: dict = None, headings=None,
               mousewheelupdown=None, newrowcontent: list = None, filename: str = None, sheetnumber: int = 0,
               modelfilename: str = None, headline: int = 0, tailline: int = 0, headname: int = -1,
               # citemeditfunc: object = None,
               # itemeditfuncargsdic: dict = {}, indexcol: int = None, visible_column_map=None,
               # col_widths=None, size=DEF_WINDOW_SIZE, cell_size=DEF_CELL_SIZE, scrollable=False,
               # col_value_scale: list = None, filter_visible=True, edit_button_visible=True,
               # visible=True,   right_click_menu=None, vertical_scroll_only=True,
               # expand_x=False, expand_y=False, select_mode=None, font=SPECIAL_TABLE_DEF_FT, pad=DEF_PAD,
               # max_rows=MAX_SPECIAL_TABLE_ROWS, max_columns=MAX_SPECIAL_TABLE_COLUMNS,
               # text_color='black', text_background_color='white',
               # head_text_color='white', head_background_color='dark blue',
               # selected_text_color='white', selected_background_color='blue',
               # disable_slider_number_display=False
               ):
        """
        本类的类似Sg update处理
        :param commandtype:                        设置的命令类型
                                            SPECIAL_TABLE_SET_ROWCOLOR = 'set_row_color'
                                            SPECIAL_TABLE_SET_CELLVALUE = 'set_cell_value'
                                            SPECIAL_TABLE_SET_ROWVALUE = 'set_row_value'
                                            SPECIAL_TABLE_SET_CURRENTROW = 'set_current_row'
                                            SPECIAL_TABLE_SET_WHOLEVALUE = 'set_whole_value'
                                            SPECIAL_TABLE_SET_BIND
                                            SPECIAL_TABLE_SET_MOUSEWHEEL
                                            SPECIAL_TABLE_SET_ADD_NEWROW
                                            SPECIAL_TABLE_SET_SAVETOFILE
                                            SPECIAL_TABLE_SET_LOADFROMFILE
        :param currentevent:                当前的事件
        :param currentwindow:               当前的窗口
        :param currentvalues:               当前的窗口values
        :param dfshowdata:                  当前传递过来的df参数，用于刷新当前数
        :param infodict:                    当前传递过来的list data dict，用于刷新当前数
        :param rownumber:                   要处理的行
        :param windowmousedict:             dict, 主窗口用于记录内部表格mouse in，leave状态的字典
        :param mousewheelupdown:            MOUSEHWEEL_UP , MOUSEWHEEL_DOWN
        :param newrowcontent:               SET_ADDNEWROW 时的行内容
        :param color:                       要设置的颜色
        :param location:                    当前要设置的表格单元坐标
        :param value:                       当前要设置的值
        :param filename:                     str = None, 保存df到文件和调入文件数据的文件名
        :param sheetnumber:                  int = 0,读取和写入的子表号
        :param modelfilename:                str = None,输出使用的模板文件
        :param headline:                     int = 0, 输入输出文件表头的行数
        :param tailline:                     int = 0, 输入输出文件表尾的行数
        :param headname:                 文件去掉表头后， 如果有列名，所处的文件行数（0开始） int = -1,
        :return:
        """
        # 更新整张表的数据
        if currentwindow is None:
            print("SpecialTable update 参数currrenwindow or currentevent 无效")
            return False

        def innersetwholevalue(myself):
            # update table data
            if dfshowdata is not None or infodict is not None and not isinstance(currentevent, tuple):
                myself.setwholevalue(currentwindow, dfdata=dfshowdata, infodict=infodict, headings=headings)

        def innersetrowcolor(myself):
            if rownumber is not None:
                if color is not None:
                    if currentwindow is not None:
                        myself.setrowcolor(myself.tablekey, currentwindow, rownumber, color)
                else:
                    # 有行号但是没有色号，设置当前行
                    myself.specialdf.setcurrentrow(rownumber)

        def innersetcellvalue(myself):
            # 设置单元格值
            if location is not None and currentwindow is not None:
                myself.setcellvalue(currentwindow, location, value)

        def innersetrowvalue(myself):
            # 设置行的值
            myself.setrowvalue(currentwindow, rownumber, newrowcontent)

        def innersetcurrentrow(myself):
            myself.specialdf.setcurrentrow(rownumber)

        def innersetbind(myself):
            myself.setbind(currentwindow)

        def innermousewheel(myself):
            myself.mousewheelupdown(currentwindow, currentvalues, mousewheelupdown)

        def inneraddnewrow(myself):
            return myself.addrow(currentevent, currentwindow, currentvalues, newrowcontent=newrowcontent)

        def innersavedatatofile(myself):
            return myself.savedatatofile(outputfilename=filename, modelfilename=modelfilename,
                                         sheetnumber=sheetnumber, headline=headline, tailline=tailline)

        def innerloaddatafromfile(myself):
            return myself.loaddatafromfile(currentwindow=currentwindow,
                                           inputfilename=filename, sheetnumber=sheetnumber,
                                           headline=headline, tailline=tailline, headname=headname)

        processdic = {
            SPECIAL_TABLE_SET_ROWCOLOR: innersetrowcolor,
            SPECIAL_TABLE_SET_CELLVALUE: innersetcellvalue,
            SPECIAL_TABLE_SET_ROWVALUE: innersetrowvalue,
            SPECIAL_TABLE_SET_CURRENTROW: innersetcurrentrow,
            SPECIAL_TABLE_SET_WHOLEVALUE: innersetwholevalue,
            SPECIAL_TABLE_SET_BIND: innersetbind,
            SPECIAL_TABLE_SET_MOUSEWHEEL: innermousewheel,
            SPECIAL_TABLE_SET_ADD_NEWROW: inneraddnewrow,
            SPECIAL_TABLE_SET_SAVETOFILE: innersavedatatofile,
            SPECIAL_TABLE_SET_LOADFROMFILE: innerloaddatafromfile
        }

        # 先处理表格内部的元组事件
        if isinstance(currentevent, tuple):
            if isinstance(currentevent[0], tuple):  # 是否鼠标事件
                if currentevent[0][0] == self.tablekey and currentevent[1] == self.MOUSE_WHEEL:
                    # print("binded mousewheel event:{}".format(currentevent))
                    self.eventprocess(currentevent, currentwindow, currentvalues)

            else:
                # print("update processing event:{}".format(currentevent))
                self.eventprocess(currentevent, currentwindow, currentvalues)

        # 此外的事件为update属性的事件
        tempfunc = processdic.get(commandtype, None)
        if tempfunc is not None:
            return tempfunc(self)

    def get(self, attribute: str = SPECIAL_TABLE_GET_SELECTED_ROWCONTENT, **kwargs):
        """
        本类的get处理
        :param attribute:                   获取的属性, 默认SPECIAL_TABLE_GET_CURRENTROW_CONTENT,get()相当于默认获取当前行内容
                                            SPECIAL_TABLE_GET_MAXROWNUMBER = 'get_max_row_number'
                                            SPECIAL_TABLE_GET_MAXCOLUMNNUMBER = 'get_max_column_number'
                                            SPECIAL_TABLE_GET_CURRENTROW = 'get_current_row'
                                            SPECIAL_TABLE_GET_CURRENTROW_CONTENT = "get_current_row_content"
                                            SPECIAL_TABLE_GET_SPECIALEVENTLIST = 'get_specail_eventlist'
                                            SPECIAL_TABLE_GET_RESULT_DICT = 'get_result_dict'
                                            SPECIAL_TABLE_GET_EVENTINFLUENCEROW = 'event_influnce_row'
                                            SPECIAL_TABLE_GET_SELECTED_ROWCOLUMN = 'get_selected_row_column'
                                            SPECIAL_TABLE_GET_SELECTED_ROWCONTENT
                                            SPECIAL_TABLE_GET_COLUMNN_NAME

        """
        # 获取attribute对应的取值函数
        tempfunction = self.getattrdict.get(attribute, None)
        # 如果函数存在则返回值，否则返回None
        return None if tempfunction is None else tempfunction(**kwargs)


@dataclass
class SpecialTableGroup(SpecialTable):
    tablegroupkey: str
    specialdflist: list = None
    dflist: list = None
    tabname: str = "表"
    tabledatadictlist: list = None
    itemeditfunc: dict = None
    itemeditfuncargsdic: dict = None
    itemeditusecolvaluescale: bool = False
    indexcol: list = None
    visible_column_map: list = None
    def_col_width: int = DEF_CELL_SIZE[0]
    auto_size_columns: bool = True
    max_col_width: int = 20
    cell_justification: str = 'r'
    col_widths: list = None
    size: tuple = DEF_WINDOW_SIZE
    cell_size: tuple = DEF_CELL_SIZE
    scrollable: bool = False
    col_value_scale: list = None
    filter_visible: bool = False
    edit_button_visible: bool = True
    filter_set_value_visible: bool = False
    border_width: int = DEF_BD
    use_inner_right_click_menu: bool = False
    move_row_button_visible: bool = True
    visible: bool = True
    headings: list = None
    right_click_menu: list = None
    vertical_scroll_only: bool = True
    expand_x: True = False
    expand_y: True = False
    select_mode: str = None
    font: tuple = SPECIAL_TABLE_DEF_FT
    pad: tuple = DEF_PAD
    max_rows: int = MAX_SPECIAL_TABLE_ROWS
    max_columns: int = MAX_SPECIAL_TABLE_COLUMNS
    text_color: str = 'black'
    text_background_color: str = 'white'
    head_text_color: str = 'white'
    show_cell_tooltips: bool = True
    head_background_color: str = 'dark blue'
    input_text_color: str = 'blue'
    input_text_background_color: str = 'white'
    selected_text_color: str = 'dark gray'
    selected_background_color: str = 'blue'
    combox_button_arrow_color: str = "blue"
    combox_button_background_color: str = "white"
    column_cell_edit_type_map: list = None
    v_slider_resolution: float = 0.1
    h_slider_resolution: float = 0.1
    slider_every_row_column: bool = True
    v_slider_show: bool = True
    h_slider_show: bool = True
    relief: str = Sg.RELIEF_FLAT
    cell_editable: bool = True
    slider_relief: str = Sg.RELIEF_FLAT
    disable_slider_number_display: bool = False
    datachangeeventtrigger: bool = False
    showcolnamewithid: bool = False
    showcolumnname: bool = True
    display_row_numbers: bool = True
    enable_events: bool = False
    enable_click_events: bool = False

    """
    表格编辑框类， 该类集成了SpecialTable，事件有两种
    1、（tablegroupkey, talblegroupkey) 该元组事件代表tab相关的事件
    2、（tablegroupkey，x， y) | (tablegroupkey , EDIT|.....) 都是父类的表格相关事件
    3、需要重载父类的eventprocess 对1、2分别处理
    4、 update和get直接对应父类的 tablegroupkey

    :param tablegroupkey:           Used with window.find_element and with return values to uniquely identify
                                    this element to uniquely identify this element
    :type tablegroupkey:            str | int | tuple | object
    :param specialdflist:           SpecialDf list, 用于显示的数据 SpecialDf 格式，优先使用
    :param dflist:                  pd.DataFrame list, 用于显示的数据
    :param tabname                  str: tab的标签前缀
    :param tabledatadictlist:        dict格式的数据，dfshowdata数据为None时，使用tabledatadict的数据
    :param itemeditfunc:             单条记录编辑时用到的界面处理函数，输入参数需要用字典方式带入。以适应可变参数
    :param itemeditfuncargsdic:      dict:  itemeditfunc函数的参数，row编辑函数，
    :param itemeditusecolvaluescale:    bool: True, itemeditfunc 需要将col_value_scale传递过去，
    :param indexcol:                  索引列序号（从0开始,目前只支持单列索引， None为没有），
                                      如果有索引列，会根据索引列判断是否有重复值并强制提醒
                                      None|int|[int,int...]
    :param visible_column_map:        是否显示该栏，
    :type visible_column_map:         List[bool]
    :param def_col_width:              默认列宽
    :param auto_size_columns:          列宽根据初始化自适应
    :param max_col_width：              最大列宽
    :param col_widths:                该栏显示宽度
    :type col_widths:                  List[int]
    :param size                        tuple(int,int) , 整个lay的尺寸
    :param size                        tuple(int,int), 表格单元的默认尺寸
    :param col_value_scale:            list , None or [[], [], None, ......]
                                       None: 没有任何限制，全部列都可以修改
                                      （二维list， 对应每列的取值范围，用于列快速填充）
                                       [None]:该列不允许批量编辑
                                       []该列任意取值范围
                                       [a,b,c]不为空，该列取值限制范围
    :param use_inner_right_click_menu       bool: True, 显示并内部右键菜单
    :param filter_visible:             bool: True, 显示筛选功能，False：关闭筛选功能
    :param filter_set_value_visible:   bool: True, 显示列赋值功能，False：关闭列赋值功能
    :param edit_button_visible:        bool: True, 各种编辑按钮，False：关闭各种编辑按钮，
    :param move_row_button_visible:    bool: True, 各种移动行按钮，False：关闭各种移动行按钮，
    :param right_click_menu:           A list of lists of Menu items to show when this element is right clicked.
    :type right_click_menu:            List[List[ List[str] | str ]]
    :param vertical_scroll_only:       if True only the vertical scrollbar will be visible
    :param select_mode:                Valid values include:
                                        TABLE_SELECT_MODE_NONE          不能选择
                                        TABLE_SELECT_MODE_BROWSE        只能单选
                                        TABLE_SELECT_MODE_EXTENDED      可以多选
                                        TABLE_LISTBOX_SELECT_MODE_SINGLE      增加一个选项 (点击后不取消）
    :param font                        窗口中使用的字体大小
    :param pad                         窗口中元素间的pad
    :param max_rows:                   最大表行数
    :param max_columns:                最大表列数
    :param expand_x:                   横向自动扩展
    :param expand_y:                   纵向自动扩展
    :param show_cell_tooltips:         cell方式下显示tooltips
    :param text_color:                 表格文本颜色
    :param text_background_color:      表格文本背景色
    :param border_width:                 表格边界线宽度
    :param head_text_color:             标题及序号文本颜色
    :param head_background_color:       标题及序号背景色
    :param selected_text_color:         选中文本颜色
    :param selected_background_color:   选中文本背景色
    :param combox_button_arrow_color:   cell为combox时，下拉箭头的颜色
    :param combox_button_background_color:  cell为combox时，下拉箭头的背景色
    :param input_text_color:            可编辑框的文本颜色
    :param input_text_background_color:  可编辑框的背景颜色
    :param column_cell_edit_type_map:       None, 所有单元格使用默认Sg.Input,
                                        list: 长度和表格数据的列数一致（该种方式下，所有列在一个max_column 页内显示，
                                        即表格页不要横行左右翻转，
                                        SPECIAL_TABLE_CELL_INPUT, SPECIAL_TABLE_CELL_COMBOX, SPECIAL_TABLE_CELL_TEXT
    :param relief:                      cell的relief
    :param cell_editable:               表格单元是否可以直接编辑
    :param slider_relief:               slider relief  style.Use constants -
                    RELIEF_RAISED  RELIEF_SUNKEN  RELIEF_FLAT RELIEF_RIDGE RELIEF_GROOVE RELIEF_SOLID
    :param v_slider_resolution:         v  slider分辨率，默认0.1
    :param h_slider_resolution:         h slider 分辨率， 默认0.1
    :param slider_every_row_column:    是否可以逐行、逐列滚动
    :param v_slider_show:               bool, 是否显示垂直滚动条
    :param h_slider_show:               bool, 是否显示水平滚动条
    :param disable_slider_number_display:          bool: True 不显示， False：显示
    :param datachangeeventtrigger      bool: 数据修改后是否向主窗口发送数据变化事件
    :param showcolnamewithid           bool: True, 显示列号， False：不显示列号
    :param showcolumnname              bool: True, 显示列名， False：显示列名
    :param display_row_numbers                bool: True, 显示序号列， False：不显示
    :param enable_events:           Turns on the element specific events. Table events happen when row is clicked
    :type enable_events:            (bool)
    :param enable_click_events:     (NO USE) Turns on the element click events that will give you (row, col) click data when the table is clicked
    :type enable_click_events:      (bool)
    :return:                           用于Sg.window的页面layout list
    """

    def __post_init__(self):
        self.enable_events = False
        self.currenttable = None
        self.tabledynamicinfo = {}
        if self.specialdflist is None or self.specialdflist == []:
            if self.dflist is not None and self.dflist != []:
                self.specialdflist = [SpecialDf(x) for x in self.dflist]
        else:
            self.dflist = [x.dfdata for x in self.specialdflist]
        super().__init__(
            tablekey=self.tablegroupkey,
            specialdf=None if self.specialdflist is None else None
            if len(self.specialdflist) == 0 else self.specialdflist[0],
            dfshowdata=None if self.dflist is None else None
            if len(self.dflist) == 0 else self.dflist[0],
            tabledatadictlist=self.tabledatadictlist,
            itemeditfunc=self.itemeditfunc,
            itemeditfuncargsdic=self.itemeditfuncargsdic,
            itemeditusecolvaluescale=self.itemeditusecolvaluescale,
            indexcol=self.indexcol,
            visible_column_map=self.visible_column_map,
            def_col_width=self.def_col_width,
            auto_size_columns=self.auto_size_columns,
            max_col_width=self.max_col_width,
            cell_justification=self.cell_justification,
            col_widths=self.col_widths,
            size=self.size,
            cell_size=self.cell_size,
            scrollable=self.scrollable,
            # col_value_scale=self.col_value_scale,
            col_value_scale=None,
            filter_visible=self.filter_visible,
            edit_button_visible=self.edit_button_visible,
            filter_set_value_visible=self.filter_set_value_visible,
            border_width=self.border_width,
            use_inner_right_click_menu=self.use_inner_right_click_menu,
            move_row_button_visible=self.move_row_button_visible,
            visible=self.visible,
            headings=self.headings,
            right_click_menu=self.right_click_menu,
            vertical_scroll_only=self.vertical_scroll_only,
            expand_x=self.expand_x,
            expand_y=self.expand_y,
            select_mode=self.select_mode,
            font=self.font,
            pad=self.pad,
            max_rows=self.max_rows,
            max_columns=self.max_columns,
            text_color=self.text_color,
            text_background_color=self.text_background_color,
            head_text_color=self.head_text_color,
            show_cell_tooltips=self.show_cell_tooltips,
            head_background_color=self.head_background_color,
            input_text_color=self.input_text_color,
            input_text_background_color=self.input_text_background_color,
            selected_text_color=self.selected_text_color,
            selected_background_color=self.selected_background_color,
            combox_button_arrow_color=self.combox_button_arrow_color,
            combox_button_background_color=self.combox_button_background_color,
            column_cell_edit_type_map=self.column_cell_edit_type_map,
            v_slider_resolution=self.v_slider_resolution,
            h_slider_resolution=self.h_slider_resolution,
            slider_every_row_column=self.slider_every_row_column,
            v_slider_show=self.v_slider_show,
            h_slider_show=self.h_slider_show,
            relief=self.relief,
            cell_editable=self.cell_editable,
            slider_relief=self.slider_relief,
            disable_slider_number_display=self.disable_slider_number_display,
            datachangeeventtrigger=self.datachangeeventtrigger,
            showcolnamewithid=self.showcolnamewithid,
            showcolumnname=self.showcolumnname,
            display_row_numbers=self.display_row_numbers,
            enable_events=self.enable_events,
            enable_click_events=self.enable_click_events
        )
        self._dftabgroup = []
        # 1、先初始化tablelayout，然后叠加tab类后返回整个layout
        # SpecailTable 总长度和表格实际显示长度的差异，需要根据以下的设置确定 :
        # filter_visible, filter_set_value_visible, edit_button_visible,  move_row_button_visible,
        # 四项设置的对应关系
        # 保存tabgroup的每个tab的layout以便后续处理
        if self.specialdflist is None:
            self.specialdflist = []
        # 针对表数建立标签组， 并将第一张表作为当前显示表格， 标签都无实际内容，使用空格[]
        # 标签的key采用（tablegroupkey, i)，
        # 当标签组产生（tablegroupkey, tablegroupkey)事件后，
        # 获取具体标签值为（tablegroupkey, i)，根据i确定切换的表specialdflist序号
        for i in range(MAX_TABLE_NUMBER):
            self._dftabgroup.append(Sg.Tab("{}[{}]".format(self.tabname, i), [],
                                           tooltip=self.specialdflist[i] if i < len(self.specialdflist) else "",
                                           visible=True if i < len(self.specialdflist) else False,
                                           k=(self.tablegroupkey, i)))
        self.currenttable = (self.tablegroupkey, 0) if len(self.specialdflist) > 0 else None

        # 形成有tab的整体layout,
        self._grouplayout = [[Sg.Frame("", [
            [Sg.TabGroup(
                [self._dftabgroup], k=(self.tablegroupkey, self.tablegroupkey), expand_x=False, expand_y=False,
                size=(self.size[0], 5), visible=self.visible, enable_events=True)],
            [Sg.Frame("", self._tablelayout, size=self.size, border_width=0)]
        ])]]

    @property
    def layout(self):
        return self._grouplayout

    def tabgroupeventprocess(self, event, window, values):
        """
        目前切换恢复页面位置，暂未考虑动态切换表格时，保留过滤项，
        :param event:
        :param window:
        :param values:
        :return:
        """
        # 获取当前点击标签，如果和上次不同，需要切换表格， 如果相同，不需要处理
        currenttable = values[event]

        # 保存现有表格的动态信息，以备恢复
        # 先判断要刷新的表格是否已经没有
        if len(self.specialdflist) == 0:
            self.currenttable = None
            # 没有任何可以显示的表格，用空值表格作为显示底图
            self.setwholevalue(currentwindow=window, dfdata=pd.DataFrame([], columns=[]))
            self.refreshcurrentshow(currentwindow=window)
            self.refreshwholegroup(event, window, values)
            return

        # 切换前，将必要的动态数据保存到SpecialDf基类中
        if self.currenttable is not None:
            # 备份当前表的动态显示项
            # 进度条位置
            self.specialdf.v_slider_position, self.specialdf.h_slider_position = \
                values[self.V_SLIDER_KEY], values[self.H_SLIDER_KEY]
            # 保留过滤条件
            if self.filter_visible:     # self.filter_visible is True, 相关的elements才会存在
                self.specialdf.filter_startrow, self.specialdf.filter_endrow, \
                self.specialdf.filter_filter_column, self.specialdf.filter_filter_like, \
                self.specialdf.filter_set_value_column, self.specialdf.filter_set_column_used_value = \
                    values[self.ROW_START], values[self.ROW_END], values[self.COLUMNFILTER], \
                    values[self.CONTENTLIKE], values[self.COLCHOOSE], values[self.COL_BAT_VALUE]
            # print("current {} dynamic info backup：{} ".format(self.currenttable, self.specialdf))

        # 开始切换
        self.currenttable = currenttable
        # 切换表格数据到新的选择
        self.specialdf = self.specialdflist[self.currenttable[1]]

        # 重置选中的标签，因为事件可能是点击产生，也可能是外部程序write_event_value
        window[self.currenttable].select()  #
        # 重新设置tooltip
        window[self.currenttable].set_tooltip("{}{}{}{}{}".format(
                "[输入] " if self.specialdf.dataimportsource != "" else "",
                self.specialdf.dataimportsource,
                " -> " if self.specialdf.dataimportsource != "" and self.specialdf.dataexporttarget != ""
                else "",
                "[输出] " if self.specialdf.dataexporttarget != "" else "",
                self.specialdf.dataexporttarget))
        # 恢复备份的当前表格的动态信息
        # print("new table {} using self backup info and specialdf: {}".format(self.currenttable, self.specialdf))
        window[self.V_SLIDER_KEY].update(value=self.specialdf.v_slider_position)
        window.write_event_value(self.V_SLIDER_KEY,  self.specialdf.v_slider_position)
        window[self.H_SLIDER_KEY].update(value=self.specialdf.h_slider_position)
        window.write_event_value(self.H_SLIDER_KEY, self.specialdf.h_slider_position)
        if self.filter_visible:
            # 恢复过滤项开始、结束行号等选择项
            rowlist = [""] + [str(a) for a in range(0, self.specialdf.maxrownumber)]
            window[self.ROW_START].update(values=rowlist, value=self.specialdf.filter_startrow)
            window[self.ROW_END].update(values=rowlist, value=self.specialdf.filter_endrow)
            window[self.COLUMNFILTER].update(
                values=self.specialdf.colnamefilterlist, value=self.specialdf.filter_filter_column)
            window[self.CONTENTLIKE].update(value=self.specialdf.filter_filter_like)
            window[self.COLCHOOSE].update(
                values=self.specialdf.colnameusedlist, value=self.specialdf.filter_set_value_column)
            window[self.COL_BAT_VALUE].update(value=self.specialdf.filter_set_column_used_value)
        self.refreshcurrentshow(currentwindow=window)
        self.resetselected(event, window, values)
        if self.specialdf.infilterstatus:
            # 如果条件不是全为空，
            window.write_event_value(self.FILTERCONDITIONOK, None)
        window.refresh()

    def refreshwholegroup(self, event, window, values):
        """
        根据当前specialdflist，刷新tablegroup的显示
        :param event:
        :param window:
        :param values:
        :return:
        """

        # 判断当前数据表格数是否大于系统设置的最大表数，如果是，提醒用户后自动增加，并建议用户扩大系统配置
        if len(self.specialdflist) > len(self._dftabgroup):
            Sg.popup_ok(
                "当前表格数{}将超过系统最大配置显示数{}，将临时动态增加，建议在系统配置中扩大！".format(
                    len(self.specialdflist), MAX_TABLE_NUMBER),
                keep_on_top=True, modal=True, icon=SYSTEM_ICON)
            for i in range(len(self._dftabgroup),
                           len(self._dftabgroup) + (len(self.specialdflist) - len(self._dftabgroup)) + 5):
                self._dftabgroup.append(Sg.Tab("{}[{}]".format(self.tabname, i), [],
                                               tooltip=self.specialdflist[i].dataimportsource
                                               if i < len(self.specialdflist) else "",
                                               visible=True if i < len(self.specialdflist) else False,
                                               k=(self.tablegroupkey, i)))
                window[(self.tablegroupkey, self.tablegroupkey)].add_tab(self._dftabgroup[-1])
        # 关闭大于specialdflist的表格显示以及tab的tooltip
        for i in range(len(self.specialdflist), len(self._dftabgroup)):
            window[(self.tablegroupkey, i)].update(visible=False)
            window[(self.tablegroupkey, i)].set_tooltip(None)
        for i in range(len(self.specialdflist)):
            window[(self.tablegroupkey, i)].update(visible=True)
            window[(self.tablegroupkey, i)].set_tooltip("{}{}{}{}{}".format(
                "[输入] " if self.specialdflist[i].dataimportsource != "" else "",
                self.specialdflist[i].dataimportsource,
                " -> " if self.specialdflist[i].dataimportsource != "" and self.specialdflist[i].dataexporttarget != ""
                else "",
                "[输入] " if self.specialdflist[i].dataexporttarget != "" else "",
                self.specialdflist[i].dataexporttarget))

        # 全面刷新后回到第0表
        if len(self.specialdflist) > 0:
            window[(self.tablegroupkey, 0)].select()  # 选择0
            # 当values没有值，说明不是内部处理，直接外部事件触发，values有值，说明是内部调用，不菜触发外部事件
            if values is None:
                window.write_event_value((self.tablegroupkey, self.tablegroupkey), (self.tablegroupkey, 0))
            else:
                values[(self.tablegroupkey, self.tablegroupkey)] = (self.tablegroupkey, 0)
                self.tabgroupeventprocess((self.tablegroupkey, self.tablegroupkey), window, values)

    def tiggertabclicked(self, event, window, values, triggertable=None):
        """

        :param window:
        :param values:
        :param triggertable:
        :return:
        """
        # 确定要激活的tab号, 没有指定表格如果表0存在就激活表0, 否则直接返回
        if len(self.specialdflist) == 0:
            print("表格为空，没有可以激活的表格！")
            return
        localtable = 0 if triggertable is None else 0 if triggertable >= len(self.specialdflist) else triggertable
        # 不采用模拟发送事件，因为会回到主程序，直接调用事件处理函数
        # window.write_event_value((self.tablegroupkey, self.tablegroupkey), (self.tablegroupkey, localtable))
        # simulate values
        values[(self.tablegroupkey, self.tablegroupkey)] = (self.tablegroupkey, localtable)
        self.tabgroupeventprocess((self.tablegroupkey, self.tablegroupkey), window, values)

    def eventprocess(self, event, window, values):
        # 判断是否点击tablegroup标签事件，重载父类方法，对于父类事件调用父类方法
        if event == (self.tablegroupkey, self.tablegroupkey):
            self.tabgroupeventprocess(event, window, values)
        else:  # 父类表格事件
            # print("tabgroup got table event: {}, currenttable:{}".format(event, self.currenttable))
            super().eventprocess(event, window, values)

    def update(self, commandtype=None, currentevent=None, currentwindow=None,
               currentvalues=None, triggertable=None, **kwargs):
        """

        :param commandtype:
        :param currentevent:
        :param currentwindow:
        :param currentvalues:
        :param triggertable:                SPECIAL_TABLE_GROUP_TRIGGER_TAB 激活得表号
        :param kwargs:
        :return:
        """
        # 分离出父类要处理的事件，目前仅fresh
        if commandtype in [SPECIAL_TABLE_GROUP_REFRESH, SPECIAL_TABLE_GROUP_TRIGGER_TAB]:
            if commandtype == SPECIAL_TABLE_GROUP_REFRESH:  # 刷新整个group
                self.refreshwholegroup(currentevent, currentwindow, currentvalues)
                return True
            elif commandtype == SPECIAL_TABLE_GROUP_TRIGGER_TAB:  # 触发点击group某张表格，没有指定表格如果表0存在就激活表0
                self.tiggertabclicked(currentevent, currentwindow, currentvalues, triggertable)
                return True
        else:
            result = super().update(commandtype=commandtype, currentevent=currentevent,
                                    currentwindow=currentwindow, currentvalues=currentvalues, **kwargs)

        # 对于子类处理后，父类需要善后的事件，单独处理
        if result is not False:
            if commandtype in [SPECIAL_TABLE_SET_LOADFROMFILE, SPECIAL_TABLE_SET_LOADFROMFILE]:
                self.refreshwholegroup(currentevent, currentwindow, currentvalues)
            return True
        else:
            return False

    def get(self, attribute=None, **kwargs):
        """
        如果没有命令，默认返回当前表号（类似TabGroup），其他情况传递给父类
        :param attribute:
        :param kwargs:
        :return:
        """
        if attribute is None:
            return self.currenttable
        else:
            return super().get(attribute=attribute, **kwargs)


def main_test():
    from PySimpleGUI import THEME_CLAM, THEME_DEFAULT, THEME_ALT, THEME_CLASSIC
    Sg.set_options(ttk_theme=THEME_DEFAULT)
    dfdata1 = pd.DataFrame(
        ["数据{}-{}".format(x, y) for y in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
        for x in range(MAX_SPECIAL_TABLE_ROWS * MAX_SPECIAL_TABLE_ROWS - 10))
    dfdata2 = pd.DataFrame(
        ["测试{}-{}".format(x, y) for y in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
        for x in range(MAX_SPECIAL_TABLE_ROWS * MAX_SPECIAL_TABLE_ROWS - 10))
    dfdata3 = pd.DataFrame(
        ["新数据{}-{}".format(x, y) for y in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
        for x in range(MAX_SPECIAL_TABLE_ROWS * MAX_SPECIAL_TABLE_ROWS - 10))
    dfdata4 = pd.DataFrame([["1", "2", "3"], ["4", "5", "6"], ['7', "8", "9"]], columns=list('abc'))
    dfdata5 = pd.DataFrame([], columns=list("ABC"))
    layoutstatic = []
    tempfuncargs = {
        "accounttypedic": test_accounttypedic,
        "functioncalldic": test_functioncalldic,
    }
    commandtypelist = ["本地文件处理", 'RPA处理', '测试文件处理']
    colvaluesecale = [
        None,  # 菜单名称，索引列，建议不要批处理修改
        None,
        None,
        None,
        None,
        None,
        [],  # 菜单配置文件
        ['是', '否'],
        commandtypelist,  # 命令类型
        []  # 图标文件
    ]

    tablegroupkey = '_table_group_'
    headings1 = ["数据列{}".format(x) for x in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
    headings2 = ["测试列{}".format(x) for x in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
    headings3 = ["新数据列{}".format(x) for x in range(MAX_SPECIAL_TABLE_COLUMNS * 2)]
    dfdata1.columns = headings1
    dfdata2.columns = headings2
    dfdata3.columns = headings3
    tablemenu = ["测试菜单", ['test1', 'test2']]
    visible_column_map = [True, True, True, True, True, True, True, True, True, True]
    # visible_column_map = [True, False, False, False, False, False, False, False, False, False]
    col_widths = [16, 16, 24, 24, 15, 18, 18, 20, 20, 20]
    column_cell_edit_type_map = [
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_INPUT,
        SPECIAL_TABLE_CELL_COMBOX,
        SPECIAL_TABLE_CELL_COMBOX,
        SPECIAL_TABLE_CELL_INPUT,
    ]
    rightclickmenu = [
        ["单表格操作", ["行操作", ["删除行::('table0', 1, -1)", "添加行", "行求和"], "行操作1", ["删除行", "添加行", "行求和"]]],
        ["列操作", ["删除列", "和并列", "列和"]],
        ["多表格操作", ["表格融合", "纵向拼接", "横向拼接"]],
        ["格式化输出"]]
    # RELIEF_RAISED  RELIEF_SUNKEN  RELIEF_FLAT RELIEF_RIDGE RELIEF_GROOVE RELIEF_SOLID
    dflist = [dfdata1, dfdata2, dfdata3]
    specialdflist = [SpecialDf(dfdata=x) for x in dflist]
    layouttable = SpecialTableGroup(
        tablegroupkey,
        specialdflist=specialdflist,
        # dflist=dflist,
        # dflist=None,
        tabname="测试表",
        max_rows=19,
        max_columns=11,
        enable_events=True,
        display_row_numbers=True,
        showcolnamewithid=True,
        showcolumnname=True,
        combox_button_arrow_color='blue',
        combox_button_background_color='white',
        use_inner_right_click_menu=True,
        input_text_color='blue',
        input_text_background_color='white',
        # tabledatadictlist=test_parminfo,
        # itemeditfunc=cmd_page_config_item_edit,
        itemeditfunc=None,
        itemeditfuncargsdic=tempfuncargs,
        v_slider_resolution=0.1,
        h_slider_resolution=0.1,
        slider_every_row_column=True,
        col_widths=col_widths,
        pad=(0, 0),
        border_width=1,
        text_color="blue",
        scrollable=False,
        size=(1400, 400),
        # col_widths=test_column_widths,
        # headings=headings,
        slider_relief=Sg.RELIEF_FLAT,
        v_slider_show=True,
        h_slider_show=True,
        relief=Sg.RELIEF_SUNKEN,
        disable_slider_number_display=True,
        # select_mode=SPECIAL_TABLE_LISTBOX_SELECT_MODE_SINGLE,
        # select_mode=SPECIAL_TABLE_SELECT_MODE_BROWSE,
        # select_mode=SPECIAL_TABLE_SELECT_MODE_NONE,
        select_mode=SPECIAL_TABLE_SELECT_MODE_EXTENDED,
        cell_editable=True,
        edit_button_visible=True,
        move_row_button_visible=True,
        indexcol=[0, 1],
        filter_visible=True,
        filter_set_value_visible=False,
        # visible_column_map=visible_column_map,
        # right_click_menu=rightclickmenu,
        column_cell_edit_type_map=column_cell_edit_type_map,
        col_value_scale=None,
    ).layout

    layouttail = [
        [
            Sg.Text("".rjust(20)),
            Sg.Button('保存&退出', k='_save_', pad=(10, 10)),
            Sg.Text("".rjust(20)),
            Sg.Button('换数据', k='_change_', pad=(10, 10)),
            Sg.Text("".rjust(20)),
            Sg.Button('导入数据', k='_loaddata_', pad=(10, 10)),
            Sg.Text("".rjust(20)),
            Sg.Button(' 关  闭 ', k='_close_', pad=(10, 20)),
        ]
    ]
    # 拼接页面配置
    layout = layoutstatic + layouttable + layouttail

    mainwindow = Sg.Window('UMD SPECIAL TABLE 功能测试',
                           layout,
                           font=SPECIAL_TABLE_DEF_FT,
                           icon=SYSTEM_ICON,
                           # return_keyboard_events=True,
                           resizable=False,
                           size=(1500, 660),
                           modal=True,
                           # keep_on_top=True,
                           finalize=True)

    # 初始化mousewheel updown相关的环境
    mainwindow[tablegroupkey].update(commandtype=SPECIAL_TABLE_SET_BIND, currentwindow=mainwindow)
    mainwindow[tablegroupkey].update(commandtype=SPECIAL_TABLE_SET_ADD_NEWROW, currentwindow=mainwindow,
                                     newrowcontent=["addrowtest{}".format(x) for x in range(MAX_SPECIAL_TABLE_COLUMNS)])

    while True:
        event, values = mainwindow.read()
        # print("getevent:{}".format(event))
        # --- Process buttons --- #
        if event in (Sg.WIN_CLOSED, '_close_'):
            break
        elif event == '_loaddata_':
            mainwindow[tablegroupkey].update(commandtype=SPECIAL_TABLE_SET_LOADFROMFILE,
                                             currentwindow=mainwindow,
                                             currentevent=event,
                                             filename="d:/temp/aaa.xlsx",
                                             headline=0,
                                             tailline=0,
                                             sheetnumber=0,
                                             headname=0)

        elif event == '_change_':
            specialdflist.pop(0)
            dflist.pop(0)
            mainwindow[tablegroupkey].update(
                commandtype=SPECIAL_TABLE_GROUP_REFRESH,
                currentevent=event, currentwindow=mainwindow, currentvalues=values)
            # 刷新tabgroup

        elif isinstance(event, tuple):
            # 兼容Sg.table和Special element方式, 已经不用考虑
            # 以及鼠标wheel处理, 目前 （tablegroupkey,x, y), "mousewheel"
            currenttable = event[0][0] if isinstance(event[0], tuple) else event[0]

            mainwindow[currenttable].update(
                commandtype=SPECIAL_TABLE_TUPLE_EVENTPROCESS,
                currentevent=event, currentwindow=mainwindow, currentvalues=values)
            continue
        elif event == "_save_":
            print(mainwindow[tablegroupkey].get(SPECIAL_TABLE_GET_RESULT_DICT))
            continue
        elif event == tablegroupkey:
            mainwindow[tablegroupkey].eventprocess(event, mainwindow, values)
            continue

        elif event == tablegroupkey:
            print("table event:{}".format(tablegroupkey))
            print("values[tablegroupkey]: {}".format(values[tablegroupkey]))
            print("mainwindow[tablegroupkey].get(): {}".format(mainwindow[tablegroupkey].get()))
            print("mainwindow[tablegroupkey].get(): {}".format(
                mainwindow[tablegroupkey].get(attribute=SPECIAL_TABLE_GET_SELECTED_ROWCONTENT)))
        else:
            # 判断是否为innerrightmenu事件，如果是就触发该事件, 当窗口中有表格设置了 use_inner_right_click_menu，
            # 需要调用isinnerrightclickmenu 进行回送事件处理
            tempresult = isinnerrightclickmenu(mainwindow, event)
            if tempresult is False:
                print("configwindow unknow event:{}".format(event))
            continue

    mainwindow.close()


if __name__ == '__main__':
    test_parminfo = [
    ]
    test_column_widths = [30, 1, 8, 3, 16, 40, 1, 6, 4, 4]
    test_visible_column_map = [True, False, True, True, True, True, False, True, True, True]
    test_columninfodic = {
        '命令': 'exe_unit',
        '账号需求': 'need_password',
        '账号类型 ': 'account_type',
        '选项': 'cmd_option',
        '项目描述': 'item_name',
        '默认文件': 'file_name',
        '文件类型': 'file_type_info',
        '文件后缀': 'file_type',
        '读/写': 'save_mode',
        '可更改': 'changable'
    }
    test_accounttypedic = {
        "JDE": {
            "default_account": "f9b531805c280b9f61e0dfdd8e69c2fd",
            "account_password": {
            }
        },
        "ERP": {
            "default_account": "f9b531805c280b9f61e0dfdd8e69c2fd",
            "account_password": {}
        },
        "EAS": {
            "default_account": "45d2a8efdea386c3b3f400c5961cf84a",
            "account_password": {}
        }
    }
    test_errdefinedic = {
        "101": "帐号为空",
        "102": "密码为空",
        "103": "密码错误",
        "104": "未知错误",
        "201": "命令所需文件不全",
        "202": "读取文件失败",
        "203": "文件格式不匹配",
        "204": "写入文件错误",
        "205": "未选择账号"
    }
    test_functioncalldic = {
        "无": "",
    }
    test_colvaluescale = [
        [],  # 命令
        ['是', '否'],  # 是否需要账号运行
        [a for a in test_accounttypedic.keys()],  # 账号类型
        [''] + ['-' + chr(i) for i in range(ord('a'), ord('z') + 1)],  # 命令选项
        [],  # 项目描述
        [],  # 默认文件
        None,  # 文件类型,与文件后缀需要互动， 因此不允许批量更改
        None,  # 文件后缀,与文件类型需要互动， 因此不允许批量更改
        ['读', '写'],  # 文件为读还是写
        ['是', '否']  # 是否可更改
    ]
    main_test()

