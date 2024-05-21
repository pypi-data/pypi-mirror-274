from string import printable
from typing import Union, Optional, List, Dict, Iterable

from deprecation import deprecated

from divinegift import main, version

import os
from datetime import datetime, timedelta, date

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException, IllegalCharacterError
    from openpyxl.styles import Border, Side, Alignment, Font
except ImportError:
    raise ImportError("openpyxl isn't installed. Run: pip install -U openpyxl")
try:
    import xlrd
except ImportError:
    pass
try:
    import xlsxwriter
except ImportError:
    raise ImportError("xlsxwriter isn't installed. Run: pip install -U xlsxwriter")
try:
    import xlwt
except ImportError:
    raise ImportError("xlwt isn't installed. Run: pip install -U xlwt")

BORDER_NONE = {
    'left': {
        'style': None,
        'color': 'FFFFFFFF'
    },
    'right': {
        'style': None,
        'color': 'FFFFFFFF'
    },
    'bottom': {
        'style': None,
        'color': 'FFFFFFFF'
    },
    'top': {
        'style': None,
        'color': 'FFFFFFFF'
    },
}

BORDER_BLACK_LINE = {
    'style': 'thin',
    'color': '000000'
}

BORDER = {
    'none': BORDER_NONE,
    'all': {
        'left': BORDER_BLACK_LINE,
        'right': BORDER_BLACK_LINE,
        'bottom': BORDER_BLACK_LINE,
        'top': BORDER_BLACK_LINE,
    },
    'left': {
        **BORDER_NONE,
        'left': BORDER_BLACK_LINE,
    },
    'right': {
        **BORDER_NONE,
        'right': BORDER_BLACK_LINE,
    },
    'bottom': {
        **BORDER_NONE,
        'bottom': BORDER_BLACK_LINE,
    },
    'top': {
        **BORDER_NONE,
        'top': BORDER_BLACK_LINE,
    }
}

FONT = {
    'default': {
        'name': 'Calibri',
        'size': 11,
        'bold': False,
        'italic': False,
        'underline': 'none',
        'strike': False,
        'color': '000000'
    }
}

NUMBER_FORMAT = {
    'date': {'number_format': 'dd.mm.yyyy'},
    'datetime': {'number_format': 'dd.mm.yyyy hh:mm:ss'},
    'other': {'number_format': 'general'}
}


class Excel:
    def __init__(self, filename: str = None):
        self.filename = filename
        self.styles = {}

        self.workbook = None
        self.worksheets = {}
        self.current_worksheet = None

        self.rows_qty: int = 0

    def get_file(self):
        raise NotImplementedError(f'Method get_file not implemented in class {self.__class__.__name__}')

    def create_file(self):
        raise NotImplementedError(f'Method create_file not implemented in class {self.__class__.__name__}')

    def open_file(self, filename: str = None):
        # Если указанной папки не существует - создаем ее (включая родительские, если нужно)
        if filename:
            self.filename = filename
        try:
            fd, fn = os.path.split(self.filename)
            if fd:
                main.check_folder_exist(fd)
        except TypeError:
            pass

    def save_file(self):
        raise NotImplementedError(f'Method save_name not implemented in class {self.__class__.__name__}')

    def add_worksheet(self, worksheet: str):
        raise NotImplementedError(f'Method add_worksheet not implemented in class {self.__class__.__name__}')

    def rename_worksheet(self, worksheet_old: str, worksheet_new: str):
        raise NotImplementedError(f'Method rename_worksheet not implemented in class {self.__class__.__name__}')

    def select_worksheet(self, worksheet):
        self.current_worksheet = self.worksheets.get(worksheet)

    def add_cell(self, row: int, col: int, value: Union[None, str, float, int, date, datetime],
                 style: Optional[Dict] = None):
        raise NotImplementedError(f'Method add_cell not implemented in class {self.__class__.__name__}')

    def add_cell_coords(self, coords: str, value: Union[None, str, float, int, date, datetime],
                        style: Optional[Dict] = None):
        raise NotImplementedError(f'Method add_cell_coords not implemented in class {self.__class__.__name__}')

    def add_row(self, row: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                start_col: int = 0, styles: Union[None, List, Dict] = None):
        raise NotImplementedError(f'Method add_row not implemented in class {self.__class__.__name__}')

    def add_column(self, col: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                   start_row: int = 1, styles: Union[None, List, Dict] = None):
        raise NotImplementedError(f'Method add_column not implemented in class {self.__class__.__name__}')

    def add_merge_cell(self, first_row: int, first_col: int, last_row: int, last_col: int,
                       value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        raise NotImplementedError(f'Method add_merge_row not implemented in class {self.__class__.__name__}')

    def add_merge_cell_coords(self, first_coords: str, last_coords,
                              value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        raise NotImplementedError(f'Method add_merge_cell_coords not implemented in class {self.__class__.__name__}')

    def add_list_rows(self, first_row: int, first_col, list_rows: List, styles: Union[None, List, Dict] = None):
        raise NotImplementedError(f'Method add_list_rows not implemented in class {self.__class__.__name__}')

    def set_cell_dimension(self, row: int, col: int, height: int = 15, width: int = 10):
        raise NotImplementedError(f'Method set_cell_dimension not implemented in class {self.__class__.__name__}')

    def set_columns_width(self, list_columns_width: List[int], start_col: int = 1):
        raise NotImplementedError(f'Method set_columns_width not implemented in class {self.__class__.__name__}')

    def set_rows_height(self, list_rows_height: List[int], start_row: int = 1):
        raise NotImplementedError(f'Method set_rows_height not implemented in class {self.__class__.__name__}')

    def create_simple_excel(self, filename, data: List[Union[Dict, List]], column_width: Optional[List] = None,
                            column_names: Optional[List] = None, styles: Union[None, List, Dict] = None):
        raise NotImplementedError(f'Method create_simple_excel not implemented in class {self.__class__.__name__}')

    def get_max_row(self):
        raise NotImplementedError(f'Method get_max_row not implemented in class {self.__class__.__name__}')

    def get_max_column(self):
        raise NotImplementedError(f'Method get_max_column not implemented in class {self.__class__.__name__}')

    def read_file(self, excel_header: Union[list, dict], sheet_name: str = None,
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
        raise NotImplementedError(f'Method read_file not implemented in class {self.__class__.__name__}')


class ExcelX(Excel):
    def open_file(self, filename: str = None, rewrite=True):
        if filename and not filename.endswith('xlsx'):
            filename += '.xlsx'

        super(ExcelX, self).open_file(filename=filename)

        if os.path.exists(self.filename) and not rewrite:
            self.get_file()
        else:
            self.create_file()

        self.worksheets = {w.title: w for w in self.workbook.worksheets}
        self.current_worksheet = self.workbook.active

    def get_file(self):
        self.workbook = openpyxl.load_workbook(self.filename)

    def create_file(self):
        self.workbook = openpyxl.workbook.Workbook()

    def add_worksheet(self, worksheet: str):
        self.worksheets[worksheet] = self.workbook.create_sheet(worksheet)

    def rename_worksheet(self, worksheet_old: str, worksheet_new: str):
        ws = self.workbook[worksheet_old]
        ws.title = worksheet_new
        self.worksheets = {w.title: w for w in self.workbook.worksheets}
        self.select_worksheet(worksheet_new)

    def select_worksheet(self, worksheet):
        if not self.worksheets.get(worksheet):
            self.add_worksheet(worksheet)
        self.current_worksheet = self.worksheets.get(worksheet)

    def save_file(self, new_filename: str = None):
        filename = self.filename if not new_filename else new_filename
        self.workbook.save(filename)

    def add_cell_coords(self, coords: str, value: Union[None, str, float, int, date, datetime],
                        style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        cell = self.current_worksheet[coords]
        self.add_cell(cell.row, cell.column, value, style)

    def add_cell(self, row: int, col: int, value: Union[None, str, float, int, date, datetime],
                 style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        try:
            cell = self.current_worksheet.cell(row, col, value)
        except IllegalCharacterError:
            value = ''.join(c for c in value if c in printable)
            cell = self.current_worksheet.cell(row, col, value)
        if style:
            if style.get('border'):
                b = style.get('border')
                left = b.get('left', {})
                right = b.get('right', {})
                bottom = b.get('bottom', {})
                top = b.get('top', {})
                cell.border = Border(left=Side(style=left.get('style', None), color=left.get('color', 'FF000000')),
                                     right=Side(style=right.get('style', None), color=right.get('color', 'FF000000')),
                                     bottom=Side(style=bottom.get('style', None),
                                                 color=bottom.get('color', 'FF000000')),
                                     top=Side(style=top.get('style', None), color=top.get('color', 'FF000000')))
            if style.get('bold'):
                cell.font = Font(bold=True)
            if style.get('italic'):
                cell.font = Font(italic=True)
            if style.get('font'):
                f = style.get('font')
                cell.font = Font(name=f.get('font', 'Calibri'),
                                 size=f.get('size', 11),
                                 bold=f.get('bold', False),
                                 italic=f.get('italic', False),
                                 underline=f.get('underline', 'none'),
                                 strike=f.get('strike', False),
                                 color=f.get('color', '000000'))
            if style.get('alignment'):
                cell.alignment = Alignment(horizontal=style.get('alignment').get('horizontal', 'general'),
                                           vertical=style.get('alignment').get('vertical', 'bottom'),
                                           wrap_text=style.get('alignment').get('wrap_text', None))
            if style.get('number_format'):
                cell.number_format = style.get('number_format', 'general')

    def add_row(self, row: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                start_col: int = 1, styles: Union[None, List, Dict] = None):
        if isinstance(list_cells, dict):
            list_cells_ = list(list_cells.values())
        else:
            list_cells_ = list_cells
        for col, cell in enumerate(list_cells_):
            style = None
            if styles:
                if isinstance(styles, list):
                    style = styles[col]
                elif isinstance(styles, dict):
                    style = styles
            self.add_cell(row, start_col + col, cell, style)
        self.rows_qty += 1

    def add_column(self, col: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                   start_row: int = 1, styles: Union[None, List, Dict] = None):
        if isinstance(list_cells, dict):
            list_cells_ = list(list_cells.values())
        else:
            list_cells_ = list_cells
        for row, cell in enumerate(list_cells_):
            style = None
            if styles:
                if isinstance(styles, list):
                    style = styles[row]
                elif isinstance(styles, dict):
                    style = styles
            self.add_cell(start_row + row, col, cell, style)
            self.rows_qty += 1

    def add_list_rows(self, first_row: int, first_col: int, list_rows: Iterable[Union[Dict, List]],
                      styles: Union[None, List, Dict] = None):
        rows_cnt = len(list_rows)
        if styles:
            styles_cnt = len(styles)
        else:
            styles_cnt = 0
        for row_index, row_ in enumerate(list_rows):
            if isinstance(row_, dict):
                row = list(row_.values())
            else:
                row = row_
            if isinstance(styles, list) and styles_cnt == rows_cnt:
                self.add_row(first_row + row_index, row, first_col, styles[row_index])
            else:
                self.add_row(first_row + row_index, row, first_col, styles)

    def add_merge_cell(self, first_row: int, first_col: int, last_row: int, last_col: int,
                       value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        self.add_cell(first_row, first_col, value, style)
        self.current_worksheet.merge_cells(start_row=first_row, start_column=first_col,
                                           end_row=last_row, end_column=last_col)

    def add_merge_cell_coords(self, first_coords: str, last_coords,
                              value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        first_cell = self.current_worksheet[first_coords]
        last_cell = self.current_worksheet[last_coords]
        self.add_merge_cell(first_cell.row, first_cell.column, last_cell.row, last_cell.column, value, style)

    def set_cell_dimension(self, row: int, col: int, height: Union[int, float] = 15, width: Union[int, float] = 10):
        cell = self.current_worksheet.cell(row, col)
        self.current_worksheet.column_dimensions[cell.column_letter].width = width
        self.current_worksheet.row_dimensions[row].height = height

    def set_rows_height(self, list_rows_height: Iterable[float], start_row: int = 1):
        for row, height in enumerate(list_rows_height):
            # cell = self.current_worksheet.cell(row + 1, 0)
            self.current_worksheet.row_dimensions[start_row + row].height = height

    def set_columns_width(self, list_columns_width: Iterable[float], start_col: int = 1):
        for col, width in enumerate(list_columns_width):
            cell = self.current_worksheet.cell(1, start_col + col)
            self.current_worksheet.column_dimensions[cell.column_letter].width = width

    def create_simple_excel(self, filename, data: Iterable[Union[Dict, List]], column_width: Optional[Iterable] = None,
                            column_names: Optional[Iterable] = None, styles: Union[None, List, Dict] = None,
                            worksheet: str = 'Sheet', rewrite: bool = True, start_row: int = 1):
        header_style = {'bold': True, 'border': BORDER.get('bottom'),
                        'alignment': {'horizontal': 'center', 'wrap_text': True}}

        self.open_file(filename, rewrite)
        if rewrite:
            self.rename_worksheet('Sheet', worksheet)
        self.select_worksheet(worksheet)

        if column_width:
            self.set_columns_width(column_width)

        _start_row = start_row

        if column_names:
            self.add_row(_start_row, column_names, styles=header_style)
            _start_row += 1

        if any(isinstance(e, dict) for e in data):
            self.add_list_rows(_start_row, 1, [list(x.values()) for x in data], styles)
        else:
            self.add_list_rows(_start_row, 1, data, styles)
        self.save_file()

    def get_max_row(self):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        return self.current_worksheet.max_row

    def get_max_column(self):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        return self.current_worksheet.max_colummn

    def read_file(self, excel_header: list, sheet_name: str = None,
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
        excel_arr = []
        columns = excel_header

        ws = self.worksheets.get(sheet_name, self.current_worksheet)
        records = [[x.value for x in row] for row in ws.iter_rows(min_row=start_row)]

        int_columns = int_columns if int_columns else []
        date_columns = date_columns if date_columns else []

        # По каждой строчке
        for r in records:
            # Сцепляем заголовки (английские) из настроек с данными из Excel-файла
            tmp_row = dict(zip(columns, r))

            # Костыль от преобразований данных экселем после ручных правок
            for c in columns:
                # Костыль для столбцов, которые должны быть int
                if tmp_row.get(c) is not None:
                    if c in int_columns and not isinstance(tmp_row.get(c), int):
                        tmp_row[c] = int(tmp_row[c])
                    # Костыль для столбцов, которые должны быть datetime
                    if c in date_columns and not isinstance(tmp_row.get(c), datetime):
                        try:
                            tmp_row[c] = main.parse_date(tmp_row[c])
                        except TypeError:
                            # Костыль, если дата в экселе не текст, а число
                            tmp_row[c] = datetime(1899, 12, 31) + timedelta(days=tmp_row[c] - 1)
                        except ValueError:
                            tmp_row[c] = None
                    # Костыль для столбцов, которые должны быть str
                    if c not in date_columns and c not in int_columns and not isinstance(tmp_row.get(c), str):
                        tmp_row[c] = str(tmp_row.get(c))
            excel_arr.append(tmp_row)

        return excel_arr


def create_excel(list_: List[Union[Dict, List]], fname: str, header: Optional[List] = None,
                 column_width: Optional[List] = None, worksheet: str = 'Sheet', start_row: int = 0, borders=False):
    """
    Создает Excel-файл в формате xlsx
    :param list_: Входной список
    :param fname: Имя файла с путем
    :param column_width: Список с шириной столбцов
    :param worksheet: Имя вкладки
    :param start_row: С какой строки начинать писать
    :return:
    """
    # Если указанной папки не существует - создаем ее (включая родительские, если нужно)
    fd, fn = os.path.split(fname)
    if fd:
        main.check_folder_exist(fd)
    # Создаем книгу
    if fname.split('.')[-1] == 'xlsx':
        create_xlsx(fname, list_, column_width, header, worksheet, start_row, borders)
    else:
        create_xls(fname, list_, column_width, header, worksheet, start_row, borders)


def create_xls(filename: str, data: List[Union[Dict, List]], column_width: Optional[List], header: Optional[List],
               sheet_name: str = 'Sheet', start_row: int = 0, add_borders: bool = False):
    if not filename.endswith('xls'):
        filename += '.xls'
    try:
        wb = xlwt.Workbook()
    except ImportError:
        raise ImportError("xlwt isn't installed. Run: pip install -U xlwt")

    ws = wb.add_sheet(sheet_name)

    common_style = xlwt.XFStyle()
    font = xlwt.Font()
    common_style.alignment.wrap = 1
    if add_borders:
        common_borders = xlwt.Borders()
        common_borders.bottom = xlwt.Borders.THIN
        common_borders.top = xlwt.Borders.THIN
        common_borders.left = xlwt.Borders.THIN
        common_borders.right = xlwt.Borders.THIN

        common_style.borders = common_borders
    common_style.font = font

    if header:
        # Устанавливаем стиль для заголовков
        header_style = xlwt.XFStyle()
        header_style.alignment.wrap = 1
        header_font = xlwt.Font()
        header_font.bold = True

        header_borders = xlwt.Borders()
        if add_borders:
            header_borders.top = xlwt.Borders.THIN
            header_borders.left = xlwt.Borders.THIN
            header_borders.right = xlwt.Borders.THIN
        header_borders.bottom = xlwt.Borders.THIN
        header_style.borders = header_borders
        header_style.font = header_font

        # Заполняем шапку
        for c_index, c_value in enumerate(header):
            ws.write(0, c_index, c_value, header_style)
        start_row += 1

    if column_width:
        # Устанавливаем ширину столбцов
        for c_index, c_width in enumerate(column_width):
            ws.col(c_index).width = 256 * c_width

    # Вставляем значения
    for r_index, row in enumerate(data):
        try:
            columns = list(row.values())
        except AttributeError:
            columns = row
        for c_index, c_value in enumerate(columns):
            # Если значение типа datetime - преобразовываем в строку вида dd.MM.yyyy HH:mm:ss
            if isinstance(c_value, datetime):
                common_style.num_format_str = 'dd.mm.yyyy hh:mm:ss'
            # Если значение типа date - преобразовываем в строку вида dd.MM.yyyy
            elif isinstance(c_value, date):
                common_style.num_format_str = 'dd.mm.yyyy'
            elif isinstance(c_value, int):
                common_style.num_format_str = 'General'
            elif isinstance(c_value, float):
                common_style.num_format_str = 'General'
            else:
                common_style.num_format_str = 'General'

            ws.write(start_row + r_index, c_index, c_value, common_style)
    wb.save(filename)


@deprecated(deprecated_in='2.6.0', current_version=version, details='Use class ExcelX instead')
def create_xlsx(filename: str, data: List[Union[Dict, List]], column_width: Optional[List], header: Optional[List],
                sheet_name: str = 'Sheet', start_row: int = 0, add_borders: bool = False):
    if not filename.endswith('xlsx'):
        filename += '.xlsx'

    try:
        wb = xlsxwriter.Workbook(filename)
    except ImportError:
        raise ImportError("xlsxwriter isn't installed. Run: pip install -U xlsxwriter")

    ws = wb.add_worksheet(sheet_name)

    common_style = wb.add_format({'text_wrap': True})
    if add_borders:
        common_style.set_border()

    if header:
        header_style = wb.add_format({'bold': 1, 'text_wrap': True})
        if add_borders:
            header_style.set_border()
        else:
            header_style.set_bottom()
        # Заполняем шапку
        for c_index, c_value in enumerate(header):
            ws.write(0, c_index, c_value, header_style)
        start_row += 1

    if column_width:
        for c_index, c_width in enumerate(column_width):
            ws.set_column(c_index, c_index, c_width)

    # Вставляем значения
    for r_index, row in enumerate(data):
        try:
            columns = list(row.values())
        except AttributeError:
            columns = row
        for c_index, c_value in enumerate(columns):
            # Если значение типа datetime - преобразовываем в строку вида dd.MM.yyyy HH:mm:ss
            if isinstance(c_value, datetime):
                common_style.num_format = 'dd.mm.yyyy hh:mm:ss'
            # Если значение типа date - преобразовываем в строку вида dd.MM.yyyy
            elif isinstance(c_value, date):
                common_style.num_format = 'dd.mm.yyyy'
            elif isinstance(c_value, int):
                common_style.num_format = 'General'
            elif isinstance(c_value, float):
                common_style.num_format = 'General'
            else:
                common_style.num_format = 'General'

            ws.write(r_index + start_row, c_index, c_value, common_style)
    wb.close()


def read_excel(filename: str, excel_header: Union[list, dict], sheet_name: str = None,
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
    # Считываем Excel-файл
    excel_arr = []
    if isinstance(excel_header, list):
        columns = excel_header
    else:
        columns = [x[1] for x in list(excel_header.values())]

    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active
        records = [[x.value for x in row] for row in ws.iter_rows(min_row=start_row)]
    else:
        try:
            wb = xlrd.open_workbook(filename)
        except ImportError:
            raise ImportError("xlrd isn't installed. Run: pip install -U xlrd")

        ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
        records = [ws.row_values(i) for i in range(start_row - 1, ws.nrows)]

    int_columns = int_columns if int_columns else []
    date_columns = date_columns if date_columns else []

    # По каждой строчке
    for r in records:
        # Сцепляем заголовки (английские) из настроек с данными из Excel-файла
        tmp_row = dict(zip(columns, r))

        # Костыль от преобразований данных экселем после ручных правок
        for c in columns:
            # Костыль для столбцов, которые должны быть int
            if tmp_row.get(c) is not None:
                if c in int_columns and not isinstance(tmp_row.get(c), int):
                    tmp_row[c] = int(tmp_row[c])
                # Костыль для столбцов, которые должны быть datetime
                if c in date_columns and not isinstance(tmp_row.get(c), datetime):
                    try:
                        tmp_row[c] = main.parse_date(tmp_row[c])
                    except TypeError:
                        # Костыль, если дата в экселе не текст, а число
                        tmp_row[c] = datetime(1899, 12, 31) + timedelta(days=tmp_row[c] - 1)
                    except ValueError:
                        tmp_row[c] = None
                # Костыль для столбцов, которые должны быть str
                if c not in date_columns and c not in int_columns and not isinstance(tmp_row.get(c), str):
                    tmp_row[c] = str(tmp_row.get(c))
        excel_arr.append(tmp_row)

    return excel_arr


if __name__ == '__main__':
    excel = ExcelX()
    excel.create_simple_excel('test.xlsx', [{'test': 'test'}], worksheet='AIMS', rewrite=True)
    excel.create_simple_excel('test.xlsx', [{'test': 'test2'}], worksheet='AIMS2', rewrite=False)
